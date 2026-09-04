from flask import Flask, request, jsonify, render_template, send_from_directory, redirect, url_for, make_response, session, flash
from flask_cors import CORS
import requests
import random
import string
import smtplib
import traceback
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import re
from datetime import datetime, date
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from flask import send_file
import base64
from werkzeug.security import generate_password_hash, check_password_hash
from db_config import execute_query, get_db_connection
import mysql.connector
import os
from datetime import datetime, timedelta


from zoneinfo import ZoneInfo  # Python 3.9+ built-in na

PH_TZ = ZoneInfo("Asia/Manila")

def ph_now():
    """Timezone-aware datetime object, Philippine time."""
    return datetime.now(PH_TZ)

def ph_now_iso():
    """Philippine time as ISO-format string."""
    return datetime.now(PH_TZ).isoformat()

def ph_now_str():
    """Philippine time as 'YYYY-MM-DD HH:MM:SS' string."""
    return datetime.now(PH_TZ).strftime("%Y-%m-%d %H:%M:%S")


# ========== ITO LANG ANG IDINAGDAG ==========
from db_config import execute_query
# ===========================================

app = Flask(__name__)


import cloudinary
import cloudinary.uploader
import cloudinary.api

# ============================================================
# CLOUDINARY CONFIGURATION
# ============================================================
cloudinary.config(
    cloud_name=os.getenv('CLOUDINARY_CLOUD_NAME'),
    api_key=os.getenv('CLOUDINARY_API_KEY'),
    api_secret=os.getenv('CLOUDINARY_API_SECRET')
)


# ============================================================
# CLOUDINARY URL HELPER FUNCTION
# ============================================================
def get_cloudinary_url(image_path, resource_type="image"):
    """Convert image path to Cloudinary URL"""
    if not image_path:
        return ''
    
    #####
    # If already a full URL
    if image_path.startswith('http'):
        return image_path
    
    # Determine resource type (image or video)
    is_video = resource_type == "video" or image_path.endswith(('.mp4', '.avi', '.mov', '.mkv'))
    upload_type = "video" if is_video else "image"
    
    # If path starts with 'cablevision/'
    if image_path.startswith('cablevision/'):
        return f"https://res.cloudinary.com/oa3fcr2b/{upload_type}/upload/{image_path}"
    
    # If path still has /shared-uploads/ (legacy)
    if image_path.startswith('/shared-uploads/'):
        cloudinary_path = image_path.replace('/shared-uploads/', 'cablevision/')
        return f"https://res.cloudinary.com/oa3fcr2b/{upload_type}/upload/{cloudinary_path}"
    
    # Default: return as is
    return image_path


# ============================================================
# CLOUDINARY HELPER FUNCTIONS
# ============================================================
def upload_to_cloudinary(file, folder="plans"):
    """Upload file to Cloudinary and return URL"""
    try:
        print(f" Uploading to Cloudinary: {file.filename}")
        
        # I-reset ang file pointer bago mag-upload
        file.stream.seek(0)  # ITO ANG SUSI!
        
        filename_without_ext = file.filename.rsplit('.', 1)[0]
        
        result = cloudinary.uploader.upload(
            file,
            folder=f"cablevision/{folder}",
            resource_type="image",
            public_id=filename_without_ext,
            overwrite=True
        )
        
        print(f" Upload successful: {result['secure_url']}")
        return result['secure_url']
        
    except Exception as e:
        print(f" Cloudinary upload error: {e}")
        import traceback
        traceback.print_exc()
        return None

def delete_from_cloudinary(image_url):
    """Delete file from Cloudinary using URL"""
    if not image_url:
        print(" No image URL to delete")
        return
    
    try:
        if 'cloudinary.com' in image_url:
            # Extract public_id from URL
            # Example: https://res.cloudinary.com/oa3fcr2b/image/upload/cablevision/plans/plan_xxx.png
            # Public ID: cablevision/plans/plan_xxx
            
            # Get everything after /upload/
            parts = image_url.split('/upload/')
            if len(parts) > 1:
                public_id_with_ext = parts[1]
                print(f" Public ID with extension: {public_id_with_ext}")
                
                # Remove version number if present (v1234567890/)
                if '/' in public_id_with_ext and public_id_with_ext.split('/')[0].startswith('v'):
                    public_id_with_ext = '/'.join(public_id_with_ext.split('/')[1:])
                    print(f" After removing version: {public_id_with_ext}")
                
                # Remove file extension (.png, .jpg, etc.)
                public_id = public_id_with_ext.rsplit('.', 1)[0]
                print(f" Final public_id: {public_id}")
                
                # Delete from Cloudinary
                result = cloudinary.uploader.destroy(public_id, resource_type="image")
                print(f" Cloudinary delete result: {result}")
                
                if result.get('result') == 'ok':
                    print(f" Successfully deleted: {public_id}")
                    return True
                else:
                    print(f" Delete result: {result}")
                    return False
                    
    except Exception as e:
        print(f" Cloudinary delete error: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return False




@app.after_request
def add_security_headers(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    return response

app.secret_key = "my_super_secure_random_key_12345"
CORS(app)


def ensure_temp_reset_password_column():
    """Make sure temp_reset can store the pending reset password hash."""
    try:
        columns = execute_query("SHOW COLUMNS FROM temp_reset", fetch=True) or []
        column_names = {col.get("Field") for col in columns if isinstance(col, dict)}

        if "new_password" not in column_names:
            execute_query("ALTER TABLE temp_reset ADD COLUMN new_password VARCHAR(255) NULL")
            print(" Added temp_reset.new_password column")
    except Exception as e:
        print(f" temp_reset schema check failed: {e}")


ensure_temp_reset_password_column()



def ensure_login_lockout_columns():
    table_columns = {
        "admins": {
            "failed_login_attempts": "INT NOT NULL DEFAULT 0",
            "locked_until": "DATETIME NULL",
            "lock_level": "INT NOT NULL DEFAULT 0",
        },
        "technicians": {
            "failed_login_attempts": "INT NOT NULL DEFAULT 0",
            "locked_until": "DATETIME NULL",
            "lock_level": "INT NOT NULL DEFAULT 0",
        },
    }
    for table, columns in table_columns.items():
        try:
            existing = execute_query(f"SHOW COLUMNS FROM {table}", fetch=True) or []
            existing_names = {column.get("Field") for column in existing}
            for column, definition in columns.items():
                if column not in existing_names:
                    execute_query(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")
        except Exception as error:
            print(f"Login lockout schema check failed for {table}: {error}")


ensure_login_lockout_columns()


def get_lockout_response(account):
    locked_until = account.get("locked_until") if account else None
    if locked_until and locked_until > datetime.now():
        minutes = max(1, int((locked_until - datetime.now()).total_seconds() / 60))
        return jsonify({
            "error": f"Your account is locked. Please contact Super Admin. Try again in {minutes} minute(s).",
            "account_locked": True,
        }), 423
    return None


def record_failed_login(table, id_column, account_id):
    account = execute_query(
        f"SELECT failed_login_attempts, lock_level FROM {table} WHERE {id_column} = %s",
        (account_id,), fetch_one=True,
    ) or {}
    attempts = int(account.get("failed_login_attempts") or 0) + 1
    lock_level = int(account.get("lock_level") or 0)
    if attempts >= 5:
        lock_hours = 24 if lock_level >= 1 else 1
        execute_query(
            f"UPDATE {table} SET failed_login_attempts = 0, lock_level = %s, locked_until = %s WHERE {id_column} = %s",
            (lock_level + 1, datetime.now() + timedelta(hours=lock_hours), account_id),
        )
        return lock_hours
    execute_query(
        f"UPDATE {table} SET failed_login_attempts = %s WHERE {id_column} = %s",
        (attempts, account_id),
    )
    return None


def reset_login_lockout(table, id_column, account_id):
    execute_query(
        f"UPDATE {table} SET failed_login_attempts = 0, lock_level = 0, locked_until = NULL WHERE {id_column} = %s",
        (account_id,),
    )

    

# ===============================
# PASSWORD HASHING HELPERS
# ===============================
def hash_password(plain_password):
    """Hash a plaintext password for storing in DB"""
    return generate_password_hash(plain_password)


def verify_password(stored_password, provided_password):
    """
    Verify a password against what's stored in DB.
    Supports both:
    - New hashed passwords (werkzeug format e.g. 'pbkdf2:sha256:...' or 'scrypt:...')
    - Old plaintext passwords already in your DB (backward compatible)
    """
    if not stored_password:
        return False

    if stored_password.startswith(('pbkdf2:', 'scrypt:')):
        try:
            return check_password_hash(stored_password, provided_password)
        except Exception:
            return False
    else:
        # Legacy plaintext password in DB - direct compare
        return stored_password == provided_password


def upgrade_password_if_needed(table, id_column, id_value, stored_password, provided_password):
    """
    If the stored password is still plaintext (legacy), replace it with a hash
    right after a successful login. Gradually migrates every account to
    hashed passwords without forcing a mass password reset.
    """
    if stored_password and not stored_password.startswith(('pbkdf2:', 'scrypt:')):
        try:
            new_hash = hash_password(provided_password)
            update_query = f"UPDATE {table} SET password = %s WHERE {id_column} = %s"
            execute_query(update_query, (new_hash, id_value))
            print(f" Upgraded {table}.{id_column}={id_value} password to hashed format")
        except Exception as e:
            print(f" Password upgrade failed for {table}.{id_column}={id_value}: {e}")

# ========== WALA NA ITO (TINANGGAL ANG FIREBASE) ==========
# cred = credentials.Certificate("cablevision-27a81-firebase-adminsdk-fbsvc-23f87c44c6.json")
# firebase_admin.initialize_app(cred, {"databaseURL": "https://cablevision-27a81-default-rtdb.firebaseio.com/"})
# =========================================================

SUPER_ADMIN_USERNAME = "super admin"
SUPER_ADMIN_PASSWORD = "cablevisioninternet"
FIREBASE_WEB_API_KEY = "YOUR_FIREBASE_WEB_API_KEY"

# ========== ITO ANG PUMALIT SA FIREBASE SUPERADMIN ==========
query = "SELECT * FROM superadmins WHERE username = %s"
existing = execute_query(query, (SUPER_ADMIN_USERNAME,), fetch_one=True)

if not existing:
    insert_query = """
        INSERT INTO superadmins (username, password, area, email, name) 
        VALUES (%s, %s, %s, %s, %s)
    """
    execute_query(insert_query, (SUPER_ADMIN_USERNAME, SUPER_ADMIN_PASSWORD, "Main Office", "superadmin@cablevision.com", "System Administrator"))


# =========================================================
# LOGIN HISTORY HELPER FUNCTIONS
# =========================================================
def ensure_login_history_table():
    try:
        execute_query("""
            CREATE TABLE IF NOT EXISTS login_history (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id VARCHAR(255) NOT NULL,
                user_type VARCHAR(50) NOT NULL DEFAULT 'admin',
                session_token VARCHAR(255) NOT NULL,
                device_info VARCHAR(255) NULL,
                device_brand VARCHAR(100) NULL,
                browser VARCHAR(100) NULL,
                os VARCHAR(100) NULL,
                ip_address VARCHAR(100) NULL,
                location VARCHAR(255) NULL,
                login_time DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_active DATETIME DEFAULT CURRENT_TIMESTAMP,
                status VARCHAR(50) DEFAULT 'Active',
                INDEX idx_user_id (user_id),
                INDEX idx_session_token (session_token)
            )
        """)
        cols = execute_query("SHOW COLUMNS FROM login_history LIKE 'user_type'", fetch=True)
        if not cols:
            execute_query("ALTER TABLE login_history ADD COLUMN user_type VARCHAR(50) NOT NULL DEFAULT 'admin' AFTER user_id")
        
        # Check if device_brand column exists
        cols = execute_query("SHOW COLUMNS FROM login_history LIKE 'device_brand'", fetch=True)
        if not cols:
            execute_query("ALTER TABLE login_history ADD COLUMN device_brand VARCHAR(100) NULL AFTER device_info")
    except Exception as e:
        print(f"Could not ensure login_history table: {e}")


def parse_user_agent(ua_string):
    if not ua_string:
        return "Web Browser", "Windows", "Desktop"
    ua_lower = ua_string.lower()

    if "windows nt 10.0" in ua_lower:
        os_name = "Windows 10/11"
    elif "windows" in ua_lower:
        os_name = "Windows"
    elif "android" in ua_lower:
        os_name = "Android"
    elif "iphone" in ua_lower or "ipad" in ua_lower:
        os_name = "iOS"
    elif "mac os x" in ua_lower or "macintosh" in ua_lower:
        os_name = "macOS"
    elif "linux" in ua_lower:
        os_name = "Linux"
    else:
        os_name = "Desktop/Mobile"

    if "edg/" in ua_lower or "edge/" in ua_lower:
        browser = "Microsoft Edge"
    elif "chrome/" in ua_lower and "edg/" not in ua_lower:
        browser = "Google Chrome"
    elif "firefox/" in ua_lower:
        browser = "Mozilla Firefox"
    elif "safari/" in ua_lower and "chrome/" not in ua_lower:
        browser = "Apple Safari"
    elif "opera/" in ua_lower or "opr/" in ua_lower:
        browser = "Opera"
    else:
        browser = "Web Browser"

    # Extract device brand/type from user agent
    device_brand = "Desktop"
    if "ipad" in ua_lower:
        device_brand = "iPad"
    elif "iphone" in ua_lower:
        device_brand = "iPhone"
    elif "ipod" in ua_lower:
        device_brand = "iPod"
    elif "android" in ua_lower:
        # Check for popular Android device brands
        if "samsung" in ua_lower or "sm-" in ua_lower:
            device_brand = "Samsung"
        elif "xiaomi" in ua_lower or "mi " in ua_lower:
            device_brand = "Xiaomi"
        elif "realme" in ua_lower:
            device_brand = "Realme"
        elif "oneplus" in ua_lower:
            device_brand = "OnePlus"
        elif "oppo" in ua_lower:
            device_brand = "OPPO"
        elif "vivo" in ua_lower:
            device_brand = "Vivo"
        elif "huawei" in ua_lower or "honor" in ua_lower:
            device_brand = "Huawei"
        elif "motorola" in ua_lower or "moto" in ua_lower:
            device_brand = "Motorola"
        elif "nokia" in ua_lower:
            device_brand = "Nokia"
        elif "asus" in ua_lower:
            device_brand = "ASUS"
        else:
            device_brand = "Android"
    elif "mac" in ua_lower or "macintosh" in ua_lower:
        device_brand = "Mac"
    elif "windows" in ua_lower:
        device_brand = "Windows PC"
    elif "linux" in ua_lower:
        device_brand = "Linux"

    return browser, os_name, device_brand


def get_client_ip():
    """Prefer the actual client IP from reverse-proxy headers when available."""
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        for ip in forwarded.split(","):
            candidate = ip.strip()
            if candidate and candidate.lower() != "unknown":
                return candidate

    for header_name in ("X-Real-IP", "True-Client-IP", "CF-Connecting-IP"):
        value = request.headers.get(header_name, "")
        if value:
            candidate = value.split(",")[0].strip()
            if candidate and candidate.lower() != "unknown":
                return candidate

    return request.remote_addr or "127.0.0.1"


def reverse_geocode_coords(lat, lng):
    """Use browser coordinates to resolve a human-readable city/area name."""
    try:
        lat = float(lat)
        lng = float(lng)
    except (TypeError, ValueError):
        return "Unknown Location"

    try:
        url = "https://nominatim.openstreetmap.org/reverse"
        response = requests.get(
            url,
            params={
                "lat": lat,
                "lon": lng,
                "format": "jsonv2",
                "addressdetails": 1,
            },
            timeout=6,
            headers={"User-Agent": "Cablevision-App/1.0"},
        )
        if response.status_code != 200:
            return "Unknown Location"

        data = response.json() or {}
        address = data.get("address") or {}
        city = (address.get("city") or address.get("town") or address.get("village") or address.get("municipality") or "").strip()
        district = (address.get("county") or address.get("state_district") or address.get("province") or "").strip()
        region = (address.get("state") or address.get("region") or "").strip()
        country = (address.get("country") or "").strip()

        if city and region and country:
            return f"{city}, {region}, {country}"
        if city and country:
            return f"{city}, {country}"
        if city and district and country:
            return f"{city}, {district}, {country}"
        if region and country:
            return f"{region}, {country}"
        if country:
            return country

        return "Unknown Location"
    except Exception:
        return "Unknown Location"


def resolve_device_location(ip_address=None, lat=None, lng=None):
    """Return the device's most accurate available location: browser coords first, then public IP."""
    if lat is not None and lng is not None:
        try:
            return reverse_geocode_coords(lat, lng)
        except Exception:
            pass

    if not ip_address:
        return "Unknown Location"

    ip_address = ip_address.strip()
    if ip_address in ("127.0.0.1", "::1", "localhost"):
        return "Local Network"

    private_prefixes = ("10.", "192.168.", "172.")
    if ip_address.startswith(private_prefixes):
        return "Local Network"

    try:
        response = requests.get(f"https://ipapi.co/{ip_address}/json/", timeout=5)
        if response.status_code != 200:
            return "Unknown Location"

        data = response.json()
        if not isinstance(data, dict):
            return "Unknown Location"

        city = (data.get("city") or "").strip()
        region = (data.get("region") or "").strip()
        country = (data.get("country_name") or data.get("country") or "").strip()

        if city and region and country:
            return f"{city}, {region}, {country}"
        if city and country:
            return f"{city}, {country}"
        if region and country:
            return f"{region}, {country}"
        if country:
            return country

        return "Unknown Location"
    except Exception:
        return "Unknown Location"


def record_login_history(user_id, user_type='admin', tab_id=None, lat=None, lng=None):
    try:
        ensure_login_history_table()
        ua_string = request.headers.get("User-Agent", "")
        ip_addr = get_client_ip()
        browser, os_name, device_brand = parse_user_agent(ua_string)
        device_info = f"{browser} on {os_name}"
        location = resolve_device_location(ip_addr, lat, lng)

        session_token = tab_id if tab_id else f"sess_{user_id}_{int(time.time())}"
        now_str = ph_now_str()

        existing = execute_query(
            "SELECT id FROM login_history WHERE user_id = %s AND user_type = %s AND session_token = %s LIMIT 1",
            (user_id, user_type, session_token),
            fetch_one=True
        )

        if existing:
            execute_query(
                "UPDATE login_history SET last_active = %s, status = 'Active', ip_address = %s, location = %s, device_info = %s, device_brand = %s WHERE id = %s",
                (now_str, ip_addr, location, device_info, device_brand, existing["id"])
            )
        else:
            execute_query(
                """INSERT INTO login_history 
                   (user_id, user_type, session_token, device_info, device_brand, browser, os, ip_address, location, login_time, last_active, status) 
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Active')""",
                (user_id, user_type, session_token, device_info, device_brand, browser, os_name, ip_addr, location, now_str, now_str)
            )
        return session_token
    except Exception as e:
        print(f"[LOGIN HISTORY ERROR] {e}")
        return None


def is_login_history_active(user_id, user_type, session_token):
    if not user_id or not user_type or not session_token:
        return False

    ensure_login_history_table()
    existing = execute_query(
        "SELECT id FROM login_history WHERE user_id = %s AND user_type = %s AND session_token = %s LIMIT 1",
        (user_id, user_type, session_token),
        fetch_one=True
    )
    return bool(existing)


def get_admin_session_user(req):
    tab_id = req.args.get("tab_id")
    if not tab_id and req.is_json:
        data = req.get_json(silent=True) or {}
        tab_id = data.get("tab_id")
    if not tab_id:
        tab_id = session.get("active_tab")

    session_data = session.get(f"admin_{tab_id}") if tab_id else None
    if session_data:
        user_id = session_data.get("user_id") or session_data.get("admin_id") or session_data.get("technician_id") or session_data.get("username")
        user_type = session_data.get("user_type", "admin")
        user_name = session_data.get("user_name") or session_data.get("admin_username") or session_data.get("technician_name") or "User"

        if is_login_history_active(user_id, user_type, tab_id):
            return user_id, user_type, user_name, tab_id

        # Invalidate stale tab session when the login history row was removed
        session.pop(f"admin_{tab_id}", None)
        if session.get("active_tab") == tab_id:
            session.pop("active_tab", None)
        return None, None, None, tab_id

    # Fallback to direct session values if no tab-specific object exists
    user_id = session.get("user_id") or session.get("admin_id") or session.get("technician_id") or session.get("username")
    user_type = session.get("user_type", "admin")
    user_name = session.get("user_name") or session.get("admin_username") or session.get("technician_name") or "User"

    if user_id and tab_id and is_login_history_active(user_id, user_type, tab_id):
        return user_id, user_type, user_name, tab_id

    if not tab_id and session.get("active_tab"):
        alt_tab_id = session.get("active_tab")
        if is_login_history_active(user_id, user_type, alt_tab_id):
            return user_id, user_type, user_name, alt_tab_id

    return None, None, None, tab_id


# ===============================
# Serve Login Page
# ===============================
@app.route("/")
def index():
    """Render the login page"""
    return render_template("login.html")

@app.route("/login")
def login_page():
    """Alias route for the login page to avoid frontend 404s."""
    return render_template("login.html")

@app.route("/api/login", methods=["POST"])
def login():
    data = request.json or {}
    identifier = data.get("username") or data.get("identifier") or data.get("email")
    password = data.get("password")
    ga_code = str(data.get("ga_code", "") or "").strip()
    tab_id = data.get("tab_id", "")
    pending_key = f"pending_login_{tab_id}" if tab_id else "pending_login"

    if not identifier or not password:
        return jsonify({"error": "Missing credentials"}), 400

    try:
        if ga_code:
            pending_login = session.get(pending_key)
            if not pending_login:
                return jsonify({"error": "Google Authenticator verification expired. Please login again.", "requires_2fa": True}), 401

            pending_user_type = pending_login.get("user_type")
            pending_identifier = pending_login.get("identifier")
            pending_password = pending_login.get("password")
            pending_user_id = pending_login.get("user_id")

            if pending_user_type == "superadmin":
                user_row = execute_query(
                    "SELECT * FROM superadmins WHERE username = %s OR email = %s OR username = %s",
                    (pending_identifier, pending_identifier, pending_user_id),
                    fetch_one=True,
                )
                secret = user_row.get("ga_secret") if user_row else None
                if not user_row or not verify_password(user_row.get('password'), pending_password):
                    return jsonify({"error": "Invalid username/ID or password"}), 401
                if not verify_ga_code(secret, ga_code):
                    return jsonify({"error": "Invalid Google Authenticator code.", "requires_2fa": True}), 401

                if tab_id:
                    session[f"admin_{tab_id}"] = {
                        'user_id': user_row.get('username'),
                        'user_type': 'superadmin',
                        'user_name': user_row.get('name', 'Super Admin'),
                        'user_area': user_row.get('area', 'All Areas'),
                        'username': user_row.get('username')
                    }
                    session["active_tab"] = tab_id

                session['user_id'] = user_row.get('username')
                session['user_type'] = 'superadmin'
                session['user_name'] = user_row.get('name', 'Super Admin')
                session['user_area'] = user_row.get('area', 'All Areas')
                session['username'] = user_row.get('username')

                record_login_history(user_row.get('username'), 'superadmin', tab_id, lat=data.get('lat'), lng=data.get('lng'))
                session.pop(pending_key, None)
                return jsonify({
                    "success": True,
                    "type": "superadmin",
                    "username": user_row.get('username'),
                    "name": user_row.get('name', 'Super Admin'),
                    "area": user_row.get('area', 'All Areas'),
                    "email": user_row.get('email', ''),
                    "redirect": "/superadmin",
                    "tab_id": tab_id
                })

            if pending_user_type == "admin":
                user_row = execute_query(
                    "SELECT * FROM admins WHERE admin_id = %s OR username = %s OR email = %s",
                    (pending_user_id, pending_identifier, pending_identifier),
                    fetch_one=True,
                )
                secret = user_row.get("ga_secret") if user_row else None
                if not user_row or not verify_password(user_row.get('password'), pending_password):
                    return jsonify({"error": "Invalid username/ID or password"}), 401
                if not verify_ga_code(secret, ga_code):
                    return jsonify({"error": "Invalid Google Authenticator code.", "requires_2fa": True}), 401

                if user_row.get('status') == "Deactivated":
                    return jsonify({"error": "Account is deactivated. Contact Super Admin."}), 403

                if tab_id:
                    session[f"admin_{tab_id}"] = {
                        'user_id': user_row.get('admin_id'),
                        'user_type': 'admin',
                        'user_name': user_row.get('username'),
                        'user_area': user_row.get('area', ''),
                        'admin_username': user_row.get('username'),
                        'admin_id': user_row.get('admin_id')
                    }
                    session["active_tab"] = tab_id

                session['user_id'] = user_row.get('admin_id')
                session['user_type'] = 'admin'
                session['user_name'] = user_row.get('username')
                session['user_area'] = user_row.get('area', '')
                session['admin_username'] = user_row.get('username')
                session['admin_id'] = user_row.get('admin_id')

                notification_id = int(datetime.now().timestamp() * 1000)
                admin_name = user_row.get('username', 'Unknown Admin')
                admin_area = user_row.get('area', 'Unknown Area')
                login_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                notif_query = """
                    INSERT INTO notifications (id, title, message, type, relatedId, timestamp, read_status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
                execute_query(notif_query, (
                    notification_id,
                    "Admin Login",
                    f"{admin_name} ({admin_area}) logged in at {login_time}",
                    "admin_login",
                    user_row.get('admin_id'),
                    datetime.now().isoformat(),
                    0
                ))

                record_login_history(user_row.get('admin_id'), 'admin', tab_id, lat=data.get('lat'), lng=data.get('lng'))
                session.pop(pending_key, None)
                return jsonify({
                    "success": True,
                    "type": "admin",
                    "username": user_row.get('username'),
                    "name": user_row.get('username'),
                    "area": user_row.get('area', ''),
                    "admin_id": user_row.get('admin_id'),
                    "redirect": "/admin",
                    "tab_id": tab_id
                })

            if pending_user_type == "technician":
                user_row = execute_query(
                    "SELECT * FROM technicians WHERE technician_id = %s OR email = %s",
                    (pending_user_id, pending_identifier),
                    fetch_one=True,
                )
                secret = user_row.get("ga_secret") if user_row else None
                if not user_row or not verify_password(user_row.get('password'), pending_password):
                    return jsonify({"error": "Invalid username/ID or password"}), 401
                if not verify_ga_code(secret, ga_code):
                    return jsonify({"error": "Invalid Google Authenticator code.", "requires_2fa": True}), 401

                if user_row.get('status') == "Deactivated":
                    return jsonify({"error": "Account is deactivated. Contact Super Admin."}), 403

                if tab_id:
                    session[f"admin_{tab_id}"] = {
                        'user_id': user_row.get('technician_id'),
                        'user_type': 'technician',
                        'user_name': user_row.get('name'),
                        'user_area': user_row.get('area', ''),
                        'technician_id': user_row.get('technician_id'),
                        'technician_name': user_row.get('name')
                    }
                    session["active_tab"] = tab_id

                session['user_id'] = user_row.get('technician_id')
                session['user_type'] = 'technician'
                session['user_name'] = user_row.get('name')
                session['user_area'] = user_row.get('area', '')
                session['technician_id'] = user_row.get('technician_id')
                session['technician_name'] = user_row.get('name')

                record_login_history(user_row.get('technician_id'), 'technician', tab_id, lat=data.get('lat'), lng=data.get('lng'))
                session.pop(pending_key, None)
                return jsonify({
                    "success": True,
                    "type": "technician",
                    "technician_id": user_row.get('technician_id'),
                    "name": user_row.get('name'),
                    "area": user_row.get('area', ''),
                    "email": user_row.get('email', ''),
                    "redirect": "/technician/dashboard",
                    "tab_id": tab_id
                })

        # ========== 1. CHECK SUPER ADMIN ==========
        query = "SELECT * FROM superadmins WHERE username = %s OR email = %s"
        superadmin = execute_query(query, (identifier, identifier), fetch_one=True)
        if superadmin and verify_password(superadmin.get('password'), password):
            upgrade_password_if_needed(
                "superadmins", "username", superadmin.get('username'),
                superadmin.get('password'), password
            )
            if bool(superadmin.get('ga_enabled')):
                session[pending_key] = {
                    'user_id': superadmin.get('username'),
                    'user_type': 'superadmin',
                    'identifier': identifier,
                    'password': password,
                }
                return jsonify({"requires_2fa": True, "user_id": superadmin.get('username'), "tab_id": tab_id, "error": "Enter the 6-digit code from Google Authenticator to continue."})

            if tab_id:
                session[f"admin_{tab_id}"] = {
                    'user_id': superadmin.get('username'),
                    'user_type': 'superadmin',
                    'user_name': superadmin.get('name', 'Super Admin'),
                    'user_area': superadmin.get('area', 'All Areas'),
                    'username': superadmin.get('username')
                }
                session["active_tab"] = tab_id
            else:
                session['user_id'] = superadmin.get('username')
                session['user_type'] = 'superadmin'
                session['user_name'] = superadmin.get('name', 'Super Admin')
                session['user_area'] = superadmin.get('area', 'All Areas')

            record_login_history(superadmin.get('username'), 'superadmin', tab_id, lat=data.get('lat'), lng=data.get('lng'))
            return jsonify({
                "success": True,
                "type": "superadmin",
                "username": superadmin.get('username'),
                "name": superadmin.get('name', 'Super Admin'),
                "area": superadmin.get('area', 'All Areas'),
                "email": superadmin.get('email', ''),
                "redirect": "/superadmin",
                "tab_id": tab_id
            })

        # ========== 2. CHECK REGULAR ADMIN ==========
        query = """
            SELECT * FROM admins
            WHERE username = %s OR email = %s OR admin_id = %s
        """
        admin = execute_query(query, (identifier, identifier, identifier), fetch_one=True)
        if admin:
            lockout_response = get_lockout_response(admin)
            if lockout_response:
                return lockout_response
            if not verify_password(admin.get('password'), password):
                lock_hours = record_failed_login("admins", "admin_id", admin.get("admin_id"))
                if lock_hours:
                    return jsonify({"error": f"Your account is locked for {lock_hours} hour(s). Please contact Super Admin.", "account_locked": True}), 423
                return jsonify({"error": "Invalid username/ID or password"}), 401
        if admin and verify_password(admin.get('password'), password):
            reset_login_lockout("admins", "admin_id", admin.get("admin_id"))
            upgrade_password_if_needed(
                "admins", "admin_id", admin.get('admin_id'),
                admin.get('password'), password
            )
            if admin.get('status') == "Deactivated":
                return jsonify({"error": "Account is deactivated. Contact Super Admin."}), 403
            if bool(admin.get('ga_enabled')):
                session[pending_key] = {
                    'user_id': admin.get('admin_id'),
                    'user_type': 'admin',
                    'identifier': identifier,
                    'password': password,
                }
                return jsonify({"requires_2fa": True, "user_id": admin.get('admin_id'), "tab_id": tab_id, "error": "Enter the 6-digit code from Google Authenticator to continue."})

            if tab_id:
                session[f"admin_{tab_id}"] = {
                    'user_id': admin.get('admin_id'),
                    'user_type': 'admin',
                    'user_name': admin.get('username'),
                    'user_area': admin.get('area', ''),
                    'admin_username': admin.get('username'),
                    'admin_id': admin.get('admin_id')
                }
                session["active_tab"] = tab_id
            else:
                session['user_id'] = admin.get('admin_id')
                session['user_type'] = 'admin'
                session['user_name'] = admin.get('username')
                session['user_area'] = admin.get('area', '')
                session['admin_username'] = admin.get('username')
                session['admin_id'] = admin.get('admin_id')

            notification_id = int(datetime.now().timestamp() * 1000)
            admin_name = admin.get('username', 'Unknown Admin')
            admin_area = admin.get('area', 'Unknown Area')
            login_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            notif_query = """
                INSERT INTO notifications (id, title, message, type, relatedId, timestamp, read_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            execute_query(notif_query, (
                notification_id,
                "Admin Login",
                f"{admin_name} ({admin_area}) logged in at {login_time}",
                "admin_login",
                admin.get('admin_id'),
                datetime.now().isoformat(),
                0
            ))

            record_login_history(admin.get('admin_id'), 'admin', tab_id, lat=data.get('lat'), lng=data.get('lng'))
            return jsonify({
                "success": True,
                "type": "admin",
                "username": admin.get('username'),
                "name": admin.get('username'),
                "area": admin.get('area', ''),
                "admin_id": admin.get('admin_id'),
                "redirect": "/admin",
                "tab_id": tab_id
            })

        # ========== 3. CHECK TECHNICIAN ==========
        query = """
            SELECT * FROM technicians
            WHERE technician_id = %s OR email = %s
            LIMIT 1
        """
        technician = execute_query(query, (identifier, identifier), fetch_one=True)
        if technician:
            lockout_response = get_lockout_response(technician)
            if lockout_response:
                return lockout_response
            if not verify_password(technician.get('password'), password):
                lock_hours = record_failed_login("technicians", "technician_id", technician.get("technician_id"))
                if lock_hours:
                    return jsonify({"error": f"Your account is locked for {lock_hours} hour(s). Please contact Super Admin.", "account_locked": True}), 423
                return jsonify({"error": "Invalid username/ID or password"}), 401
        if technician and verify_password(technician.get('password'), password):
            reset_login_lockout("technicians", "technician_id", technician.get("technician_id"))
            upgrade_password_if_needed(
                "technicians", "technician_id", technician.get('technician_id'),
                technician.get('password'), password
            )
            if technician.get('status') == "Deactivated":
                return jsonify({"error": "Account is deactivated. Contact Super Admin."}), 403
            if bool(technician.get('ga_enabled')):
                session[pending_key] = {
                    'user_id': technician.get('technician_id'),
                    'user_type': 'technician',
                    'identifier': identifier,
                    'password': password,
                }
                return jsonify({"requires_2fa": True, "user_id": technician.get('technician_id'), "tab_id": tab_id, "error": "Enter the 6-digit code from Google Authenticator to continue."})

            if tab_id:
                session[f"admin_{tab_id}"] = {
                    'user_id': technician.get('technician_id'),
                    'user_type': 'technician',
                    'user_name': technician.get('name'),
                    'user_area': technician.get('area', ''),
                    'technician_id': technician.get('technician_id'),
                    'technician_name': technician.get('name')
                }
                session["active_tab"] = tab_id
                session['technician_id'] = technician.get('technician_id')
                session['user_type'] = 'technician'
            else:
                session['user_id'] = technician.get('technician_id')
                session['user_type'] = 'technician'
                session['user_name'] = technician.get('name')
                session['user_area'] = technician.get('area', '')
                session['technician_id'] = technician.get('technician_id')
                session['technician_name'] = technician.get('name')

            record_login_history(technician.get('technician_id'), 'technician', tab_id, lat=data.get('lat'), lng=data.get('lng'))
            return jsonify({
                "success": True,
                "type": "technician",
                "technician_id": technician.get('technician_id'),
                "name": technician.get('name'),
                "area": technician.get('area', ''),
                "email": technician.get('email', ''),
                "redirect": "/technician/dashboard",
                "tab_id": tab_id
            })

        return jsonify({"error": "Invalid username/ID or password"}), 401

    except Exception as e:
        print("Login error:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/verify-session", methods=["GET"])
def verify_admin_session():
    """Verify if the session is valid using tab_id"""
    tab_id = request.args.get("tab_id")
    if not tab_id:
        return jsonify({"valid": False, "error": "No tab_id provided"}), 401

    user_session = session.get(f"admin_{tab_id}")
    if user_session:
        user_id = user_session.get("user_id") or user_session.get("admin_id") or user_session.get("technician_id") or user_session.get("username")
        user_type = user_session.get("user_type", "admin")
        if is_login_history_active(user_id, user_type, tab_id):
            return jsonify({"valid": True, "user": user_session.get("user_name") or user_session.get("admin_username") or user_session.get("username")})

    return jsonify({"valid": False, "error": "Invalid session"}), 401


@app.route("/api/superadmin/verify-session", methods=["GET"])
def verify_superadmin_session():
    """Verify superadmin session using tab_id."""
    tab_id = request.args.get("tab_id")
    if not tab_id:
        return jsonify({"valid": False, "error": "No tab_id provided"}), 401

    user_session = session.get(f"admin_{tab_id}")
    if user_session and user_session.get("user_type") == "superadmin":
        user_id = user_session.get("user_id") or user_session.get("username")
        if is_login_history_active(user_id, "superadmin", tab_id):
            return jsonify({"valid": True, "user": user_session.get("user_name") or user_session.get("user_id")})

    return jsonify({"valid": False, "error": "Invalid session"}), 401


@app.route("/api/technician/verify-session", methods=["GET"])
def verify_technician_session():
    """Verify technician session using tab_id."""
    tab_id = request.args.get("tab_id")
    if not tab_id:
        return jsonify({"valid": False, "error": "No tab_id provided"}), 401

    user_session = session.get(f"admin_{tab_id}")
    if user_session and user_session.get("user_type") == "technician":
        user_id = user_session.get("user_id") or user_session.get("technician_id")
        if is_login_history_active(user_id, "technician", tab_id):
            return jsonify({"valid": True, "user": user_session.get("technician_name") or user_session.get("user_name")})

    return jsonify({"valid": False, "error": "Invalid session"}), 401


# ===============================
# FORGOT PASSWORD - WITH TECHNICIAN SUPPORT
# ===============================
@app.route("/api/admin/forgot-password", methods=["POST"])
def forgot_password():
    data = request.json
    identifier = data.get("identifier")
    
    print(f" Searching for identifier: {identifier}")
    
    if not identifier:
        return jsonify({"error": "Identifier required"}), 400

    # ========== CHECK SUPERADMIN ==========
    query = "SELECT * FROM superadmins WHERE email = %s OR username = %s"
    superadmin = execute_query(query, (identifier, identifier), fetch_one=True)
    
    if superadmin:
        print(f" Found SUPERADMIN: {superadmin.get('username')}")
        email = superadmin.get('email')
        username = superadmin.get('username')
        user_type = "superadmin"
        area = superadmin.get('area', 'All Areas')
        name = superadmin.get('name', 'Super Admin')
        
        # I-check kung may email
        if not email:
            return jsonify({"error": "No email address found for this account"}), 400
        
        otp_code = str(random.randint(100000, 999999))
        expiry = datetime.now().timestamp() + 300
        
        insert_query = """
            INSERT INTO temp_reset (email, otp, expiry, user_type, area, username)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        execute_query(insert_query, (email, otp_code, expiry, user_type, area, username))
        
        if send_universal_otp_email(email, name, otp_code, user_type):
            return jsonify({
                "message": "OTP sent successfully", 
                "username": username,
                "type": user_type
            })
        else:
            return jsonify({"error": "Failed to send OTP email"}), 500

    # ========== CHECK REGULAR ADMIN ==========
    query = """
        SELECT * FROM admins 
        WHERE email = %s OR username = %s OR admin_id = %s
    """
    admin = execute_query(query, (identifier, identifier, identifier), fetch_one=True)
    
    if admin:
        print(f" Found ADMIN: {admin.get('username')}")
        username = admin.get('username')
        email = admin.get('email')
        user_type = "admin"
        area = admin.get('area', '')
        name = admin.get('username')
        
        # I-check kung may email
        if not email:
            return jsonify({"error": "No email address found for this account"}), 400
        
        otp_code = str(random.randint(100000, 999999))
        expiry = datetime.now().timestamp() + 300
        
        insert_query = """
            INSERT INTO temp_reset (email, otp, expiry, user_type, area, username)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        execute_query(insert_query, (email, otp_code, expiry, user_type, area, username))
        
        if send_universal_otp_email(email, name, otp_code, user_type):
            return jsonify({
                "message": "OTP sent successfully", 
                "username": username,
                "type": user_type
            })
        else:
            return jsonify({"error": "Failed to send OTP email"}), 500

    # ========== CHECK TECHNICIAN ==========
    query = """
        SELECT * FROM technicians 
        WHERE technician_id = %s OR email = %s
        LIMIT 1
    """
    technician = execute_query(query, (identifier, identifier), fetch_one=True)
    
    if technician:
        print(f" Found TECHNICIAN: {technician.get('technician_id')}")
        username = technician.get('technician_id')
        email = technician.get('email')
        user_type = "technician"
        area = technician.get('area', '')
        name = technician.get('name')
        
        # I-check kung may email
        if not email:
            return jsonify({"error": "No email address found for this account"}), 400
        
        otp_code = str(random.randint(100000, 999999))
        expiry = datetime.now().timestamp() + 300
        
        insert_query = """
            INSERT INTO temp_reset (email, otp, expiry, user_type, area, username)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        execute_query(insert_query, (email, otp_code, expiry, user_type, area, username))
        
        if send_universal_otp_email(email, name, otp_code, user_type):
            return jsonify({
                "message": "OTP sent successfully", 
                "username": username,
                "type": user_type
            })
        else:
            return jsonify({"error": "Failed to send OTP email"}), 500

    # Kung walang nahanap sa lahat
    print(f" No user found with identifier: {identifier}")
    return jsonify({"error": "Email not found"}), 404

# ===============================
# RESET PASSWORD - WITH TECHNICIAN SUPPORT
# ===============================
@app.route("/api/admin/reset-password", methods=["POST"])
def reset_password():

    data = request.get_json(silent=True) or {}

    username = data.get("username")
    identifier = data.get("identifier")
    code = data.get("code")
    new_password = data.get("new_password")

    print("======= RESET PASSWORD DEBUG =======")
    print(f"Username: {username}")
    print(f"Identifier: {identifier}")
    print(f"Code: {code}")
    print("====================================")

    # ===============================
    # VALIDATION
    # ===============================
    if not code or not new_password:
        return jsonify({
            "error": "All fields are required"
        }), 400

    if not username and not identifier:
        return jsonify({
            "error": "Account identifier is missing"
        }), 400

    if len(new_password) < 8:
        return jsonify({
            "error": "Password must be at least 8 characters"
        }), 400

    # ===============================
    # FIND OTP
    # ===============================
    temp_query = """
        SELECT *
        FROM temp_reset
        WHERE (username = %s OR email = %s)
        AND otp = %s
        ORDER BY expiry DESC
        LIMIT 1
    """

    temp_data = execute_query(
        temp_query,
        (
            username,
            identifier,
            code
        ),
        fetch_one=True
    )

    if not temp_data:
        return jsonify({
            "error": "Invalid verification code"
        }), 400

    print(f" Found temp_data: {temp_data}")

    # ===============================
    # CHECK OTP EXPIRATION
    # ===============================
    current_time = datetime.now().timestamp()
    expiry_time = temp_data.get("expiry", 0)

    if current_time > expiry_time:

        execute_query(
            "DELETE FROM temp_reset WHERE id = %s",
            (temp_data.get("id"),)
        )

        return jsonify({
            "error": "Verification code expired"
        }), 400

    # ===============================
    # GET RESET INFORMATION
    # ===============================
    user_type = temp_data.get("user_type")
    area = temp_data.get("area", "")
    actual_username = temp_data.get("username")

    if not actual_username:
        return jsonify({
            "error": "Account information not found"
        }), 400

    print(f" User Type: {user_type}")
    print(f" Username/ID: {actual_username}")

    # ===============================
    # HASH NEW PASSWORD
    # ===============================
    hashed_new_password = hash_password(new_password)

    print(
        f" Hashed password: "
        f"{hashed_new_password[:50]}..."
    )

    # ===============================
    # STEP 1
    # SAVE NEW PASSWORD TO temp_reset
    # ===============================
    update_temp_query = """
        UPDATE temp_reset
        SET new_password = %s
        WHERE id = %s
    """

    update_temp_result = execute_query(
        update_temp_query,
        (
            hashed_new_password,
            temp_data.get("id")
        )
    )

    print(
        f" Update temp_reset result: "
        f"{update_temp_result}"
    )

    # ===============================
    # VERIFY temp_reset
    # ===============================
    verify_temp = execute_query(
        """
        SELECT new_password
        FROM temp_reset
        WHERE id = %s
        """,
        (temp_data.get("id"),),
        fetch_one=True
    )

    print(
        f" Verified new_password: "
        f"{'SAVED' if verify_temp and verify_temp.get('new_password') else 'NULL'}"
    )

    # ===============================
    # STEP 2
    # UPDATE ACTUAL USER PASSWORD
    # ===============================
    update_rows = 0

    if user_type == "superadmin":

        update_query = """
            UPDATE superadmins
            SET password = %s
            WHERE username = %s
        """

        update_rows = execute_query(
            update_query,
            (
                hashed_new_password,
                actual_username
            )
        )

        print(
            f" Superadmin password updated: "
            f"{update_rows} rows affected"
        )

    elif user_type == "admin":

        update_query = """
            UPDATE admins
            SET password = %s
            WHERE username = %s
        """

        update_rows = execute_query(
            update_query,
            (
                hashed_new_password,
                actual_username
            )
        )

        print(
            f" Admin password updated: "
            f"{update_rows} rows affected"
        )

    elif user_type == "technician":

        update_query = """
            UPDATE technicians
            SET password = %s
            WHERE technician_id = %s
        """

        update_rows = execute_query(
            update_query,
            (
                hashed_new_password,
                actual_username
            )
        )

        print(
            f" Technician password updated: "
            f"{update_rows} rows affected"
        )

    else:

        return jsonify({
            "error": "Invalid user type"
        }), 400

    # ===============================
    # CHECK PASSWORD UPDATE
    # ===============================
    if update_rows is None or update_rows == 0:

        return jsonify({
            "error": "Failed to update password. Please try again."
        }), 400

    # ===============================
    # AUTO-LOGIN AFTER RESET
    # ===============================

    tab_id = (
        data.get("tab_id", "")
        or session.get("active_tab", "")
    )

    # Generate tab_id if none exists
    if not tab_id:

        import time

        tab_id = (
            f"tab_{int(time.time() * 1000)}_"
            f"{random.randint(1000, 9999)}"
        )

    print(
        f" Auto-login after reset for "
        f"{actual_username} "
        f"(type: {user_type})"
    )

    user_data = None
    user_id = actual_username
    redirect_url = ""

    # ===============================
    # SUPERADMIN AUTO-LOGIN
    # ===============================
    if user_type == "superadmin":

        user_query = """
            SELECT *
            FROM superadmins
            WHERE username = %s
            LIMIT 1
        """

        user_data = execute_query(
            user_query,
            (actual_username,),
            fetch_one=True
        )

        if not user_data:
            return jsonify({
                "error": "Superadmin account not found"
            }), 404

        user_id = user_data.get("username")

        # TAB SESSION
        session[f"admin_{tab_id}"] = {
            "user_id": user_data.get("username"),
            "user_type": "superadmin",
            "user_name": user_data.get(
                "name",
                "Super Admin"
            ),
            "user_area": user_data.get(
                "area",
                "All Areas"
            ),
            "username": user_data.get("username")
        }

        # NORMAL SESSION
        session["active_tab"] = tab_id

        session["user_id"] = user_data.get("username")
        session["user_type"] = "superadmin"
        session["user_name"] = user_data.get(
            "name",
            "Super Admin"
        )
        session["user_area"] = user_data.get(
            "area",
            "All Areas"
        )
        session["username"] = user_data.get("username")

        redirect_url = (
            f"/superadmin?tab_id={tab_id}"
        )

        # IMPORTANT:
        # Create active login history just like
        # normal login.
        record_login_history(
            user_id,
            "superadmin",
            tab_id,
            lat=data.get("lat"),
            lng=data.get("lng")
        )

        print(
            f" Superadmin session created "
            f"for tab: {tab_id}"
        )

    # ===============================
    # ADMIN AUTO-LOGIN
    # ===============================
    elif user_type == "admin":

        user_query = """
            SELECT *
            FROM admins
            WHERE username = %s
            LIMIT 1
        """

        user_data = execute_query(
            user_query,
            (actual_username,),
            fetch_one=True
        )

        if not user_data:
            return jsonify({
                "error": "Admin account not found"
            }), 404

        user_id = user_data.get("admin_id")

        # TAB SESSION
        session[f"admin_{tab_id}"] = {
            "user_id": user_data.get("admin_id"),
            "user_type": "admin",
            "user_name": user_data.get(
                "username"
            ),
            "user_area": user_data.get(
                "area",
                ""
            ),
            "admin_username": user_data.get(
                "username"
            ),
            "admin_id": user_data.get(
                "admin_id"
            )
        }

        # NORMAL SESSION
        session["active_tab"] = tab_id

        session["user_id"] = user_data.get(
            "admin_id"
        )
        session["user_type"] = "admin"
        session["user_name"] = user_data.get(
            "username"
        )
        session["user_area"] = user_data.get(
            "area",
            ""
        )
        session["admin_username"] = user_data.get(
            "username"
        )
        session["admin_id"] = user_data.get(
            "admin_id"
        )
        session["username"] = user_data.get(
            "username"
        )

        redirect_url = (
            f"/admin?tab_id={tab_id}"
        )

        # IMPORTANT:
        # Create active login history just like
        # normal login.
        record_login_history(
            user_id,
            "admin",
            tab_id,
            lat=data.get("lat"),
            lng=data.get("lng")
        )

        print(
            f" Admin session created "
            f"for tab: {tab_id}"
        )

    # ===============================
    # TECHNICIAN AUTO-LOGIN
    # ===============================
    elif user_type == "technician":

        user_query = """
            SELECT *
            FROM technicians
            WHERE technician_id = %s
            LIMIT 1
        """

        user_data = execute_query(
            user_query,
            (actual_username,),
            fetch_one=True
        )

        if not user_data:
            return jsonify({
                "error": "Technician account not found"
            }), 404

        user_id = user_data.get(
            "technician_id"
        )

        # TAB SESSION
        session[f"admin_{tab_id}"] = {
            "user_id": user_data.get(
                "technician_id"
            ),
            "user_type": "technician",
            "user_name": user_data.get(
                "name"
            ),
            "user_area": user_data.get(
                "area",
                ""
            ),
            "technician_id": user_data.get(
                "technician_id"
            ),
            "technician_name": user_data.get(
                "name"
            )
        }

        # NORMAL SESSION
        session["active_tab"] = tab_id

        session["user_id"] = user_data.get(
            "technician_id"
        )
        session["user_type"] = "technician"
        session["user_name"] = user_data.get(
            "name"
        )
        session["user_area"] = user_data.get(
            "area",
            ""
        )
        session["technician_id"] = user_data.get(
            "technician_id"
        )
        session["technician_name"] = user_data.get(
            "name"
        )
        session["username"] = user_data.get(
            "technician_id"
        )

        redirect_url = (
            f"/technician/dashboard?tab_id={tab_id}"
        )

        # IMPORTANT:
        # Create active login history just like
        # normal login.
        record_login_history(
            user_id,
            "technician",
            tab_id,
            lat=data.get("lat"),
            lng=data.get("lng")
        )

        print(
            f" Technician session created "
            f"for tab: {tab_id}"
        )

    # ===============================
    # UNKNOWN USER TYPE
    # ===============================
    else:

        return jsonify({
            "error": "Unable to determine account type"
        }), 400

    # ===============================
    # FORCE SESSION SAVE
    # ===============================
    session.modified = True

    # ===============================
    # DEBUG SESSION
    # ===============================
    print("=========================================")
    print(" PASSWORD RESET SUCCESSFUL")
    print(" AUTO-LOGIN SESSION CREATED")
    print(f" User ID: {user_id}")
    print(f" Username: {actual_username}")
    print(f" User Type: {user_type}")
    print(f" Tab ID: {tab_id}")
    print(f" Active Tab: {session.get('active_tab')}")
    print(
        f" Tab Session Exists: "
        f"{bool(session.get(f'admin_{tab_id}'))}"
    )
    print(
        f" Session User Type: "
        f"{session.get('user_type')}"
    )
    print("=========================================")

    # ===============================
    # RETURN SUCCESS
    # ===============================
    return jsonify({
        "success": True,
        "message": "Password updated successfully",
        "username": actual_username,
        "type": user_type,
        "area": area,
        "redirect": redirect_url,
        "tab_id": tab_id,
        "user_id": user_id,
        "name": (
            user_data.get("name")
            if user_data
            else actual_username
        ),
        "admin_id": (
            user_data.get("admin_id")
            if user_data
            else None
        ),
        "technician_id": (
            user_data.get("technician_id")
            if user_data
            else None
        )
    }), 200



# ===============================
# SEND UNIVERSAL OTP EMAIL
# ADMIN / SUPERADMIN / TECHNICIAN
# BREVO API
# ===============================

def send_universal_otp_email(to_email, name, otp_code, user_type):
    """
    Sends password reset OTP email using Brevo HTTP API.
    Supports:
        - Super Admin
        - Admin
        - Technician

    Returns:
        True  = email sent successfully
        False = email sending failed
    """

    import requests

    # ===============================
    # BREVO CONFIGURATION
    # ===============================

    brevo_api_key = os.getenv("BREVO_API_KEY")

    sender_email = os.getenv(
        "SMTP_FROM",
        "noreply@cablevisioncableinternet.com"
    )

    sender_name = "Cablevision Systems Corporation"

    # ===============================
    # CHECK BREVO API KEY
    # ===============================

    if not brevo_api_key:
        print(" BREVO_API_KEY is not configured!")
        return False

    # ===============================
    # USER TYPE DISPLAY
    # ===============================

    type_display = {
        "superadmin": "Super Admin",
        "admin": "Admin",
        "technician": "Technician"
    }.get(user_type, "User")

    # ===============================
    # EMAIL SUBJECT
    # ===============================

    subject = f"CableVision {type_display} Password Reset OTP"

    # ===============================
    # HTML EMAIL
    # ===============================

    html_body = f"""
    <!DOCTYPE html>
    <html>

    <head>
        <meta charset="UTF-8">
        <meta name="viewport"
              content="width=device-width, initial-scale=1.0">

        <title>
            CableVision {type_display} Password Reset
        </title>
    </head>

    <body style="
        font-family: Arial, sans-serif;
        background-color: #f0f4f8;
        padding: 20px;
        margin: 0;
    ">

        <div style="
            background-color: #ffffff;
            padding: 30px;
            border-radius: 12px;
            max-width: 500px;
            margin: auto;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        ">

            <h2 style="
                color: #003d73;
                margin-top: 0;
            ">
                CableVision {type_display} Password Reset
            </h2>

            <p>
                Hello <strong>{name}</strong>,
            </p>

            <p>
                You requested to reset your CableVision
                {type_display} account password.
            </p>

            <p>
                Please use the One-Time Password (OTP) below:
            </p>

            <div style="
                text-align: center;
                margin: 30px 0;
            ">

                <div style="
                    display: inline-block;
                    font-size: 32px;
                    font-weight: bold;
                    letter-spacing: 6px;
                    color: #001f3f;
                    background-color: #f0f7ff;
                    padding: 18px 25px;
                    border-radius: 8px;
                    font-family: monospace;
                ">
                    {otp_code}
                </div>

            </div>

            <p style="
                color: #555;
                font-size: 14px;
            ">
                This verification code will expire in
                <strong>5 minutes</strong>.
            </p>

            <p style="
                color: #555;
                font-size: 14px;
            ">
                If you did not request a password reset,
                please ignore this email.
            </p>

            <hr style="
                margin: 30px 0;
                border: none;
                border-top: 1px solid #eeeeee;
            ">

            <p style="
                font-size: 12px;
                color: #666;
                text-align: center;
                margin-bottom: 0;
            ">
                Cablevision Systems Corporation<br>
                Sta. Cruz, Laguna, Philippines
            </p>

        </div>

    </body>

    </html>
    """

    # ===============================
    # PLAIN TEXT EMAIL
    # ===============================

    plain_body = (
        f"CableVision {type_display} Password Reset\n\n"
        f"Hello {name},\n\n"
        f"Your OTP code is: {otp_code}\n\n"
        f"This code expires in 5 minutes.\n\n"
        f"If you did not request a password reset, "
        f"please ignore this email."
    )

    # ===============================
    # BREVO API PAYLOAD
    # ===============================

    payload = {
        "sender": {
            "name": sender_name,
            "email": sender_email
        },

        "to": [
            {
                "email": to_email
            }
        ],

        "subject": subject,

        "htmlContent": html_body,

        "textContent": plain_body
    }

    # ===============================
    # BREVO API HEADERS
    # ===============================

    headers = {
        "accept": "application/json",
        "api-key": brevo_api_key,
        "content-type": "application/json"
    }

    # ===============================
    # SEND EMAIL
    # ===============================

    try:

        print(
            f" Sending {type_display} password reset OTP "
            f"via Brevo to {to_email}..."
        )

        response = requests.post(
            "https://api.brevo.com/v3/smtp/email",
            headers=headers,
            json=payload,
            timeout=30
        )

        # ===============================
        # CHECK BREVO RESPONSE
        # ===============================

        if response.status_code not in (200, 201):

            print(
                f" Brevo API error "
                f"({response.status_code}): "
                f"{response.text}"
            )

            return False

        # ===============================
        # GET BREVO MESSAGE ID
        # ===============================

        try:

            brevo_response = response.json()

            message_id = brevo_response.get("messageId")

            if message_id:

                print(
                    f" Brevo Message ID: {message_id}"
                )

        except Exception:
            pass

        print(
            f" {type_display} password reset OTP "
            f"sent successfully to {to_email}"
        )

        return True

    # ===============================
    # ERROR HANDLING
    # ===============================

    except requests.exceptions.Timeout:

        print(
            " Brevo API request timed out"
        )

        return False

    except requests.exceptions.RequestException as e:

        print(
            f" Brevo API request error: {e}"
        )

        return False

    except Exception as e:

        print(
            f" Error sending password reset OTP: {e}"
        )

        import traceback
        traceback.print_exc()

        return False



# ===============================
# LOGOUT - UNIVERSAL
# ===============================
@app.route("/api/logout", methods=["POST"])
def api_logout():
    data = request.json or {}
    tab_id = data.get("tab_id", "")

    if tab_id:
        session.pop(f"admin_{tab_id}", None)
        if session.get("active_tab") == tab_id:
            session.pop("active_tab", None)
    else:
        for key in ["user_id", "user_type", "user_name", "user_area",
                    "admin_username", "admin_id",
                    "technician_id", "technician_name"]:
            session.pop(key, None)

    return jsonify({"success": True, "redirect": "/"})


# ===============================
# Serve Super Admin Dashboard
# ===============================
@app.route("/superadmin")
def superadmin_dashboard():
    return render_template("superadmin-dashboard.html")

# ===============================
# CREATE NOTIFICATION - SUPERADMIN (CONVERTED TO MYSQL)
# ===============================
@app.route("/api/superadmin/notifications", methods=["POST"])
def create_notification():
    """Create a new notification (called from admin side)"""
    try:
        data = request.json
        import time
        notification_id = int(time.time() * 1000)
        timestamp = ph_now_iso()
        
        query = """
            INSERT INTO notifications (id, title, message, type, relatedId, timestamp, read_status)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        params = (
            notification_id,
            data.get("title"),
            data.get("message"),
            data.get("type"),
            data.get("relatedId"),
            data.get("timestamp", timestamp),
            0  # read_status = 0 (unread)
        )
        
        execute_query(query, params)
        return jsonify({"message": "Notification created", "id": notification_id}), 201
        
    except Exception as e:
        print(f"Error creating notification: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/superadmin/notifications", methods=["GET"])
def get_notifications():
    """Get all notifications for superadmin"""
    try:
        query = """
            SELECT id, title, message, type, relatedId, timestamp, read_status
            FROM notifications 
            ORDER BY timestamp DESC
        """
        # Use fetch=True instead of fetch_all=True
        notifications_list = execute_query(query, fetch=True)
        
        if not notifications_list:
            return jsonify([])
        
        formatted = []
        for n in notifications_list:
            formatted.append({
                "id": n["id"],
                "title": n["title"],
                "message": n["message"],
                "type": n["type"] if n["type"] else "info",
                "relatedId": n["relatedId"],
                "timestamp": n["timestamp"],
                "read": n["read_status"] == 1
            })
        
        return jsonify(formatted)
        
    except Exception as e:
        print(f"Error: {e}")
        return jsonify([])




# ===============================
# MARK NOTIFICATION AS READ - SUPERADMIN (CONVERTED TO MYSQL)
# ===============================
@app.route("/api/superadmin/notifications/<int:notification_id>/read", methods=["PUT"])
def mark_notification_read(notification_id):
    """Mark a notification as read"""
    try:
        query = "UPDATE notifications SET read_status = 1 WHERE id = %s"
        execute_query(query, (notification_id,))
        return jsonify({"message": "Notification marked as read"})
        
    except Exception as e:
        print(f"Error marking notification as read: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# MARK ALL NOTIFICATIONS AS READ - SUPERADMIN (CONVERTED TO MYSQL)
# ===============================
@app.route("/api/superadmin/notifications/read-all", methods=["PUT"])
def mark_all_notifications_read():
    """Mark all notifications as read"""
    try:
        query = "UPDATE notifications SET read_status = 1"
        execute_query(query)
        return jsonify({"message": "All notifications marked as read"})
        
    except Exception as e:
        print(f"Error marking all notifications as read: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# DELETE NOTIFICATION - SUPERADMIN (OPTIONAL)
# ===============================
@app.route("/api/superadmin/notifications/<int:notification_id>", methods=["DELETE"])
def delete_notification(notification_id):
    """Delete a notification"""
    try:
        query = "DELETE FROM notifications WHERE id = %s"
        execute_query(query, (notification_id,))
        return jsonify({"message": "Notification deleted"})
        
    except Exception as e:
        print(f"Error deleting notification: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# GET UNREAD NOTIFICATION COUNT ONLY - SUPERADMIN
# ===============================
@app.route("/api/superadmin/notifications/unread/count", methods=["GET"])
def get_unread_notification_count():
    """Get only the unread notification count"""
    try:
        query = "SELECT COUNT(*) as unread_count FROM notifications WHERE read_status = 0"
        result = execute_query(query, fetch_one=True)
        unread_count = result['unread_count'] if result else 0
        return jsonify({"unread_count": unread_count})
        
    except Exception as e:
        print(f"Error getting unread count: {e}")
        return jsonify({"unread_count": 0})
    



# ===============================
# My Admins Page
# ===============================
@app.route("/superadmin/admins")
def superadmin_admins():
    return render_template("superadmin-admins.html")

# ===============================
# HELPER FUNCTION: Generate Next Admin ID
# ===============================
def generate_next_admin_id():
    """Generate the next available admin ID (ACV-XXXX format)"""
    try:
        # Get all admin IDs from MySQL
        query = "SELECT admin_id FROM admins WHERE admin_id LIKE 'ACV-%'"
        results = execute_query(query, fetch=True) or []
        
        # Extract numbers from existing IDs
        existing_numbers = set()
        for row in results:
            admin_id = row.get('admin_id', '')
            if admin_id.startswith("ACV-"):
                try:
                    num = int(admin_id.split("-")[1])
                    existing_numbers.add(num)
                except:
                    pass
        
        # Find the smallest missing number starting from 1
        next_number = 1
        while next_number in existing_numbers:
            next_number += 1
        
        # Generate admin ID with zero-padded 4 digits
        admin_id = f"ACV-{str(next_number).zfill(4)}"
        print(f" Generated new admin ID: {admin_id}")
        return admin_id
        
    except Exception as e:
        print(f"Error generating admin ID: {e}")
        return f"ACV-{int(datetime.now().timestamp())}"

# ===============================
# CREATE ADMIN (Super Admin Only) - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/admins", methods=["POST"])
def create_admin():
    data = request.json
    username = data.get("username")
    email = data.get("email")
    area = data.get("area")

    print(f" Received - Username: {username}, Email: {email}, Area: {area}")

    if not username or not email or not area:
        return jsonify({"error": "All fields are required"}), 400

    # USERNAME VALIDATION
    username_pattern = re.compile(r"^[a-zA-Z0-9_-]{4,20}$")
    if not username_pattern.match(username):
        return jsonify({"error": "Invalid username. Use 4-20 characters (letters, numbers, _, -)"}), 400

    # EMAIL VALIDATION
    email_pattern = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")
    if not email_pattern.match(email):
        return jsonify({"error": "Invalid email address"}), 400

    try:
        # ========== CHECK USERNAME IN ADMINS TABLE ==========
        check_username_query = "SELECT username FROM admins WHERE username = %s"
        username_exists = execute_query(check_username_query, (username,), fetch_one=True)
        
        if username_exists:
            return jsonify({"error": "Username already exists"}), 400

        # ========== CHECK DUPLICATE AREA ==========
        area_exists_query = "SELECT admin_id FROM admins WHERE area = %s LIMIT 1"
        area_exists = execute_query(area_exists_query, (area,), fetch_one=True)
        if area_exists:
            return jsonify({"error": "This area already has an administrator assigned. Delete the existing admin first before creating another one for the same area."}), 400

        # ========== CHECK EMAIL IN ALL TABLES ==========
        check_all_query = """
            SELECT 
                (SELECT COUNT(*) FROM technicians WHERE email = %s) as tech_count,
                (SELECT COUNT(*) FROM admins WHERE email = %s) as admin_count,
                (SELECT COUNT(*) FROM superadmins WHERE email = %s) as superadmin_count
        """
        result = execute_query(check_all_query, (email, email, email), fetch_one=True)
        
        tech_exists = result.get('tech_count', 0) > 0 if result else False
        admin_exists = result.get('admin_count', 0) > 0 if result else False
        superadmin_exists = result.get('superadmin_count', 0) > 0 if result else False

        # GENERIC MESSAGE - EMAIL LANG ANG NAKALAGAY
        if tech_exists or admin_exists or superadmin_exists:
            return jsonify({"error": f"Email '{email}' already exists"}), 400

        # Generate next available Admin ID
        admin_id = generate_next_admin_id()
        
        # ✅ GUMAWA NG RANDOM PASSWORD (8 CHARACTERS - letters and numbers)
        default_password = generate_secure_password(8)
        created_at = ph_now_iso()

        # Insert into MySQL
        insert_query = """
            INSERT INTO admins 
            (admin_id, username, email, password, area, status, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        
        params = (
            admin_id,
            username,
            email,
            hash_password(default_password),
            area,
            "Active",
            created_at
        )
        
        execute_query(insert_query, params)
        
        print(f" Admin saved successfully: {admin_id}")
        print(f" Password: {default_password}")
        
        # Send email (with plain password, not hashed)
        try:
            send_admin_email(email, admin_id, username, default_password)
        except Exception as e:
            print(f" Email error but admin was created: {e}")

        return jsonify({
            "message": "Admin created successfully",
            "admin_id": admin_id,
            "password": default_password  # ✅ I-RETURN ANG PASSWORD
        }), 201

    except Exception as e:
        print(f" Error creating admin: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ===============================
# SEND ADMIN ACCOUNT EMAIL
# BREVO API
# ===============================

def send_admin_email(to_email, admin_id, username, password):
    """
    Sends newly created Admin account credentials via Brevo.

    Includes:
        - Admin ID
        - Username
        - Temporary Password

    Returns:
        True  = email sent successfully
        False = email sending failed
    """

    import requests
    import html

    # ===============================
    # BREVO CONFIGURATION
    # ===============================

    brevo_api_key = os.getenv("BREVO_API_KEY")
    sender_email = os.getenv(
        "SMTP_FROM",
        "noreply@cablevisioncableinternet.com"
    )
    sender_name = "Cablevision Systems Corporation"

    # ===============================
    # CHECK CONFIGURATION
    # ===============================

    if not brevo_api_key:
        print(" BREVO_API_KEY is not configured!")
        return False

    if not to_email:
        print(" Cannot send admin email: recipient email is empty.")
        return False

    # ===============================
    # HTML ESCAPE
    # ===============================

    safe_admin_id = html.escape(str(admin_id or "N/A"))
    safe_username = html.escape(str(username or "N/A"))
    safe_password = html.escape(str(password or "N/A"))

    # ===============================
    # EMAIL SUBJECT
    # ===============================

    subject = "Your Admin Account - CableVision"

    # ===============================
    # HTML EMAIL
    # ===============================

    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>CableVision Admin Account</title>
    </head>
    <body style="
        margin:0;
        padding:0;
        background:#eef2f6;
        font-family:Arial, Helvetica, sans-serif;
    ">
        <div style="
            width:100%;
            padding:30px 0;
        ">
            <div style="
                max-width:580px;
                margin:0 auto;
                background:#ffffff;
                border-radius:24px;
                overflow:hidden;
                box-shadow:0 10px 30px rgba(0,0,0,0.10);
            ">
                <!-- HEADER -->
                <div style="
                    background:#001f3f;
                    padding:30px 25px;
                    text-align:center;
                ">
                    <div style="
                        font-size:28px;
                        font-weight:700;
                        color:#ffffff;
                    ">
                        Cablevision
                    </div>
                    <div style="
                        margin-top:6px;
                        font-size:13px;
                        color:#93c5fd;
                    ">
                        Internet Service Provider
                    </div>
                    <div style="
                        margin-top:16px;
                        display:inline-block;
                        background:rgba(255,255,255,0.12);
                        color:#dbeafe;
                        padding:6px 14px;
                        border-radius:20px;
                        font-size:11px;
                        font-weight:700;
                    ">
                        ADMIN ACCOUNT
                    </div>
                </div>

                <!-- CONTENT -->
                <div style="padding:28px;">
                    <h2 style="
                        margin:0 0 8px 0;
                        font-size:22px;
                        color:#0f172a;
                    ">
                        Welcome to Cablevision!
                    </h2>
                    <p style="
                        margin:0 0 20px 0;
                        font-size:15px;
                        line-height:1.6;
                        color:#475569;
                    ">
                        Your administrator account has been successfully created.
                    </p>

                    <!-- ACCOUNT DETAILS -->
                    <div style="
                        background:#f8fafc;
                        border-radius:16px;
                        padding:20px;
                        border:1px solid #e2e8f0;
                    ">
                        <div style="
                            font-size:11px;
                            font-weight:700;
                            color:#64748b;
                            margin-bottom:5px;
                        ">
                            ADMIN ID
                        </div>
                        <div style="
                            font-size:18px;
                            font-weight:700;
                            color:#0f172a;
                            margin-bottom:16px;
                        ">
                            {safe_admin_id}
                        </div>

                        <div style="
                            border-top:1px solid #e2e8f0;
                            padding-top:14px;
                            margin-bottom:16px;
                        ">
                            <div style="
                                font-size:11px;
                                font-weight:700;
                                color:#64748b;
                                margin-bottom:5px;
                            ">
                                USERNAME
                            </div>
                            <div style="
                                font-size:17px;
                                font-weight:700;
                                color:#0f172a;
                            ">
                                {safe_username}
                            </div>
                        </div>

                        <div style="
                            border-top:1px solid #e2e8f0;
                            padding-top:14px;
                        ">
                            <div style="
                                font-size:11px;
                                font-weight:700;
                                color:#64748b;
                                margin-bottom:5px;
                            ">
                                TEMPORARY PASSWORD
                            </div>
                            <div style="
                                display:inline-block;
                                background:#eff6ff;
                                color:#001f3f;
                                padding:10px 14px;
                                border-radius:8px;
                                font-family:monospace;
                                font-size:18px;
                                font-weight:700;
                                letter-spacing:1px;
                            ">
                                {safe_password}
                            </div>
                        </div>
                    </div>

                    <!-- SECURITY NOTICE -->
                    <div style="
                        margin-top:20px;
                        padding:16px;
                        background:#fff7ed;
                        border:1px solid #fed7aa;
                        border-radius:14px;
                    ">
                        <div style="
                            font-size:14px;
                            font-weight:700;
                            color:#9a3412;
                            margin-bottom:6px;
                        ">
                            Security Notice
                        </div>
                        <div style="
                            font-size:13px;
                            line-height:1.6;
                            color:#7c2d12;
                        ">
                            This is your temporary password.
                            Please change your password after your
                            first login and do not share your
                            account credentials with anyone.
                        </div>
                    </div>

                    <!-- FOOTER MESSAGE -->
                    <div style="
                        margin-top:28px;
                        padding-top:20px;
                        border-top:1px solid #e2e8f0;
                        text-align:center;
                    ">
                        <div style="
                            font-size:13px;
                            color:#64748b;
                            line-height:1.5;
                        ">
                            Welcome to the Cablevision team!
                        </div>
                    </div>
                </div>

                <!-- FOOTER -->
                <div style="
                    background:#f1f5f9;
                    padding:16px 20px;
                    text-align:center;
                ">
                    <div style="
                        font-size:11px;
                        color:#64748b;
                    ">
                        2026 Cablevision Systems Corporation. All rights reserved.
                    </div>
                    <div style="
                        margin-top:4px;
                        font-size:11px;
                        color:#94a3b8;
                    ">
                        Sta. Cruz, Laguna, Philippines
                    </div>
                </div>
            </div>
        </div>
    </body>
    </html>
    """

    # ===============================
    # PLAIN TEXT EMAIL
    # ===============================

    plain_body = f"""
Cablevision Systems Corporation

ADMIN ACCOUNT CREATED

Welcome to Cablevision!

Your administrator account has been successfully created.

Admin ID:
{admin_id}

Username:
{username}

Temporary Password:
{password}

Please change your password after your first login.

Do not share your account credentials with anyone.

Welcome to the Cablevision team!

Cablevision Systems Corporation
Sta. Cruz, Laguna, Philippines
"""

    # ===============================
    # BREVO API PAYLOAD
    # ===============================

    payload = {
        "sender": {
            "name": sender_name,
            "email": sender_email
        },
        "to": [
            {
                "email": to_email
            }
        ],
        "subject": subject,
        "htmlContent": html_body,
        "textContent": plain_body
    }

    # ===============================
    # BREVO API HEADERS
    # ===============================

    headers = {
        "accept": "application/json",
        "api-key": brevo_api_key,
        "content-type": "application/json"
    }

    # ===============================
    # SEND EMAIL
    # ===============================

    try:
        print("========================================")
        print(" ADMIN ACCOUNT EMAIL")
        print("========================================")
        print(f" From: {sender_email}")
        print(f" To: {to_email}")
        print(f" Admin ID: {admin_id}")
        print(f" Username: {username}")
        print("========================================")

        response = requests.post(
            "https://api.brevo.com/v3/smtp/email",
            headers=headers,
            json=payload,
            timeout=30
        )

        # ===============================
        # CHECK BREVO RESPONSE
        # ===============================

        if response.status_code not in (200, 201):
            print(f" Brevo API error ({response.status_code}): {response.text}")
            return False

        # ===============================
        # GET BREVO MESSAGE ID
        # ===============================

        try:
            brevo_response = response.json()
            message_id = brevo_response.get("messageId")
            if message_id:
                print(f" Brevo Message ID: {message_id}")
        except Exception:
            pass

        print(f" Admin account email sent successfully to {to_email}")
        return True

    # ===============================
    # ERROR HANDLING
    # ===============================

    except requests.exceptions.Timeout:
        print(" Brevo API request timed out.")
        return False

    except requests.exceptions.RequestException as e:
        print(f" Brevo API request error: {e}")
        return False

    except Exception as e:
        print(f" Error sending admin account email: {e}")
        import traceback
        traceback.print_exc()
        return False

# ===============================
# LIST ALL ADMINS - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/admins", methods=["GET"])
def list_admins():
    try:
        query = """
                 SELECT admin_id, username, email, area, status, created_at,
                     failed_login_attempts, locked_until, lock_level,
                     CASE WHEN locked_until > NOW() THEN 1 ELSE 0 END AS login_locked
            FROM admins 
            ORDER BY admin_id
        """
        admins_list = execute_query(query, fetch=True) or []
        
        print(f"📋 Listing {len(admins_list)} admins")
        return jsonify(admins_list)
        
    except Exception as e:
        print(f"Error listing admins: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# GET SINGLE ADMIN BY ID - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/admins/<admin_id>", methods=["GET"])
def get_admin(admin_id):
    try:
        query = """
            SELECT admin_id, username, email, area, status, created_at
            FROM admins 
            WHERE admin_id = %s
        """
        admin_data = execute_query(query, (admin_id,), fetch_one=True)
        
        if not admin_data:
            return jsonify({"error": "Admin not found"}), 404
        
        # Don't return password for security
        if 'password' in admin_data:
            del admin_data['password']
        
        return jsonify(admin_data)
        
    except Exception as e:
        print(f"Error getting admin: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# UPDATE ADMIN (PUT) - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/admins/<admin_id>", methods=["PUT"])
def update_admin(admin_id):
    try:
        data = request.json
        username = data.get("username")
        email = data.get("email")
        area = data.get("area")
        
        # Check if admin exists
        check_query = "SELECT admin_id FROM admins WHERE admin_id = %s"
        exists = execute_query(check_query, (admin_id,), fetch_one=True)
        
        if not exists:
            return jsonify({"error": "Admin not found"}), 404
        
        # Check for duplicate username/email (excluding current admin)
        duplicate_query = """
            SELECT username, email, area FROM admins 
            WHERE (username = %s OR email = %s OR area = %s) AND admin_id != %s
        """
        duplicate = execute_query(duplicate_query, (username, email, area, admin_id), fetch_one=True)
        
        if duplicate:
            if duplicate.get('username') == username:
                return jsonify({"error": "Username already exists"}), 400
            if duplicate.get('email') == email:
                return jsonify({"error": "Email already exists"}), 400
            if duplicate.get('area') == area:
                return jsonify({"error": "This area already has an administrator assigned. Choose a different area."}), 400
        
        # Update admin
        update_query = """
            UPDATE admins 
            SET username = %s, email = %s, area = %s
            WHERE admin_id = %s
        """
        execute_query(update_query, (username, email, area, admin_id))
        
        return jsonify({"message": "Admin updated successfully"})
        
    except Exception as e:
        print(f"Error updating admin: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# DELETE ADMIN BY ID - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/admins/<admin_id>", methods=["DELETE"])
def delete_admin(admin_id):
    try:
        print(f" Attempting to delete admin with ID: {admin_id}")
        
        # Get admin info first (for logging)
        get_query = "SELECT username, email FROM admins WHERE admin_id = %s"
        admin_data = execute_query(get_query, (admin_id,), fetch_one=True)
        
        if not admin_data:
            print(f" Admin not found with ID: {admin_id}")
            return jsonify({"error": "Admin not found"}), 404
        
        username = admin_data.get("username")
        print(f" Found admin: {username} (ID: {admin_id})")
        
        # Delete from MySQL
        delete_query = "DELETE FROM admins WHERE admin_id = %s"
        execute_query(delete_query, (admin_id,))
        
        print(f" Admin '{username}' (ID: {admin_id}) deleted successfully")
        return jsonify({"message": f"Admin '{username}' deleted successfully"})
        
    except Exception as e:
        print(f"Error deleting admin: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# UPDATE ADMIN STATUS BY ID - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/admins/<admin_id>/status", methods=["POST"])
def update_admin_status(admin_id):
    try:
        data = request.get_json()
        new_status = data.get("status")
        
        print(f" Updating admin {admin_id} status to: {new_status}")
        
        if new_status not in ["Active", "Deactivated"]:
            return jsonify({"error": "Invalid status. Use 'Active' or 'Deactivated'"}), 400
        
        # Check if admin exists
        check_query = "SELECT username FROM admins WHERE admin_id = %s"
        admin_data = execute_query(check_query, (admin_id,), fetch_one=True)
        
        if not admin_data:
            return jsonify({"error": "Admin not found"}), 404
        
        # Update status
        update_query = "UPDATE admins SET status = %s WHERE admin_id = %s"
        execute_query(update_query, (new_status, admin_id))
        
        username = admin_data.get("username")
        print(f" Admin '{username}' status updated to {new_status}")
        return jsonify({"message": f"Admin '{username}' status updated to {new_status}"})
        
    except Exception as e:
        print(f"Error updating status: {e}")
        return jsonify({"error": str(e)}), 500



# ===============================
# unlock_admin 
# ===============================
@app.route("/api/superadmin/admins/<admin_id>/unlock", methods=["POST"])
def unlock_admin(admin_id):
    try:
        admin = execute_query("SELECT username FROM admins WHERE admin_id = %s", (admin_id,), fetch_one=True)
        if not admin:
            return jsonify({"error": "Admin not found"}), 404
        reset_login_lockout("admins", "admin_id", admin_id)
        return jsonify({"message": f"Admin '{admin.get('username')}' is allowed to log in again"})
    except Exception as error:
        return jsonify({"error": str(error)}), 500



# ===============================
# TEAMS MANAGEMENT (Super Admin Only)
# ===============================

# Helper function: Get team prefix based on area
def get_team_prefix(area):
    """Get team ID prefix based on area"""
    if not area:
        return 'TEAM'
    area_clean = area.upper().strip()
    
    if area_clean == 'SANTA CRUZ':
        return 'TMSC'
    elif area_clean == 'PAGSANJAN':
        return 'TMPG'
    elif area_clean == 'MAGDALENA':
        return 'TMMG'
    elif area_clean == 'PILA':
        return 'TMPL'
    else:
        return 'TEAM'

# Helper function: Generate next team ID
def generate_next_team_id(area):
    """Generate the next available team ID"""
    try:
        prefix = get_team_prefix(area)
        
        if prefix == 'TEAM':
            return f"TEAM-{int(datetime.now().timestamp())}"
        
        query = "SELECT team_id FROM teams WHERE team_id LIKE %s"
        pattern = f"{prefix}-%"
        results = execute_query(query, (pattern,), fetch=True) or []
        
        existing_numbers = set()
        for row in results:
            team_id = row.get('team_id', '')
            if team_id and team_id.startswith(f"{prefix}-"):
                try:
                    num = int(team_id.split("-")[1])
                    existing_numbers.add(num)
                except:
                    pass
        
        next_number = 1
        while next_number in existing_numbers:
            next_number += 1
        
        return f"{prefix}-{str(next_number).zfill(4)}"
        
    except Exception as e:
        print(f" Error generating team ID: {e}")
        return f"TEAM-{int(datetime.now().timestamp())}"

# ===============================
# CREATE TEAM
# ===============================
@app.route("/api/superadmin/teams", methods=["POST"])
def create_team():
    """Create a new team"""
    data = request.json
    team_name = data.get("team_name")
    area = data.get("area")
    team_leader_id = data.get("team_leader_id")
    status = data.get("status", "Active")
    
    print("=" * 60)
    print(" DEBUGGING CREATE TEAM")
    print(f" Team Name: '{team_name}'")
    print(f" Area: '{area}'")
    print(f" Team Leader: '{team_leader_id}'")
    print("=" * 60)
    
    if not team_name or not area:
        return jsonify({"error": "Team name and area are required"}), 400
    
    if len(team_name) < 2 or len(team_name) > 100:
        return jsonify({"error": "Team name must be 2-100 characters"}), 400
    
    try:
        # Check duplicate team name
        check_query = "SELECT team_id FROM teams WHERE team_name = %s"
        existing = execute_query(check_query, (team_name,), fetch_one=True)
        if existing:
            return jsonify({"error": "Team name already exists"}), 400
        
        # Generate team_id
        team_id = generate_next_team_id(area)
        
        # ==============================================
        # BAGO: I-verify ang team leader
        # ==============================================
        if team_leader_id:
            # Check if technician exists and is active
            tech_check = """
                SELECT technician_id, status, team_id 
                FROM technicians 
                WHERE technician_id = %s
            """
            tech = execute_query(tech_check, (team_leader_id,), fetch_one=True)
            
            if not tech:
                return jsonify({"error": "Selected team leader not found"}), 400
            
            if tech.get('status') != 'Active':
                return jsonify({"error": "Selected technician is not active"}), 400
            
            if tech.get('team_id'):
                return jsonify({"error": "Selected technician is already assigned to a team"}), 400
        
        # ==============================================
        # 1. INSERT TEAM
        # ==============================================
        insert_query = """
            INSERT INTO teams (team_id, team_name, area, team_leader_id, status, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
        """
        params = (team_id, team_name, area, team_leader_id, status)
        execute_query(insert_query, params)
        
        print(f" Team created successfully: {team_id}")
        
        # ==============================================
        # BAGO: 2. I-ADD ANG TEAM LEADER AS MEMBER
        # ==============================================
        if team_leader_id:
            # I-update ang technician's team_id
            update_tech_query = """
                UPDATE technicians 
                SET team_id = %s 
                WHERE technician_id = %s
            """
            execute_query(update_tech_query, (team_id, team_leader_id))
            print(f" Team leader {team_leader_id} added as member of team {team_id}")
        
        return jsonify({
            "message": "Team created successfully",
            "team_id": team_id
        }), 201
        
    except Exception as e:
        print(f" Error creating team: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ===============================
# LIST ALL TEAMS
# ===============================
@app.route("/api/superadmin/teams", methods=["GET"])
def list_teams():
    """Get all teams"""
    try:
        query = """
            SELECT team_id, team_name, area, team_leader_id, status, created_at, updated_at
            FROM teams 
            ORDER BY team_name
        """
        teams = execute_query(query, fetch=True) or []
        return jsonify(teams)
    except Exception as e:
        print(f"Error listing teams: {e}")
        return jsonify({"error": str(e)}), 500




# ===============================
# UPDATE TEAM STATUS
# ===============================
@app.route("/api/superadmin/teams/<team_id>/status", methods=["POST"])
def update_team_status(team_id):
    """Activate or deactivate a team"""
    try:
        data = request.get_json()
        new_status = data.get("status")
        
        if new_status not in ["Active", "Inactive"]:
            return jsonify({"error": "Invalid status. Use 'Active' or 'Inactive'"}), 400
        
        check_query = "SELECT team_name FROM teams WHERE team_id = %s"
        team = execute_query(check_query, (team_id,), fetch_one=True)
        if not team:
            return jsonify({"error": "Team not found"}), 404
        
        update_query = "UPDATE teams SET status = %s, updated_at = NOW() WHERE team_id = %s"
        execute_query(update_query, (new_status, team_id))
        
        return jsonify({"message": f"Team '{team['team_name']}' status updated to {new_status}"})
        
    except Exception as e:
        print(f"Error updating status: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# GET SINGLE TEAM
# ===============================
@app.route("/api/superadmin/teams/<team_id>", methods=["GET"])
def get_team(team_id):
    """Get a single team by ID"""
    try:
        query = """
            SELECT team_id, team_name, area, team_leader_id, status, created_at, updated_at
            FROM teams 
            WHERE team_id = %s
        """
        team = execute_query(query, (team_id,), fetch_one=True)
        
        if not team:
            return jsonify({"error": "Team not found"}), 404
            
        return jsonify(team)
        
    except Exception as e:
        print(f"Error getting team: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# UPDATE TEAM
# ===============================
@app.route("/api/superadmin/teams/<team_id>", methods=["PUT"])
def update_team(team_id):
    """Update team details"""
    data = request.json
    team_name = data.get("team_name")
    area = data.get("area")
    team_leader_id = data.get("team_leader_id")
    status = data.get("status")
    
    if not team_name or not area:
        return jsonify({"error": "Team name and area are required"}), 400
    
    try:
        # Check if team exists
        check_query = "SELECT team_id, team_leader_id, area FROM teams WHERE team_id = %s"
        existing_team = execute_query(check_query, (team_id,), fetch_one=True)
        if not existing_team:
            return jsonify({"error": "Team not found"}), 404
        
        # Check duplicate team name (exclude current team)
        check_name_query = "SELECT team_id FROM teams WHERE team_name = %s AND team_id != %s"
        duplicate = execute_query(check_name_query, (team_name, team_id), fetch_one=True)
        if duplicate:
            return jsonify({"error": "Team name already exists"}), 400
        
        # ==============================================
        # I-DETECT KUNG NAGBAGO ANG AREA
        # ==============================================
        old_area = existing_team.get('area')
        area_changed = old_area != area
        print(f" Area changed: {old_area} → {area} (Changed: {area_changed})")
        
        # ==============================================
        # I-HANDLE ANG TEAM LEADER CHANGES
        # ==============================================
        old_leader_id = existing_team.get('team_leader_id')
        
        if team_leader_id and team_leader_id != old_leader_id:
            tech_check = """
                SELECT technician_id, status, team_id 
                FROM technicians 
                WHERE technician_id = %s
            """
            tech = execute_query(tech_check, (team_leader_id,), fetch_one=True)
            
            if not tech:
                return jsonify({"error": "Selected team leader not found"}), 400
            
            if tech.get('status') != 'Active':
                return jsonify({"error": "Selected technician is not active"}), 400
            
            if tech.get('team_id') and tech.get('team_id') != team_id:
                return jsonify({"error": "Selected technician is already assigned to another team"}), 400
        
        # ==============================================
        # 1. UPDATE TEAM
        # ==============================================
        update_query = """
            UPDATE teams 
            SET team_name = %s, area = %s, team_leader_id = %s, status = %s, updated_at = NOW()
            WHERE team_id = %s
        """
        params = (team_name, area, team_leader_id, status, team_id)
        execute_query(update_query, params)
        
        # ==============================================
        # 2. KUNG NAGBAGO ANG AREA, I-UPDATE ANG LAHAT NG MEMBERS
        # ==============================================
        if area_changed:
            print(f" Updating all members of team {team_id} from '{old_area}' to '{area}'")
            
            # I-update ang area ng lahat ng technicians na member ng team
            update_members_query = """
                UPDATE technicians 
                SET area = %s 
                WHERE team_id = %s
            """
            execute_query(update_members_query, (area, team_id))
            
            # I-count kung ilan ang na-update
            count_query = "SELECT COUNT(*) as count FROM technicians WHERE team_id = %s"
            count_result = execute_query(count_query, (team_id,), fetch_one=True)
            member_count = count_result.get('count', 0) if count_result else 0
            
            print(f" Updated {member_count} technicians' area from '{old_area}' to '{area}'")
        
        # ==============================================
        # 3. I-HANDLE ANG TEAM LEADER CHANGES
        # ==============================================
        if team_leader_id and team_leader_id != old_leader_id:
            # I-add ang bagong leader sa team (kung hindi pa member)
            check_member = """
                SELECT technician_id FROM technicians 
                WHERE technician_id = %s AND team_id = %s
            """
            is_member = execute_query(check_member, (team_leader_id, team_id), fetch_one=True)
            
            if not is_member:
                update_tech = "UPDATE technicians SET team_id = %s WHERE technician_id = %s"
                execute_query(update_tech, (team_id, team_leader_id))
                print(f" New team leader {team_leader_id} added as member")
        
        print(f" Team updated successfully: {team_id}")
        
        return jsonify({
            "message": "Team updated successfully",
            "area_changed": area_changed,
            "old_area": old_area,
            "new_area": area,
            "members_updated": member_count if area_changed else 0
        })
        
    except Exception as e:
        print(f"Error updating team: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ===============================
# ADD MEMBER TO TEAM
# ===============================
@app.route("/api/superadmin/teams/<team_id>/add-member", methods=["POST"])
def add_team_member(team_id):
    """Add a technician to a team"""
    data = request.json
    technician_id = data.get("technician_id")
    
    print("=" * 60)
    print(f" ADD MEMBER - Team: {team_id}, Technician: {technician_id}")
    print("=" * 60)
    
    if not technician_id:
        return jsonify({"error": "Technician ID is required"}), 400
    
    try:
        # Check if team exists
        team_query = "SELECT team_id FROM teams WHERE team_id = %s"
        team = execute_query(team_query, (team_id,), fetch_one=True)
        if not team:
            print(f" Team {team_id} not found")
            return jsonify({"error": "Team not found"}), 404
        
        # Check if technician exists
        tech_query = "SELECT technician_id, team_id FROM technicians WHERE technician_id = %s"
        tech = execute_query(tech_query, (technician_id,), fetch_one=True)
        if not tech:
            print(f" Technician {technician_id} not found")
            return jsonify({"error": "Technician not found"}), 404
        
        # Check if technician is already in a team
        if tech.get('team_id'):
            print(f" Technician {technician_id} already in team {tech.get('team_id')}")
            return jsonify({"error": "Technician is already assigned to a team"}), 400
        
        # Update technician's team
        update_query = "UPDATE technicians SET team_id = %s WHERE technician_id = %s"
        execute_query(update_query, (team_id, technician_id))
        
        print(f" Successfully added {technician_id} to {team_id}")
        return jsonify({"message": "Member added to team successfully"})
        
    except Exception as e:
        print(f" Error adding member: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# REMOVE MEMBER FROM TEAM
# ===============================
@app.route("/api/superadmin/teams/<team_id>/remove-member", methods=["POST"])
def remove_team_member(team_id):
    """Remove a technician from a team"""
    data = request.json
    technician_id = data.get("technician_id")
    
    print("=" * 60)
    print(f" REMOVE MEMBER - Team: {team_id}, Technician: {technician_id}")
    print("=" * 60)
    
    if not technician_id:
        return jsonify({"error": "Technician ID is required"}), 400
    
    try:
        # Check if technician exists and is in this team
        check_query = "SELECT team_id FROM technicians WHERE technician_id = %s"
        current = execute_query(check_query, (technician_id,), fetch_one=True)
        if not current:
            print(f" Technician {technician_id} not found")
            return jsonify({"error": "Technician not found"}), 404
        
        if current.get('team_id') != team_id:
            print(f" Technician {technician_id} is in team {current.get('team_id')}, not {team_id}")
            return jsonify({"error": "Technician is not in this team"}), 400
        
        # Remove technician from team
        update_query = "UPDATE technicians SET team_id = NULL WHERE technician_id = %s"
        execute_query(update_query, (technician_id,))
        
        print(f" Successfully removed {technician_id} from {team_id}")
        return jsonify({"message": "Member removed from team successfully"})
        
    except Exception as e:
        print(f" Error removing member: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ===============================
# GET ALL TECHNICIANS (with team_id)
# ===============================
@app.route("/api/superadmin/technicians", methods=["GET"])
def get_all_technicians():
    """Get all technicians with their team info"""
    try:
        query = """
            SELECT 
                technician_id, 
                name, 
                email, 
                contact_number,
                area, 
                team_id, 
                status, 
                created_at,
                profile_photo,
                failed_login_attempts, locked_until, lock_level,
                CASE WHEN locked_until > NOW() THEN 1 ELSE 0 END AS login_locked
            FROM technicians 
            ORDER BY created_at DESC
        """
        technicians = execute_query(query, fetch=True) or []
        return jsonify(technicians)
    except Exception as e:
        print(f"❌ Error getting technicians: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# TECHNICIANS PAGE
# ===============================
@app.route("/superadmin/technicians")
def superadmin_technicians():
    """Render technicians management page"""
    return render_template("superadmin-technicians.html")


# ===============================
# TECHNICIAN MANAGEMENT (Super Admin Only)
# ===============================
# Helper function: Get prefix based on area
def get_technician_prefix(area):
    """Get technician ID prefix based on area"""
    # Handle None or empty
    if not area:
        print(" Area is None or empty")
        return 'TECH'
    
    # Convert to uppercase and strip whitespace
    area_clean = area.upper().strip()
    print(f" get_technician_prefix - Cleaned area: '{area_clean}'")
    
    # Simple if-else
    if area_clean == 'SANTA CRUZ':
        print(" Match: SANTA CRUZ -> TSTC")
        return 'TSTC'
    elif area_clean == 'PAGSANJAN':
        print(" Match: PAGSANJAN -> TPSN")
        return 'TPSN'
    elif area_clean == 'MAGDALENA':
        print(" Match: MAGDALENA -> TMAG")
        return 'TMAG'
    elif area_clean == 'PILA':
        print(" Match: PILA -> TPIL")
        return 'TPIL'
    else:
        print(f" No match for '{area_clean}', using default TECH")
        return 'TECH'


# Helper function: Generate next technician ID
def generate_next_technician_id(area):
    """Generate the next available technician ID based on area"""
    try:
        prefix = get_technician_prefix(area)
        
        # Kung prefix ay TECH, ibig sabihin hindi recognized ang area
        if prefix == 'TECH':
            print(f" Area '{area}' not recognized, using timestamp fallback")
            return f"TECH-{int(datetime.now().timestamp())}"
        
        # Get all technician IDs with this prefix
        query = "SELECT technician_id FROM technicians WHERE technician_id LIKE %s"
        pattern = f"{prefix}-%"
        results = execute_query(query, (pattern,), fetch=True) or []
        
        # Extract numbers from existing IDs
        existing_numbers = set()
        for row in results:
            tech_id = row.get('technician_id', '')
            if tech_id and tech_id.startswith(f"{prefix}-"):
                try:
                    num = int(tech_id.split("-")[1])
                    existing_numbers.add(num)
                except:
                    pass
        
        # Find next number
        next_number = 1
        while next_number in existing_numbers:
            next_number += 1
        
        technician_id = f"{prefix}-{str(next_number).zfill(4)}"
        print(f" Generated: {technician_id}")
        return technician_id
        
    except Exception as e:
        print(f" Error: {e}")
        return f"TECH-{int(datetime.now().timestamp())}"



# ===============================
# CREATE TECHNICIAN (Super Admin Only)
# ===============================
@app.route("/api/superadmin/technicians", methods=["POST"])
def create_technician():
    """Create a new technician account"""
    data = request.json
    name = data.get("name")
    email = data.get("email")
    area = data.get("area")
    team_id = data.get("team_id")

    print("=" * 60)
    print(" DEBUGGING CREATE TECHNICIAN")
    print(f" Name: '{name}'")
    print(f" Email: '{email}'")
    print(f" Area: '{area}'")
    print(f" Team ID: '{team_id}'")
    print("=" * 60)

    if not name or not email or not area:
        return jsonify({"error": "Name, email, and area are required"}), 400

    # NAME VALIDATION
    name_pattern = re.compile(r"^[a-zA-Z\s\'-]{2,100}$")
    if not name_pattern.match(name):
        return jsonify({"error": "Invalid name. Use 2-100 characters (letters, spaces, apostrophe, hyphen)"}), 400

    # EMAIL VALIDATION
    email_pattern = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")
    if not email_pattern.match(email):
        return jsonify({"error": "Invalid email address"}), 400

    try:
        # ========== CHECK EMAIL IN ALL TABLES ==========
        check_all_query = """
            SELECT 
                (SELECT COUNT(*) FROM technicians WHERE email = %s) as tech_count,
                (SELECT COUNT(*) FROM admins WHERE email = %s) as admin_count,
                (SELECT COUNT(*) FROM superadmins WHERE email = %s) as superadmin_count
        """
        result = execute_query(check_all_query, (email, email, email), fetch_one=True)
        
        tech_exists = result.get('tech_count', 0) > 0 if result else False
        admin_exists = result.get('admin_count', 0) > 0 if result else False
        superadmin_exists = result.get('superadmin_count', 0) > 0 if result else False

        # GENERIC MESSAGE - EMAIL LANG ANG NAKALAGAY
        if tech_exists or admin_exists or superadmin_exists:
            return jsonify({"error": f"Email '{email}' already exists"}), 400

        # Validate team_id if provided
        if team_id:
            team_check = "SELECT team_id FROM teams WHERE team_id = %s AND status = 'Active'"
            team_exists = execute_query(team_check, (team_id,), fetch_one=True)
            if not team_exists:
                return jsonify({"error": "Invalid or inactive team selected"}), 400

        # Generate technician_id
        technician_id = generate_next_technician_id(area)
        
        # ✅ GUMAWA NG RANDOM PASSWORD (8 CHARACTERS - letters and numbers)
        default_password = generate_secure_password(8)
        created_at = ph_now_iso()

        # Insert into MySQL
        insert_query = """
            INSERT INTO technicians 
            (technician_id, name, email, password, area, team_id, status, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        params = (
            technician_id,
            name,
            email,
            hash_password(default_password),
            area,
            team_id,
            "Active",
            created_at
        )
        
        execute_query(insert_query, params)
        print(f" Technician saved successfully: {technician_id}")
        print(f" Password: {default_password}")
        
        # Send email
        try:
            send_technician_email(email, technician_id, name, default_password, area, team_id)
        except Exception as e:
            print(f" Email error: {e}")

        return jsonify({
            "message": "Technician created successfully",
            "technician_id": technician_id,
            "password": default_password  # ✅ I-RETURN ANG PASSWORD
        }), 201

    except Exception as e:
        print(f" Error creating technician: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ===============================
# SEND TECHNICIAN ACCOUNT EMAIL
# BREVO API
# ===============================

def send_technician_email(
    to_email,
    technician_id,
    name,
    password,
    area,
    team_id=None
):
    """
    Sends newly created Technician account credentials via Brevo.

    Includes:
        - Technician ID
        - Name
        - Email
        - Area
        - Team ID
        - Temporary Password

    Returns:
        True  = email sent successfully
        False = email sending failed
    """

    import requests
    import html

    # ===============================
    # BREVO CONFIGURATION
    # ===============================

    brevo_api_key = os.getenv("BREVO_API_KEY")
    sender_email = os.getenv(
        "SMTP_FROM",
        "noreply@cablevisioncableinternet.com"
    )
    sender_name = "Cablevision Systems Corporation"

    # ===============================
    # CHECK CONFIGURATION
    # ===============================

    if not brevo_api_key:
        print(" BREVO_API_KEY is not configured!")
        return False

    if not to_email:
        print(" Cannot send technician email: recipient email is empty.")
        return False

    # ===============================
    # HTML ESCAPE
    # ===============================

    safe_technician_id = html.escape(str(technician_id or "N/A"))
    safe_name = html.escape(str(name or "N/A"))
    safe_email = html.escape(str(to_email or "N/A"))
    safe_area = html.escape(str(area or "N/A"))
    safe_team_id = html.escape(str(team_id or "Not Assigned"))
    safe_password = html.escape(str(password or "N/A"))

    # ===============================
    # EMAIL SUBJECT
    # ===============================

    subject = "Your Technician Account - CableVision"

    # ===============================
    # HTML EMAIL
    # ===============================

    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>CableVision Technician Account</title>
    </head>
    <body style="
        margin:0;
        padding:0;
        background:#eef2f6;
        font-family:Arial, Helvetica, sans-serif;
    ">
        <div style="
            width:100%;
            padding:30px 0;
        ">
            <div style="
                max-width:580px;
                margin:0 auto;
                background:#ffffff;
                border-radius:24px;
                overflow:hidden;
                box-shadow:0 10px 30px rgba(0,0,0,0.10);
            ">
                <!-- HEADER -->
                <div style="
                    background:#001f3f;
                    padding:30px 25px;
                    text-align:center;
                ">
                    <div style="
                        font-size:28px;
                        font-weight:700;
                        color:#ffffff;
                    ">
                        Cablevision
                    </div>
                    <div style="
                        margin-top:6px;
                        font-size:13px;
                        color:#93c5fd;
                    ">
                        Internet Service Provider
                    </div>
                    <div style="
                        margin-top:16px;
                        display:inline-block;
                        background:rgba(255,255,255,0.12);
                        color:#dbeafe;
                        padding:6px 14px;
                        border-radius:20px;
                        font-size:11px;
                        font-weight:700;
                    ">
                        TECHNICIAN ACCOUNT
                    </div>
                </div>

                <!-- CONTENT -->
                <div style="padding:28px;">
                    <h2 style="
                        margin:0 0 8px 0;
                        font-size:22px;
                        color:#0f172a;
                    ">
                        Welcome to Cablevision!
                    </h2>
                    <p style="
                        margin:0 0 20px 0;
                        font-size:15px;
                        line-height:1.6;
                        color:#475569;
                    ">
                        Hello <strong>{safe_name}</strong>,
                        your technician account has been successfully created.
                    </p>

                    <!-- ACCOUNT DETAILS -->
                    <div style="
                        background:#f8fafc;
                        border-radius:16px;
                        padding:20px;
                        border:1px solid #e2e8f0;
                    ">
                        <div style="
                            font-size:11px;
                            font-weight:700;
                            color:#64748b;
                            margin-bottom:5px;
                        ">
                            TECHNICIAN ID
                        </div>
                        <div style="
                            font-size:18px;
                            font-weight:700;
                            color:#0f172a;
                            margin-bottom:16px;
                        ">
                            {safe_technician_id}
                        </div>

                        <div style="
                            border-top:1px solid #e2e8f0;
                            padding-top:14px;
                            margin-bottom:16px;
                        ">
                            <div style="
                                font-size:11px;
                                font-weight:700;
                                color:#64748b;
                                margin-bottom:5px;
                            ">
                                NAME
                            </div>
                            <div style="
                                font-size:17px;
                                font-weight:700;
                                color:#0f172a;
                            ">
                                {safe_name}
                            </div>
                        </div>

                        <div style="
                            border-top:1px solid #e2e8f0;
                            padding-top:14px;
                            margin-bottom:16px;
                        ">
                            <div style="
                                font-size:11px;
                                font-weight:700;
                                color:#64748b;
                                margin-bottom:5px;
                            ">
                                EMAIL
                            </div>
                            <div style="
                                font-size:15px;
                                color:#0f172a;
                            ">
                                {safe_email}
                            </div>
                        </div>

                        <div style="
                            border-top:1px solid #e2e8f0;
                            padding-top:14px;
                            margin-bottom:16px;
                        ">
                            <div style="
                                font-size:11px;
                                font-weight:700;
                                color:#64748b;
                                margin-bottom:5px;
                            ">
                                ASSIGNED AREA
                            </div>
                            <div style="
                                font-size:16px;
                                font-weight:700;
                                color:#0f172a;
                            ">
                                {safe_area}
                            </div>
                        </div>

                        <div style="
                            border-top:1px solid #e2e8f0;
                            padding-top:14px;
                            margin-bottom:16px;
                        ">
                            <div style="
                                font-size:11px;
                                font-weight:700;
                                color:#64748b;
                                margin-bottom:5px;
                            ">
                                TEAM ID
                            </div>
                            <div style="
                                font-size:16px;
                                font-weight:700;
                                color:#0f172a;
                            ">
                                {safe_team_id}
                            </div>
                        </div>

                        <div style="
                            border-top:1px solid #e2e8f0;
                            padding-top:14px;
                        ">
                            <div style="
                                font-size:11px;
                                font-weight:700;
                                color:#64748b;
                                margin-bottom:5px;
                            ">
                                TEMPORARY PASSWORD
                            </div>
                            <div style="
                                display:inline-block;
                                background:#eff6ff;
                                color:#001f3f;
                                padding:10px 14px;
                                border-radius:8px;
                                font-family:monospace;
                                font-size:18px;
                                font-weight:700;
                                letter-spacing:1px;
                            ">
                                {safe_password}
                            </div>
                        </div>
                    </div>

                    <!-- SECURITY NOTICE -->
                    <div style="
                        margin-top:20px;
                        padding:16px;
                        background:#fff7ed;
                        border:1px solid #fed7aa;
                        border-radius:14px;
                    ">
                        <div style="
                            font-size:14px;
                            font-weight:700;
                            color:#9a3412;
                            margin-bottom:6px;
                        ">
                            Security Notice
                        </div>
                        <div style="
                            font-size:13px;
                            line-height:1.6;
                            color:#7c2d12;
                        ">
                            This is your temporary password.
                            Please change your password after your
                            first login and do not share your
                            account credentials with anyone.
                        </div>
                    </div>

                    <!-- FOOTER MESSAGE -->
                    <div style="
                        margin-top:28px;
                        padding-top:20px;
                        border-top:1px solid #e2e8f0;
                        text-align:center;
                    ">
                        <div style="
                            font-size:13px;
                            color:#64748b;
                            line-height:1.5;
                        ">
                            Welcome to the Cablevision team!
                        </div>
                    </div>
                </div>

                <!-- FOOTER -->
                <div style="
                    background:#f1f5f9;
                    padding:16px 20px;
                    text-align:center;
                ">
                    <div style="
                        font-size:11px;
                        color:#64748b;
                    ">
                        2026 Cablevision Systems Corporation. All rights reserved.
                    </div>
                    <div style="
                        margin-top:4px;
                        font-size:11px;
                        color:#94a3b8;
                    ">
                        Sta. Cruz, Laguna, Philippines
                    </div>
                </div>
            </div>
        </div>
    </body>
    </html>
    """

    # ===============================
    # PLAIN TEXT EMAIL
    # ===============================

    plain_body = f"""
Cablevision Systems Corporation

TECHNICIAN ACCOUNT CREATED

Welcome to Cablevision!

Hello {name},

Your technician account has been successfully created.

Technician ID:
{technician_id}

Name:
{name}

Email:
{to_email}

Assigned Area:
{area}

Team ID:
{team_id or 'Not Assigned'}

Temporary Password:
{password}

Please change your password after your first login.

Do not share your account credentials with anyone.

Welcome to the Cablevision team!

Cablevision Systems Corporation
Sta. Cruz, Laguna, Philippines
"""

    # ===============================
    # BREVO API PAYLOAD
    # ===============================

    payload = {
        "sender": {
            "name": sender_name,
            "email": sender_email
        },
        "to": [
            {
                "email": to_email
            }
        ],
        "subject": subject,
        "htmlContent": html_body,
        "textContent": plain_body
    }

    # ===============================
    # BREVO API HEADERS
    # ===============================

    headers = {
        "accept": "application/json",
        "api-key": brevo_api_key,
        "content-type": "application/json"
    }

    # ===============================
    # SEND EMAIL
    # ===============================

    try:
        print("========================================")
        print(" TECHNICIAN ACCOUNT EMAIL")
        print("========================================")
        print(f" From: {sender_email}")
        print(f" To: {to_email}")
        print(f" Technician ID: {technician_id}")
        print(f" Name: {name}")
        print(f" Area: {area}")
        print(f" Team ID: {team_id}")
        print("========================================")

        response = requests.post(
            "https://api.brevo.com/v3/smtp/email",
            headers=headers,
            json=payload,
            timeout=30
        )

        # ===============================
        # CHECK BREVO RESPONSE
        # ===============================

        if response.status_code not in (200, 201):
            print(f" Brevo API error ({response.status_code}): {response.text}")
            return False

        # ===============================
        # GET BREVO MESSAGE ID
        # ===============================

        try:
            brevo_response = response.json()
            message_id = brevo_response.get("messageId")
            if message_id:
                print(f" Brevo Message ID: {message_id}")
        except Exception:
            pass

        print(f" Technician account email sent successfully to {to_email}")
        return True

    # ===============================
    # ERROR HANDLING
    # ===============================

    except requests.exceptions.Timeout:
        print(" Brevo API request timed out.")
        return False

    except requests.exceptions.RequestException as e:
        print(f" Brevo API request error: {e}")
        return False

    except Exception as e:
        print(f" Error sending technician account email: {e}")
        import traceback
        traceback.print_exc()
        return False



# ===============================
# GET SINGLE TECHNICIAN BY ID
# ===============================
@app.route("/api/superadmin/technicians/<technician_id>", methods=["GET"])
def get_technician(technician_id):
    """Get a single technician by ID"""
    try:
        query = """
            SELECT technician_id, name, email, area, team_id, status, created_at
            FROM technicians 
            WHERE technician_id = %s
        """
        technician_data = execute_query(query, (technician_id,), fetch_one=True)
        
        if not technician_data:
            return jsonify({"error": "Technician not found"}), 404
        
        return jsonify(technician_data)
        
    except Exception as e:
        print(f"Error getting technician: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# UPDATE TECHNICIAN
# ===============================
@app.route("/api/superadmin/technicians/<technician_id>", methods=["PUT"])
def update_technician(technician_id):
    """Update technician information"""
    try:
        data = request.json
        name = data.get("name")
        email = data.get("email")
        area = data.get("area")
        team_id = data.get("team_id")  # ← BAGO
        
        # Check if technician exists
        check_query = "SELECT technician_id FROM technicians WHERE technician_id = %s"
        exists = execute_query(check_query, (technician_id,), fetch_one=True)
        
        if not exists:
            return jsonify({"error": "Technician not found"}), 404
        
        # Build update query dynamically
        updates = []
        params = []
        
        if name is not None:
            updates.append("name = %s")
            params.append(name)
        if email is not None:
            # Check for duplicate email
            duplicate_query = "SELECT email FROM technicians WHERE email = %s AND technician_id != %s"
            duplicate = execute_query(duplicate_query, (email, technician_id), fetch_one=True)
            if duplicate:
                return jsonify({"error": "Email already exists"}), 400
            updates.append("email = %s")
            params.append(email)
        if area is not None:
            updates.append("area = %s")
            params.append(area)
        if team_id is not None:
            # If team_id is empty string, set to NULL
            if team_id == "":
                team_id = None
            # Validate team if provided
            if team_id:
                team_check = "SELECT team_id FROM teams WHERE team_id = %s AND status = 'Active'"
                team_exists = execute_query(team_check, (team_id,), fetch_one=True)
                if not team_exists:
                    return jsonify({"error": "Invalid or inactive team selected"}), 400
            updates.append("team_id = %s")
            params.append(team_id)
        
        if not updates:
            return jsonify({"error": "No fields to update"}), 400
        
        params.append(technician_id)
        update_query = f"UPDATE technicians SET {', '.join(updates)} WHERE technician_id = %s"
        execute_query(update_query, tuple(params))
        
        print(f" Technician {technician_id} updated successfully")
        return jsonify({"message": "Technician updated successfully"})
        
    except Exception as e:
        print(f"Error updating technician: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# DELETE TEAM
# ===============================
@app.route("/api/superadmin/teams/<team_id>", methods=["DELETE"])
def delete_team(team_id):
    """Delete a team"""
    try:
        # Check if team exists
        check_query = "SELECT team_name FROM teams WHERE team_id = %s"
        team = execute_query(check_query, (team_id,), fetch_one=True)
        if not team:
            return jsonify({"error": "Team not found"}), 404
        
        # Remove team_id from technicians before deleting
        update_techs = "UPDATE technicians SET team_id = NULL WHERE team_id = %s"
        execute_query(update_techs, (team_id,))
        
        # Delete team
        delete_query = "DELETE FROM teams WHERE team_id = %s"
        execute_query(delete_query, (team_id,))
        
        return jsonify({"message": f"Team '{team['team_name']}' deleted successfully"})
        
    except Exception as e:
        print(f"Error deleting team: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# UPDATE TECHNICIAN STATUS
# ===============================
@app.route("/api/superadmin/technicians/<technician_id>/status", methods=["POST"])
def update_technician_status(technician_id):
    """Activate or deactivate a technician"""
    try:
        data = request.get_json()
        new_status = data.get("status")
        
        print(f" Updating technician {technician_id} status to: {new_status}")
        
        if new_status not in ["Active", "Deactivated"]:
            return jsonify({"error": "Invalid status. Use 'Active' or 'Deactivated'"}), 400
        
        # Check if technician exists
        check_query = "SELECT name FROM technicians WHERE technician_id = %s"
        technician_data = execute_query(check_query, (technician_id,), fetch_one=True)
        
        if not technician_data:
            return jsonify({"error": "Technician not found"}), 404
        
        # Update status
        update_query = "UPDATE technicians SET status = %s WHERE technician_id = %s"
        execute_query(update_query, (new_status, technician_id))
        
        name = technician_data.get("name")
        print(f" Technician '{name}' status updated to {new_status}")
        return jsonify({"message": f"Technician '{name}' status updated to {new_status}"})
        
    except Exception as e:
        print(f"Error updating status: {e}")
        return jsonify({"error": str(e)}), 500



# ===============================
# Unlock Technician
# ===============================
@app.route("/api/superadmin/technicians/<technician_id>/unlock", methods=["POST"])
def unlock_technician(technician_id):
    try:
        technician = execute_query("SELECT name FROM technicians WHERE technician_id = %s", (technician_id,), fetch_one=True)
        if not technician:
            return jsonify({"error": "Technician not found"}), 404
        reset_login_lockout("technicians", "technician_id", technician_id)
        return jsonify({"message": f"Technician '{technician.get('name')}' is allowed to log in again"})
    except Exception as error:
        return jsonify({"error": str(error)}), 500



# ===============================
# GET TECHNICIANS BY AREA
# ===============================
@app.route("/api/superadmin/technicians/by-area/<area>", methods=["GET"])
def get_technicians_by_area(area):
    """Get technicians filtered by area"""
    try:
        query = """
            SELECT technician_id, name, email, area, status, created_at
            FROM technicians 
            WHERE area = %s
            ORDER BY name
        """
        technicians_list = execute_query(query, (area,), fetch=True) or []
        
        return jsonify(technicians_list)
        
    except Exception as e:
        print(f"Error getting technicians by area: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# GET TECHNICIAN STATISTICS
# ===============================
@app.route("/api/superadmin/technicians/statistics", methods=["GET"])
def get_technician_statistics():
    """Get technician statistics"""
    try:
        # Total technicians
        total_query = "SELECT COUNT(*) as total FROM technicians"
        total_result = execute_query(total_query, fetch_one=True)
        total = total_result.get('total', 0) if total_result else 0
        
        # Active vs Deactivated
        active_query = "SELECT COUNT(*) as active FROM technicians WHERE status = 'Active'"
        active_result = execute_query(active_query, fetch_one=True)
        active = active_result.get('active', 0) if active_result else 0
        
        # By area
        area_query = """
            SELECT area, COUNT(*) as count 
            FROM technicians 
            GROUP BY area 
            ORDER BY count DESC
        """
        by_area = execute_query(area_query, fetch=True) or []
        
        return jsonify({
            "total": total,
            "active": active,
            "deactivated": total - active,
            "by_area": by_area
        })
        
    except Exception as e:
        print(f"Error getting technician statistics: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# DELETE TECHNICIAN - I-ADD ITO (MISSING)
# ===============================
@app.route("/api/superadmin/technicians/<technician_id>", methods=["DELETE"])
def delete_technician(technician_id):
    """Delete a technician"""
    try:
        # Check if technician exists
        check_query = "SELECT name FROM technicians WHERE technician_id = %s"
        tech = execute_query(check_query, (technician_id,), fetch_one=True)
        if not tech:
            return jsonify({"error": "Technician not found"}), 404
        
        # Delete technician
        delete_query = "DELETE FROM technicians WHERE technician_id = %s"
        execute_query(delete_query, (technician_id,))
        
        return jsonify({"message": f"Technician '{tech['name']}' deleted successfully"})
        
    except Exception as e:
        print(f"Error deleting technician: {e}")
        return jsonify({"error": str(e)}), 500
    

@app.route('/api/superadmin/napbox-slots', methods=['GET'])
def get_super_napbox_slots():
    """Get all NAP boxes and slots for super admin monitoring (read-only)"""
    try:
        # FIXED: Gamitin ang tamang column name na 'napbox_name' hindi 'name'
        napboxes_query = """
            SELECT id, napbox_name as name, location, latitude, longitude, area, barangay, coverage_radius 
            FROM napboxes 
            ORDER BY area, napbox_name
        """
        napboxes = execute_query(napboxes_query, fetch=True) or []
        
        # FIXED: Gamitin ang tamang column name sa JOIN
        slots_query = """
            SELECT 
                ns.id, 
                ns.napbox_id, 
                ns.slot_number, 
                ns.status, 
                ns.customer_name, 
                ns.customer_phone, 
                ns.barangay, 
                ns.updated_at,
                n.napbox_name as napbox_name, 
                n.area
            FROM napbox_slots ns
            LEFT JOIN napboxes n ON ns.napbox_id = n.id
            ORDER BY n.area, n.napbox_name, CAST(ns.slot_number AS UNSIGNED)
        """
        slots = execute_query(slots_query, fetch=True) or []
        
        # Convert datetime to string for JSON serialization
        for slot in slots:
            if slot.get('updated_at'):
                if hasattr(slot['updated_at'], 'isoformat'):
                    slot['updated_at'] = slot['updated_at'].isoformat()
                else:
                    slot['updated_at'] = str(slot['updated_at'])
        
        # Convert napboxes to list - gamitin ang napbox_name bilang name
        napboxes_list = []
        for nb in napboxes:
            napboxes_list.append({
                "id": nb.get('id'),
                "name": nb.get('name'),  # galing sa alias na 'name'
                "latitude": float(nb.get('latitude')) if nb.get('latitude') else None,
                "longitude": float(nb.get('longitude')) if nb.get('longitude') else None,
                "area": nb.get('area'),
                "barangay": nb.get('barangay'),
                "coverage_radius": nb.get('coverage_radius')
            })
        
        # Convert slots to list
        slots_list = []
        for slot in slots:
            slots_list.append({
                "id": slot.get('id'),
                "napbox_id": slot.get('napbox_id'),
                "slot_number": slot.get('slot_number'),
                "status": slot.get('status'),
                "customer_name": slot.get('customer_name'),
                "customer_phone": slot.get('customer_phone'),
                "barangay": slot.get('barangay'),
                "updated_at": slot.get('updated_at'),
                "napbox_name": slot.get('napbox_name'),
                "area": slot.get('area')
            })
        
        print(f" Superadmin NAP Box Slots: {len(napboxes_list)} napboxes, {len(slots_list)} slots")
        
        return jsonify({
            'success': True,
            'napboxes': napboxes_list,
            'slots': slots_list
        })
        
    except Exception as e:
        print(f" Error in get_super_napbox_slots: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'napboxes': [], 'slots': []}), 500

# =========================
# ADD AREA PAGE
# =========================
@app.route("/superadmin/area")
def superadmin_area():
    return render_template("superadmin-area.html")

# =========================
# ADD AREA (WITH DUPLICATE VALIDATION + ZIP) - CONVERTED TO MYSQL
# =========================
@app.route("/api/superadmin/area", methods=["POST"])
def add_area():
    data = request.json

    province = data.get("province")
    city = data.get("city")
    barangay = data.get("barangay")
    zip_code = data.get("zip_code")

    # VALIDATION
    if not province or not city or not barangay or not zip_code:
        return jsonify({"error": "Missing fields"}), 400

    # FORMAT (para consistent)
    province = province.upper().strip()
    city = city.upper().strip()
    barangay = barangay.upper().strip()
    zip_code = str(zip_code).strip()

    try:
        # CHECK DUPLICATE (same province + city + barangay)
        check_query = """
            SELECT id FROM areas 
            WHERE province = %s AND city = %s AND barangay = %s
        """
        existing = execute_query(check_query, (province, city, barangay), fetch_one=True)
        
        if existing:
            return jsonify({
                "error": "Duplicate area already exists",
                "duplicate": True,
                "existing_id": existing['id']
            }), 409

        # INSERT INTO MYSQL
        insert_query = """
            INSERT INTO areas (province, city, barangay, zip, created_at)
            VALUES (%s, %s, %s, %s, NOW())
        """
        area_id = execute_query(insert_query, (province, city, barangay, zip_code))
        
        return jsonify({
            "success": True,
            "id": area_id,
            "message": "Area added successfully"
        })

    except Exception as e:
        print(f"Error adding area: {e}")
        return jsonify({"error": str(e)}), 500

# =========================
# GET AREAS - CONVERTED TO MYSQL
# =========================
@app.route("/api/superadmin/areas")
def get_areas():
    try:
        query = """
            SELECT id, province, city, barangay, zip, created_at
            FROM areas 
            ORDER BY province, city, barangay
        """
        areas = execute_query(query, fetch=True) or []
        
        # Format response to match frontend expectations
        result = []
        for area in areas:
            result.append({
                "id": area['id'],
                "province": area['province'],
                "city": area['city'],
                "barangay": area['barangay'],
                "zip": area.get('zip', ''),
                "created_at": area.get('created_at', '')
            })
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error getting areas: {e}")
        return jsonify({"error": str(e)}), 500

# =========================
# DELETE AREA - CONVERTED TO MYSQL
# =========================
@app.route("/api/superadmin/area/<int:area_id>", methods=["DELETE"])
def delete_area(area_id):
    try:
        # Check if area exists before deleting
        check_query = "SELECT id, city, barangay FROM areas WHERE id = %s"
        area = execute_query(check_query, (area_id,), fetch_one=True)
        
        if not area:
            return jsonify({"error": "Area not found"}), 404
        
        # Delete from MySQL
        delete_query = "DELETE FROM areas WHERE id = %s"
        execute_query(delete_query, (area_id,))
        
        return jsonify({
            "success": True, 
            "message": f"Area '{area['barangay']}, {area['city']}' deleted successfully"
        })
        
    except Exception as e:
        print(f"Error deleting area: {e}")
        return jsonify({"error": str(e)}), 500

# =========================
# UPDATE AREA - NEW ENDPOINT
# =========================
@app.route("/api/superadmin/area/<int:area_id>", methods=["PUT"])
def update_area(area_id):
    data = request.json
    
    province = data.get("province")
    city = data.get("city")
    barangay = data.get("barangay")
    zip_code = data.get("zip_code")
    
    if not province or not city or not barangay:
        return jsonify({"error": "Missing required fields"}), 400
    
    try:
        # Check if area exists
        check_query = "SELECT id FROM areas WHERE id = %s"
        exists = execute_query(check_query, (area_id,), fetch_one=True)
        
        if not exists:
            return jsonify({"error": "Area not found"}), 404
        
        # Check for duplicate (excluding current area)
        duplicate_query = """
            SELECT id FROM areas 
            WHERE province = %s AND city = %s AND barangay = %s AND id != %s
        """
        duplicate = execute_query(duplicate_query, (province, city, barangay, area_id), fetch_one=True)
        
        if duplicate:
            return jsonify({"error": "Duplicate area already exists"}), 409
        
        # Update area
        update_query = """
            UPDATE areas 
            SET province = %s, city = %s, barangay = %s, zip = %s
            WHERE id = %s
        """
        execute_query(update_query, (province, city, barangay, zip_code, area_id))
        
        return jsonify({
            "success": True,
            "message": "Area updated successfully"
        })
        
    except Exception as e:
        print(f"Error updating area: {e}")
        return jsonify({"error": str(e)}), 500

# =========================
# BULK ADD AREAS (with duplicate handling) - CONVERTED TO MYSQL
# =========================
@app.route("/api/superadmin/areas/bulk", methods=["POST"])
def add_areas_bulk():
    data = request.json
    areas_to_add = data.get("areas", [])
    
    if not areas_to_add:
        return jsonify({"error": "No areas provided"}), 400
    
    added = []
    duplicates = []
    failed = []
    
    for area in areas_to_add:
        province = area.get("province", "").upper().strip()
        city = area.get("city", "").upper().strip()
        barangay = area.get("barangay", "").upper().strip()
        zip_code = area.get("zip", "").strip()
        
        if not province or not city or not barangay:
            failed.append({"area": area, "reason": "Missing fields"})
            continue
        
        try:
            # Check for duplicate
            check_query = """
                SELECT id FROM areas 
                WHERE province = %s AND city = %s AND barangay = %s
            """
            existing = execute_query(check_query, (province, city, barangay), fetch_one=True)
            
            if existing:
                duplicates.append({
                    "province": province, 
                    "city": city, 
                    "barangay": barangay
                })
            else:
                # Insert new area
                insert_query = """
                    INSERT INTO areas (province, city, barangay, zip, created_at)
                    VALUES (%s, %s, %s, %s, NOW())
                """
                area_id = execute_query(insert_query, (province, city, barangay, zip_code))
                added.append({
                    "id": area_id, 
                    "province": province, 
                    "city": city, 
                    "barangay": barangay
                })
        except Exception as e:
            print(f"Error adding area {barangay}: {e}")
            failed.append({"area": area, "reason": str(e)})
    
    return jsonify({
        "success": True,
        "added_count": len(added),
        "added": added,
        "duplicate_count": len(duplicates),
        "duplicates": duplicates,
        "failed_count": len(failed),
        "failed": failed,
        "message": f"Added {len(added)} areas, {len(duplicates)} duplicates skipped, {len(failed)} failed"
    })

# =========================
# GET MISSING BARANGAYS - CONVERTED TO MYSQL
# =========================
@app.route("/api/superadmin/missing-barangays/<city>")
def get_missing_barangays(city):
    """Get list of barangays that are not yet added for a specific city"""
    city_upper = city.upper().strip()
    
    # All valid barangays per city
    barangay_database = {
      "SANTA CRUZ": ["ALIPIT", "BAGUMBAYAN", "BUBUKAL", "CALIOS", "DUHAT", "GATID", "JASAAN", "LABUIN", "MALINAO", "OOGONG", "PAGSAWITAN", "PALASAN", "PATIMBAO", "POBLACION I", "POBLACION II", "POBLACION III", "POBLACION IV", "POBLACION V", "SAN JOSE", "SAN JUAN", "SAN PABLO NORTE", "SAN PABLO SUR", "SANTISIMA CRUZ", "SANTO ANGEL CENTRAL", "SANTO ANGEL NORTE", "SANTO ANGEL SUR"],
       "PAGSANJAN": ["ANIBONG", "BARANGAY I (POBLACION)", "BARANGAY II (POBLACION)", "BIÑAN", "BUBOY", "CABANBANAN", "CALUSICHE", "DINGIN", "LAMBAC", "LAYUGAN", "MAGDAPIO", "MAULAWIN", "PINAGSANJAN", "SABANG", "SAMPALOC", "SAN ISIDRO"],
        "MAGDALENA": ["MALAKING AMBLING", "MUNTING AMBLING", "BUCAL", "BUENAVISTA", "CIGARAS", "IBABANG ATINGAY", "IBABANG BUTNONG", "ILAYANG ATINGAY", "ILAYANG BUTNONG", "POBLACION", "SABANG", "SALASAD", "TIPUNAN", "ALIPIT", "BANAAN", "BALANAC", "BUNGKOL", "BUO", "BURLUNGAN", "HALAYHAYIN", "ILOG", "MALINAO", "MARAVILLA", "TANAWAN"],
        "PILA": ["APLAYA", "BAGONG POOK", "BUKAL", "BULILAN NORTE (POBLACION)", "BULILAN SUR (POBLACION)", "CONCEPCION", "LABUIN", "LINGA", "MASICO", "MOJON", "PANSOL", "PINAGBAYANAN", "SAN ANTONIO", "SAN MIGUEL", "SANTA CLARA NORTE (POBLACION)", "SANTA CLARA SUR (POBLACION)", "TUBUAN"]
    }
    
    # Check if city exists
    if city_upper not in barangay_database:
        return jsonify({
            "error": f"'{city}' is not a valid city",
            "valid_cities": list(barangay_database.keys())
        }), 404
    
    all_barangays = set(barangay_database[city_upper])
    
    try:
        # Get already added barangays from MySQL
        query = """
            SELECT DISTINCT barangay FROM areas 
            WHERE city = %s
        """
        existing_areas = execute_query(query, (city_upper,), fetch=True) or []
        
        added_barangays = set([area['barangay'].upper() for area in existing_areas])
        
        # Find missing barangays (not yet added)
        missing_barangays = list(all_barangays - added_barangays)
        missing_barangays.sort()
        
        # Get province from existing area or default
        province_query = "SELECT province FROM areas WHERE city = %s LIMIT 1"
        province_result = execute_query(province_query, (city_upper,), fetch_one=True)
        province = province_result['province'] if province_result else "LAGUNA"
        
        return jsonify({
            "city": city_upper,
            "province": province,
            "total_barangays": len(all_barangays),
            "added_count": len(added_barangays),
            "missing_count": len(missing_barangays),
            "missing_barangays": missing_barangays
        })
        
    except Exception as e:
        print(f"Error getting missing barangays: {e}")
        return jsonify({"error": str(e)}), 500

# =========================
# GET AREA STATISTICS - NEW ENDPOINT
# =========================
@app.route("/api/superadmin/areas/stats", methods=["GET"])
def get_area_stats():
    try:
        # Get total count of areas
        total_query = "SELECT COUNT(*) as total FROM areas"
        total_result = execute_query(total_query, fetch_one=True)
        total_areas = total_result['total'] if total_result else 0
        
        # Get count per city
        city_query = """
            SELECT city, COUNT(*) as count 
            FROM areas 
            GROUP BY city 
            ORDER BY city
        """
        city_stats = execute_query(city_query, fetch=True) or []
        
        return jsonify({
            "total_areas": total_areas,
            "cities": city_stats
        })
        
    except Exception as e:
        print(f"Error getting area stats: {e}")
        return jsonify({"error": str(e)}), 500



# ===============================
# USER MANAGEMENT PAGE
# ===============================
@app.route("/superadmin/users")
def superadmin_users():
    return render_template("superadmin-users.html")

# ===============================
# GET ALL USERS - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/users", methods=["GET"])
def get_users():
    try:
        # Query users table for customer role (with pending request flag and pending reassignment flag)
        users_query = """
            SELECT u.user_id, u.email, u.status, u.connection_status, 
                   u.first_name, u.last_name, u.middle_name, u.suffix,
                   u.customer_id, u.application_number,
                   CASE WHEN r_pending.request_id IS NOT NULL THEN 1 ELSE 0 END AS has_pending_request,
                   CASE WHEN u.status = 'Terminated' AND (c.installation_status IN ('Pending', 'Slot Assigned', 'Ongoing')) THEN 1 ELSE 0 END AS pending_reassignment
            FROM users u
            LEFT JOIN reconnect_requests r_pending 
                   ON r_pending.user_id = u.user_id AND r_pending.status = 'Pending'
            LEFT JOIN customers c 
                   ON (c.application_number = u.application_number OR c.email = u.email OR c.contract_number = u.contract_number)
            WHERE u.role = 'customer'
        """
        users_data = execute_query(users_query, fetch=True) or []
        
        users = []
        
        for user in users_data:
            user_id = user.get('user_id')
            
            # Build full name from users table fields
            first = user.get('first_name', '')
            middle = user.get('middle_name', '')
            last = user.get('last_name', '')
            suffix = user.get('suffix', '')
            
            full_name_parts = [first]
            if middle:
                full_name_parts.append(middle)
            if last:
                full_name_parts.append(last)
            if suffix:
                full_name_parts.append(suffix)
            
            full_name = " ".join(filter(None, full_name_parts)) if full_name_parts else "N/A"
            
            users.append({
                "user_id": user_id,
                "full_name": full_name,
                "email": user.get('email', ''),
                "status": user.get('status', 'Active'),
                "connection_status": user.get('connection_status', 'Disconnected'),
                "has_pending_request": bool(user.get('has_pending_request', 0)),
                "pending_reassignment": bool(user.get('pending_reassignment', 0))
            })

        
        return jsonify(users)
        
    except Exception as e:
        print("Get users error:", e)
        return jsonify({"error": str(e)}), 500

# ===============================
# UPDATE USER STATUS - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/users/<user_id>/status", methods=["POST"])
def update_user_status(user_id):
    try:
        data = request.get_json()
        new_status = data.get("status")
        balance = data.get("balance", 0)  # KUNIN ANG BALANCE
        
        print(f" Received: user_id={user_id}, new_status={new_status}, balance={balance}")
        
        # TANGGAPIN ANG "Active" at "Inactive"
        if new_status not in ["Active", "Inactive"]:
            return jsonify({"error": "Invalid status. Use 'Active' or 'Inactive'"}), 400
        
        # Check if user exists
        check_query = "SELECT user_id, status, connection_status FROM users WHERE user_id = %s"
        user = execute_query(check_query, (user_id,), fetch_one=True)
        
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        current_status = user.get('status')
        current_connection = user.get('connection_status')
        
        print(f" Current: Status='{current_status}', Connection='{current_connection}'")
        print(f" New: Status='{new_status}', Balance='{balance}'")
        
        # I-UPDATE ANG STATUS, CONNECTION_STATUS, AT BALANCE
        if new_status == "Active":
            update_query = """
                UPDATE users 
                SET status = %s, 
                    connection_status = 'Connected',
                    balance = %s
                WHERE user_id = %s
            """
            execute_query(update_query, (new_status, balance, user_id))
            print(f" User {user_id} ACTIVATED: Status=Active, Connection=Connected, Balance={balance}")
            
            title = "Account Activated"
            message = "Your account has been successfully activated. You can now log in to your dashboard."
            notif_type = "account_activated"
            conn_status = "Connected"
            
        elif new_status == "Inactive":
            update_query = """
                UPDATE users 
                SET status = %s, 
                    connection_status = 'Disconnected',
                    balance = %s
                WHERE user_id = %s
            """
            execute_query(update_query, (new_status, balance, user_id))
            print(f" User {user_id} DEACTIVATED: Status=Inactive, Connection=Disconnected, Balance={balance}")
            
            title = "Account Deactivated"
            message = f"Your account has been deactivated. Your internet connection has been disconnected. Balance: ₱{balance:.2f}"
            notif_type = "account_deactivated"
            conn_status = "Disconnected"
        
        # ========== CREATE NOTIFICATION ==========
        try:
            user_query = """
                SELECT email, first_name, last_name 
                FROM users 
                WHERE user_id = %s
            """
            user_details = execute_query(user_query, (user_id,), fetch_one=True)
            
            notification_id = int(datetime.now().timestamp() * 1000)
            user_name = f"{user_details.get('first_name', '')} {user_details.get('last_name', '')}".strip() or 'User'
            user_email = user_details.get('email', '')
            
            notif_query = """
                INSERT INTO user_notifications 
                (id, title, message, type, relatedId, user_id, user_email, user_name, connection_status, timestamp, read_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            execute_query(notif_query, (
                notification_id,
                title,
                message,
                notif_type,
                user_id,
                user_id,
                user_email,
                user_name,
                conn_status,
                ph_now_iso(),
                0
            ))
            print(f" Notification sent to user {user_id}")
            
        except Exception as notif_error:
            print(f" Error creating notification: {notif_error}")
        
        # I-VERIFY ANG UPDATE
        verify_query = "SELECT status, connection_status, balance FROM users WHERE user_id = %s"
        verified = execute_query(verify_query, (user_id,), fetch_one=True)
        print(f" VERIFIED: Status='{verified.get('status')}', Connection='{verified.get('connection_status')}', Balance='{verified.get('balance')}'")
        
        return jsonify({
            "message": f"User {user_id} updated to {new_status}",
            "status": new_status,
            "connection_status": verified.get('connection_status', 'Disconnected'),
            "balance": verified.get('balance', 0)
        })
        
    except Exception as e:
        print(" Update user status error:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ===============================
# UPDATE USER CONNECTION STATUS - WITH NOTIFICATION (XAMPP/MYSQL VERSION)
# ===============================
@app.route("/api/superadmin/users/<user_id>/connection", methods=["POST"])
def update_user_connection(user_id):
    try:
        data = request.get_json()
        new_status = data.get("connection_status")
        
        if new_status not in ["Connected", "Disconnected"]:
            return jsonify({"error": "Invalid connection status"}), 400
        
        # Check if user exists and get details
        check_query = """
            SELECT user_id, email, first_name, last_name, connection_status 
            FROM users 
            WHERE user_id = %s
        """
        user = execute_query(check_query, (user_id,), fetch_one=True)
        
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        # Update the connection status
        update_query = "UPDATE users SET connection_status = %s WHERE user_id = %s"
        execute_query(update_query, (new_status, user_id))
        
        # ========== CREATE NOTIFICATION FOR THE USER ==========
        try:
            notification_id = int(datetime.now().timestamp() * 1000)
            user_name = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip() or 'User'
            user_email = user.get('email', '')
            
            # Create message based on new status
            if new_status == "Connected":
                title = "Internet Connection Activated"
                message = f"Your internet connection has been successfully activated. You can now enjoy our high-speed internet service."
                notif_type = "connection_activated"
            else:
                title = "Internet Connection Deactivated"
                message = f"Your internet connection has been deactivated. Please contact support for assistance."
                notif_type = "connection_deactivated"
            
            # Insert notification into user_notifications table
            notif_query = """
                INSERT INTO user_notifications 
                (id, title, message, type, relatedId, user_id, user_email, user_name, 
                 connection_status, timestamp, read_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            execute_query(notif_query, (
                notification_id,
                title,
                message,
                notif_type,
                user_id,  # relatedId
                user_id,  # user_id
                user_email,
                user_name,
                new_status,
                ph_now_iso(),
                0  # unread
            ))
            print(f" Notification sent to user {user_id} about connection {new_status}")
            
        except Exception as notif_error:
            print(f" Error creating user notification: {notif_error}")
            import traceback
            traceback.print_exc()
        
        return jsonify({"message": f"User {user_id} connection updated to {new_status}"})
        
    except Exception as e:
        print("Update user connection error:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ===============================
# GET SINGLE USER DETAILS - NEW ENDPOINT
# ===============================
@app.route("/api/superadmin/users/<user_id>", methods=["GET"])
def get_user_details(user_id):
    try:
        query = """
            SELECT user_id, email, status, connection_status, 
                   first_name, last_name, middle_name, suffix,
                   contact_number, address, created_at
            FROM users 
            WHERE user_id = %s
        """
        user = execute_query(query, (user_id,), fetch_one=True)
        
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        # Build full name
        first = user.get('first_name', '')
        middle = user.get('middle_name', '')
        last = user.get('last_name', '')
        suffix = user.get('suffix', '')
        
        full_name_parts = [first]
        if middle:
            full_name_parts.append(middle)
        if last:
            full_name_parts.append(last)
        if suffix:
            full_name_parts.append(suffix)
        
        full_name = " ".join(filter(None, full_name_parts)) if full_name_parts else "N/A"
        
        return jsonify({
            "user_id": user.get('user_id'),
            "full_name": full_name,
            "email": user.get('email', ''),
            "status": user.get('status', 'Active'),
            "connection_status": user.get('connection_status', 'Disconnected'),
            "contact_number": user.get('contact_number', ''),
            "address": user.get('address', ''),
            "created_at": user.get('created_at', '')
        })
        
    except Exception as e:
        print("Get user details error:", e)
        return jsonify({"error": str(e)}), 500


# ===============================
# DELETE TERMINATED USER - PERMANENT DELETION
# ===============================
@app.route("/api/superadmin/users/<user_id>/delete", methods=["DELETE"])
def delete_terminated_user(user_id):
    """Permanently delete a terminated user and all related data"""
    try:
        # Check if user exists and is terminated
        check_query = "SELECT user_id, status, application_number FROM users WHERE user_id = %s"
        user = execute_query(check_query, (user_id,), fetch_one=True)
        
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        if user.get('status') != 'Terminated':
            return jsonify({"error": "User is not terminated. Only terminated users can be deleted."}), 400
        
        application_number = user.get('application_number')
        
        print(f" Deleting terminated user: {user_id}")
        print(f" Application number: {application_number}")
        
        # ========== DELETE CONTRACT (if exists) ==========
        if application_number:
            # Get contract number from customers table
            contract_query = "SELECT contract_number FROM customers WHERE application_number = %s"
            contract = execute_query(contract_query, (application_number,), fetch_one=True)
            
            if contract and contract.get('contract_number'):
                contract_number = contract.get('contract_number')
                # Delete from contracts table
                delete_contract = "DELETE FROM contracts WHERE contract_number = %s"
                execute_query(delete_contract, (contract_number,))
                print(f" Deleted contract: {contract_number}")
        
        # ========== DELETE CUSTOMER (if exists) ==========
        if application_number:
            delete_customer = "DELETE FROM customers WHERE application_number = %s"
            execute_query(delete_customer, (application_number,))
            print(f" Deleted customer record for: {application_number}")
        
        # ========== DELETE APPLICATION (if exists) ==========
        if application_number:
            delete_application = "DELETE FROM applications WHERE application_number = %s"
            execute_query(delete_application, (application_number,))
            print(f" Deleted application: {application_number}")
        
        # ========== DELETE USER NOTIFICATIONS ==========
        delete_notifications = "DELETE FROM user_notifications WHERE user_id = %s"
        execute_query(delete_notifications, (user_id,))
        print(f" Deleted notifications for user: {user_id}")
        
        # ========== DELETE USER ==========
        delete_user = "DELETE FROM users WHERE user_id = %s"
        execute_query(delete_user, (user_id,))
        print(f" Deleted user: {user_id}")
        
        return jsonify({
            "success": True,
            "message": f"User {user_id} has been permanently deleted."
        })
        
    except Exception as e:
        print(f" Error deleting user: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# GET PENDING REQUEST DETAILS FOR A USER
# ===============================
@app.route("/api/superadmin/users/<user_id>/pending-request", methods=["GET"])
def get_pending_request(user_id):
    try:
        query = """
            SELECT request_id, user_id, application_number, change_plan,
                   new_plan_id, new_plan_name, current_plan_name, 
                   current_plan_speed, current_plan_price,
                   first_name, middle_name, last_name, suffix,
                   contact_number, email, address, status, created_at
            FROM reconnect_requests
            WHERE user_id = %s AND status = 'Pending'
            ORDER BY created_at DESC
            LIMIT 1
        """
        request_data = execute_query(query, (user_id,), fetch_one=True)
        
        if not request_data:
            return jsonify({"error": "No pending request found"}), 404
        
        full_name = " ".join(filter(None, [
            request_data.get('first_name', ''),
            request_data.get('middle_name', ''),
            request_data.get('last_name', ''),
            request_data.get('suffix', '')
        ]))
        
        return jsonify({
            "request_id": request_data.get('request_id'),
            "user_id": request_data.get('user_id'),
            "application_number": request_data.get('application_number'),
            "change_plan": bool(request_data.get('change_plan')),
            "new_plan_id": request_data.get('new_plan_id'),
            "new_plan_name": request_data.get('new_plan_name'),
            "current_plan_name": request_data.get('current_plan_name'),
            "current_plan_speed": request_data.get('current_plan_speed'),
            "current_plan_price": request_data.get('current_plan_price'),
            "full_name": full_name or "N/A",
            "contact_number": request_data.get('contact_number', ''),
            "email": request_data.get('email', ''),
            "address": request_data.get('address', ''),
            "status": request_data.get('status'),
            "created_at": request_data.get('created_at').isoformat() if request_data.get('created_at') else None
        })
        
    except Exception as e:
        print("Get pending request error:", e)
        return jsonify({"error": str(e)}), 500


# ===============================
# APPROVE RECONNECT / CHANGE-PLAN REQUEST - FIXED
# ===============================
@app.route("/api/superadmin/requests/<request_id>/approve", methods=["POST"])
def approve_reconnect_request(request_id):
    try:
        req_query = """
            SELECT request_id, user_id, application_number, change_plan,
                   new_plan_id, new_plan_name, status
            FROM reconnect_requests
            WHERE request_id = %s
        """
        req = execute_query(req_query, (request_id,), fetch_one=True)
        
        if not req:
            return jsonify({"error": "Request not found"}), 404
        
        if req.get('status') != 'Pending':
            return jsonify({"error": "Request is not pending"}), 400
        
        user_id = req.get('user_id')
        application_number = req.get('application_number')
        change_plan = req.get('change_plan')
        new_plan_id = req.get('new_plan_id')
        new_plan_name = req.get('new_plan_name')
        
        # 1. GET USER DETAILS - CHECK CURRENT STATUS
        user_info = execute_query(
            "SELECT status, application_number, contract_number, email, first_name, last_name FROM users WHERE user_id = %s", 
            (user_id,), fetch_one=True
        )
        
        if not user_info:
            return jsonify({"error": "User not found"}), 404
        
        # CHECK KUNG ANO ANG CURRENT STATUS NG USER
        current_user_status = user_info.get('status', '')
        req_app_num = application_number or user_info.get("application_number")
        req_contract_num = user_info.get("contract_number")
        user_email = user_info.get("email") or ""
        user_full_name = f"{user_info.get('first_name', '')} {user_info.get('last_name', '')}".strip() or "User"
        
        # KUNG INACTIVE LANG ANG USER, DI NA KAILANGAN NG SLOT REASSIGNMENT
        is_inactive_only = (current_user_status == 'Inactive')
        
        customer_city = None
        
        if req_app_num:
            cust_info = execute_query(
                "SELECT contract_number, city, barangay, first_name, last_name, email FROM customers WHERE application_number = %s", 
                (req_app_num,), fetch_one=True
            )
            if cust_info:
                if not req_contract_num:
                    req_contract_num = cust_info.get("contract_number")
                customer_city = cust_info.get("city") or cust_info.get("barangay")
                if not user_email:
                    user_email = cust_info.get("email") or ""

        # PAG INACTIVE LANG: DI NA KAILANGAN NG NAP BOX SLOT CHECK
        if is_inactive_only:
            print(f" User {user_id} is INACTIVE only - reactivating directly...")
            
            # I-ACTIVATE AGAD ANG USER
            update_user_query = """
                UPDATE users 
                SET status = 'Active', 
                    connection_status = 'Connected',
                    has_pending_reconnect = 0
                WHERE user_id = %s
            """
            execute_query(update_user_query, (user_id,))
            print(f" User {user_id} reactivated: Active / Connected")
            
            if req_app_num:
                execute_query("UPDATE customers SET status = 'Approved', installation_status = 'Installed' WHERE application_number = %s", (req_app_num,))
                execute_query("UPDATE applications SET installation_status = 'Installed' WHERE application_number = %s", (req_app_num,))
            
            user_connection_status_for_notif = "Connected"
            user_notif_message = "Your reconnection request has been approved. Your account is now active and connected."
            is_slot_available = True
            
        else:
            # ORIGINAL LOGIC PARA SA TERMINATED USERS
            # CHECK DATING NAP BOX SLOT
            former_slot = None
            if req_app_num or req_contract_num:
                former_slot_query = """
                    SELECT id, status, application_number, contract_number, customer_name 
                    FROM napbox_slots 
                    WHERE (application_number IS NOT NULL AND application_number != '' AND application_number = %s)
                       OR (contract_number IS NOT NULL AND contract_number != '' AND contract_number = %s)
                    LIMIT 1
                """
                former_slot = execute_query(former_slot_query, (req_app_num, req_contract_num), fetch_one=True)

            is_slot_available = False
            if former_slot:
                if former_slot.get('status') == 'available' and (former_slot.get('application_number') == req_app_num or former_slot.get('contract_number') == req_contract_num):
                    is_slot_available = True

            if is_slot_available:
                # CASE A: Former slot is still available!
                update_slot_query = """
                    UPDATE napbox_slots 
                    SET status = 'occupied', 
                        updated_at = NOW()
                    WHERE id = %s
                """
                execute_query(update_slot_query, (former_slot['id'],))
                print(f" Former NAP Box slot #{former_slot['id']} set back to 'occupied' for user {user_id}")

                update_user_query = """
                    UPDATE users 
                    SET status = 'Active', 
                        connection_status = 'Connected',
                        has_pending_reconnect = 0
                    WHERE user_id = %s
                """
                execute_query(update_user_query, (user_id,))
                print(f" User {user_id} reactivated: Active / Connected")

                if req_app_num:
                    execute_query("UPDATE customers SET status = 'Approved', installation_status = 'Installed' WHERE application_number = %s", (req_app_num,))
                    execute_query("UPDATE applications SET installation_status = 'Installed' WHERE application_number = %s", (req_app_num,))

                user_notif_message = "Your reconnection request has been approved. Your account is now active."
                user_connection_status_for_notif = "Connected"
                
            else:
                # CASE B: Former slot is occupied by a NEW owner or missing!
                print(f" Former NAP Box slot for user {user_id} is occupied by another customer or missing.")
                
                # KUNG TERMINATED ANG USER, KEEP AS TERMINATED
                update_user_query = """
                    UPDATE users 
                    SET status = 'Terminated', 
                        connection_status = 'Disconnected',
                        has_pending_reconnect = 1
                    WHERE user_id = %s
                """
                execute_query(update_user_query, (user_id,))

                if req_app_num:
                    execute_query("UPDATE customers SET installation_status = 'Pending' WHERE application_number = %s", (req_app_num,))
                    execute_query("UPDATE applications SET installation_status = 'Pending' WHERE application_number = %s", (req_app_num,))
                    print(f" Customer {req_app_num} installation_status set to 'Pending' for technician reassignment.")

                if customer_city:
                    create_technician_notifications_by_area(
                        area=customer_city,
                        title="Slot Reassignment Needed",
                        message=f"Reconnection request for {user_full_name} (App #{req_app_num}) was approved, but their former NAP Box slot is occupied by a new owner. Please reassign a new NAP Box slot for this customer.",
                        notif_type="reassign_slot",
                        related_id=request_id,
                        application_number=req_app_num,
                        customer_name=user_full_name
                    )
                    print(f" Technician notification sent for area: {customer_city}")

                user_notif_message = "Your reconnection request has been approved by the superadmin. A technician will reassign a new NAP Box slot for your connection soon."
                user_connection_status_for_notif = "Disconnected"

        # 2. IF CHANGE PLAN, UPDATE CUSTOMERS TABLE
        if change_plan and application_number:
            plan_query = "SELECT name, speed, price FROM plans WHERE id = %s"
            plan_details = execute_query(plan_query, (new_plan_id,), fetch_one=True)
            
            if plan_details:
                update_customer_query = """
                    UPDATE customers 
                    SET plan = %s, plan_speed = %s, plan_price = %s
                    WHERE application_number = %s
                """
                execute_query(update_customer_query, (
                    plan_details.get('name'),
                    plan_details.get('speed'),
                    str(plan_details.get('price')),
                    application_number
                ))
                print(f" Customer {application_number} plan updated to {plan_details.get('name')} ({plan_details.get('speed')} - ₱{plan_details.get('price')})")
            else:
                update_customer_query = """
                    UPDATE customers 
                    SET plan = %s
                    WHERE application_number = %s
                """
                execute_query(update_customer_query, (new_plan_name, application_number))
                print(f" Plan id {new_plan_id} not found in plans table — only updated plan name to {new_plan_name}")
        
        # 3. MARK REQUEST AS APPROVED
        update_request_query = """
            UPDATE reconnect_requests 
            SET status = 'Approved'
            WHERE request_id = %s
        """
        execute_query(update_request_query, (request_id,))
        
        # 4. NOTIFY THE USER
        try:
            notification_id = int(datetime.now().timestamp() * 1000)
            title = "Request Approved"
            if change_plan:
                message = f"Your reconnection and plan change request has been approved. Your new plan: {new_plan_name}."
            else:
                message = user_notif_message
            
            notif_query = """
                INSERT INTO user_notifications 
                (id, title, message, type, relatedId, user_id, user_email, user_name, connection_status, timestamp, read_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            execute_query(notif_query, (
                notification_id, title, message, "request_approved",
                request_id, user_id, user_email, user_full_name,
                user_connection_status_for_notif, ph_now_iso(), 0
            ))
        except Exception as notif_error:
            print(f" Error creating notification: {notif_error}")
        
        return jsonify({
            "success": True,
            "message": "Request approved successfully",
            "user_id": user_id,
            "change_plan": bool(change_plan)
        })
        
    except Exception as e:
        print(" Approve request error:", e)
        import traceback    
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# REJECT RECONNECT REQUEST
# ===============================
@app.route("/api/superadmin/requests/<request_id>/reject", methods=["POST"])
def reject_reconnect_request(request_id):
    try:
        req_query = """
            SELECT request_id, user_id, application_number, change_plan,
                   new_plan_name, status
            FROM reconnect_requests
            WHERE request_id = %s
        """
        req = execute_query(req_query, (request_id,), fetch_one=True)
        
        if not req:
            return jsonify({"error": "Request not found"}), 404
        
        if req.get('status') != 'Pending':
            return jsonify({"error": "Request is not pending"}), 400
        
        user_id = req.get('user_id')
        application_number = req.get('application_number')
        
        # 1. I-UPDATE ANG USER - REMOVE PENDING FLAG
        update_user_query = """
            UPDATE users 
            SET has_pending_reconnect = 0
            WHERE user_id = %s
        """
        execute_query(update_user_query, (user_id,))
        print(f" User {user_id} pending reconnect flag cleared")
        
        # 2. MARK REQUEST AS REJECTED
        update_request_query = """
            UPDATE reconnect_requests 
            SET status = 'Rejected'
            WHERE request_id = %s
        """
        execute_query(update_request_query, (request_id,))
        print(f" Request {request_id} marked as Rejected")
        
        # 3. NOTIFY THE USER
        try:
            user_details_query = "SELECT email, first_name, last_name FROM users WHERE user_id = %s"
            user_details = execute_query(user_details_query, (user_id,), fetch_one=True)
            
            notification_id = int(datetime.now().timestamp() * 1000)
            user_name = f"{user_details.get('first_name', '')} {user_details.get('last_name', '')}".strip() or 'User'
            user_email = user_details.get('email', '')
            
            title = "Reconnection Request Rejected"
            message = "Your reconnection request has been rejected. Please contact support for more information."
            
            notif_query = """
                INSERT INTO user_notifications 
                (id, title, message, type, relatedId, user_id, user_email, user_name, connection_status, timestamp, read_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            execute_query(notif_query, (
                notification_id, title, message, "request_rejected",
                request_id, user_id, user_email, user_name,
                "Disconnected", ph_now_iso(), 0
            ))
        except Exception as notif_error:
            print(f" Error creating notification: {notif_error}")
        
        return jsonify({
            "success": True,
            "message": "Request rejected successfully",
            "user_id": user_id
        })
        
    except Exception as e:
        print(" Reject request error:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    

# ===============================
# GLOBAL STATISTICS - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/statistics", methods=["GET"])
def superadmin_statistics():
    try:
        # ========== GET APPLICATIONS COUNT ==========
        # Count total applications (excluding rejected? keep original logic)
        apps_query = "SELECT COUNT(*) as total FROM applications"
        apps_result = execute_query(apps_query, fetch_one=True)
        total_applicants = apps_result['total'] if apps_result else 0
        
        # ========== GET ADMINS COUNT ==========
        admins_query = "SELECT COUNT(*) as total FROM admins"
        admins_result = execute_query(admins_query, fetch_one=True)
        total_admins = admins_result['total'] if admins_result else 0
        
        # ========== GET ACTIVE ADMINS COUNT ==========
        active_admins_query = "SELECT COUNT(*) as total FROM admins WHERE status = 'Active'"
        active_admins_result = execute_query(active_admins_query, fetch_one=True)
        active_admins = active_admins_result['total'] if active_admins_result else 0
        
        # ========== GET POPULAR PLANS ==========
        plans_query = """
            SELECT plan, COUNT(*) as count 
            FROM applications 
            WHERE plan IS NOT NULL AND plan != ''
            GROUP BY plan 
            ORDER BY count DESC
        """
        popular_plans_result = execute_query(plans_query, fetch=True) or []
        popular_plans = {}
        for row in popular_plans_result:
            popular_plans[row['plan']] = row['count']
        
        # ========== GET COVERAGE GROWTH (applications per city/area) ==========
        coverage_query = """
            SELECT city, COUNT(*) as count 
            FROM applications 
            WHERE city IS NOT NULL AND city != ''
            GROUP BY city 
            ORDER BY count DESC
        """
        coverage_result = execute_query(coverage_query, fetch=True) or []
        coverage_growth = {}
        for row in coverage_result:
            coverage_growth[row['city']] = row['count']
        
        # ========== GET TOTAL ACTIVE APPLICANTS (non-rejected) - optional ==========
        active_applicants_query = """
            SELECT COUNT(*) as total 
            FROM applications 
            WHERE status != 'Rejected'
        """
        active_applicants_result = execute_query(active_applicants_query, fetch_one=True)
        total_active_applicants = active_applicants_result['total'] if active_applicants_result else 0
        
        return jsonify({
            "total_applicants": total_applicants,
            "total_active_applicants": total_active_applicants,
            "popular_plans": popular_plans,
            "coverage_growth": coverage_growth,
            "total_admins": total_admins,
            "active_admins": active_admins
        })
        
    except Exception as e:
        print(f"Error in superadmin_statistics: {e}")
        return jsonify({"error": str(e)}), 500



# ===============================
# SUPERADMIN - EXPORT ALL CUSTOMERS DATA TO EXCEL
# ===============================
@app.route("/api/superadmin/export-all-customers-excel", methods=["GET"])
def superadmin_export_all_customers_excel():
    """Export all customers data to Excel (for superadmin - all areas) with auto column width"""
    
    try:
        # Get filters
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        area_filter = request.args.get("area")
        
        print(f" Export filters - Start: {start_date}, End: {end_date}, Area: {area_filter}")
        
        # Build query for customers
        query = """
            SELECT 
                application_number, contract_number, first_name, last_name,
                middle_name, suffix, email, mobile, address, barangay,
                city, province, zip, plan, plan_speed, plan_price,
                status, installation_status, 
                DATE(approval_date) as approval_date,
                DATE(date_installed) as date_installed,
                billing_date,
                DATE(created_at) as created_at
            FROM customers 
            WHERE 1=1
        """
        params = []
        
        # Add area filter
        if area_filter and area_filter != "":
            query += " AND city = %s"
            params.append(area_filter)
            print(f" Filtering by area: {area_filter}")
        
        # Add date filters
        if start_date and end_date:
            query += " AND DATE(approval_date) >= DATE(%s) AND DATE(approval_date) <= DATE(%s)"
            params.append(start_date)
            params.append(end_date)
            print(f" Date range: {start_date} to {end_date}")
        elif start_date:
            query += " AND DATE(approval_date) >= DATE(%s)"
            params.append(start_date)
            print(f" Start date: {start_date}")
        elif end_date:
            query += " AND DATE(approval_date) <= DATE(%s)"
            params.append(end_date)
            print(f" End date: {end_date}")
        
        query += " ORDER BY approval_date DESC"
        
        customers = execute_query(query, params, fetch=True) or []
        
        print(f" Found {len(customers)} customers matching filters")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        area_name = area_filter if area_filter else "all_areas"
        ws.title = f"Customers_{area_name}"
        
        # Define headers
        headers = [
            "Application Number", "Contract Number", "Customer Name", "First Name",
            "Last Name", "Middle Name", "Suffix", "Email", "Mobile", "Plan",
            "Speed", "Price", "Status", "Installation Status", "Address",
            "Barangay", "City", "Province", "Zip Code", "Approval Date",
            "Date Installed", "Billing Date", "Created At"
        ]
        
        # Add headers with styling
        from openpyxl.styles import Font, Alignment, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0047ab", end_color="0047ab", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row_idx, customer in enumerate(customers, 2):
            full_name = f"{customer.get('first_name', '')} {customer.get('last_name', '')}".strip()
            
            # Get date values
            approval_date = customer.get('approval_date')
            date_installed = customer.get('date_installed')
            created_at = customer.get('created_at')
            
            ws.cell(row=row_idx, column=1, value=str(customer.get('application_number', '')))
            ws.cell(row=row_idx, column=2, value=customer.get('contract_number', ''))
            ws.cell(row=row_idx, column=3, value=full_name)
            ws.cell(row=row_idx, column=4, value=customer.get('first_name', ''))
            ws.cell(row=row_idx, column=5, value=customer.get('last_name', ''))
            ws.cell(row=row_idx, column=6, value=customer.get('middle_name', ''))
            ws.cell(row=row_idx, column=7, value=customer.get('suffix', ''))
            ws.cell(row=row_idx, column=8, value=customer.get('email', ''))
            ws.cell(row=row_idx, column=9, value=customer.get('mobile', ''))
            ws.cell(row=row_idx, column=10, value=customer.get('plan', ''))
            ws.cell(row=row_idx, column=11, value=customer.get('plan_speed', ''))
            ws.cell(row=row_idx, column=12, value=customer.get('plan_price', ''))
            ws.cell(row=row_idx, column=13, value=customer.get('status', 'Approved'))
            ws.cell(row=row_idx, column=14, value=customer.get('installation_status', 'Pending'))
            ws.cell(row=row_idx, column=15, value=customer.get('address', ''))
            ws.cell(row=row_idx, column=16, value=customer.get('barangay', ''))
            ws.cell(row=row_idx, column=17, value=customer.get('city', ''))
            ws.cell(row=row_idx, column=18, value=customer.get('province', ''))
            ws.cell(row=row_idx, column=19, value=customer.get('zip', ''))
            ws.cell(row=row_idx, column=20, value=str(approval_date) if approval_date else '')
            ws.cell(row=row_idx, column=21, value=str(date_installed) if date_installed else '')
            ws.cell(row=row_idx, column=22, value=customer.get('billing_date', ''))
            ws.cell(row=row_idx, column=23, value=str(created_at) if created_at else '')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        from datetime import datetime
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        date_now = ph_now().strftime("%Y%m%d_%H%M%S")
        filename = f"customers_{area_name}_{date_now}.xlsx"
        
        from flask import send_file
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error in superadmin_export_all_customers_excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# GET SUPERADMIN PROFILE (XAMPP/MYSQL) - SECURE VERSION
# ===============================
@app.route("/api/superadmin/profile", methods=["GET"])
def get_superadmin_profile():
    try:
        tab_id = request.args.get('tab_id', '')
        session_data = session.get(f"admin_{tab_id}")
        
        if not session_data or session_data.get('user_type') != 'superadmin':
            print(f" Unauthorized access attempt. tab_id: {tab_id}")
            return jsonify({"error": "Unauthorized"}), 403
        
        username = session_data.get('username')
        
        if not username:
            return jsonify({"error": "Not logged in"}), 401
        
        print(f" Getting profile for username: {username} (tab_id: {tab_id})")
        
        query = "SELECT username, name, email, area, status, ga_enabled, ga_secret FROM superadmins WHERE username = %s"
        superadmin = execute_query(query, (username,), fetch_one=True)
        
        if superadmin:
            display_name = session_data.get('admin_display_name') or superadmin.get("name") or superadmin.get("username")
            
            # I-CONVERT SA BOOL
            ga_enabled = bool(superadmin.get("ga_enabled", 0))
            ga_secret = superadmin.get("ga_secret")
            
            # I-UPDATE ANG SESSION KUNG MAGKAIBA
            if session_data.get('ga_enabled') != ga_enabled:
                session_data['ga_enabled'] = ga_enabled
                session[f"admin_{tab_id}"] = session_data
            
            # Generate QR code URI if not enabled
            ga_setup_uri = None
            if not ga_enabled:
                if not ga_secret:
                    # Generate new secret
                    ga_secret = generate_ga_secret()
                    execute_query("UPDATE superadmins SET ga_secret = %s WHERE username = %s", (ga_secret, username))
                ga_setup_uri = generate_ga_provisioning_uri(username, ga_secret)
            
            return jsonify({
                "username": superadmin.get("username"),
                "name": display_name,
                "email": superadmin.get("email", ""),
                "area": superadmin.get("area", "Sta. Cruz"),
                "status": superadmin.get("status", "Active"),
                "ga_enabled": ga_enabled,  # BOOL NA ITO
                "ga_secret": ga_secret,
                "ga_setup_uri": ga_setup_uri
            })
        
        # Fallback
        return jsonify({
            "username": username,
            "name": username,
            "email": "",
            "area": "Sta. Cruz",
            "status": "Active",
            "ga_enabled": False,
            "ga_secret": None,
            "ga_setup_uri": None
        })
        
    except Exception as e:
        print(f"Get superadmin profile error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ===============================
# UPDATE SUPERADMIN PROFILE - SECURE VERSION
# ===============================
@app.route("/api/update-superadmin-profile", methods=["POST"])
def update_superadmin_profile():
    try:
        data = request.get_json()
        name = data.get("name")
        email = data.get("email")
        password = data.get("password")
        current_password = data.get("current_password")
        area = data.get("area")
        
        # 1. KUNIN ANG TAB_ID MULA SA REQUEST (or from session)
        tab_id = request.args.get('tab_id', '') or data.get('tab_id', '')
        
        # 2. I-VERIFY ANG SESSION GAMIT ANG TAB_ID
        session_data = session.get(f"admin_{tab_id}")
        
        if not session_data or session_data.get('user_type') != 'superadmin':
            print(f" Unauthorized update attempt. tab_id: {tab_id}")
            return jsonify({"error": "Unauthorized"}), 403
        
        # 3. KUNIN ANG USERNAME MULA SA SESSION
        username = session_data.get('username')
        
        print(f" Updating profile for username: {username} (tab_id: {tab_id})")
        
        if not username:
            return jsonify({"error": "Not logged in"}), 401

        # 4. I-VERIFY KUNG EXIST ANG USER SA DATABASE
        check_query = "SELECT username FROM superadmins WHERE username = %s"
        check_result = execute_query(check_query, (username,), fetch_one=True)
        
        if not check_result:
            print(f" Username '{username}' not found in database!")
            return jsonify({"error": f"User '{username}' not found in database"}), 404
        
        if password and len(password) >= 8:
            current_user = execute_query("SELECT password FROM superadmins WHERE username = %s", (username,), fetch_one=True)
            if not current_user:
                return jsonify({"error": "Account not found"}), 404
            stored_password = current_user.get('password')
            if not current_password or not verify_password(stored_password, current_password):
                return jsonify({"error": "Current password is incorrect"}), 400

        # 5. BUILD UPDATE QUERY
        update_fields = []
        params = []
        
        if name:
            update_fields.append("name = %s")
            params.append(name)
        if email:
            update_fields.append("email = %s")
            params.append(email)
        if area:
            update_fields.append("area = %s")
            params.append(area)
        if password and len(password) >= 8:
            update_fields.append("password = %s")
            params.append(hash_password(password))
        
        if not update_fields:
            return jsonify({"error": "No fields to update"}), 400
        
        params.append(username)
        update_query = f"UPDATE superadmins SET {', '.join(update_fields)} WHERE username = %s"
        
        print(f" UPDATE QUERY: {update_query}")
        print(f" PARAMS: {params}")
        
        # 6. EXECUTE UPDATE
        result = execute_query(update_query, params)
        print(f" Update result: {result}")
        
        # 7. I-UPDATE ANG SESSION DISPLAY NAME KUNG MAY NAME CHANGE
        if name:
            # I-update ang session data
            session[f"admin_{tab_id}"]['admin_display_name'] = name
        
        return jsonify({
            "success": True,
            "message": "Profile updated successfully"
        }), 200

    except Exception as e:
        print(f"Update error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# SERVE SUPERADMIN PROFILE PAGE
# ===============================
@app.route("/superadmin/profile")
def superadmin_profile():
    """Render superadmin profile page with GA info"""
    tab_id = request.args.get('tab_id') or session.get('active_tab')
    
    if not tab_id:
        flash("Invalid session.", "warning")
        return redirect(url_for("login"))
    
    session_data = session.get(f"admin_{tab_id}")
    
    if not session_data or session_data.get('user_type') != 'superadmin':
        flash("Unauthorized access.", "danger")
        return redirect(url_for("login"))
    
    username = session_data.get('username')
    
    # Get GA info from database
    query = "SELECT ga_enabled, ga_secret FROM superadmins WHERE username = %s"
    user_data = execute_query(query, (username,), fetch_one=True)
    
    ga_enabled = bool(user_data.get("ga_enabled")) if user_data else False
    ga_secret = user_data.get("ga_secret") if user_data else None
    
    # Generate QR code if not enabled
    ga_setup_uri = None
    if not ga_enabled:
        if not ga_secret:
            ga_secret = generate_ga_secret()
            execute_query("UPDATE superadmins SET ga_secret = %s WHERE username = %s", (ga_secret, username))
        ga_setup_uri = generate_ga_provisioning_uri(username, ga_secret)
    
    return render_template(
        "superadmin-profile.html",
        ga_enabled=ga_enabled,
        ga_secret=ga_secret,
        ga_setup_uri=ga_setup_uri
    )

import os
from werkzeug.utils import secure_filename
from datetime import datetime

# ==================== SHARED UPLOADS CONFIGURATION ====================
# Shared folder sa labas ng project (sa XAMPP htdocs)
SHARED_UPLOADS_BASE = r"C:\xampp\htdocs\cablevision_uploads"

# Plans subfolder
UPLOAD_FOLDER_PLANS = os.path.join(SHARED_UPLOADS_BASE, 'plans')
ALLOWED_EXTENSIONS_PLANS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# Create upload folder if not exists
os.makedirs(UPLOAD_FOLDER_PLANS, exist_ok=True)

# NOTE: Ang /shared-uploads route ay nasa channel logos na
# HUWAG nang i-duplicate dito!

def allowed_plan_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS_PLANS

def validate_plan_image_orientation(image_file):
    """Ensure uploaded plan image is landscape (width > height)."""
    if not image_file:
        return False, "Image file is required"

    from PIL import Image

    try:
        image_file.stream.seek(0)
        with Image.open(image_file.stream) as img:
            width, height = img.size
            if width <= height:
                return False, "Image must be landscape (width greater than height)"
            return True, None
    except Exception as e:
        print(f"Plan image orientation validation failed: {e}")
        return False, "Failed to read image"


def save_plan_image(image_file):
    """Save plan image to shared folder and return URL path"""
    if not image_file or not allowed_plan_file(image_file.filename):
        return None
    
    image_file.stream.seek(0)
    filename = secure_filename(f"plan_{int(datetime.now().timestamp())}_{image_file.filename}")
    # Store relative path from shared uploads base
    relative_path = os.path.join('plans', filename)
    full_path = os.path.join(SHARED_UPLOADS_BASE, relative_path)
    
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    image_file.save(full_path)
    
    # Return URL path (for HTML access via shared-uploads route)
    return f"/shared-uploads/{relative_path.replace(os.sep, '/')}"

def delete_plan_image(image_url):
    """Delete plan image file from shared folder if exists"""
    if image_url and image_url.startswith('/shared-uploads/'):
        # Extract relative path from URL
        relative_path = image_url.replace('/shared-uploads/', '')
        full_path = os.path.join(SHARED_UPLOADS_BASE, relative_path.replace('/', os.sep))
        if os.path.exists(full_path):
            os.remove(full_path)
            print(f"Deleted plan image: {full_path}")

# ===============================
# SUPERADMIN PLANS PAGE
# ===============================
@app.route("/superadmin/plans")
def superadmin_plans():
    return render_template("superadmin-plans.html")

# ===============================
# GET ALL PLANS - SHARED UPLOADS
# ===============================
@app.route("/api/superadmin/plans", methods=["GET"])
def get_plans():
    try:
        query = """
            SELECT id, name, speed, price, image_path, created_at
            FROM plans 
            ORDER BY id DESC
        """
        plans = execute_query(query, fetch=True) or []
        
        plan_list = []
        for plan in plans:
            image_path = plan.get('image_path', '')
            
            # If already Cloudinary URL, keep it
            # If local path, convert to Cloudinary URL
            if image_path and not image_path.startswith('http'):
                image_path = get_cloudinary_url(image_path)
            
            plan_list.append({
                "id": plan['id'],
                "name": plan['name'],
                "speed": plan['speed'],
                "price": float(plan['price']) if plan['price'] else 0,
                "image": image_path
            })
        
        return jsonify(plan_list)
        
    except Exception as e:
        print(f"Get plans error: {e}")
        return jsonify([])

# ===============================
# CREATE PLAN - SHARED UPLOADS
# ===============================
@app.route("/api/superadmin/plans", methods=["POST"])
def create_plan():
    try:
        name = request.form.get("name")
        speed = request.form.get("speed")
        price = request.form.get("price")
        image_file = request.files.get("image")
        
        if not name or not speed or not price:
            return jsonify({"error": "Name, speed, and price are required"}), 400
        
        if not image_file or not allowed_plan_file(image_file.filename):
            return jsonify({"error": "Valid image file is required"}), 400

        # I-reset ang file pointer bago mag-validation
        image_file.stream.seek(0)
        
        # Validate orientation
        is_valid, error_message = validate_plan_image_orientation(image_file)
        if not is_valid:
            return jsonify({"error": error_message}), 400
        
        # I-reset ang file pointer BAGO mag-upload sa Cloudinary!
        image_file.stream.seek(0)  # ITO ANG SUSI!
        
        # Upload to Cloudinary
        image_url = upload_to_cloudinary(image_file)
        
        if not image_url:
            return jsonify({"error": "Failed to upload image to Cloudinary"}), 500
        
        # Save to database
        insert_query = """
            INSERT INTO plans (name, speed, price, image_path, created_at)
            VALUES (%s, %s, %s, %s, NOW())
        """
        plan_id = execute_query(insert_query, (name, speed, float(price), image_url))
        
        return jsonify({
            "message": "Plan created successfully", 
            "id": plan_id,
            "image_path": image_url
        })
        
    except Exception as e:
        print(f"Create plan error: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# UPDATE PLAN - SHARED UPLOADS
# ===============================
@app.route("/api/superadmin/plans/<int:plan_id>", methods=["PUT"])
def update_plan(plan_id):
    try:
        name = request.form.get("name")
        speed = request.form.get("speed")
        price = request.form.get("price")
        
        if not name or not speed or not price:
            return jsonify({"error": "Name, speed, and price are required"}), 400
        
        check_query = "SELECT id, image_path FROM plans WHERE id = %s"
        existing = execute_query(check_query, (plan_id,), fetch_one=True)
        
        if not existing:
            return jsonify({"error": "Plan not found"}), 404
        
        image_file = request.files.get("image")
        image_url = existing.get('image_path')
        
        if image_file and allowed_plan_file(image_file.filename):
            is_valid, error_message = validate_plan_image_orientation(image_file)
            if not is_valid:
                return jsonify({"error": error_message}), 400

            # Delete old image from Cloudinary
            if image_url and 'cloudinary.com' in image_url:
                delete_from_cloudinary(image_url)
            
            # Upload new image to Cloudinary
            image_url = upload_to_cloudinary(image_file)
        
        # Update plan
        if image_url and image_url != existing.get('image_path'):
            update_query = """
                UPDATE plans 
                SET name = %s, speed = %s, price = %s, image_path = %s
                WHERE id = %s
            """
            execute_query(update_query, (name, speed, float(price), image_url, plan_id))
        else:
            update_query = """
                UPDATE plans 
                SET name = %s, speed = %s, price = %s
                WHERE id = %s
            """
            execute_query(update_query, (name, speed, float(price), plan_id))
        
        return jsonify({"message": "Plan updated successfully"})
        
    except Exception as e:
        print(f"Update plan error: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# DELETE PLAN - SHARED UPLOADS
# ===============================
@app.route("/api/superadmin/plans/<int:plan_id>", methods=["DELETE"])
def delete_plan(plan_id):
    try:
        check_query = "SELECT id, image_path FROM plans WHERE id = %s"
        plan = execute_query(check_query, (plan_id,), fetch_one=True)
        
        if not plan:
            return jsonify({"error": "Plan not found"}), 404
        
        image_url = plan.get('image_path')
        
        # Delete from Cloudinary
        if image_url and 'cloudinary.com' in image_url:
            delete_from_cloudinary(image_url)
        
        delete_query = "DELETE FROM plans WHERE id = %s"
        execute_query(delete_query, (plan_id,))
        
        return jsonify({"message": "Plan deleted successfully"})
        
    except Exception as e:
        print(f"Delete plan error: {e}")
        return jsonify({"error": str(e)}), 500
    
# ===============================
# Internet Applications Page
# ===============================
@app.route("/superadmin/internet-applications")
def superadmin_internet_applications_page():
    return render_template("superadmin-internet-applications.html")


# ===============================
# GET APPLICATIONS - CONVERTED TO MYSQL (WITH ARCHIVE FILTER)
# ===============================
@app.route("/api/superadmin/applications", methods=["GET"])
def superadmin_get_all_applications():
    try:
        limit = int(request.args.get("limit", 50))
        include_archived = request.args.get("include_archived", "false").lower() == "true"
        city_filter = request.args.get("city", "")

        sql_params = [limit]
        query = """
            SELECT application_number, first_name, last_name, email, plan,
                   date_submitted, time_submitted, barangay, city, birthdate, 
                   status, rejection_reason, is_archived
            FROM applications
            WHERE 1 = 1
        """

        if not include_archived:
            query += " AND (is_archived = 0 OR is_archived IS NULL)"

        if city_filter and city_filter.lower() != "all":
            query += " AND city = %s"
            sql_params = [city_filter, limit]

        query += " ORDER BY timestamp DESC LIMIT %s"

        applications = execute_query(query, tuple(sql_params), fetch=True) or []

        # Format response
        apps = []
        for app in applications:
            # Combine date and time submitted for display
            datetime_submitted = None
            if app.get("date_submitted") and app.get("time_submitted"):
                datetime_submitted = f"{app.get('date_submitted')} {app.get('time_submitted')}"
            elif app.get("date_submitted"):
                datetime_submitted = app.get("date_submitted")
            
            apps.append({
                "id": app.get("application_number", ""),
                "application_number": app.get("application_number", ""),
                "first_name": app.get("first_name", ""),
                "last_name": app.get("last_name", ""),
                "email": app.get("email", ""),
                "plan": app.get("plan", "") or "",
                "date_submitted": datetime_submitted,
                "barangay": app.get("barangay", ""),
                "city": app.get("city", ""),
                "birthdate": app.get("birthdate", ""),
                "status": app.get("status", "Pending"),
                "rejection_reason": app.get("rejection_reason", ""),
                "is_archived": app.get("is_archived", 0) == 1
            })

        return jsonify(apps)

    except Exception as e:
        print("Superadmin get applications error:", e)
        return jsonify({"error": str(e)}), 500


# ===============================
# View Single Application Page (Superadmin)
# ===============================
@app.route("/superadmin/view-application/<app_id>")
def superadmin_view_application(app_id):
    return render_template("superadmin-view_application.html", app_id=app_id)


# ===============================
# GET SINGLE APPLICATION - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/application/<string:app_id>", methods=["GET"])
def superadmin_get_single_application(app_id):
    try:
        query = """
            SELECT 
                application_number, address, approval_date, barangay, 
                billing_address, billing_date, birthdate, business_address, 
                business_phone, citizenship, city, civil_status, contract_number,
                date_submitted, email, employer, father_name, first_name,
                home_ownership, house_number, id_back, id_front,
                installation_address, installation_fee, installation_phone,
                installation_status, landmark, last_name, latitude, longitude,
                middle_name, mobile, mother_maiden_name, occupation, phone,
                place_of_birth, plan, plan_price, plan_speed, profile_photo,
                proof_billing, province, secondary_mobile, service_type, sex,
                signature, spouse_employer, spouse_name, spouse_occupation,
                spouse_phone, status, suffix, terms_agreed, time_submitted,
                timestamp, rejection_reason, user_created, user_created_at,
                user_id, zip, tv_qty, tv_brand, tv_type,
                assigned_team_id, installation_date, reapplied_count,
                reapply_requested, reapply_requested_at, reapply_message
            FROM applications 
            WHERE application_number = %s
        """
        application = execute_query(query, (app_id,), fetch_one=True)
        
        if not application:
            return jsonify({"error": "Application not found"}), 404
        
        # Convert image paths to Cloudinary URLs
        image_fields = ['profile_photo', 'id_front', 'id_back', 'proof_billing', 'signature']
        for field in image_fields:
            if application.get(field):
                application[field] = get_cloudinary_url(application[field])
                print(f" Converted {field}: {application[field]}")
        
        # Parse JSON fields (tv_qty, tv_brand, tv_type are stored as JSON strings)
        if application.get('tv_qty'):
            try:
                application['tv_qty'] = json.loads(application['tv_qty'])
            except:
                application['tv_qty'] = []
        
        if application.get('tv_brand'):
            try:
                application['tv_brand'] = json.loads(application['tv_brand'])
            except:
                application['tv_brand'] = []
        
        if application.get('tv_type'):
            try:
                application['tv_type'] = json.loads(application['tv_type'])
            except:
                application['tv_type'] = []
        
        # Keep id field for frontend compatibility
        application['id'] = application.get('application_number')
        
        print(f" Reapply data from DB: reapply_requested={application.get('reapply_requested')}, reapply_requested_at={application.get('reapply_requested_at')}")
        
        return jsonify(application)
        
    except Exception as e:
        print("Get single application error:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    
    
# ===============================
# SUPERADMIN DOWNLOAD - CONVERTED TO MYSQL
# ===============================
@app.route('/superadmin/download/pdf/<app_id>')
def superadmin_download_pdf(app_id):
    try:
        # Get application data from MySQL using application_number (app_id)
        query = """
            SELECT application_number FROM applications 
            WHERE application_number = %s
        """
        data = execute_query(query, (app_id,), fetch_one=True)

        if not data:
            return "Application not found", 404

        # Get application_number
        application_number = data.get("application_number")

        if not application_number:
            return "Application number missing", 400

        # Use existing PDF generator
        return download_pdf(application_number)

    except Exception as e:
        print(f"PDF download error: {e}")
        return str(e), 500
    
@app.route('/download/pdf/<application_number>')
def download_pdf(application_number):
    import io, base64, os
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.utils import ImageReader
    from flask import send_file
    import requests
    import json
    import re
    from PIL import Image
    import tempfile

    # Get application data from MySQL
    query = """
        SELECT * FROM applications 
        WHERE application_number = %s
    """
    data = execute_query(query, (application_number,), fetch_one=True)
    
    if not data:
        return "Application not found", 404
    
    # ================= GET APPLICATION NUMBER AS FOLDER NAME =================
    app_folder = str(application_number)
    
    # Parse JSON fields
    if data.get('tv_qty'):
        try:
            data['tv_qty'] = json.loads(data['tv_qty'])
        except:
            data['tv_qty'] = []
    
    if data.get('tv_brand'):
        try:
            data['tv_brand'] = json.loads(data['tv_brand'])
        except:
            data['tv_brand'] = []
    
    if data.get('tv_type'):
        try:
            data['tv_type'] = json.loads(data['tv_type'])
        except:
            data['tv_type'] = []

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # ================= DEBUG: Print image info =================
    print("=" * 80)
    print(" IMAGE DATA FROM DATABASE:")
    print(f"Application Number (folder): {app_folder}")
    print(f"id_front: {data.get('id_front', 'None')}")
    print(f"id_back: {data.get('id_back', 'None')}")
    print(f"signature: {data.get('signature', 'None')}")
    print(f"proof_billing: {data.get('proof_billing', 'None')}")
    print("=" * 80)

    # ================= FIXED: Get image from Cloudinary =================
    def get_image_from_cloudinary(image_path):
        """Get image from Cloudinary URL or local path"""
        if not image_path:
            return None
        
        # Convert to Cloudinary URL
        cloudinary_url = get_cloudinary_url(image_path)
        
        if cloudinary_url and cloudinary_url.startswith('http'):
            try:
                print(f" Downloading from Cloudinary: {cloudinary_url}")
                response = requests.get(cloudinary_url, timeout=30)
                if response.status_code == 200:
                    print(f" Downloaded {len(response.content)} bytes")
                    return response.content
                else:
                    print(f" Cloudinary download failed: {response.status_code}")
            except Exception as e:
                print(f" Error downloading from Cloudinary: {e}")
        
        # Fallback: Try local path (for development)
        SHARED_UPLOADS_BASE = r"C:\xampp\htdocs\cablevision_uploads"
        
        # Extract filename from path
        filename = os.path.basename(image_path)
        
        # Try to find in application_uploads folder
        full_path = os.path.join(SHARED_UPLOADS_BASE, 'application_uploads', app_folder, filename)
        if os.path.exists(full_path):
            print(f" Found locally: {full_path}")
            with open(full_path, 'rb') as f:
                return f.read()
        
        # Try alternative paths
        alt_paths = [
            os.path.join(SHARED_UPLOADS_BASE, filename),
            os.path.join(SHARED_UPLOADS_BASE, image_path.lstrip('/'))
        ]
        
        for alt_path in alt_paths:
            if os.path.exists(alt_path):
                print(f" Found locally: {alt_path}")
                with open(alt_path, 'rb') as f:
                    return f.read()
        
        print(f" Image not found: {image_path}")
        return None

    # ================= HELPER: Load and convert image =================
    def load_and_convert_image(image_data):
        """Load image and convert to RGB format for PDF"""
        if not image_data:
            print(f" No image data provided")
            return None
        
        img_bytes = None
        
        # Try to get from Cloudinary
        if isinstance(image_data, str):
            img_bytes = get_image_from_cloudinary(image_data)
        
        # Try base64 decoding if file loading failed
        if not img_bytes and isinstance(image_data, str):
            if 'base64,' in image_data or 'data:image' in image_data:
                try:
                    if 'base64,' in image_data:
                        image_data = image_data.split('base64,')[1]
                    elif 'data:image' in image_data:
                        match = re.search(r'data:image/(png|jpeg|jpg|gif);base64,(.+)', image_data)
                        if match:
                            image_data = match.group(2)
                    
                    image_data = image_data.strip()
                    img_bytes = base64.b64decode(image_data)
                    print(f" Decoded base64 image ({len(img_bytes)} bytes)")
                except Exception as e:
                    print(f" Error decoding base64: {e}")
        
        if not img_bytes:
            print(f" No image bytes loaded")
            return None
        
        # Convert image to RGB format using PIL
        try:
            img = Image.open(io.BytesIO(img_bytes))
            print(f" Image opened: {img.format}, {img.size}, {img.mode}")
            
            if img.mode != 'RGB':
                img = img.convert('RGB')
                print(f" Converted to RGB")
            
            output = io.BytesIO()
            img.save(output, format='JPEG', quality=90)
            output.seek(0)
            
            print(f" Converted to JPEG ({output.getbuffer().nbytes} bytes)")
            return output.getvalue()
            
        except Exception as e:
            print(f" Error converting image: {e}")
            return img_bytes

    # ================= HELPER: Draw image safely =================
    def draw_image_safe(p, image_data, x, y, width, height, label="Image"):
        """Safely draw an image on the PDF"""
        try:
            print(f" Drawing {label}...")
            img_bytes = load_and_convert_image(image_data)
            if img_bytes:
                with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as tmp_file:
                    tmp_file.write(img_bytes)
                    tmp_path = tmp_file.name
                    print(f" Temp file: {tmp_path}")
                
                try:
                    img = ImageReader(tmp_path)
                    p.drawImage(img, x, y, width, height, preserveAspectRatio=True, mask='auto')
                    print(f" Drew {label} successfully")
                    return True
                except Exception as e:
                    print(f" Error in drawImage for {label}: {e}")
                    return False
                finally:
                    try:
                        os.unlink(tmp_path)
                        print(f" Deleted temp file: {tmp_path}")
                    except:
                        pass
            else:
                print(f" No image bytes for {label}")
                return False
        except Exception as e:
            print(f" Error drawing {label}: {e}")
            return False

    # ================= MAX PAGES =================
    MAX_PAGES = 5
    current_page = 1
    y = height - 120

    # ================= HEADER =================
    def draw_header():
        nonlocal y
        try:
            logo = ImageReader("static/logo1.png")
            p.drawImage(logo, 40, height - 90, width=60, height=60, mask='auto')
        except:
            pass

        p.setFont("Helvetica-Bold", 16)
        p.drawString(110, height - 60, "APPLICATION FORM")

        p.setFont("Helvetica-Bold", 10)
        p.drawRightString(width - 50, height - 60, f"Application No: {application_number}")

        p.setFont("Helvetica", 8)
        p.drawString(110, height - 75, "Sitio Sampaguita, Brgy. Pagsawitan, Santa Cruz, 4009 Laguna")
        p.drawString(110, height - 87, "Tel: (049) 501-1495 | Fax: (049) 501-0229 | Mobile: 0917 501 0341")

        y = height - 110

    def draw_page_number():
        p.setFont("Helvetica-Bold", 10)
        p.setFillColorRGB(0.4, 0.4, 0.4)
        p.drawRightString(width - 25, 20, str(current_page))
        p.setFillColorRGB(0, 0, 0)

    def new_page():
        nonlocal y, current_page
        if current_page >= MAX_PAGES:
            return False
        p.showPage()
        current_page += 1
        draw_header()
        draw_page_number()
        y = height - 110
        return True

    def ensure_space(required):
        nonlocal y
        if y - required < 50:
            return new_page()
        return True

    def draw_section_title(title):
        nonlocal y
        ensure_space(30)
        p.setFont("Helvetica-Bold", 12)
        p.setFillColorRGB(0, 0.4, 0.6)
        p.drawString(50, y, title)
        p.setFillColorRGB(0, 0, 0)
        y -= 22

    def draw_section_title_centered(title):
        nonlocal y
        ensure_space(30)
        p.setFont("Helvetica-Bold", 14)
        p.setFillColorRGB(0, 0.4, 0.6)
        p.drawCentredString(width / 2, y, title)
        p.setFillColorRGB(0, 0, 0)
        y -= 25

    def draw_two_columns(fields):
        nonlocal y
        col1_x = 50
        col2_x = 310
        label_width = 120
        value_x = col1_x + label_width + 5
        
        for i in range(0, len(fields), 2):
            ensure_space(20)
            label1, value1 = fields[i]
            p.setFont("Helvetica-Bold", 9)
            p.drawString(col1_x, y, f"{label1}:")
            p.setFont("Helvetica", 9)
            val1_str = str(value1) if value1 and value1 != "-" and value1 != "none" else "___________________"
            if len(val1_str) > 35:
                val1_str = val1_str[:32] + "..."
            p.drawString(value_x, y, val1_str)
            
            if i + 1 < len(fields):
                label2, value2 = fields[i + 1]
                p.setFont("Helvetica-Bold", 9)
                p.drawString(col2_x, y, f"{label2}:")
                p.setFont("Helvetica", 9)
                val2_str = str(value2) if value2 and value2 != "-" and value2 != "none" else "___________________"
                if len(val2_str) > 30:
                    val2_str = val2_str[:27] + "..."
                p.drawString(col2_x + label_width + 5, y, val2_str)
            
            y -= 18
        y -= 5

    def draw_images_top_bottom(label1, img1_data, label2, img2_data, img_width=280, img_height=190):
        nonlocal y
        
        ensure_space(img_height + 60)
        p.setFont("Helvetica-Bold", 11)
        p.drawCentredString(width / 2, y, label1)
        y -= 22
        
        if draw_image_safe(p, img1_data, (width - img_width) / 2, y - img_height, img_width, img_height, label1):
            y -= img_height + 35
        else:
            p.setFont("Helvetica", 9)
            p.setFillColorRGB(0.5, 0.5, 0.5)
            p.drawCentredString(width / 2, y, "No image provided")
            p.setFillColorRGB(0, 0, 0)
            y -= 25
        
        ensure_space(img_height + 60)
        p.setFont("Helvetica-Bold", 11)
        p.drawCentredString(width / 2, y, label2)
        y -= 22
        
        if draw_image_safe(p, img2_data, (width - img_width) / 2, y - img_height, img_width, img_height, label2):
            y -= img_height + 35
        else:
            p.setFont("Helvetica", 9)
            p.setFillColorRGB(0.5, 0.5, 0.5)
            p.drawCentredString(width / 2, y, "No image provided")
            p.setFillColorRGB(0, 0, 0)
            y -= 25

    def draw_signature_section(signature_img, full_name):
        nonlocal y
        
        y -= 15
        
        sig_width = 250
        sig_height = 85
        
        if draw_image_safe(p, signature_img, (width - sig_width) / 2, y - sig_height, sig_width, sig_height, "Signature"):
            y -= sig_height + 20
        else:
            p.setFont("Helvetica", 9)
            p.setFillColorRGB(0.5, 0.5, 0.5)
            p.drawCentredString(width / 2, y, "No signature provided")
            p.setFillColorRGB(0, 0, 0)
            y -= 25
        
        p.setFont("Helvetica", 10)
        p.drawCentredString(width / 2, y, full_name if full_name else "_________________________")
        y -= 20
        
        p.setFont("Helvetica", 8)
        p.setFillColorRGB(0.4, 0.4, 0.4)
        p.drawCentredString(width / 2, y, "signature over printed name")
        p.setFillColorRGB(0, 0, 0)
        y -= 30

    # ================= START BUILDING =================
    draw_header()
    draw_page_number()

    # ================= PAGE 1: PERSONAL INFORMATION =================
    draw_section_title("I. PERSONAL INFORMATION")
    draw_two_columns([
        ("Last Name", data.get("last_name")),
        ("First Name", data.get("first_name")),
        ("Middle Name", data.get("middle_name")),
        ("Suffix", data.get("suffix")),
        ("Date of Birth", data.get("birthdate")),
        ("Place of Birth", data.get("place_of_birth")),
        ("Sex", data.get("sex")),
        ("Civil Status", data.get("civil_status")),
        ("Citizenship", data.get("citizenship")),
        ("Occupation", data.get("occupation")),
    ])

    draw_section_title("II. FAMILY DETAILS")
    draw_two_columns([
        ("Mother's Maiden Name", data.get("mother_maiden_name")),
        ("Father's Name", data.get("father_name")),
    ])

    draw_section_title("III. CONTACT & ADDRESS")
    draw_two_columns([
        ("Mobile Number", data.get("mobile")),
        ("Email Address", data.get("email")),
        ("Home Ownership", data.get("home_ownership")),
        ("House No./Unit", data.get("house_number")),
        ("Nearest Landmark", data.get("landmark")),
        ("Street/Village", data.get("address")),
    ])

    # Billing Address
    p.setFont("Helvetica-Bold", 9)
    p.drawString(50, y, "Billing Address:")
    p.setFont("Helvetica", 9)
    
    billing_address = data.get("billing_address", "")
    if not billing_address or billing_address == "-" or billing_address == "none":
        billing_address = "_________________________"
    
    from reportlab.pdfbase.pdfmetrics import stringWidth
    max_width = 400
    words = billing_address.split()
    lines = []
    current_line = ""
    
    for word in words:
        test_line = current_line + (" " if current_line else "") + word
        if stringWidth(test_line, "Helvetica", 9) <= max_width:
            current_line = test_line
        else:
            lines.append(current_line)
            current_line = word
    
    if current_line:
        lines.append(current_line)
    
    for line in lines:
        p.drawString(170, y, line)
        y -= 14
    y -= 5

    draw_section_title("IV. EMPLOYMENT DETAILS")
    draw_two_columns([
        ("Employer / Company", data.get("employer")),
        ("Business Phone", data.get("business_phone")),
        ("Business Address", data.get("business_address")),
        ("", ""),
    ])

    civil_status = data.get("civil_status", "")
    if civil_status and civil_status.lower() in ["married", "Married"]:
        draw_section_title("V. SPOUSE INFORMATION")
        draw_two_columns([
            ("Spouse Full Name", data.get("spouse_name")),
            ("Spouse Occupation", data.get("spouse_occupation")),
            ("Spouse Employer", data.get("spouse_employer")),
            ("Spouse Phone", data.get("spouse_phone")),
        ])

    draw_section_title("VI. SERVICE PLAN")
    draw_two_columns([
        ("Service Type / Plan", data.get("service_type")),
        ("Installation Fee", data.get("installation_fee")),
    ])

    # Installation Phone
    p.setFont("Helvetica-Bold", 9)
    p.drawString(50, y, "Installation Phone:")
    p.setFont("Helvetica", 9)
    installation_phone = data.get("installation_phone", "")
    if not installation_phone or installation_phone == "-" or installation_phone == "none":
        installation_phone = "_________________________"
    p.drawString(170, y, installation_phone)
    y -= 18
    y -= 5

    # Installation Address
    p.setFont("Helvetica-Bold", 9)
    p.drawString(50, y, "Installation Address:")
    p.setFont("Helvetica", 9)
    
    installation_address = data.get("installation_address", "")
    if not installation_address or installation_address == "-" or installation_address == "none":
        installation_address = "_________________________"
    
    words = installation_address.split()
    lines = []
    current_line = ""
    
    for word in words:
        test_line = current_line + (" " if current_line else "") + word
        if stringWidth(test_line, "Helvetica", 9) <= max_width:
            current_line = test_line
        else:
            lines.append(current_line)
            current_line = word
    
    if current_line:
        lines.append(current_line)
    
    for line in lines:
        p.drawString(170, y, line)
        y -= 14
    y -= 5

    # TV SET DETAILS
    tv_qty = data.get("tv_qty", [])
    tv_brand = data.get("tv_brand", [])
    tv_type = data.get("tv_type", [])
    
    if tv_qty and any(tv_qty):
        draw_section_title("VII. TV SET DETAILS")
        ensure_space(40)
        
        p.setFont("Helvetica-Bold", 9)
        p.drawString(50, y, "QTY")
        p.drawString(120, y, "BRAND / MODEL")
        p.drawString(320, y, "TYPE (HD/REGULAR)")
        y -= 15
        
        p.setFont("Helvetica", 9)
        for i in range(min(len(tv_qty), 5)):
            if y < 120:
                break
            qty = str(tv_qty[i]) if i < len(tv_qty) else "-"
            brand = tv_brand[i] if i < len(tv_brand) else "-"
            if len(brand) > 25:
                brand = brand[:22] + "..."
            tv_t = tv_type[i] if i < len(tv_type) else "-"
            
            p.drawString(50, y, qty)
            p.drawString(120, y, brand)
            p.drawString(320, y, tv_t)
            y -= 16
        y -= 5

    draw_section_title("VIII. SUBMISSION DETAILS")
    draw_two_columns([
        ("Date Submitted", data.get("date_submitted")),
        ("Time Submitted", data.get("time_submitted")),
    ])

    full_name = f"{data.get('first_name', '')} {data.get('last_name', '')}".strip()
    draw_signature_section(data.get("signature"), full_name)

    # ================= PAGE 2: MAP =================
    new_page()
    draw_section_title_centered("INSTALLATION LOCATION MAP")

    lat = data.get("latitude")
    lng = data.get("longitude")
    google_maps_url = None
    google_maps_direction_url = None
    map_img = None

    try:
        if lat and lng:
            lat = float(lat)
            lng = float(lng)
            
            google_maps_url = f"https://www.google.com/maps?q={lat},{lng}"
            google_maps_direction_url = f"https://www.google.com/maps/dir//{lat},{lng}"
            
            map_url = f"https://maps.locationiq.com/v3/staticmap?key=pk.0fdad07272d959e4de881139988b0883&center={lat},{lng}&zoom=17&size=600x400&markers=icon:large-red-cutout|{lat},{lng}"
            response = requests.get(map_url, timeout=10)
            if response.status_code == 200:
                map_img = ImageReader(io.BytesIO(response.content))
    except Exception as e:
        print("Map error:", e)

    draw_two_columns([
        ("Street/Village", data.get("address")),
        ("Barangay/City", f"{data.get('barangay', '-')}, {data.get('city', '-')}"),
        ("Latitude", lat if lat else "-"),
        ("Longitude", lng if lng else "-"),
    ])

    if google_maps_direction_url:
        ensure_space(25)
        p.setFont("Helvetica-Bold", 12)
        p.setFillColorRGB(0, 0.5, 0)
        p.drawCentredString(width / 2, y, " GET DIRECTIONS from your current location to this address")
        text_width = p.stringWidth(" GET DIRECTIONS from your current location to this address", "Helvetica-Bold", 12)
        p.linkURL(google_maps_direction_url, ((width - text_width) / 2, y - 2, (width + text_width) / 2, y + 12), relative=0)
        p.setFillColorRGB(0, 0, 0)
        y -= 20
        
        p.setFont("Helvetica", 8)
        p.setFillColorRGB(0.5, 0.5, 0.5)
        p.drawCentredString(width / 2, y, " Click above to see distance from YOUR location, travel time, and turn-by-turn directions")
        p.setFillColorRGB(0, 0, 0)
        y -= 20

    if google_maps_url:
        p.setFont("Helvetica", 9)
        p.setFillColorRGB(0, 0, 1)
        p.drawCentredString(width / 2, y, "Or click here to view location on Google Maps")
        text_width = p.stringWidth("Or click here to view location on Google Maps", "Helvetica", 9)
        p.linkURL(google_maps_url, ((width - text_width) / 2, y - 2, (width + text_width) / 2, y + 8), relative=0)
        p.setFillColorRGB(0, 0, 0)
        y -= 30

    if map_img:
        ensure_space(380)
        img_width = 500
        img_height = 320
        x_center = (width - img_width) / 2
        p.drawImage(map_img, x_center, y - img_height, img_width, img_height)
        y -= img_height + 20
        
        p.setFont("Helvetica", 8)
        p.setFillColorRGB(0.5, 0.5, 0.5)
        p.drawCentredString(width / 2, y, " Tip: Click the green 'GET DIRECTIONS' link above to see distance from your current location")
        p.setFillColorRGB(0, 0, 0)
        y -= 15
    else:
        draw_two_columns([("Map Status", "Not available")])

    # ================= PAGE 3: FRONT AND BACK ID =================
    new_page()
    draw_section_title_centered("VALID IDENTIFICATION")
    
    draw_images_top_bottom(
        "VALID ID (FRONT)", data.get("id_front"),
        "VALID ID (BACK)", data.get("id_back"),
        img_width=320, img_height=220
    )

    # ================= PAGE 4: PROOF OF BILLING =================
    new_page()
    draw_section_title_centered("PROOF OF BILLING")
    
    proof = data.get("proof_billing")
    if draw_image_safe(p, proof, (width - 500) / 2, y - 580, 500, 580, "Proof of Billing"):
        y -= 580 + 30
    else:
        p.setFont("Helvetica", 9)
        p.setFillColorRGB(0.5, 0.5, 0.5)
        p.drawCentredString(width / 2, y, "No proof of billing provided")
        p.setFillColorRGB(0, 0, 0)
        y -= 30

    p.save()
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name="Application_Form.pdf")




# ==================== UPDATE APPLICATION STATUS - CONVERTED TO MYSQL ====================
# ==================== UPDATE APPLICATION STATUS - FIXED ====================
@app.route("/api/superadmin/application/<string:app_id>/status", methods=["PUT"])
def update_internet_application_status(app_id):
    conn = None
    cursor = None
    try:
        data = request.get_json()
        status = data.get("status")
        reason = data.get("reason", "")
        contract_number = data.get("contract_number", None)
        billing_date = data.get("billing_date", None)
        first_installment_date = data.get("first_installment_date", None)
        last_installment_date = data.get("last_installment_date", None)
        assigned_team_id = data.get("assigned_team_id", None)
        installation_date = data.get("installation_date", None)

        print(f" DEBUG - Received request for app_id: {app_id}")
        print(f" DEBUG - Status: {status}")

        # ALLOW BOTH APPROVED, REJECTED, AND PENDING (for restore)
        if status not in ["Approved", "Rejected", "Pending"]:
            return jsonify({"error": "Invalid status"}), 400

        conn = get_db_connection()

        if not conn:
            return jsonify({"error": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)

        # Check if application exists
        check_query = "SELECT * FROM applications WHERE application_number = %s"
        cursor.execute(check_query, (app_id,))
        app_data = cursor.fetchone()

        if not app_data:
            return jsonify({"error": "Application not found"}), 404

        # Generate contract number if approved and not provided
        if status == "Approved" and not contract_number:
            import random
            import string
            date_part = ph_now().strftime("%Y%m%d")
            random_part = ''.join(random.choices(string.digits, k=4))
            contract_number = f"CV-{date_part}-{random_part}"
            
            if not billing_date:
                billing_date = "15th"
            print(f" Generated contract number: {contract_number}")

        # ========== UPDATE APPLICATIONS TABLE ==========
        if status == "Approved":
            update_query = """
                UPDATE applications SET 
                    status = %s,
                    contract_number = %s,
                    billing_date = %s,
                    approval_date = %s,
                    installation_status = %s,
                    rejection_reason = %s,
                    assigned_team_id = %s,
                    installation_date = %s
                WHERE application_number = %s
            """
            params = (
                "Approved",
                contract_number,
                billing_date,
                ph_now().strftime("%Y-%m-%d %H:%M:%S"),
                "Pending",
                "",
                assigned_team_id,
                installation_date,
                app_id
            )
            
            print(f" UPDATE QUERY: {update_query[:100]}...")
            print(f" PARAMS: {params}")
            
            cursor.execute(update_query, params)
            conn.commit()
            print(f" Application {app_id} updated - Rows affected: {cursor.rowcount}")
            
            # I-VERIFY AGAD
            cursor.execute("SELECT status, contract_number, billing_date, approval_date FROM applications WHERE application_number = %s", (app_id,))
            verify = cursor.fetchone()
            print(f" VERIFIED - Status: {verify.get('status') if verify else 'Not found'}")
            print(f" VERIFIED - Contract: {verify.get('contract_number') if verify else 'Not found'}")
            print(f" VERIFIED - Billing Date: {verify.get('billing_date') if verify else 'Not found'}")
            
        elif status == "Pending":
            # FOR RESTORE - Update status to Pending
            update_query = """
                UPDATE applications SET 
                    status = %s,
                    rejection_reason = NULL,
                    approval_date = NULL,
                    contract_number = NULL,
                    billing_date = NULL,
                    assigned_team_id = NULL,
                    installation_date = NULL,
                    is_archived = 0
                WHERE application_number = %s
            """
            params = ("Pending", app_id)
            
            cursor.execute(update_query, params)
            conn.commit()
            print(f" Application {app_id} restored to Pending")
            
        else:  # Rejected
            update_query = """
                UPDATE applications SET 
                    status = %s,
                    rejection_reason = %s,
                    approval_date = NULL,
                    contract_number = NULL,
                    billing_date = NULL,
                    assigned_team_id = NULL,
                    installation_date = NULL,
                    reapply_requested = 0,
                    reapply_requested_at = NULL,
                    reapply_message = NULL
                WHERE application_number = %s
            """
            params = ("Rejected", reason, app_id)
            
            cursor.execute(update_query, params)
            conn.commit()
            print(f" Application {app_id} rejected — reapply flags reset")

        # ========== IF APPROVED, INSERT INTO CUSTOMERS TABLE ==========
        if status == "Approved":
            print(f" DEBUG - Processing customer record for {app_id}")
            
            customer_data = {
                "application_number": app_id,
                "first_name": app_data.get("first_name"),
                "last_name": app_data.get("last_name"),
                "middle_name": app_data.get("middle_name"),
                "suffix": app_data.get("suffix"),
                "email": app_data.get("email"),
                "mobile": app_data.get("mobile"),
                "address": app_data.get("address"),
                "barangay": app_data.get("barangay"),
                "city": app_data.get("city"),
                "province": app_data.get("province"),
                "zip": app_data.get("zip"),
                "plan": app_data.get("plan"),
                "plan_speed": app_data.get("plan_speed"),
                "plan_price": app_data.get("plan_price"),
                "status": "Approved",
                "installation_status": "Pending",
                "contract_number": contract_number,
                "billing_date": billing_date,
                "approval_date": ph_now().strftime("%Y-%m-%d %H:%M:%S"),
                "date_pending": ph_now().strftime("%Y-%m-%d %H:%M:%S"),
                "assigned_team_id": assigned_team_id,
                "installation_date": installation_date,
                "latitude": app_data.get("latitude"),
                "longitude": app_data.get("longitude")
            }
            
            # Remove None values
            customer_data = {k: v for k, v in customer_data.items() if v is not None}
            
            # Check if customer already exists
            cursor.execute("SELECT application_number FROM customers WHERE application_number = %s", (app_id,))
            existing_customer = cursor.fetchone()
            
            if not existing_customer:
                columns = ', '.join(customer_data.keys())
                placeholders = ', '.join(['%s'] * len(customer_data))
                insert_query = f"INSERT INTO customers ({columns}) VALUES ({placeholders})"
                print(f" INSERT Query: {insert_query}")
                cursor.execute(insert_query, tuple(customer_data.values()))
                conn.commit()
                print(f" Customer INSERTED for {app_id}")
            else:
                update_fields = []
                update_params = []
                for key, value in customer_data.items():
                    update_fields.append(f"{key} = %s")
                    update_params.append(value)
                update_params.append(app_id)
                update_customer_query = f"UPDATE customers SET {', '.join(update_fields)} WHERE application_number = %s"
                print(f" UPDATE Query: {update_customer_query}")
                cursor.execute(update_customer_query, tuple(update_params))
                conn.commit()
                print(f" Customer UPDATED for {app_id}")
            
            # ========== INSERT INTO CONTRACTS TABLE ==========
            try:
                print(f" CONTRACT - Insert for {app_id}")
                
                full_name = ' '.join(filter(None, [
                    app_data.get('first_name', ''),
                    app_data.get('middle_name', ''),
                    app_data.get('last_name', ''),
                    app_data.get('suffix', '')
                ])).strip()
                
                address_parts = []
                if app_data.get('house_number') and app_data.get('house_number') != 'none':
                    address_parts.append(app_data.get('house_number'))
                if app_data.get('address') and app_data.get('address') != 'none':
                    address_parts.append(app_data.get('address'))
                if app_data.get('barangay') and app_data.get('barangay') != 'none':
                    address_parts.append(f"Barangay {app_data.get('barangay')}")
                if app_data.get('city') and app_data.get('city') != 'none':
                    address_parts.append(app_data.get('city'))
                if app_data.get('province') and app_data.get('province') != 'none':
                    address_parts.append(app_data.get('province'))
                address = ', '.join(address_parts) if address_parts else 'Not provided'
                
                # Check if contract already exists
                cursor.execute("SELECT contract_number FROM contracts WHERE application_id = %s", (app_id,))
                existing_contract = cursor.fetchone()
                
                age_value = calculate_age(app_data.get('birthdate', ''))
                
                if not existing_contract:
                    contract_insert_query = """
                        INSERT INTO contracts (
                            contract_number, application_id, first_name, middle_name, last_name, suffix,
                            full_name, age, civil_status, address, barangay, city, province,
                            billing_date, date_submitted, status, created_at,
                            is_installment_plan, first_installment_date, last_installment_date,
                            installation_fee, application_data, assigned_team_id, installation_date
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    
                    contract_params = (
                        contract_number,
                        app_id,
                        app_data.get('first_name', ''),
                        app_data.get('middle_name', ''),
                        app_data.get('last_name', ''),
                        app_data.get('suffix', ''),
                        full_name,
                        age_value if age_value else None,
                        app_data.get('civil_status', ''),
                        address,
                        app_data.get('barangay', ''),
                        app_data.get('city', ''),
                        app_data.get('province', ''),
                        billing_date,
                        app_data.get('date_submitted', ''),
                        "Active",
                        ph_now_iso(),
                        1 if first_installment_date else 0,
                        first_installment_date,
                        last_installment_date,
                        app_data.get('installation_fee', ''),
                        json.dumps(app_data, default=str),
                        assigned_team_id,
                        installation_date
                    )
                    
                    cursor.execute(contract_insert_query, contract_params)
                    conn.commit()
                    print(f" Contract {contract_number} INSERTED!")
                else:
                    print(f" Contract already exists for {app_id}")
                
            except Exception as contract_err:
                print(f" Contract error: {contract_err}")
                import traceback
                traceback.print_exc()
                # Don't rollback here, we want to keep the application update

                # ========== SEND EMAIL NOTIFICATION (Background Thread) ==========
        try:
            customer_email = app_data.get("email")
            first_name = app_data.get("first_name")
            application_number = app_data.get("application_number", "N/A")
            reapplied_count = app_data.get("reapplied_count", 0)

            if customer_email:
                # GAMITIN ANG THREAD PARA HINDI MAG-TIMEOUT ANG WORKER
                import threading
                email_thread = threading.Thread(
                    target=send_application_status_email,
                    args=(
                        customer_email,
                        first_name,
                        status,
                        application_number,
                    ),
                    kwargs={
                        'reason': reason if status == "Rejected" else None,
                        'contract_number': contract_number if status == "Approved" else None,
                        'billing_date': billing_date if status == "Approved" else None,
                        'application_id': app_id,
                        'reapplied_count': reapplied_count
                    }
                )
                email_thread.start()
                print(f" Email thread started for {customer_email}")
                # Huwag hintayin ang thread, hayaan itong tumakbo sa background.
                # Ibalik agad ang response sa frontend para hindi mag-timeout.
            else:
                print(f" No email address for {app_id}")
        except Exception as email_err:
            print(f" Email thread error: {email_err}")
            import traceback
            traceback.print_exc()
            # Huwag i-fail ang request kung mag-fail ang email

        # ALWAYS RETURN JSON
        return jsonify({
            "message": "Status updated successfully",
            "status": status,
            "contract_number": contract_number if status == "Approved" else None
        }), 200

    except mysql.connector.Error as db_err:
        print(f" Database error: {db_err}")
        if conn:
            conn.rollback()
        return jsonify({"error": str(db_err)}), 500
    except Exception as e:
        print(f" Update status error: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
            print(" Database connection closed")


import requests  # I-ADD SA PINAKA-ITAAS NG FILE

# ===============================
# SEND APPLICATION STATUS EMAIL
# ===============================

def send_application_status_email(
    to_email,
    first_name,
    status,
    app_id,
    reason=None,
    contract_number=None,
    billing_date=None,
    application_id=None,
    reapplied_count=0
):
    """Send application approval/rejection email via Brevo API."""

    import base64
    import html
    import requests

    # ===============================
    # BREVO CONFIGURATION
    # ===============================

    api_key = os.getenv("BREVO_API_KEY", "")

    if not api_key:
        print(" Brevo API key not configured!")
        return False

    print(f" Sending application status email via Brevo to {to_email}...")

    subject = "Cablevision - Application Status Update"

    # ===============================
    # HELPER
    # ===============================

    def escape_html(text):
        if text is None:
            return ""
        return html.escape(str(text))

    # ===============================
    # STATUS SETTINGS
    # ===============================

    is_approved = status == "Approved"
    status_color = "#10b981" if is_approved else "#ef4444"
    status_bg = "#ecfdf5" if is_approved else "#fef2f2"
    status_icon = "" if is_approved else ""

    # ===============================
    # MAIN MESSAGE
    # ===============================

    if is_approved:
        message = f"Congratulations, {escape_html(first_name)}!"
        message_sub = "Your application has been approved successfully."
    else:
        message = f"Application Update, {escape_html(first_name)}"
        message_sub = "We regret to inform you about your application status."

    # ===============================
    # CONTRACT SECTION
    # ===============================

    contract_section = ""

    if is_approved and contract_number:
        contract_section = f"""
        <div style="margin: 20px 0; padding: 20px;
                    background: linear-gradient(135deg, #ecfdf5 0%, #d1fae5 100%);
                    border-radius: 16px; text-align: center;">
            <div style="font-size: 14px; color: #047857;
                        margin-bottom: 8px; letter-spacing: 1px;">
                CONTRACT NUMBER
            </div>
            <div style="font-size: 28px; font-weight: 700;
                        color: #065f46; letter-spacing: 2px;
                        font-family: monospace;">
                {escape_html(contract_number)}
            </div>
            <div style="font-size: 11px; color: #059669;
                        margin-top: 8px;">
                ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            </div>
            <div style="font-size: 11px; color: #047857;
                        margin-top: 6px;">
                Please keep this number for future reference
            </div>
        </div>
        """

    # ===============================
    # BILLING SECTION
    # ===============================

    billing_section = ""

    if is_approved and billing_date:
        billing_section = f"""
        <div style="margin: 20px 0; padding: 16px;
                    background: #eff6ff; border-radius: 12px;">
            <div style="font-size: 12px; font-weight: 600;
                        color: #1e40af; margin-bottom: 4px;">
                BILLING INFORMATION
            </div>
            <div style="font-size: 16px; font-weight: 600;
                        color: #1e3a8a;">
                Every {escape_html(billing_date)} of the month
            </div>
            <div style="font-size: 11px; color: #3b82f6;
                        margin-top: 4px;">
                Your monthly bill will be generated on this date
            </div>
        </div>
        """

    # ===============================
    # EXTRA MESSAGE
    # ===============================

    if is_approved:
        if contract_number:
            next_message = """
            Please find attached your application PDF with contract details.<br>
            Our team will contact you soon for installation scheduling.<br>
            For inquiries, please contact our support team.
            """
        else:
            next_message = """
            Please find attached your application PDF.<br>
            Our team will contact you soon for installation scheduling.<br>
            For inquiries, please contact our support team.
            """

        extra_message = f"""
        <div style="margin-top: 20px; padding: 16px;
                    background: #f0fdf4; border-radius: 12px;">
            <p style="margin: 0 0 8px 0; color: #166534;">
                <strong>What's Next?</strong>
            </p>
            <p style="margin: 0; color: #14532d; font-size: 14px;">
                {next_message}
            </p>
        </div>
        """
    else:
        rejection_reason = escape_html(reason) if reason else "No specific reason provided."

        extra_message = f"""
        <div style="margin-top: 20px; padding: 16px;
                    background: #fef2f2; border-radius: 12px;">
            <p style="margin: 0 0 8px 0; color: #991b1b;">
                <strong>Reason for Rejection</strong>
            </p>
            <p style="margin: 0; color: #7f1d1d; font-size: 14px;">
                {rejection_reason}
            </p>
            <p style="margin-top: 12px; color: #7f1d1d; font-size: 13px;">
                Our team will reach out via email with instructions
                if a re-application is possible.
            </p>
        </div>
        """

    # ===============================
    # HTML EMAIL BODY
    # ===============================

    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Cablevision Email</title>
    </head>
    <body style="
        margin: 0;
        padding: 0;
        font-family: 'Segoe UI', 'Inter', -apple-system, BlinkMacSystemFont, Arial, sans-serif;
        background-color: #eef2ff;
    ">
        <div style="
            max-width: 580px;
            margin: 0 auto;
            padding: 30px 20px;
        ">
            <div style="
                background: #ffffff;
                border-radius: 32px;
                overflow: hidden;
                box-shadow: 0 20px 35px -12px rgba(0, 0, 0, 0.15);
            ">
                <!-- HEADER SECTION -->
                <div style="
                    background: linear-gradient(135deg, #001f3f 0%, #002b5c 100%);
                    padding: 32px 28px;
                    text-align: center;
                ">
                    <div style="
                        position: absolute;
                        top: 20px;
                        right: 25px;
                    ">
                        <span style="
                            background: rgba(255,255,255,0.15);
                            padding: 6px 14px;
                            border-radius: 50px;
                            font-size: 11px;
                            font-weight: 600;
                            color: #a5f3fc;
                        ">
                            STATUS UPDATE
                        </span>
                    </div>
                    <h1 style="
                        margin: 0;
                        font-size: 26px;
                        font-weight: 700;
                        color: #ffffff;
                    ">
                        Cablevision
                    </h1>
                    <p style="
                        margin: 6px 0 0 0;
                        color: #93c5fd;
                        font-size: 13px;
                    ">
                        Internet Service Provider
                    </p>
                </div>

                <!-- STATUS BADGE -->
                <div style="
                    padding: 20px 28px 0 28px;
                    text-align: center;
                ">
                    <div style="
                        display: inline-block;
                        background: {status_bg};
                        padding: 8px 24px;
                        border-radius: 60px;
                    ">
                        <span style="
                            font-size: 14px;
                            font-weight: 600;
                            color: {status_color};
                        ">
                            {status_icon} APPLICATION {escape_html(status).upper()}
                        </span>
                    </div>
                </div>

                <!-- CONTENT SECTION -->
                <div style="padding: 20px 28px 32px 28px;">
                    <h2 style="
                        margin: 0 0 8px 0;
                        font-size: 22px;
                        font-weight: 700;
                        color: #0f172a;
                    ">
                        {message}
                    </h2>
                    <p style="
                        margin: 0 0 20px 0;
                        font-size: 15px;
                        color: #475569;
                    ">
                        {message_sub}
                    </p>

                    <!-- APPLICATION DETAILS -->
                    <div style="
                        background: #f8fafc;
                        border-radius: 20px;
                        padding: 18px;
                        margin-bottom: 16px;
                    ">
                        <div style="
                            margin-bottom: 16px;
                            padding-bottom: 12px;
                            border-bottom: 1px solid #e2e8f0;
                        ">
                            <div style="
                                font-size: 11px;
                                font-weight: 600;
                                color: #64748b;
                            ">
                                Application Number
                            </div>
                            <div style="
                                font-size: 18px;
                                font-weight: 700;
                                color: #0f172a;
                                font-family: monospace;
                            ">
                                {escape_html(app_id)}
                            </div>
                        </div>
                        <div>
                            <div style="
                                font-size: 11px;
                                font-weight: 600;
                                color: #64748b;
                            ">
                                Status
                            </div>
                            <div style="
                                font-size: 16px;
                                font-weight: 700;
                                color: {status_color};
                            ">
                                {escape_html(status)}
                            </div>
                        </div>
                    </div>

                    {contract_section}
                    {billing_section}
                    {extra_message}

                    <!-- THANK YOU -->
                    <div style="
                        margin-top: 28px;
                        padding-top: 20px;
                        text-align: center;
                        border-top: 1px solid #e2e8f0;
                    ">
                        <p style="
                            margin: 0;
                            font-size: 12px;
                            color: #94a3b8;
                        ">
                            Thank you for choosing Cablevision!
                        </p>
                    </div>
                </div>

                <!-- FOOTER -->
                <div style="
                    background: #f1f5f9;
                    padding: 16px 28px;
                    text-align: center;
                ">
                    <div style="
                        font-size: 11px;
                        color: #64748b;
                    ">
                        2026 Cablevision Internet Service Provider. All rights reserved.
                    </div>
                </div>
            </div>
        </div>
    </body>
    </html>
    """

    # ===============================
    # SEND VIA BREVO API
    # ===============================

    try:
        url = "https://api.brevo.com/v3/smtp/email"

        headers = {
            "accept": "application/json",
            "api-key": api_key,
            "content-type": "application/json"
        }

        data = {
            "sender": {
                "name": "Cablevision Systems Corp.",
                "email": "cablevision.cableinternet@gmail.com"
            },
            "to": [
                {
                    "email": to_email,
                    "name": first_name
                }
            ],
            "subject": subject,
            "htmlContent": html_body
        }

        # ===============================
        # ADD PDF ATTACHMENT IF APPROVED
        # ===============================

        if is_approved:
            try:
                pdf_app_key = application_id if application_id else app_id

                if pdf_app_key:
                    query = """
                        SELECT *
                        FROM applications
                        WHERE application_number = %s
                    """

                    app_data = execute_query(
                        query,
                        (pdf_app_key,),
                        fetch_one=True
                    )

                    if app_data:
                        pdf_buffer = generate_application_pdf(
                            pdf_app_key,
                            app_data,
                            contract_number
                        )

                        if pdf_buffer:
                            pdf_content = base64.b64encode(
                                pdf_buffer.read()
                            ).decode("utf-8")

                            data["attachment"] = [
                                {
                                    "content": pdf_content,
                                    "name": f"Application_{app_id}.pdf"
                                }
                            ]

                            print(f" PDF attached for application {app_id}")
            except Exception as pdf_err:
                print(f" PDF attachment error: {pdf_err}")

        # ===============================
        # SEND REQUEST
        # ===============================

        response = requests.post(url, json=data, headers=headers, timeout=30)

        if response.status_code in [200, 201]:
            print(f" Application status email sent to {to_email}")
            return True

        print(f" Brevo API error: {response.status_code} - {response.text}")
        return False

    except Exception as e:
        print(f" Email API error: {e}")
        import traceback
        traceback.print_exc()
        return False


def calculate_age(birthdate_str):
    if not birthdate_str or birthdate_str == 'none':
        return None
    try:
        from datetime import datetime
        birthdate = datetime.strptime(birthdate_str, '%Y-%m-%d')
        today = ph_now()
        age = today.year - birthdate.year
        if (today.month, today.day) < (birthdate.month, birthdate.day):
            age -= 1
        return age
    except:
        return None

# ==================== CHECK CONTRACT NUMBER UNIQUENESS - CONVERTED TO MYSQL ====================
@app.route("/api/superadmin/check-contract-number/<contract_number>", methods=["GET"])
def check_contract_number(contract_number):
    try:
        # Check in applications table
        app_query = "SELECT contract_number FROM applications WHERE contract_number = %s LIMIT 1"
        app_result = execute_query(app_query, (contract_number,), fetch_one=True)
        
        if app_result:
            return jsonify({"exists": True})
        
        # Check in customers table
        customer_query = "SELECT contract_number FROM customers WHERE contract_number = %s LIMIT 1"
        customer_result = execute_query(customer_query, (contract_number,), fetch_one=True)
        
        if customer_result:
            return jsonify({"exists": True})
        
        return jsonify({"exists": False})
        
    except Exception as e:
        print("Error checking contract number:", e)
        return jsonify({"error": str(e)}), 500
    


@app.route("/api/superadmin/contracts/<contract_number>", methods=["POST"])
def save_contract(contract_number):
    print(f"🟢🟢🟢 SAVE_CONTRACT CALLED! Contract: {contract_number}")
    
    try:
        data = request.get_json()
        print(f"🟢 Data received: {data}")
        
        if not data:
            print(" No data received!")
            return jsonify({"error": "No data provided"}), 400
        
        # I-print ang lahat ng keys at values para makita kung may mali
        print(f"🟢 Data keys: {list(data.keys())}")
        
        # Check if contract already exists
        check_query = "SELECT contract_number FROM contracts WHERE contract_number = %s"
        existing = execute_query(check_query, (contract_number,), fetch_one=True)
        print(f"🟢 Existing contract: {existing}")
        
        # Prepare data for insert
        insert_data = data.copy()
        
        # Convert application_data to JSON if it's a dict
        if 'application_data' in insert_data and isinstance(insert_data['application_data'], dict):
            insert_data['application_data'] = json.dumps(insert_data['application_data'])
            print(f"🟢 Converted application_data to JSON")
        
        # Remove any keys that might cause issues
        if 'id' in insert_data:
            del insert_data['id']
        
        # I-print ang final data bago i-insert
        print(f"🟢 Final data for insert: {insert_data}")
        
        # Build INSERT query
        columns = ', '.join(insert_data.keys())
        placeholders = ', '.join(['%s'] * len(insert_data))
        insert_query = f"INSERT INTO contracts ({columns}) VALUES ({placeholders})"
        print(f"🟢 Insert query: {insert_query}")
        print(f"🟢 Values: {list(insert_data.values())}")
        
        # Execute insert
        result = execute_query(insert_query, list(insert_data.values()))
        print(f"🟢 Execute result: {result}")
        
        # Verify if inserted
        verify_query = "SELECT * FROM contracts WHERE contract_number = %s"
        verify_result = execute_query(verify_query, (contract_number,), fetch_one=True)
        print(f"🟢 Verification result: {verify_result}")
        
        if verify_result:
            print(f" Contract {contract_number} successfully saved to MySQL!")
            return jsonify({"success": True, "message": "Contract saved successfully"})
        else:
            print(f" Contract {contract_number} was NOT saved!")
            return jsonify({"error": "Contract was not saved"}), 500
        
    except Exception as e:
        print(f" Error saving contract: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

import json
# ===============================
# GET CONTRACT BY NUMBER - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/contracts/<contract_number>", methods=["GET"])
def get_contract(contract_number):
    try:
        query = "SELECT * FROM contracts WHERE contract_number = %s"
        contract = execute_query(query, (contract_number,), fetch_one=True)
        
        if not contract:
            return jsonify({"error": "Contract not found"}), 404
        
        # Parse application_data back to dict if it's a JSON string
        if contract.get('application_data') and isinstance(contract.get('application_data'), str):
            try:
                contract['application_data'] = json.loads(contract['application_data'])
            except:
                pass
        
        return jsonify(contract)
        
    except Exception as e:
        print("Error getting contract:", e)
        return jsonify({"error": str(e)}), 500
    
# ===============================
# DOWNLOAD CONTRACT PDF - CONVERTED TO MYSQL
# ===============================
@app.route("/superadmin/download/contract/<app_id>/<contract_number>")
def download_contract_pdf(app_id, contract_number):
    """Generate and download contract PDF - with Addendum and Installment on second page"""
    try:
        # ========== 1. FETCH DATA FROM MYSQL ==========
        # Get application data
        app_query = "SELECT * FROM applications WHERE application_number = %s"
        application_data = execute_query(app_query, (app_id,), fetch_one=True)
        
        if not application_data:
            return "Application not found", 404
        
        # Get contract data
        contract_query = "SELECT * FROM contracts WHERE contract_number = %s"
        contract_data = execute_query(contract_query, (contract_number,), fetch_one=True)
        
        # Parse JSON fields
        if application_data.get('tv_qty'):
            try:
                application_data['tv_qty'] = json.loads(application_data['tv_qty'])
            except:
                application_data['tv_qty'] = []
        
        if application_data.get('tv_brand'):
            try:
                application_data['tv_brand'] = json.loads(application_data['tv_brand'])
            except:
                application_data['tv_brand'] = []
        
        if application_data.get('tv_type'):
            try:
                application_data['tv_type'] = json.loads(application_data['tv_type'])
            except:
                application_data['tv_type'] = []
        
        signature_data = application_data.get('signature')
        
        # If no contract data, create from application data
        if not contract_data:
            first_name = application_data.get('first_name', '')
            middle_name = application_data.get('middle_name', '')
            last_name = application_data.get('last_name', '')
            full_name = ' '.join(filter(None, [first_name, middle_name, last_name])).strip()
            contract_data = {
                'full_name': full_name,
                'first_name': first_name,
                'middle_name': middle_name,
                'last_name': last_name,
                'age': calculate_age(application_data.get('birthdate', '')),
                'civil_status': application_data.get('civil_status', ''),
                'address': f"{application_data.get('barangay', '')}, {application_data.get('city', '')}, {application_data.get('province', '')}".strip(', '),
                'billing_date': application_data.get('billing_date', ''),
                'date_submitted': application_data.get('date_submitted', ''),
                'contract_number': contract_number,
                'signature': signature_data,
                'plan': application_data.get('plan', ''),
                'plan_speed': application_data.get('plan_speed', ''),
                'installation_fee': application_data.get('installation_fee', ''),
                'first_installment_date': application_data.get('first_installment_date', ''),
                'last_installment_date': application_data.get('last_installment_date', '')
            }
        
        full_name = contract_data.get('full_name', '')
        if not full_name:
            first = contract_data.get('first_name', '') or application_data.get('first_name', '')
            middle = contract_data.get('middle_name', '') or application_data.get('middle_name', '')
            last = contract_data.get('last_name', '') or application_data.get('last_name', '')
            full_name = ' '.join(filter(None, [first, middle, last])).strip()
        if not full_name:
            full_name = "Customer Name Not Available"
        
        age = contract_data.get('age', '') or calculate_age(application_data.get('birthdate', ''))
        civil_status = contract_data.get('civil_status', '') or application_data.get('civil_status', '')
        address = contract_data.get('address', '') or f"{application_data.get('barangay', '')}, {application_data.get('city', '')}, {application_data.get('province', '')}".strip(', ')
        billing_date = contract_data.get('billing_date', '') or application_data.get('billing_date', '')
        date_submitted = contract_data.get('date_submitted', '') or application_data.get('date_submitted', '')
        plan_name = contract_data.get('plan', '') or application_data.get('plan', '')
        plan_speed = contract_data.get('plan_speed', '') or application_data.get('plan_speed', '')
        installation_fee = contract_data.get('installation_fee', '') or application_data.get('installation_fee', '')
        first_installment = contract_data.get('first_installment_date', '') or application_data.get('first_installment_date', '')
        last_installment = contract_data.get('last_installment_date', '') or application_data.get('last_installment_date', '')
        
        # Check if installment plan
        is_installment = False
        if installation_fee:
            fee_lower = installation_fee.lower()
            if 'installment' in fee_lower:
                is_installment = True
        
        # Format dates for display
        def format_month_year(date_str):
            if not date_str:
                return '_____________'
            try:
                year, month = date_str.split('-')
                month_names = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
                return f"{month_names[int(month) - 1]} {year}"
            except:
                return '_____________'
        
        first_installment_formatted = format_month_year(first_installment)
        last_installment_formatted = format_month_year(last_installment)
        approval_date = ph_now().strftime('%B %d, %Y')
        
        # ========== 2. PDF SETUP ==========
        from reportlab.lib.pagesizes import LEGAL
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        import io, base64, requests, os, re
        from flask import current_app, send_file
        from PIL import Image as PILImage
        import tempfile
        
        buffer = io.BytesIO()
        
        doc = SimpleDocTemplate(buffer, pagesize=LEGAL,
                                rightMargin=36, leftMargin=36,
                                topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        story = []
        
        # ========== SHARED UPLOADS CONFIG ==========
        SHARED_UPLOADS_BASE = r"C:\xampp\htdocs\cablevision_uploads"

        def get_image_bytes_from_cloudinary(image_path):
            """Get image bytes from Cloudinary URL"""
            if not image_path:
                return None
            
            # Convert to Cloudinary URL
            cloudinary_url = get_cloudinary_url(image_path)
            
            if cloudinary_url and cloudinary_url.startswith('http'):
                try:
                    print(f" Downloading from Cloudinary: {cloudinary_url}")
                    response = requests.get(cloudinary_url, timeout=30)
                    if response.status_code == 200:
                        print(f" Downloaded {len(response.content)} bytes")
                        return response.content
                    else:
                        print(f" Cloudinary download failed: {response.status_code}")
                        return None
                except Exception as e:
                    print(f" Error downloading from Cloudinary: {e}")
                    return None
            
            return None
        
        # ========== HELPER: Get image bytes from file path ==========
        def get_image_bytes(image_path_or_data):
            """Get image bytes from file path or base64 data or Cloudinary"""
            if not image_path_or_data:
                return None
            
            # TRY CLOUDINARY FIRST
            img_bytes = get_image_bytes_from_cloudinary(image_path_or_data)
            if img_bytes:
                return img_bytes
            
            # Check if it's a file path
            if isinstance(image_path_or_data, str):
                # Pattern: /shared-uploads/application_uploads/3482179683/signature_1783904528.jpg
                filename = os.path.basename(image_path_or_data)
                
                # Build path using app_id
                app_folder = str(app_id)
                full_path = os.path.join(SHARED_UPLOADS_BASE, 'application_uploads', app_folder, filename)
                
                print(f" Looking for signature: {full_path}")
                
                if os.path.exists(full_path):
                    try:
                        with open(full_path, 'rb') as f:
                            img_bytes = f.read()
                        print(f" Found signature: {full_path} ({len(img_bytes)} bytes)")
                        return img_bytes
                    except Exception as e:
                        print(f" Error reading signature: {e}")
                        return None
                
                # Try extracting app_id from path
                match = re.search(r'/application_uploads/(\d+)/', image_path_or_data)
                if match:
                    extracted_folder = match.group(1)
                    alt_path = os.path.join(SHARED_UPLOADS_BASE, 'application_uploads', extracted_folder, filename)
                    print(f" Trying extracted: {alt_path}")
                    if os.path.exists(alt_path):
                        try:
                            with open(alt_path, 'rb') as f:
                                img_bytes = f.read()
                            print(f" Found signature: {alt_path} ({len(img_bytes)} bytes)")
                            return img_bytes
                        except Exception as e:
                            print(f" Error reading signature: {e}")
                            return None
                
                # Try base64 decoding
                if 'base64,' in image_path_or_data or 'data:image' in image_path_or_data:
                    try:
                        if 'base64,' in image_path_or_data:
                            image_path_or_data = image_path_or_data.split('base64,')[1]
                        elif 'data:image' in image_path_or_data:
                            match = re.search(r'data:image/(png|jpeg|jpg|gif);base64,(.+)', image_path_or_data)
                            if match:
                                image_path_or_data = match.group(2)
                        image_path_or_data = image_path_or_data.strip()
                        img_bytes = base64.b64decode(image_path_or_data)
                        print(f" Decoded base64 signature ({len(img_bytes)} bytes)")
                        return img_bytes
                    except Exception as e:
                        print(f" Error decoding base64: {e}")
                        return None
            
            return None
        
        # ========== HELPER: Convert image for ReportLab ==========
        def convert_image_for_reportlab(img_bytes, max_width=180, max_height=50):
            """Convert image to format compatible with ReportLab"""
            if not img_bytes:
                return None
            
            try:
                img = PILImage.open(io.BytesIO(img_bytes))
                print(f" Signature opened: {img.format}, {img.size}, {img.mode}")
                
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                
                if img.width > max_width or img.height > max_height:
                    img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
                    print(f" Signature resized to: {img.size}")
                
                output = io.BytesIO()
                img.save(output, format='JPEG', quality=85)
                output.seek(0)
                
                img_reportlab = Image(output)
                img_reportlab.drawWidth = min(img.width, max_width)
                img_reportlab.drawHeight = min(img.height, max_height)
                print(f" Signature converted for ReportLab")
                return img_reportlab
                
            except Exception as e:
                print(f" Error converting signature: {e}")
                return None
        
        # ========== GET SIGNATURE IMAGE ==========
        signature_img = None
        print(f" Signature data from DB: {signature_data}")
        
        img_bytes = get_image_bytes(signature_data)
        if img_bytes:
            signature_img = convert_image_for_reportlab(img_bytes, max_width=180, max_height=60)
        
        if not signature_img:
            print(" No signature image available")
        
        # ========== 3. STYLES ==========
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=11,
            alignment=1,
            spaceAfter=2,
            fontName='Times-Bold'
        )
        heading_style = ParagraphStyle(
            'HeadingStyle',
            parent=styles['Normal'],
            fontSize=11,
            alignment=1,
            spaceBefore=8,
            spaceAfter=6,
            fontName='Times-Bold'
        )
        contract_style = ParagraphStyle(
            'ContractStyle',
            parent=styles['Normal'],
            fontSize=8.5,
            leading=10,
            alignment=4,
            spaceAfter=2
        )
        addendum_style = ParagraphStyle(
            'AddendumStyle',
            parent=styles['Normal'],
            fontSize=9,
            leading=12,
            alignment=4,
            spaceAfter=4
        )
        signature_name_style = ParagraphStyle(
            'SignatureNameStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,
            fontName='Helvetica',
            textDecoration='underline'
        )
        signature_label_style = ParagraphStyle(
            'SignatureLabelStyle',
            parent=styles['Normal'],
            fontSize=7,
            alignment=1,
            fontName='Helvetica',
            textColor=colors.grey
        )
        
        # ========== PAGE 1: HEADER AND MAIN CONTRACT ==========
        left_logo_path = os.path.join(current_app.root_path, 'static', 'logo.png')
        right_logo_path = os.path.join(current_app.root_path, 'static', 'logo_right.png')
        
        left_logo_exists = os.path.exists(left_logo_path)
        right_logo_exists = os.path.exists(right_logo_path)
        
        left_logo = None
        right_logo = None
        
        if left_logo_exists:
            left_logo = Image(left_logo_path, width=60, height=60)
        if right_logo_exists:
            right_logo = Image(right_logo_path, width=60, height=60)
        
        title_1 = Paragraph("CABLE TELEVISION/CABLE ONLY/OR", header_style)
        title_2 = Paragraph("CABLE &amp; INTERNET SERVICE CONTRACT", header_style)
        title_3 = Paragraph(f"NO. <u>{contract_number}</u>", header_style)
        
        center_text = Table(
            [[title_1], [title_2], [Spacer(1, 2)], [title_3]],
            colWidths=[360]
        )
        center_text.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]))
        
        if left_logo_exists and right_logo_exists:
            header_table = Table([[left_logo, center_text, right_logo]], colWidths=[70, 360, 70])
        elif left_logo_exists and not right_logo_exists:
            header_table = Table([[left_logo, center_text, '']], colWidths=[70, 360, 70])
        elif not left_logo_exists and right_logo_exists:
            header_table = Table([['', center_text, right_logo]], colWidths=[70, 360, 70])
        else:
            header_table = Table([[center_text]], colWidths=[500])
        
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'CENTER'),
            ('ALIGN', (2,0), (2,0), 'RIGHT'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        
        story.append(header_table)
        story.append(Spacer(1, 12))
        story.append(Paragraph("CONTRACT TERMS AND CONDITIONS", heading_style))
        story.append(Spacer(1, 3))
        
        # Opening statement
        story.append(Paragraph(
            f"I, <strong>{full_name}</strong>, legal age, <strong>{age}</strong> years old, {civil_status} "
            f"and residing at <strong>{address}</strong> hereby apply and subscribed for the service of "
            f"CABLE &amp; INTERNET and agree to the following terms and conditions:",
            contract_style
        ))
        story.append(Spacer(1, 4))
        
        # Payment
        payment_text = (
            f"<strong>Payment:</strong> The subscriber shall pay a Non-Refundable connection fee of P 1800 and "
            f"cable in excess of 100 meters at P10.00 per meter. For CABLE/INTERNET BUNDLE subscriber, a one (1) "
            f"month subscription fee of P800 shall be paid upon installation and activation of the service. "
            f"Succeeding monthly subscription fee is due and payable every <strong>{billing_date}</strong> of each month. "
            f"Failure to pay the monthly subscription fee on due date and after the grace period of 7 days will mean "
            f"automatic disconnection of cable/internet service. "
            f"<strong>For CABLE SUBSCRIBER ONLY</strong>, subscriber shall pay a monthly subscription fee of P per month "
            f"from the period the TV set is activated on or before of each month. "
            f"The company shall have the right to discontinue/terminate/cancel and effect disconnection of Cable TV services "
            f"in case of default or non-payment of accounts for two (2) succeeding payments."
        )
        story.append(Paragraph(payment_text, contract_style))
        story.append(Spacer(1, 2))
        
        # Deposit
        story.append(Paragraph(
            "<strong>Deposit:</strong> Subscriber, who leases his/her house or does not own the house where service "
            "will be installed, shall pay a DEPOSIT upon installation. A deposit equivalent to one (1) month subscription fee "
            "for CABLE/INTERNET BUNDLE subscriber while two (2) months subscription fee for CABLE SUBSCRIBER ONLY. "
            "The said deposit cannot be applied to the monthly fee and shall only be refunded upon termination of the contract "
            "and upon pull out of all equipment installed in the premises of the subscriber. Should the subscriber wishes to apply "
            "for reconnection, a reconnection fee of P500.00 shall be paid plus the Deposit and the one (1) month advance "
            "subscription fee for CABLE/INTERNET BUNDLE subscriber. For CABLE SUBSCRIBER ONLY, a reconnection fee of P300.00 "
            "plus the DEPOSIT shall be paid.",
            contract_style
        ))
        story.append(Spacer(1, 2))
        
        # Access to the Premises
        story.append(Paragraph(
            "<strong>Access to the Premises:</strong> The subscriber authorizes our employees, contractors and "
            "representatives to enter your premise in order to install, maintain, inspect, repair, remove and replace "
            "Equipment at a time mutually agreeable upon by both parties.",
            contract_style
        ))
        story.append(Spacer(1, 2))
        
        # Subscriber Usage
        story.append(Paragraph(
            "<strong>Subscriber Usage:</strong> The subscriber shall not in any way use his subscription for commercial purposes. "
            "Transmission of any Internet content which violates national or international law is prohibited. This includes but "
            "not limited to copyrighted materials, those legally adjudged to be threat to national security, or intruding into the "
            "privacy of individuals, offensive on moral, religious, racial or political grounds; abusive, indecent, obscene or "
            "menacing nature of material or information, infringement of intellectual property rights of any person as well as trade secrets.",
            contract_style
        ))
        story.append(Spacer(1, 2))
        
        # Relocating Equipment
        story.append(Paragraph(
            "<strong>Relocating Equipment:</strong> The subscriber is not allowed to relocate equipment installed in their premises. "
            "However, equipment may be relocated by the company's authorized representatives upon the request of the subscriber at a time "
            "mutually agreeable to both parties. Applicable fees and charges may apply.",
            contract_style
        ))
        story.append(Spacer(1, 2))
        
        # Cable Modem and Setup Box
        story.append(Paragraph(
            "<strong>Cable Modem and Setup Box:</strong> The subscriber will be given FREE USE of a Cable Modem and Set Top Box. "
            "This equipment will remain the property of CABLEVISION SYSTEMS CORP. For any Cable TV Extension the subscriber will have to pay "
            "for the cost of the SET TOP BOX amounting to 1400 and a HUB amounting to 420. There will be no additional cost on the monthly "
            "subscription. All equipment has one (1) year warranty against factory defects. If the defect was due to improper use and mishandling "
            "by the user during the warranty period, the cost of replacement will be chargeable to the account of the subscriber. If cable modem "
            "or Set Top Box becomes defective after the warranty period, cost of the new equipment is chargeable to the subscriber.",
            contract_style
        ))
        story.append(Spacer(1, 2))
        
        # Termination/Suspension
        story.append(Paragraph(
            "<strong>Termination/Suspension of Service:</strong> The company reserves the right to suspend or terminate this contract without "
            "prior notice and pull out equipment provided at the subscriber's premises due to non-payment of all applicable fees and charges within "
            "the period and shall not be held liable for any damage; or loss which the Subscriber may incur by reason of suspension and/or termination "
            "of services based on this agreement.",
            contract_style
        ))
        story.append(Spacer(1, 2))
        
        # Disclaimer
        disclaimer_text = (
            "<strong>Disclaimer:</strong> Cablevision Systems Corp./MyCv Broadband shall not be held liable for any damages or delay in business transaction "
            "or communication of the subscriber or whatsoever, the subscriber may suffer or may have suffered due to the use of myCv Broadband Services. "
            "This includes but not limited to any loss of profits, incidental or consequential damages arising out of the Costumer's use of or inability to use; "
            "any loss of information howsoever caused whether as a result of any interruption, suspension, or termination of the Service or otherwise, or for the "
            "contents, accuracy or quality of information available, received or transmitted through the Service; or for failure of the Subscriber to comply with "
            "applicable laws, rules and regulations and all the terms prescribed by the Philippine National Telecommunications Commission for the use of any "
            "telecommunication systems, service or equipment. myCv Broadband shall not be liable for any delay or failure in the performance of service under "
            "this agreement resulting from acts beyond its control, including without limitation, acts of God, acts or regulations of any government or national authority, "
            "war or national emergency, accident, fire, electric power failure, temporary loss of signal not attributed to myCv Broadband, lightning, strikes, lock-outs, "
            "industrial disputes whether or not involving myCv Broadband employees."
        )
        story.append(Paragraph(disclaimer_text, contract_style))
        story.append(Spacer(1, 2))
        
        # Right to modify
        story.append(Paragraph(
            "myCv Broadband reserves the right to adjust, modify, amend or supplements these terms and condition as the service may require. "
            "myCv Broadband will advise SUBSCRIBER of any change by sending him notice setting out these changes.",
            contract_style
        ))
        story.append(Spacer(1, 2))
        
        # Governing Law
        story.append(Paragraph(
            "<strong>Governing Law and Jurisdiction:</strong> The Laws of the Republic of the Philippines governs this Agreement and the Subscriber and myCv Broadband "
            "hereby submit to the exclusive jurisdiction of the courts of Sta. Cruz, Laguna, Philippines.",
            contract_style
        ))
        story.append(Spacer(1, 6))
        
        # Acknowledgment
        story.append(Paragraph(
            "I hereby acknowledge that I have read and understood all the terms and conditions herein and that I voluntarily sign this agreement with full knowledge "
            "and consent of everything this Agreement contains, implies and entails.",
            contract_style
        ))
        story.append(Spacer(1, 10))
        
        # Top Signature Section
        if signature_img:
            top_left_data = [
                [signature_img],
                [Spacer(1, 3)],
                [Paragraph(f"<u>{full_name}</u>", signature_name_style)],
                [Spacer(1, 2)],
                [Paragraph("Subscriber's Signature Over Printed Name", signature_label_style)]
            ]
        else:
            top_left_data = [
                [Paragraph("_________________________", signature_label_style)],
                [Spacer(1, 3)],
                [Paragraph(f"<u>{full_name}</u>", signature_name_style)],
                [Spacer(1, 2)],
                [Paragraph("Subscriber's Signature Over Printed Name", signature_label_style)]
            ]
        
        top_right_data = [
            [Spacer(1, 50)],
            [Paragraph(f"<u>{date_submitted}</u>", signature_name_style)],
            [Spacer(1, 2)],
            [Paragraph("Date", signature_label_style)]
        ]
        
        top_left_table = Table(top_left_data, colWidths=[220])
        top_left_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        
        top_right_table = Table(top_right_data, colWidths=[220])
        top_right_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        
        top_signature_table = Table([[top_left_table, top_right_table]], colWidths=[220, 220])
        top_signature_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        
        story.append(top_signature_table)
        
        # ========== PAGE BREAK ==========
        story.append(PageBreak())
        
        # ========== PAGE 2: ADDENDUM AND INSTALLMENT SECTIONS ==========
        story.append(Spacer(1, 30))
        story.append(Paragraph("<strong>CABLEVISION SYSTEMS CORPORATION</strong>", heading_style))
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"<strong>ADDENDUM TO CONTRACT NUMBER {contract_number}</strong>", heading_style))
        story.append(Spacer(1, 15))
        
        addendum_text = (
            f"That I, <strong>{full_name}</strong> holder of CONTRACT Number <strong>{contract_number}</strong> dated <strong>{approval_date}</strong> "
            f"wishes to avail of your INTERNET SERVICE under <strong>{plan_name} ({plan_speed})</strong>. To take effect on <strong>_________________________</strong>."
        )
        story.append(Paragraph(addendum_text, addendum_style))
        story.append(Spacer(1, 8))
        
        story.append(Paragraph(
            "This is also to acknowledge that I have to pay in advance the monthly dues corresponding to the plan that I choose and it is understood that the "
            "TERMS AND CONDITIONS on the original contract remain.",
            addendum_style
        ))
        story.append(Spacer(1, 25))
        
        # INSTALLMENT SECTION
        story.append(Paragraph("<strong>AGREEMENT TO PAY ON INSTALLMENT</strong>", heading_style))
        story.append(Paragraph("<strong>FOR THE INSTALLATION FEE AND/OR SET TOP BOX FOR TV EXTENSION</strong>", heading_style))
        story.append(Spacer(1, 15))
        
        if is_installment:
            display_full_name = full_name
            display_contract_number = contract_number
            display_first_date = first_installment_formatted
            display_last_date = last_installment_formatted
        else:
            display_full_name = '_____________'
            display_contract_number = '_____________'
            display_first_date = '_____________'
            display_last_date = '_____________'
        
        installment_text = (
            f"That I, <strong>{display_full_name}</strong> holder of contract no. <strong>{display_contract_number}</strong> wishes to avail of the INSTALLMENT PLAN "
            f"for the INSTALLATION FEE starting <strong>{display_first_date}</strong> up to <strong>{display_last_date}</strong> "
            f"and the SET TOP BOX for our <strong>_________</strong> TV Extension/s for five (5) months."
        )
        
        story.append(Paragraph(installment_text, addendum_style))
        story.append(Spacer(1, 12))
        
        story.append(Paragraph(
            "<strong>NOTE:</strong> In the event that the account is disconnected during the said period, the remaining installment shall be paid in full.",
            addendum_style
        ))
        story.append(Spacer(1, 40))
        
        # Bottom Signature Section
        if signature_img:
            bottom_signature_data = [
                [signature_img],
                [Spacer(1, 5)],
                [Paragraph(f"<u>{full_name}</u>", signature_name_style)],
                [Spacer(1, 2)],
                [Paragraph("Signature over printed name", signature_label_style)]
            ]
        else:
            bottom_signature_data = [
                [Paragraph("_________________________", signature_label_style)],
                [Spacer(1, 5)],
                [Paragraph(f"<u>{full_name}</u>", signature_name_style)],
                [Spacer(1, 2)],
                [Paragraph("Signature over printed name", signature_label_style)]
            ]
        
        bottom_signature_table = Table(bottom_signature_data, colWidths=[250])
        bottom_signature_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        
        right_aligned_table = Table([[bottom_signature_table]], colWidths=[500])
        right_aligned_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        
        story.append(right_aligned_table)
        
        # ========== BUILD PDF ==========
        doc.build(story)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"Service_Contract_{contract_number}.pdf",
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"Error generating contract PDF: {e}")
        import traceback
        traceback.print_exc()
        return f"Error generating PDF: {str(e)}", 500


# ===============================
# CALCULATE AGE - CONVERTED TO MYSQL (No changes needed)
# ===============================
def calculate_age(birthdate):
    if not birthdate:
        return ''
    try:
        from datetime import datetime
        birth = datetime.strptime(birthdate, "%Y-%m-%d")
        today = ph_now()
        age = today.year - birth.year
        if (today.month, today.day) < (birth.month, birth.day):
            age -= 1
        return str(age)
    except:
        return ''
    
    
def generate_application_pdf(application_number, application_data=None, contract_number=None):
    """Generate PDF for an application using MySQL data with Cloudinary images"""
    import io, base64, os, re
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import requests
    import json
    from PIL import Image as PILImage
    import tempfile

    # If data not provided, fetch from MySQL
    if application_data is None:
        query = """
            SELECT * FROM applications 
            WHERE application_number = %s
        """
        application_data = execute_query(query, (application_number,), fetch_one=True)
    
    if not application_data:
        print(f" Application not found for: {application_number}")
        return None
    
    # Parse JSON fields (tv_qty, tv_brand, tv_type)
    if application_data.get('tv_qty'):
        try:
            application_data['tv_qty'] = json.loads(application_data['tv_qty'])
        except:
            application_data['tv_qty'] = []
    
    if application_data.get('tv_brand'):
        try:
            application_data['tv_brand'] = json.loads(application_data['tv_brand'])
        except:
            application_data['tv_brand'] = []
    
    if application_data.get('tv_type'):
        try:
            application_data['tv_type'] = json.loads(application_data['tv_type'])
        except:
            application_data['tv_type'] = []
    
    # Add contract number to data if provided
    if contract_number:
        application_data["contract_number"] = contract_number

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # ================= CLOUDINARY HELPER =================
    def get_image_from_cloudinary(image_path):
        """Get image from Cloudinary URL or convert relative path"""
        if not image_path:
            return None
        
        # Convert to Cloudinary URL if not already
        if not image_path.startswith('http'):
            cloudinary_url = get_cloudinary_url(image_path)
        else:
            cloudinary_url = image_path
        
        if cloudinary_url and cloudinary_url.startswith('http'):
            try:
                print(f" Downloading from Cloudinary: {cloudinary_url[:80]}...")
                response = requests.get(cloudinary_url, timeout=30)
                if response.status_code == 200:
                    print(f" Downloaded {len(response.content)} bytes")
                    return response.content
                else:
                    print(f" Cloudinary download failed: {response.status_code}")
                    return None
            except Exception as e:
                print(f" Error downloading from Cloudinary: {e}")
                return None
        
        return None

    # ================= HELPER: Get image bytes =================
    def get_image_bytes(image_path_or_data, app_id=None):
        """Get image bytes from Cloudinary or local path"""
        if not image_path_or_data:
            return None
        
        # TRY CLOUDINARY FIRST
        cloudinary_bytes = get_image_from_cloudinary(image_path_or_data)
        if cloudinary_bytes:
            return cloudinary_bytes
        
        # If app_id not provided, use application_number
        if app_id is None:
            app_id = application_number
        
        # Check if it's a file path
        if isinstance(image_path_or_data, str):
            # Extract filename
            filename = os.path.basename(image_path_or_data)
            
            # Try local path (fallback)
            SHARED_UPLOADS_BASE = r"C:\xampp\htdocs\cablevision_uploads"
            full_path = os.path.join(SHARED_UPLOADS_BASE, 'application_uploads', str(app_id), filename)
            
            print(f" Looking for image locally: {full_path}")
            
            if os.path.exists(full_path):
                try:
                    with open(full_path, 'rb') as f:
                        img_bytes = f.read()
                    print(f" Found image locally: {full_path} ({len(img_bytes)} bytes)")
                    return img_bytes
                except Exception as e:
                    print(f" Error reading image: {e}")
                    return None
            
            # Try extracting app_id from path
            match = re.search(r'/application_uploads/(\d+)/', image_path_or_data)
            if match:
                extracted_folder = match.group(1)
                alt_path = os.path.join(SHARED_UPLOADS_BASE, 'application_uploads', extracted_folder, filename)
                print(f" Trying extracted: {alt_path}")
                if os.path.exists(alt_path):
                    try:
                        with open(alt_path, 'rb') as f:
                            img_bytes = f.read()
                        print(f" Found image: {alt_path} ({len(img_bytes)} bytes)")
                        return img_bytes
                    except Exception as e:
                        print(f" Error reading image: {e}")
                        return None
            
            # Try base64 decoding
            if 'base64,' in image_path_or_data or 'data:image' in image_path_or_data:
                try:
                    if 'base64,' in image_path_or_data:
                        image_path_or_data = image_path_or_data.split('base64,')[1]
                    elif 'data:image' in image_path_or_data:
                        match = re.search(r'data:image/(png|jpeg|jpg|gif);base64,(.+)', image_path_or_data)
                        if match:
                            image_path_or_data = match.group(2)
                    image_path_or_data = image_path_or_data.strip()
                    img_bytes = base64.b64decode(image_path_or_data)
                    print(f" Decoded base64 image ({len(img_bytes)} bytes)")
                    return img_bytes
                except Exception as e:
                    print(f" Error decoding base64: {e}")
                    return None
        
        return None

    # ================= HELPER: Convert image for ReportLab =================
    def convert_image_for_reportlab(img_bytes, max_width=None, max_height=None):
        """Convert image to format compatible with ReportLab"""
        if not img_bytes:
            return None
        
        try:
            # Open with PIL
            img = PILImage.open(io.BytesIO(img_bytes))
            print(f" Image opened: {img.format}, {img.size}, {img.mode}")
            
            # Convert to RGB if needed
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Resize if dimensions provided
            if max_width and max_height:
                img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
                print(f" Image resized to: {img.size}")
            
            # Save to BytesIO as JPEG
            output = io.BytesIO()
            img.save(output, format='JPEG', quality=85)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            print(f" Error converting image: {e}")
            return None

    # ================= HELPER: Draw image safely =================
    def draw_image_safe(p, image_data, x, y, width, height, label="Image", app_id=None):
        """Safely draw an image on the PDF"""
        try:
            print(f" Drawing {label}...")
            img_bytes = get_image_bytes(image_data, app_id)
            if img_bytes:
                converted = convert_image_for_reportlab(img_bytes, max_width=width, max_height=height)
                if converted:
                    # Use temp file for ReportLab
                    with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as tmp_file:
                        tmp_file.write(converted)
                        tmp_path = tmp_file.name
                    
                    try:
                        img = ImageReader(tmp_path)
                        p.drawImage(img, x, y, width, height, preserveAspectRatio=True, mask='auto')
                        print(f" Drew {label} successfully")
                        return True
                    except Exception as e:
                        print(f" Error in drawImage for {label}: {e}")
                        return False
                    finally:
                        try:
                            os.unlink(tmp_path)
                        except:
                            pass
            else:
                print(f" No image data for {label}")
                return False
        except Exception as e:
            print(f" Error drawing {label}: {e}")
            return False

    # ================= MAX PAGES =================
    MAX_PAGES = 4
    current_page = 1
    y = height - 120

    # ================= HEADER =================
    def draw_header():
        nonlocal y
        try:
            logo = ImageReader("static/logo1.png")
            p.drawImage(logo, 40, height - 90, width=60, height=60, mask='auto')
        except:
            pass

        p.setFont("Helvetica-Bold", 16)
        p.drawString(110, height - 60, "APPLICATION FORM")

        p.setFont("Helvetica-Bold", 10)
        p.drawRightString(width - 50, height - 60, f"Application No: {application_number}")

        p.setFont("Helvetica", 8)
        p.drawString(110, height - 75, "Sitio Sampaguita, Brgy. Pagsawitan, Santa Cruz, 4009 Laguna")
        p.drawString(110, height - 87, "Tel: (049) 501-1495 | Fax: (049) 501-0229 | Mobile: 0917 501 0341")

        y = height - 110

    def draw_page_number():
        p.setFont("Helvetica-Bold", 10)
        p.setFillColorRGB(0.4, 0.4, 0.4)
        p.drawRightString(width - 25, 20, str(current_page))
        p.setFillColorRGB(0, 0, 0)

    def new_page():
        nonlocal y, current_page
        if current_page >= MAX_PAGES:
            return False
        p.showPage()
        current_page += 1
        draw_header()
        draw_page_number()
        y = height - 110
        return True

    def ensure_space(required):
        nonlocal y
        if y - required < 50:
            return new_page()
        return True

    def draw_section_title(title):
        nonlocal y
        ensure_space(30)
        p.setFont("Helvetica-Bold", 12)
        p.setFillColorRGB(0, 0.4, 0.6)
        p.drawString(50, y, title)
        p.setFillColorRGB(0, 0, 0)
        y -= 22

    def draw_section_title_centered(title):
        nonlocal y
        ensure_space(30)
        p.setFont("Helvetica-Bold", 14)
        p.setFillColorRGB(0, 0.4, 0.6)
        p.drawCentredString(width / 2, y, title)
        p.setFillColorRGB(0, 0, 0)
        y -= 25

    # Two-column field drawer
    def draw_two_columns(fields):
        nonlocal y
        col1_x = 50
        col2_x = 310
        label_width = 120
        value_x = col1_x + label_width + 5
        
        for i in range(0, len(fields), 2):
            ensure_space(20)
            # Left column
            label1, value1 = fields[i]
            p.setFont("Helvetica-Bold", 9)
            p.drawString(col1_x, y, f"{label1}:")
            p.setFont("Helvetica", 9)
            val1_str = str(value1) if value1 and value1 != "-" and value1 != "none" else "___________________"
            if len(val1_str) > 35:
                val1_str = val1_str[:32] + "..."
            p.drawString(value_x, y, val1_str)
            
            # Right column
            if i + 1 < len(fields):
                label2, value2 = fields[i + 1]
                p.setFont("Helvetica-Bold", 9)
                p.drawString(col2_x, y, f"{label2}:")
                p.setFont("Helvetica", 9)
                val2_str = str(value2) if value2 and value2 != "-" and value2 != "none" else "___________________"
                if len(val2_str) > 30:
                    val2_str = val2_str[:27] + "..."
                p.drawString(col2_x + label_width + 5, y, val2_str)
            
            y -= 18
        y -= 5

    # ================= Draw images from Cloudinary =================
    def draw_images_top_bottom(label1, img1_data, label2, img2_data, img_width=280, img_height=190):
        nonlocal y
        
        # Front ID
        ensure_space(img_height + 60)
        p.setFont("Helvetica-Bold", 11)
        p.drawCentredString(width / 2, y, label1)
        y -= 22
        
        if draw_image_safe(p, img1_data, (width - img_width) / 2, y - img_height, img_width, img_height, label1, application_number):
            y -= img_height + 35
        else:
            p.setFont("Helvetica", 9)
            p.setFillColorRGB(0.5, 0.5, 0.5)
            p.drawCentredString(width / 2, y, "No image provided")
            p.setFillColorRGB(0, 0, 0)
            y -= 25
        
        # Back ID
        ensure_space(img_height + 60)
        p.setFont("Helvetica-Bold", 11)
        p.drawCentredString(width / 2, y, label2)
        y -= 22
        
        if draw_image_safe(p, img2_data, (width - img_width) / 2, y - img_height, img_width, img_height, label2, application_number):
            y -= img_height + 35
        else:
            p.setFont("Helvetica", 9)
            p.setFillColorRGB(0.5, 0.5, 0.5)
            p.drawCentredString(width / 2, y, "No image provided")
            p.setFillColorRGB(0, 0, 0)
            y -= 25

    # ================= Signature section =================
    def draw_signature_section(signature_img, full_name):
        nonlocal y
        
        y -= 15
        
        sig_width = 250
        sig_height = 85
        
        if draw_image_safe(p, signature_img, (width - sig_width) / 2, y - sig_height, sig_width, sig_height, "Signature", application_number):
            y -= sig_height + 20
        else:
            p.setFont("Helvetica", 9)
            p.setFillColorRGB(0.5, 0.5, 0.5)
            p.drawCentredString(width / 2, y, "No signature provided")
            p.setFillColorRGB(0, 0, 0)
            y -= 25
        
        p.setFont("Helvetica", 10)
        p.drawCentredString(width / 2, y, full_name if full_name else "_________________________")
        y -= 20
        
        p.setFont("Helvetica", 8)
        p.setFillColorRGB(0.4, 0.4, 0.4)
        p.drawCentredString(width / 2, y, "signature over printed name")
        p.setFillColorRGB(0, 0, 0)
        y -= 30

    # ================= START BUILDING =================
    draw_header()
    draw_page_number()

    # ================= PAGE 1 =================
    draw_section_title("I. PERSONAL INFORMATION")
    draw_two_columns([
        ("Last Name", application_data.get("last_name")),
        ("First Name", application_data.get("first_name")),
        ("Middle Name", application_data.get("middle_name")),
        ("Suffix", application_data.get("suffix")),
        ("Date of Birth", application_data.get("birthdate")),
        ("Place of Birth", application_data.get("place_of_birth")),
        ("Sex", application_data.get("sex")),
        ("Civil Status", application_data.get("civil_status")),
        ("Citizenship", application_data.get("citizenship")),
        ("Occupation", application_data.get("occupation")),
    ])

    draw_section_title("II. FAMILY DETAILS")
    draw_two_columns([
        ("Mother's Maiden Name", application_data.get("mother_maiden_name")),
        ("Father's Name", application_data.get("father_name")),
    ])

    draw_section_title("III. CONTACT & ADDRESS")
    draw_two_columns([
        ("Mobile Number", application_data.get("mobile")),
        ("Email Address", application_data.get("email")),
        ("Home Ownership", application_data.get("home_ownership")),
        ("House No./Unit", application_data.get("house_number")),
        ("Nearest Landmark", application_data.get("landmark")),
        ("Street/Village", application_data.get("address")),
    ])

    # Billing Address
    p.setFont("Helvetica-Bold", 9)
    p.drawString(50, y, "Billing Address:")
    p.setFont("Helvetica", 9)
    
    billing_address = application_data.get("billing_address", "")
    if not billing_address or billing_address == "-" or billing_address == "none":
        billing_address = "_________________________"
    
    from reportlab.pdfbase.pdfmetrics import stringWidth
    max_width = 400
    words = billing_address.split()
    lines = []
    current_line = ""
    
    for word in words:
        test_line = current_line + (" " if current_line else "") + word
        if stringWidth(test_line, "Helvetica", 9) <= max_width:
            current_line = test_line
        else:
            lines.append(current_line)
            current_line = word
    
    if current_line:
        lines.append(current_line)
    
    for line in lines:
        p.drawString(170, y, line)
        y -= 14
    y -= 5

    draw_section_title("IV. EMPLOYMENT DETAILS")
    draw_two_columns([
        ("Employer / Company", application_data.get("employer")),
        ("Business Phone", application_data.get("business_phone")),
        ("Business Address", application_data.get("business_address")),
        ("", ""),
    ])

    civil_status = application_data.get("civil_status", "")
    if civil_status and civil_status.lower() in ["married", "Married"]:
        draw_section_title("V. SPOUSE INFORMATION")
        draw_two_columns([
            ("Spouse Full Name", application_data.get("spouse_name")),
            ("Spouse Occupation", application_data.get("spouse_occupation")),
            ("Spouse Employer", application_data.get("spouse_employer")),
            ("Spouse Phone", application_data.get("spouse_phone")),
        ])

    draw_section_title("VI. SERVICE PLAN")
    draw_two_columns([
        ("Service Type / Plan", application_data.get("service_type")),
        ("Installation Fee", application_data.get("installation_fee")),
    ])

    # Installation Phone
    p.setFont("Helvetica-Bold", 9)
    p.drawString(50, y, "Installation Phone:")
    p.setFont("Helvetica", 9)
    installation_phone = application_data.get("installation_phone", "")
    if not installation_phone or installation_phone == "-" or installation_phone == "none":
        installation_phone = "_________________________"
    p.drawString(170, y, installation_phone)
    y -= 18
    y -= 5

    # Installation Address
    p.setFont("Helvetica-Bold", 9)
    p.drawString(50, y, "Installation Address:")
    p.setFont("Helvetica", 9)
    
    installation_address = application_data.get("installation_address", "")
    if not installation_address or installation_address == "-" or installation_address == "none":
        installation_address = "_________________________"
    
    words = installation_address.split()
    lines = []
    current_line = ""
    
    for word in words:
        test_line = current_line + (" " if current_line else "") + word
        if stringWidth(test_line, "Helvetica", 9) <= max_width:
            current_line = test_line
        else:
            lines.append(current_line)
            current_line = word
    
    if current_line:
        lines.append(current_line)
    
    for line in lines:
        p.drawString(170, y, line)
        y -= 14
    y -= 5

    # TV SET DETAILS
    tv_qty = application_data.get("tv_qty", [])
    tv_brand = application_data.get("tv_brand", [])
    tv_type = application_data.get("tv_type", [])
    
    if tv_qty and any(tv_qty):
        draw_section_title("VII. TV SET DETAILS")
        ensure_space(40)
        
        p.setFont("Helvetica-Bold", 9)
        p.drawString(50, y, "QTY")
        p.drawString(120, y, "BRAND / MODEL")
        p.drawString(320, y, "TYPE (HD/REGULAR)")
        y -= 15
        
        p.setFont("Helvetica", 9)
        for i in range(min(len(tv_qty), 5)):
            if y < 120:
                break
            qty = str(tv_qty[i]) if i < len(tv_qty) else "-"
            brand = tv_brand[i] if i < len(tv_brand) else "-"
            if len(brand) > 25:
                brand = brand[:22] + "..."
            tv_t = tv_type[i] if i < len(tv_type) else "-"
            
            p.drawString(50, y, qty)
            p.drawString(120, y, brand)
            p.drawString(320, y, tv_t)
            y -= 16
        y -= 5

    draw_section_title("VIII. SUBMISSION DETAILS")
    draw_two_columns([
        ("Date Submitted", application_data.get("date_submitted")),
        ("Time Submitted", application_data.get("time_submitted")),
    ])

    full_name = f"{application_data.get('first_name', '')} {application_data.get('last_name', '')}".strip()
    draw_signature_section(application_data.get("signature"), full_name)

    # ================= PAGE 2: MAP =================
    new_page()
    draw_section_title_centered("INSTALLATION LOCATION MAP")

    lat = application_data.get("latitude")
    lng = application_data.get("longitude")
    google_maps_url = None
    google_maps_direction_url = None
    map_img = None

    try:
        if lat and lng:
            lat = float(lat)
            lng = float(lng)
            
            google_maps_url = f"https://www.google.com/maps?q={lat},{lng}"
            google_maps_direction_url = f"https://www.google.com/maps/dir//{lat},{lng}"
            
            map_url = f"https://maps.locationiq.com/v3/staticmap?key=pk.0fdad07272d959e4de881139988b0883&center={lat},{lng}&zoom=17&size=600x400&markers=icon:large-red-cutout|{lat},{lng}"
            response = requests.get(map_url, timeout=10)
            if response.status_code == 200:
                map_img = ImageReader(io.BytesIO(response.content))
    except Exception as e:
        print("Map error:", e)

    draw_two_columns([
        ("Street/Village", application_data.get("address")),
        ("Barangay/City", f"{application_data.get('barangay', '-')}, {application_data.get('city', '-')}"),
        ("Latitude", lat if lat else "-"),
        ("Longitude", lng if lng else "-"),
    ])

    if google_maps_direction_url:
        ensure_space(25)
        p.setFont("Helvetica-Bold", 12)
        p.setFillColorRGB(0, 0.5, 0)
        p.drawCentredString(width / 2, y, " GET DIRECTIONS from your current location to this address")
        text_width = p.stringWidth(" GET DIRECTIONS from your current location to this address", "Helvetica-Bold", 12)
        p.linkURL(google_maps_direction_url, ((width - text_width) / 2, y - 2, (width + text_width) / 2, y + 12), relative=0)
        p.setFillColorRGB(0, 0, 0)
        y -= 20
        
        p.setFont("Helvetica", 8)
        p.setFillColorRGB(0.5, 0.5, 0.5)
        p.drawCentredString(width / 2, y, " Click above to see distance from YOUR location, travel time, and turn-by-turn directions")
        p.setFillColorRGB(0, 0, 0)
        y -= 20

    if google_maps_url:
        p.setFont("Helvetica", 9)
        p.setFillColorRGB(0, 0, 1)
        p.drawCentredString(width / 2, y, "Or click here to view location on Google Maps")
        text_width = p.stringWidth("Or click here to view location on Google Maps", "Helvetica", 9)
        p.linkURL(google_maps_url, ((width - text_width) / 2, y - 2, (width + text_width) / 2, y + 8), relative=0)
        p.setFillColorRGB(0, 0, 0)
        y -= 30

    if map_img:
        ensure_space(380)
        img_width = 500
        img_height = 320
        x_center = (width - img_width) / 2
        p.drawImage(map_img, x_center, y - img_height, img_width, img_height)
        y -= img_height + 20
        
        p.setFont("Helvetica", 8)
        p.setFillColorRGB(0.5, 0.5, 0.5)
        p.drawCentredString(width / 2, y, " Tip: Click the green 'GET DIRECTIONS' link above to see distance from your current location")
        p.setFillColorRGB(0, 0, 0)
        y -= 15
    else:
        draw_two_columns([("Map Status", "Not available")])

    # ================= PAGE 3: FRONT AND BACK ID =================
    new_page()
    draw_section_title_centered("VALID IDENTIFICATION")
    
    draw_images_top_bottom(
        "VALID ID (FRONT)", application_data.get("id_front"),
        "VALID ID (BACK)", application_data.get("id_back"),
        img_width=320, img_height=220
    )

    # ================= PAGE 4: PROOF OF BILLING =================
    new_page()
    draw_section_title_centered("PROOF OF BILLING")
    
    proof = application_data.get("proof_billing")
    if draw_image_safe(p, proof, (width - 500) / 2, y - 580, 500, 580, "Proof of Billing", application_number):
        y -= 580 + 30
    else:
        p.setFont("Helvetica", 9)
        p.setFillColorRGB(0.5, 0.5, 0.5)
        p.drawCentredString(width / 2, y, "No proof of billing provided")
        p.setFillColorRGB(0, 0, 0)
        y -= 30

    p.save()
    buffer.seek(0)
    return buffer


# ===============================
# ARCHIVE APPLICATION
# ===============================
@app.route("/api/superadmin/application/<string:app_id>/archive", methods=["PUT"])
def archive_application(app_id):
    try:
        # Check if application exists and is rejected
        query = "SELECT status FROM applications WHERE application_number = %s"
        app = execute_query(query, (app_id,), fetch_one=True)
        
        if not app:
            return jsonify({"error": "Application not found"}), 404
            
        if app.get("status") != "Rejected":
            return jsonify({"error": "Only rejected applications can be archived"}), 400
        
        # Update is_archived to 1
        update_query = "UPDATE applications SET is_archived = 1 WHERE application_number = %s"
        execute_query(update_query, (app_id,))
        
        return jsonify({
            "success": True,
            "message": "Application archived successfully"
        })
        
    except Exception as e:
        print("Archive error:", e)
        return jsonify({"error": str(e)}), 500



# ===============================
# DELETE APPLICATION - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/application/<string:app_id>", methods=["DELETE"])
def delete_application(app_id):
    try:
        # Check if application exists
        check_query = "SELECT application_number FROM applications WHERE application_number = %s"
        existing = execute_query(check_query, (app_id,), fetch_one=True)
        
        if not existing:
            return jsonify({"error": "Application not found"}), 404
        
        # Delete the application from applications table
        delete_query = "DELETE FROM applications WHERE application_number = %s"
        execute_query(delete_query, (app_id,))
        
        # Also delete from customers table if exists (to keep data consistent)
        delete_customer_query = "DELETE FROM customers WHERE application_number = %s"
        execute_query(delete_customer_query, (app_id,))
        
        print(f" Application {app_id} deleted successfully from MySQL")
        
        return jsonify({"message": "Application deleted successfully"})
        
    except Exception as e:
        print("Delete application error:", e)
        return jsonify({"error": str(e)}), 500





# ==================== RESTORE APPLICATION - REJECTED -> PENDING, CANCELLED -> APPROVED ====================
@app.route("/api/superadmin/application/<string:app_id>/restore", methods=["PUT"])
def restore_application(app_id):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()

        if not conn:
            return jsonify({"error": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)

        # START TRANSACTION
        conn.start_transaction()

        # Check if application exists
        check_query = "SELECT * FROM applications WHERE application_number = %s"
        cursor.execute(check_query, (app_id,))
        app_data = cursor.fetchone()

        if not app_data:
            return jsonify({"error": "Application not found"}), 404

        current_status = app_data.get("status")
        contract_number = app_data.get("contract_number")
        
        print(f" Current status: {current_status}")
        print(f" Contract number: {contract_number}")
        
        # ALLOW BOTH REJECTED AND CANCELLED
        if current_status not in ["Rejected", "Cancelled"]:
            return jsonify({"error": f"Only rejected or cancelled applications can be restored. Current status: {current_status}"}), 400

        # DETERMINE TARGET STATUS
        is_cancelled = current_status == "Cancelled"
        target_status = "Approved" if is_cancelled else "Pending"
        
        # GET TEAM AND INSTALLATION DATE FROM REQUEST BODY
        request_data = request.get_json() or {}
        assigned_team_id = request_data.get('assigned_team_id')
        installation_date = request_data.get('installation_date')
        
        print(f" Request data: assigned_team_id={assigned_team_id}, installation_date={installation_date}")
        
        # ========== UPDATE APPLICATIONS TABLE ==========
        if is_cancelled:
            update_query = """
                UPDATE applications SET 
                    status = 'Approved',
                    installation_status = 'Pending',
                    is_archived = 0,
                    rejection_reason = NULL,
                    assigned_team_id = %s,
                    installation_date = %s,
                    updated_at = NOW()
                WHERE application_number = %s
            """
            params = (assigned_team_id, installation_date, app_id)
        else:
            update_query = """
                UPDATE applications SET 
                    status = 'Pending',
                    is_archived = 0,
                    rejection_reason = NULL,
                    approval_date = NULL,
                    contract_number = NULL,
                    billing_date = NULL,
                    assigned_team_id = NULL,
                    installation_date = NULL,
                    updated_at = NOW()
                WHERE application_number = %s
            """
            params = (app_id,)
        
        print(f" SQL Query: {update_query}")
        print(f" Params: {params}")
        
        cursor.execute(update_query, params)
        rows_affected = cursor.rowcount
        print(f" Rows affected in applications: {rows_affected}")
        
        if rows_affected == 0:
            conn.rollback()
            return jsonify({"error": "Failed to restore application - no rows updated"}), 500

        # ========== UPDATE CUSTOMERS TABLE (NO updated_at) ==========
        if is_cancelled:
            try:
                # CHECK MUNA KUNG MAY CUSTOMER RECORD
                cursor.execute("SELECT * FROM customers WHERE application_number = %s", (app_id,))
                existing_customer = cursor.fetchone()
                
                if existing_customer:
                    print(f" Existing customer found:")
                    print(f"  - status: {existing_customer.get('status')}")
                    print(f"  - installation_status: {existing_customer.get('installation_status')}")
                    print(f"  - assigned_team_id: {existing_customer.get('assigned_team_id')}")
                    print(f"  - installation_date: {existing_customer.get('installation_date')}")
                    
                    # UPDATE - WALANG updated_at (hindi existing ang column)
                    update_customer_sql = """
                        UPDATE customers 
                        SET 
                            status = 'Approved',
                            installation_status = 'Pending',
                            assigned_team_id = %s,
                            installation_date = %s
                        WHERE application_number = %s
                    """
                    cursor.execute(update_customer_sql, (assigned_team_id, installation_date, app_id))
                    customer_rows = cursor.rowcount
                    print(f" Customer rows affected: {customer_rows}")
                    
                    if customer_rows == 0:
                        # TRY USING contract_number
                        if contract_number:
                            print(f" Trying update with contract_number: {contract_number}")
                            update_customer_sql2 = """
                                UPDATE customers 
                                SET 
                                    status = 'Approved',
                                    installation_status = 'Pending',
                                    assigned_team_id = %s,
                                    installation_date = %s
                                WHERE contract_number = %s
                            """
                            cursor.execute(update_customer_sql2, (assigned_team_id, installation_date, contract_number))
                            customer_rows2 = cursor.rowcount
                            print(f" Customer rows affected (via contract): {customer_rows2}")
                else:
                    print(f" No customer record found for {app_id}")
                    
                    # INSERT NEW CUSTOMER RECORD (WALANG updated_at)
                    print(f" Inserting new customer record for {app_id}")
                    insert_sql = """
                        INSERT INTO customers (
                            application_number, first_name, last_name, middle_name, suffix,
                            email, mobile, address, barangay, city, province, zip,
                            plan, status, installation_status, contract_number,
                            assigned_team_id, installation_date, billing_date, approval_date,
                            date_pending, date_ongoing, date_installed, plan_speed, plan_price,
                            latitude, longitude, created_at
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                            %s, 'Approved', 'Pending', %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s, %s, NOW()
                        )
                    """
                    insert_params = (
                        app_id,
                        app_data.get('first_name'),
                        app_data.get('last_name'),
                        app_data.get('middle_name'),
                        app_data.get('suffix'),
                        app_data.get('email'),
                        app_data.get('mobile'),
                        app_data.get('address'),
                        app_data.get('barangay'),
                        app_data.get('city'),
                        app_data.get('province'),
                        app_data.get('zip'),
                        app_data.get('plan'),
                        app_data.get('contract_number'),
                        assigned_team_id,
                        installation_date,
                        app_data.get('billing_date'),
                        app_data.get('approval_date'),
                        app_data.get('date_pending'),
                        app_data.get('date_ongoing'),
                        app_data.get('date_installed'),
                        app_data.get('plan_speed'),
                        app_data.get('plan_price'),
                        app_data.get('latitude'),
                        app_data.get('longitude')
                    )
                    cursor.execute(insert_sql, insert_params)
                    print(f" New customer record inserted")
                    
            except mysql.connector.Error as customer_err:
                print(f" Customer error: {customer_err}")
                # Don't rollback - continue

        # COMMIT TRANSACTION
        conn.commit()
        print(" Transaction committed successfully")

        # FINAL VERIFICATION
        if is_cancelled:
            cursor.execute("SELECT status, installation_status, assigned_team_id, installation_date FROM customers WHERE application_number = %s", (app_id,))
            final_check = cursor.fetchone()
            if final_check:
                print(f" FINAL CUSTOMER VERIFICATION:")
                print(f"  - status: {final_check.get('status')}")
                print(f"  - installation_status: {final_check.get('installation_status')}")
                print(f"  - assigned_team_id: {final_check.get('assigned_team_id')}")
                print(f"  - installation_date: {final_check.get('installation_date')}")
            else:
                print(f" NO CUSTOMER RECORD FOUND AFTER RESTORE")

        # ========== SEND EMAIL NOTIFICATION ==========
        try:
            customer_email = app_data.get("email")
            first_name = app_data.get("first_name")
            if customer_email:
                send_restore_email(
                    to_email=customer_email,
                    first_name=first_name,
                    app_id=app_id,
                    application_data=app_data,
                    target_status=target_status,
                    is_cancelled=is_cancelled,
                    assigned_team_id=assigned_team_id,
                    installation_date=installation_date
                )
                print(f" Restore email sent to {customer_email}")
        except Exception as email_err:
            print(f" Email error: {email_err}")

        return jsonify({
            "success": True,
            "message": f"Application restored to {target_status} successfully",
            "status": target_status
        })

    except mysql.connector.Error as db_err:
        print(f" Database error: {db_err}")
        if conn:
            conn.rollback()
        return jsonify({"error": str(db_err)}), 500
    except Exception as e:
        print(f" Restore error: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
            print(" Database connection closed")


# ===============================
# SEND RESTORE EMAIL - CANCELLED → APPROVED
# ===============================
def send_restore_email(to_email, first_name, app_id, application_data=None, target_status="Approved", is_cancelled=True, assigned_team_id=None, installation_date=None):
    import os, html, requests
    from datetime import datetime

    api_key = os.getenv("BREVO_API_KEY")
    sender_email = os.getenv("BREVO_SENDER_EMAIL", "cablevision.cableinternet@gmail.com")
    sender_name = os.getenv("BREVO_SENDER_NAME", "Cablevision")

    if not api_key:
        print(" BREVO_API_KEY is not configured.")
        return False

    def escape_html(text):
        return html.escape(str(text)) if text is not None else ""

    status_color = "#059669"
    status_bg = "#d1fae5"
    subject = "Cablevision Application Approved"

    formatted_date = "Not set"
    if installation_date:
        try:
            if isinstance(installation_date, str):
                date_obj = datetime.strptime(installation_date.split(" ")[0], "%Y-%m-%d")
                formatted_date = date_obj.strftime("%B %d, %Y")
            else:
                formatted_date = installation_date.strftime("%B %d, %Y")
        except Exception:
            formatted_date = str(installation_date)

    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Cablevision Application Approved</title>
    </head>
    <body style="margin:0;padding:0;font-family:'Segoe UI','Inter',Arial,sans-serif;background:#eef2ff;">
        <div style="max-width:580px;margin:0 auto;padding:30px 20px;">
            <div style="background:#fff;border-radius:32px;overflow:hidden;box-shadow:0 20px 35px -12px rgba(0,0,0,.15);">

                <div style="background:linear-gradient(135deg,#001f3f 0%,#002b5c 100%);padding:32px 28px;text-align:center;">
                    <h1 style="margin:0;font-size:26px;font-weight:700;color:#fff;">Cablevision</h1>
                    <p style="margin:6px 0 0;color:#93c5fd;font-size:13px;">Internet Service Provider</p>
                </div>

                <div style="padding:20px 28px 0;text-align:center;">
                    <div style="display:inline-block;background:{status_bg};padding:8px 24px;border-radius:60px;">
                        <span style="font-size:14px;font-weight:600;color:{status_color};"> APPLICATION APPROVED</span>
                    </div>
                </div>

                <div style="padding:20px 28px 32px;">
                    <h2 style="margin:0 0 8px;font-size:22px;font-weight:700;color:#0f172a;">
                        Hello, {escape_html(first_name)}!
                    </h2>

                    <p style="margin:0 0 20px;font-size:15px;color:#475569;">
                        Your cancelled application has been <strong>restored</strong> and is now <strong>Approved</strong>.
                    </p>

                    <div style="background:#f8fafc;border-radius:20px;padding:18px;margin-bottom:16px;">
                        <div style="margin-bottom:16px;padding-bottom:12px;border-bottom:1px solid #e2e8f0;">
                            <div style="font-size:11px;font-weight:600;color:#64748b;">Application Number</div>
                            <div style="font-size:18px;font-weight:700;color:#0f172a;font-family:monospace;">
                                {escape_html(app_id)}
                            </div>
                        </div>

                        <div>
                            <div style="font-size:11px;font-weight:600;color:#64748b;">Status</div>
                            <div style="font-size:16px;font-weight:700;color:{status_color};">Approved</div>
                        </div>

                        <div style="margin-top:12px;padding-top:12px;border-top:1px solid #e2e8f0;">
                            <div style="font-size:11px;font-weight:600;color:#64748b;">Installation Date</div>
                            <div style="font-size:14px;font-weight:500;color:#0f172a;">
                                {escape_html(formatted_date)}
                            </div>
                        </div>
                    </div>

                    <div style="margin:20px 0;padding:16px;background:{status_bg};border-radius:12px;border-left:4px solid {status_color};">
                        <p style="margin:0 0 8px;color:#0f172a;">
                            <strong> What happens next?</strong>
                        </p>
                        <ul style="margin:0;padding-left:20px;color:#1e293b;font-size:14px;line-height:1.6;">
                            <li>Your application has been approved and restored.</li>
                            <li>Your installation is scheduled for <strong>{escape_html(formatted_date)}</strong>.</li>
                            <li>Please prepare your location for the scheduled installation.</li>
                            <li>You will receive further updates from Cablevision.</li>
                        </ul>
                    </div>

                    <div style="margin-top:28px;padding-top:20px;text-align:center;border-top:1px solid #e2e8f0;">
                        <p style="margin:0;font-size:12px;color:#94a3b8;">Thank you for choosing Cablevision!</p>
                        <p style="margin:4px 0 0;font-size:11px;color:#94a3b8;">
                            If you have any questions, please contact our support team.
                        </p>
                    </div>
                </div>

                <div style="background:#f1f5f9;padding:16px 28px;text-align:center;">
                    <div style="font-size:11px;color:#64748b;">
                        2026 Cablevision Internet Service Provider. All rights reserved.
                    </div>
                </div>

            </div>
        </div>
    </body>
    </html>
    """

    payload = {
        "sender": {"name": sender_name, "email": sender_email},
        "to": [{"email": to_email, "name": first_name or "Customer"}],
        "subject": subject,
        "htmlContent": html_body
    }

    try:
        print(f" Sending restore email to {to_email} via Brevo API...")

        response = requests.post(
            "https://api.brevo.com/v3/smtp/email",
            headers={
                "accept": "application/json",
                "api-key": api_key,
                "content-type": "application/json"
            },
            json=payload,
            timeout=30
        )

        if response.ok:
            result = response.json()
            print(f" Restore email sent to {to_email}")
            print(f" Brevo Message ID: {result.get('messageId')}")
            return True

        print(f" Brevo restore email failed: {response.status_code} - {response.text}")
        return False

    except requests.RequestException as e:
        print(f" Brevo connection error: {e}")
        return False
    except Exception as e:
        print(f" Restore email failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    

@app.route("/api/superadmin/application/<string:app_id>/request-reapply", methods=["POST"])
def request_reapply(app_id):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()

        if not conn:
            return jsonify({"error": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM applications WHERE application_number = %s", (app_id,))
        app_data = cursor.fetchone()

        if not app_data:
            return jsonify({"error": "Application not found"}), 404

        if app_data.get("status") != "Rejected":
            return jsonify({"error": "Only rejected applications can request a re-application."}), 400

        # ONCE-ONLY CHECK
        if app_data.get("reapply_requested"):
            return jsonify({"error": "A reapply request has already been sent for this application."}), 400

        request_data = request.get_json() or {}
        admin_message = (request_data.get("message") or "").strip()

        if not admin_message:
            return jsonify({"error": "Message is required"}), 400

        customer_email = app_data.get("email")
        first_name = app_data.get("first_name")
        rejection_reason = app_data.get("rejection_reason") or "Not specified"
        reapplied_count = app_data.get("reapplied_count") or 0

        if not customer_email:
            return jsonify({"error": "Customer email not found for this application."}), 400

        sent = send_reapply_request_email(
            to_email=customer_email,
            first_name=first_name,
            app_id=app_id,
            rejection_reason=rejection_reason,
            admin_message=admin_message,
            application_id=app_id,
            reapplied_count=reapplied_count
        )

        if not sent:
            return jsonify({"error": "Failed to send email"}), 500

        # SET THE FLAG + SAVE MESSAGE + TIMESTAMP
        update_query = """
            UPDATE applications
            SET reapply_requested = 1,
                reapply_requested_at = NOW(),
                reapply_message = %s
            WHERE application_number = %s
        """
        cursor.execute(update_query, (admin_message, app_id))
        conn.commit()

        return jsonify({
            "success": True,
            "message": "Reapply request email sent successfully",
            "reapply_requested_at": app_data.get("reapply_requested_at")  # optional, front-end may reload anyway
        })

    except mysql.connector.Error as db_err:
        print(f" Database error: {db_err}")
        if conn:
            conn.rollback()
        return jsonify({"error": str(db_err)}), 500
    except Exception as e:
        print(f" Request reapply error: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def send_reapply_request_email(
    to_email,
    first_name,
    app_id,
    rejection_reason,
    admin_message,
    application_id,
    reapplied_count=0
):
    import requests
    import html as html_lib

    # ==========================================
    # BREVO CONFIGURATION
    # ==========================================

    api_key = os.getenv("BREVO_API_KEY", "")

    if not api_key:
        print(" Brevo API key not configured!")
        return False

    print(f" Sending reapply request email via Brevo to {to_email}...")

    # ==========================================
    # PRODUCTION USER WEBSITE
    # ==========================================

    BASE_URL = "https://cablevisioncableinternet.com"

    # ==========================================
    # ESCAPE USER/ADMIN INPUT
    # ==========================================

    safe_reason = html_lib.escape(rejection_reason or "Not specified")
    safe_message = html_lib.escape(admin_message or "")
    safe_first_name = html_lib.escape(first_name or "")
    safe_app_id = html_lib.escape(app_id or "")

    # ==========================================
    # RE-APPLICATION LIMIT
    # ==========================================

    if reapplied_count < 2:
        remaining = 2 - reapplied_count
        reapply_url = f"{BASE_URL}/reapply/{application_id}"

        reapply_section = f"""
        <div style="margin: 28px 0; text-align: center;">
            <a href="{reapply_url}"
               style="
                    display: inline-block;
                    background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%);
                    color: #ffffff;
                    text-decoration: none;
                    padding: 15px 36px;
                    border-radius: 50px;
                    font-weight: 600;
                    font-size: 15px;
                    box-shadow: 0 4px 12px rgba(37, 99, 235, 0.30);
               ">
                Re-apply Now
            </a>
            <p style="
                font-size: 12px;
                color: #6b7280;
                margin-top: 12px;
            ">
                You have <strong>{remaining}</strong> re-application(s) left.
            </p>
        </div>
        """
    else:
        reapply_section = """
        <div style="
            margin: 20px 0;
            padding: 14px;
            background: #fef2f2;
            border-radius: 12px;
            text-align: center;
        ">
            <p style="
                margin: 0;
                color: #991b1b;
                font-size: 14px;
                font-weight: 600;
            ">
                You have reached the maximum number of re-applications (2).
                Further re-applications are not allowed.
            </p>
        </div>
        """

    # ==========================================
    # EMAIL HTML
    # ==========================================

    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Cablevision - Re-application Request</title>
    </head>
    <body style="
        margin:0;
        padding:0;
        font-family:'Segoe UI','Inter',-apple-system,BlinkMacSystemFont,Arial,sans-serif;
        background-color:#eef2ff;
    ">
        <div style="
            max-width:580px;
            margin:0 auto;
            padding:30px 20px;
        ">
            <div style="
                background:#ffffff;
                border-radius:32px;
                overflow:hidden;
                box-shadow:0 20px 35px -12px rgba(0,0,0,0.15);
            ">
                <!-- HEADER -->
                <div style="
                    background:linear-gradient(135deg, #001f3f 0%, #002b5c 100%);
                    padding:32px 28px;
                    text-align:center;
                ">
                    <h1 style="
                        margin:0;
                        font-size:26px;
                        font-weight:700;
                        color:#ffffff;
                    ">
                        Cablevision
                    </h1>
                    <p style="
                        margin:6px 0 0 0;
                        color:#93c5fd;
                        font-size:13px;
                    ">
                        Internet Service Provider
                    </p>
                </div>

                <!-- STATUS BADGE -->
                <div style="
                    padding:20px 28px 0 28px;
                    text-align:center;
                ">
                    <div style="
                        display:inline-block;
                        background:#dbeafe;
                        padding:8px 24px;
                        border-radius:60px;
                    ">
                        <span style="
                            font-size:14px;
                            font-weight:600;
                            color:#1d4ed8;
                        ">
                            WE'D LIKE YOU TO RE-APPLY
                        </span>
                    </div>
                </div>

                <!-- CONTENT -->
                <div style="padding:20px 28px 32px 28px;">
                    <h2 style="
                        margin:0 0 8px 0;
                        font-size:22px;
                        font-weight:700;
                        color:#0f172a;
                    ">
                        Hello, {safe_first_name}!
                    </h2>
                    <p style="
                        margin:0 0 20px 0;
                        font-size:15px;
                        color:#475569;
                        line-height:1.6;
                    ">
                        Our team reviewed your rejected application
                        and would like to invite you to re-apply
                        with corrected information.
                    </p>

                    <!-- APPLICATION DETAILS -->
                    <div style="
                        background:#f8fafc;
                        border-radius:20px;
                        padding:18px;
                        margin-bottom:16px;
                    ">
                        <div style="
                            margin-bottom:16px;
                            padding-bottom:12px;
                            border-bottom:1px solid #e2e8f0;
                        ">
                            <div style="
                                font-size:11px;
                                font-weight:600;
                                color:#64748b;
                            ">
                                Application Number
                            </div>
                            <div style="
                                font-size:18px;
                                font-weight:700;
                                color:#0f172a;
                                font-family:monospace;
                                margin-top:4px;
                            ">
                                {safe_app_id}
                            </div>
                        </div>
                        <div>
                            <div style="
                                font-size:11px;
                                font-weight:600;
                                color:#64748b;
                                margin-bottom:5px;
                            ">
                                Original Rejection Reason
                            </div>
                            <div style="
                                font-size:14px;
                                font-weight:500;
                                color:#991b1b;
                                line-height:1.5;
                            ">
                                {safe_reason}
                            </div>
                        </div>
                    </div>

                    <!-- ADMIN MESSAGE -->
                    <div style="
                        margin:20px 0;
                        padding:16px;
                        background:#eff6ff;
                        border-radius:12px;
                        border-left:4px solid #2563eb;
                    ">
                        <p style="
                            margin:0 0 8px 0;
                            color:#1e3a8a;
                        ">
                            <strong>Message from our team</strong>
                        </p>
                        <p style="
                            margin:0;
                            color:#1e293b;
                            font-size:14px;
                            line-height:1.6;
                            white-space:pre-wrap;
                        ">
                            {safe_message}
                        </p>
                    </div>

                    <!-- RE-APPLY BUTTON -->
                    {reapply_section}

                    <!-- FOOTER MESSAGE -->
                    <div style="
                        margin-top:28px;
                        padding-top:20px;
                        text-align:center;
                        border-top:1px solid #e2e8f0;
                    ">
                        <p style="
                            margin:0;
                            font-size:12px;
                            color:#94a3b8;
                        ">
                            Thank you for choosing Cablevision!
                        </p>
                    </div>
                </div>

                <!-- FOOTER -->
                <div style="
                    background:#f1f5f9;
                    padding:16px 28px;
                    text-align:center;
                ">
                    <div style="
                        font-size:11px;
                        color:#64748b;
                    ">
                        2026 Cablevision Internet Service Provider. All rights reserved.
                    </div>
                </div>
            </div>
        </div>
    </body>
    </html>
    """

    # ==========================================
    # BREVO API
    # ==========================================

    try:
        url = "https://api.brevo.com/v3/smtp/email"

        headers = {
            "accept": "application/json",
            "api-key": api_key,
            "content-type": "application/json"
        }

        email_data = {
            "sender": {
                "name": "Cablevision Systems Corp.",
                "email": "cablevision.cableinternet@gmail.com"
            },
            "to": [
                {
                    "email": to_email,
                    "name": first_name
                }
            ],
            "subject": "Cablevision - Re-application Requested",
            "htmlContent": html_body
        }

        response = requests.post(url, json=email_data, headers=headers, timeout=30)

        if response.status_code in [200, 201, 202]:
            print(f" Reapply request email sent successfully to {to_email}")
            print(f" Reapply URL: {BASE_URL}/reapply/{application_id}")
            return True
        else:
            print(f" Brevo API error: {response.status_code} - {response.text}")
            return False

    except Exception as e:
        print(f" Reapply email API error: {e}")
        import traceback
        traceback.print_exc()
        return False

# ===============================
# GET ALL APPROVAL REQUESTS - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/approval-requests", methods=["GET"])
def get_approval_requests():
    try:
        query = """
            SELECT id, request_id, app_id, requested_by, requested_status, 
                   status, admin_id, admin_area, admin_city, reason,
                   contract_number, billing_date, date_requested, processed_at
            FROM approval_requests 
            ORDER BY id DESC
        """
        data = execute_query(query, fetch=True) or []

        requests = []
        for req in data:
            requests.append({
                "id": req.get('request_id') or str(req.get('id')),
                "app_id": req.get("app_id"),
                "requested_by": req.get("requested_by"),
                "requested_status": req.get("requested_status"),
                "status": req.get("status"),
                "admin_id": req.get("admin_id"),
                "admin_area": req.get("admin_area"),
                "admin_city": req.get("admin_city"),
                "reason": req.get("reason"),
                "contract_number": req.get("contract_number"),
                "billing_date": req.get("billing_date"),
                "date_requested": req.get("date_requested"),
                "processed_at": req.get("processed_at")
            })

        return jsonify(requests)
        
    except Exception as e:
        print(f"Error getting approval requests: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# GET SINGLE APPROVAL REQUEST - CONVERTED TO MYSQL
# ===============================
@app.route("/api/superadmin/approval-request/<string:request_id>", methods=["GET"])
def get_approval_request(request_id):
    try:
        print(f"Looking for approval request with ID: {request_id}")
        
        query = """
            SELECT id, request_id, app_id, requested_by, requested_status, 
                   status, admin_id, admin_area, admin_city, reason,
                   contract_number, billing_date, date_requested, processed_at
            FROM approval_requests 
            WHERE request_id = %s OR id = %s
        """
        request_data = execute_query(query, (request_id, request_id), fetch_one=True)
        
        print(f"Found data: {request_data}")
        
        if request_data:
            result = {
                "id": request_data.get('request_id') or str(request_data.get('id')),
                "app_id": request_data.get("app_id"),
                "requested_by": request_data.get("requested_by"),
                "requested_status": request_data.get("requested_status"),
                "status": request_data.get("status"),
                "admin_id": request_data.get("admin_id"),
                "admin_area": request_data.get("admin_area"),
                "admin_city": request_data.get("admin_city"),
                "date_requested": request_data.get("date_requested"),
                "reason": request_data.get("reason"),
                "contract_number": request_data.get("contract_number"),
                "billing_date": request_data.get("billing_date"),
                "processed_at": request_data.get("processed_at")
            }
            return jsonify(result)
        else:
            return jsonify({"error": "Request not found"}), 404
            
    except Exception as e:
        print(f"Error getting approval request: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/superadmin/approval-request/<string:req_id>", methods=["PUT"])
def approve_request(req_id):
    conn = None
    cursor = None
    try:
        print("=" * 60)
        print(" APPROVE REQUEST STARTED")
        print(f" Request ID: {req_id}")

        request_data = request.get_json() or {}
        print(f" Request data: {request_data}")

        contract_number = request_data.get("contract_number", None)
        billing_date = request_data.get("billing_date", None)
        first_installment_date = request_data.get("first_installment_date", None)
        last_installment_date = request_data.get("last_installment_date", None)
        assigned_team_id = request_data.get("assigned_team_id", None)
        installation_date = request_data.get("installation_date", None)

        import mysql.connector
        conn = get_db_connection()

        if not conn:
            return jsonify({"error": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        print(" Database connected")

        # Get the approval request
        req_query = """
            SELECT id, request_id, app_id, requested_by, requested_status, status,
                admin_id, admin_area, admin_city, reason
            FROM approval_requests
            WHERE request_id = %s
            LIMIT 1
        """
        cursor.execute(req_query, (req_id,))
        req = cursor.fetchone()
        print(f" Approval request found: {req}")

        if not req:
            return jsonify({"error": "Request not found"}), 404

        app_id = req.get("app_id")
        requested_status = req.get("requested_status")
        requested_by = req.get("requested_by", "Unknown Admin")
        reason = req.get("reason", "")
        admin_id = req.get("admin_id")
        admin_area = req.get("admin_area")
        admin_city = req.get("admin_city")

        print(f" App ID: {app_id}, Requested Status: {requested_status}")

        # Check if request is already processed
        if req.get("status") == "Done":
            return jsonify({"error": "This request has already been processed"}), 400

        # Get application data
        app_query = "SELECT * FROM applications WHERE application_number = %s"
        cursor.execute(app_query, (app_id,))
        app_data = cursor.fetchone()
        print(f" Application data status: {app_data.get('status') if app_data else 'Not found'}")

        if not app_data:
            return jsonify({"error": "Application not found"}), 404

        if app_data.get("status") != "Request Sent":
            return jsonify({"error": f"Application status is '{app_data.get('status')}', cannot process this request"}), 400

        # ========== HANDLE REAPPLY REQUEST ==========
        if requested_status == "Reapply":
            print(" Processing REAPPLY request from admin")

            if app_data.get("reapply_requested"):
                return jsonify({"error": "A reapply request has already been sent to the customer"}), 400

            customer_email = app_data.get("email")
            first_name = app_data.get("first_name")
            rejection_reason = app_data.get("rejection_reason") or "Not specified"
            reapplied_count = app_data.get("reapplied_count") or 0

            if not customer_email:
                return jsonify({"error": "Customer email not found"}), 400

            # SEND REAPPLY EMAIL VIA BREVO (HTTP API — hindi naaapektuhan ng SMTP port blocking sa Railway)
            email_sent = send_reapply_request_email(
                to_email=customer_email,
                first_name=first_name,
                app_id=app_id,
                rejection_reason=rejection_reason,
                admin_message=reason,
                application_id=app_id,
                reapplied_count=reapplied_count
            )

            if not email_sent:
                return jsonify({"error": "Failed to send email"}), 500

            # UPDATE APPLICATION - SET REAPPLY FLAGS
            update_query = """
                UPDATE applications
                SET reapply_requested = 1,
                    reapply_requested_at = NOW(),
                    reapply_message = %s,
                    status = 'Rejected'
                WHERE application_number = %s
            """
            cursor.execute(update_query, (reason, app_id))

            # CREATE ADMIN NOTIFICATION
            admin_notification_id = int(datetime.now().timestamp() * 1000)
            admin_notif_title = "Admin Reapply Request - Approved"
            admin_notif_message = f"Your request to send a reapply invitation to {app_data.get('first_name', '')} {app_data.get('last_name', '')}'s application ({app_id}) has been APPROVED by superadmin. The customer has been notified."

            cursor.execute("""
                INSERT INTO admin_notifications
                (id, title, message, type, relatedId, request_id, timestamp, read_status,
                 admin_id, admin_area, admin_city, requested_by, requested_status, application_city, action_taken_by, action_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                admin_notification_id,
                admin_notif_title,
                admin_notif_message,
                "request_approved",
                app_id, req_id, ph_now_iso(), 0,
                admin_id, admin_area, admin_city, requested_by, requested_status,
                app_data.get('city', ''), "superadmin", "Approved"
            ))

            # MARK REQUEST AS DONE
            cursor.execute("""
                UPDATE approval_requests
                SET status = 'Done', processed_at = %s
                WHERE request_id = %s
            """, (ph_now_iso(), req_id))

            conn.commit()

            return jsonify({
                "message": "Reapply request approved and email sent to customer",
                "status": "Done",
                "reapply_requested": True
            })

        # ========== HANDLE OTHER STATUSES (Approved, Rejected, Pending) ==========
        if requested_status == "Approved":
            if not contract_number or contract_number.strip() == "":
                import random
                import string
                date_part = ph_now().strftime("%Y%m%d")
                random_part = ''.join(random.choices(string.digits, k=4))
                contract_number = f"CV-{date_part}-{random_part}"
                print(f" Auto-generated contract number: {contract_number}")

            if not billing_date or billing_date.strip() == "":
                billing_date = "15th"
                print(f" Default billing date set to: {billing_date}")

        # ========== 1. UPDATE APPLICATIONS TABLE ==========
        update_fields = ["status = %s"]
        params = [requested_status]

        if requested_status == "Approved":
            update_fields.append("contract_number = %s")
            params.append(contract_number)
            update_fields.append("billing_date = %s")
            params.append(billing_date)
            update_fields.append("approval_date = %s")
            params.append(ph_now().strftime("%Y-%m-%d %H:%M:%S"))
            update_fields.append("installation_status = %s")
            params.append("Pending")
            update_fields.append("assigned_team_id = %s")
            params.append(assigned_team_id)
            update_fields.append("installation_date = %s")
            params.append(installation_date)
        elif requested_status == "Rejected":
            update_fields.append("rejection_reason = %s")
            params.append(reason if reason else "Not specified")
        elif requested_status == "Pending":
            update_fields.append("rejection_reason = %s")
            params.append(None)

        params.append(app_id)
        update_query = f"UPDATE applications SET {', '.join(update_fields)} WHERE application_number = %s"
        cursor.execute(update_query, params)

        # ========== 2. INSERT/UPDATE CUSTOMERS TABLE ==========
        if requested_status == "Approved":
            current_datetime = ph_now().strftime("%Y-%m-%d %H:%M:%S")

            customer_data = {
                "application_number": app_data.get("application_number"),
                "first_name": app_data.get("first_name"),
                "last_name": app_data.get("last_name"),
                "middle_name": app_data.get("middle_name"),
                "suffix": app_data.get("suffix"),
                "email": app_data.get("email"),
                "mobile": app_data.get("mobile"),
                "address": app_data.get("address"),
                "barangay": app_data.get("barangay"),
                "city": app_data.get("city"),
                "province": app_data.get("province"),
                "zip": app_data.get("zip"),
                "plan": app_data.get("plan"),
                "plan_speed": app_data.get("plan_speed"),
                "plan_price": app_data.get("plan_price"),
                "status": "Approved",
                "installation_status": "Pending",
                "contract_number": contract_number,
                "billing_date": billing_date,
                "approval_date": current_datetime,
                "date_pending": current_datetime,
                "assigned_team_id": assigned_team_id,
                "installation_date": installation_date,
                "latitude": app_data.get("latitude"),
                "longitude": app_data.get("longitude")
            }

            customer_data = {k: v for k, v in customer_data.items() if v is not None and v != 'none'}

            print(" CUSTOMER DATA:", customer_data)

            app_number = app_data.get("application_number")
            cursor.execute("SELECT application_number FROM customers WHERE application_number = %s", (app_number,))
            existing_customer = cursor.fetchone()

            if not existing_customer:
                columns = ', '.join(customer_data.keys())
                placeholders = ', '.join(['%s'] * len(customer_data))
                insert_query = f"INSERT INTO customers ({columns}) VALUES ({placeholders})"
                cursor.execute(insert_query, list(customer_data.values()))
                print(f" Customer record INSERTED for {app_number}")
            else:
                update_customer_fields = []
                update_params = []
                for key, value in customer_data.items():
                    update_customer_fields.append(f"{key} = %s")
                    update_params.append(value)
                update_params.append(app_number)
                update_customer_query = f"UPDATE customers SET {', '.join(update_customer_fields)} WHERE application_number = %s"
                cursor.execute(update_customer_query, update_params)
                print(f" Customer record UPDATED for {app_number}")

        # ========== 3. SAVE TO CONTRACTS TABLE (if approved) ==========
        if requested_status == "Approved":
            fullName = ' '.join(filter(None, [
                app_data.get('first_name', ''),
                app_data.get('middle_name', ''),
                app_data.get('last_name', ''),
                app_data.get('suffix', '')
            ])).strip()

            date_submitted_val = app_data.get('date_submitted')
            if date_submitted_val:
                if hasattr(date_submitted_val, 'strftime'):
                    date_submitted_str = date_submitted_val.strftime("%Y-%m-%d")
                else:
                    date_submitted_str = str(date_submitted_val)
            else:
                date_submitted_str = ph_now().strftime("%Y-%m-%d")

            contract_data = {
                "contract_number": contract_number,
                "application_id": app_id,
                "first_name": app_data.get('first_name'),
                "middle_name": app_data.get('middle_name'),
                "last_name": app_data.get('last_name'),
                "suffix": app_data.get('suffix'),
                "full_name": fullName,
                "age": calculate_age(app_data.get('birthdate', '')),
                "civil_status": app_data.get('civil_status'),
                "address": f"{app_data.get('barangay', '')}, {app_data.get('city', '')}, {app_data.get('province', '')}".strip(', '),
                "barangay": app_data.get('barangay'),
                "city": app_data.get('city'),
                "province": app_data.get('province'),
                "billing_date": billing_date,
                "date_submitted": date_submitted_str,
                "status": "Active",
                "created_at": ph_now_iso(),
                "is_installment_plan": 1 if first_installment_date else 0,
                "first_installment_date": first_installment_date,
                "last_installment_date": last_installment_date,
                "installation_fee": app_data.get('installation_fee'),
                "application_data": json.dumps(app_data, default=str)
            }

            contract_data = {k: v for k, v in contract_data.items() if v is not None and v != ''}

            for key, value in contract_data.items():
                if hasattr(value, 'isoformat'):
                    contract_data[key] = value.isoformat()
                elif hasattr(value, 'strftime'):
                    contract_data[key] = value.strftime("%Y-%m-%d %H:%M:%S")

            cursor.execute("SELECT contract_number FROM contracts WHERE contract_number = %s", (contract_number,))
            existing_contract = cursor.fetchone()

            if not existing_contract:
                columns = ', '.join(contract_data.keys())
                placeholders = ', '.join(['%s'] * len(contract_data))
                insert_query = f"INSERT INTO contracts ({columns}) VALUES ({placeholders})"
                cursor.execute(insert_query, list(contract_data.values()))
                print(f" Contract {contract_number} INSERTED")
            else:
                update_contract_fields = []
                update_params = []
                for key, value in contract_data.items():
                    if key != 'contract_number':
                        update_contract_fields.append(f"{key} = %s")
                        update_params.append(value)
                update_params.append(contract_number)
                update_contract_query = f"UPDATE contracts SET {', '.join(update_contract_fields)} WHERE contract_number = %s"
                cursor.execute(update_contract_query, update_params)
                print(f" Contract {contract_number} UPDATED")

        # ========== 4. CREATE NOTIFICATION FOR ADMIN ==========
        applicant_name = f"{app_data.get('first_name', '')} {app_data.get('last_name', '')}".strip()
        application_number = app_data.get('application_number', 'N/A')
        action_status = requested_status

        admin_notification_id = int(datetime.now().timestamp() * 1000)

        if requested_status == "Pending":
            admin_notif_title = "Request Accepted - Application Restored"
            admin_notif_message = f"Your request to restore {applicant_name}'s application ({application_number}) to Pending status has been ACCEPTED by superadmin."
        else:
            admin_notif_title = f"Request {requested_status} - Application {application_number}"
            admin_notif_message = f"Your request to {requested_status.lower()} {applicant_name}'s application ({application_number}) has been {requested_status.upper()} by superadmin."

        admin_notif_query = """
            INSERT INTO admin_notifications
            (id, title, message, type, relatedId, request_id, timestamp, read_status,
             admin_id, admin_area, admin_city, requested_by, requested_status,
             application_city, action_taken_by, action_status, contract_number, billing_date)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        admin_notif_params = (
            admin_notification_id,
            admin_notif_title,
            admin_notif_message,
            "request_approved" if requested_status in ("Approved", "Pending") else "request_rejected",
            app_id, req_id, ph_now_iso(), 0,
            admin_id, admin_area, admin_city, requested_by, requested_status,
            app_data.get('city', ''), "superadmin", action_status,
            contract_number if requested_status == "Approved" else None,
            billing_date if requested_status == "Approved" else None
        )

        cursor.execute(admin_notif_query, admin_notif_params)

        # ========== 5. CREATE NOTIFICATION FOR TECHNICIANS (if approved) ==========
        if requested_status == "Approved":
            try:
                application_city = app_data.get("city", "")
                print(f" Creating technician notification for area: {application_city}")

                tech_query = "SELECT technician_id FROM technicians WHERE UPPER(area) = UPPER(%s) AND status = 'Active'"
                cursor.execute(tech_query, (application_city,))
                technicians = cursor.fetchall()

                print(f" Found {len(technicians)} technicians in area: {application_city}")

                if technicians:
                    import time
                    success_count = 0
                    base_id = int(time.time() * 1000)

                    for idx, tech in enumerate(technicians):
                        technician_id = tech.get('technician_id')
                        if technician_id:
                            notification_id = base_id + idx + 1

                            tech_notif_query = """
                                INSERT INTO technician_notifications
                                (id, technician_id, technician_area, title, message, type, relatedId,
                                 application_number, customer_name, timestamp, read_status, created_at)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, NOW())
                            """
                            cursor.execute(tech_notif_query, (
                                notification_id,
                                technician_id,
                                application_city,
                                "New Application Approved (via Admin Request)",
                                f"New application from {applicant_name} (Application: {app_id}) has been APPROVED and is ready for slot assignment.",
                                "new_approved_application",
                                app_id,
                                app_id,
                                applicant_name,
                                ph_now_iso()
                            ))
                            success_count += 1

                    print(f" Created {success_count} technician notifications in area: {application_city}")
                else:
                    print(f" No active technicians found in area: {application_city}")

            except Exception as tech_err:
                print(f" Technician notification error: {tech_err}")
                import traceback
                traceback.print_exc()

        # ========== 6. MARK REQUEST AS DONE ==========
        processed_at_str = ph_now_iso()
        cursor.execute("""
            UPDATE approval_requests
            SET status = 'Done',
                contract_number = %s,
                billing_date = %s,
                processed_at = %s
            WHERE request_id = %s
        """, (
            contract_number if requested_status == "Approved" else None,
            billing_date if requested_status == "Approved" else None,
            processed_at_str,
            req_id
        ))
        print(f" Request {req_id} marked as Done")

        # ========== 7. CREATE GENERAL NOTIFICATION ==========
        general_notification_id = int(datetime.now().timestamp() * 1000) + 1
        general_notif_query = """
            INSERT INTO notifications
            (id, title, message, type, relatedId, timestamp, read_status)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(general_notif_query, (
            general_notification_id,
            f"Admin Request {requested_status}",
            f"Superadmin {requested_status.upper()} {requested_by}'s request to {requested_status.lower()} {applicant_name}'s application ({application_number})",
            "superadmin_action",
            app_id,
            ph_now_iso(),
            0
        ))
        print(f" General notification created")

        # ========== 8. COMMIT ALL CHANGES ==========
        conn.commit()
        print(" All changes COMMITTED to database")

        # ========== 9. VERIFY UPDATES ==========
        cursor.execute("SELECT status, contract_number FROM applications WHERE application_number = %s", (app_id,))
        app_verified = cursor.fetchone()
        print(f" APPLICATION VERIFIED - Status: {app_verified.get('status') if app_verified else 'Not found'}")

        if requested_status == "Approved":
            cursor.execute("SELECT application_number FROM customers WHERE application_number = %s", (app_id,))
            customer_verified = cursor.fetchone()
            print(f" CUSTOMER VERIFIED - Exists: {customer_verified is not None}")

            cursor.execute("SELECT contract_number FROM contracts WHERE contract_number = %s", (contract_number,))
            contract_verified = cursor.fetchone()
            print(f" CONTRACT VERIFIED - Exists: {contract_verified is not None}")

        # ========== 10. SEND EMAIL TO CUSTOMER ==========
        try:
            applicant_email = app_data.get("email")
            if applicant_email:
                if requested_status == "Pending":
                    send_restore_email(
                        to_email=applicant_email,
                        first_name=applicant_name,
                        app_id=application_number
                    )
                    print(f" Restore email sent to {applicant_email}")
                else:
                    send_application_status_email(
                        to_email=applicant_email,
                        first_name=applicant_name,
                        status=requested_status,
                        app_id=application_number,
                        reason=reason if requested_status == "Rejected" else None,
                        contract_number=contract_number if requested_status == "Approved" else None,
                        billing_date=billing_date if requested_status == "Approved" else None,
                        application_id=app_id,
                        reapplied_count=app_data.get("reapplied_count", 0)
                    )
                    print(f" Email sent to {applicant_email}")
        except Exception as email_err:
            print(f" Error sending email: {email_err}")

        # ========== 11. RETURN SUCCESS RESPONSE ==========
        response_data = {
            "message": "Application restored to Pending status successfully" if requested_status == "Pending" else f"Request {requested_status} successfully",
            "contract_number": contract_number if requested_status == "Approved" else None,
            "billing_date": billing_date if requested_status == "Approved" else None,
            "status": requested_status,
            "request_status": "Done"
        }
        response_data = {k: v for k, v in response_data.items() if v is not None}
        return jsonify(response_data)

    except mysql.connector.Error as db_err:
        print(f" Database error: {db_err}")
        if conn:
            conn.rollback()
        return jsonify({"error": str(db_err)}), 500
    except Exception as e:
        print(f" Error in approve_request: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
            print(" Database connection closed")



# ===============================
# SUPERADMIN REJECTS REQUEST (PATCH) - USING DIRECT CONNECTION
# ===============================
@app.route("/api/superadmin/approval-request/<string:req_id>/reject", methods=["PATCH"])
def reject_request(req_id):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()

        if not conn:
            return jsonify({"error": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        
        req_query = """
            SELECT id, request_id, app_id, requested_by, requested_status, status, 
                admin_id, admin_area, admin_city, reason 
            FROM approval_requests  
            WHERE request_id = %s
            LIMIT 1
        """
        cursor.execute(req_query, (req_id,))
        req_data = cursor.fetchone()
        
        if not req_data:
            return jsonify({"error": "Request not found"}), 404

        if req_data.get("status") == "Done":
            return jsonify({"error": "This request has already been processed"}), 400

        app_id = req_data.get("app_id")
        requested_status = req_data.get("requested_status", "Rejected")
        requested_by = req_data.get("requested_by", "Unknown Admin")
        reason = req_data.get("reason", "No specific reason provided")
        admin_id = req_data.get("admin_id")
        admin_area = req_data.get("admin_area")
        admin_city = req_data.get("admin_city")
        
        cursor.execute("SELECT * FROM applications WHERE application_number = %s", (app_id,))
        app_data = cursor.fetchone()
        
        if not app_data:
            return jsonify({"error": "Application not found"}), 404
        
        cursor.execute("""
            UPDATE approval_requests  
            SET status = 'Rejected', 
                processed_at = %s 
            WHERE request_id = %s
        """, (
            ph_now_iso(),
            req_id
        ))
        
        # DETERMINE REVERT STATUS BASED ON REQUESTED STATUS
        if requested_status == "Pending":
            revert_status = "Rejected"
        elif requested_status == "Reapply":
            revert_status = "Rejected"  # Stay Rejected
        else:
            revert_status = "Pending"
        
        cursor.execute("UPDATE applications SET status = %s WHERE application_number = %s", (revert_status, app_id))
        print(f" Application {app_id} reverted to {revert_status} status")
        
        applicant_name = f"{app_data.get('first_name', '')} {app_data.get('last_name', '')}".strip()
        application_number = app_data.get('application_number', 'N/A')
        application_city = app_data.get('city', '') if app_data else ''
        
        admin_notification_id = int(datetime.now().timestamp() * 1000)
        
        # MESSAGE WORDING
        if requested_status == "Pending":
            admin_notif_message = f"Your request to restore {applicant_name}'s application ({application_number}) has been REJECTED by superadmin. The application remains Rejected.\nReason: {reason}"
        elif requested_status == "Reapply":
            admin_notif_message = f"Your request to send a reapply invitation to {applicant_name}'s application ({application_number}) has been REJECTED by superadmin. The application remains Rejected.\nReason: {reason}"
        else:
            admin_notif_message = f"Your request to {requested_status.lower()} {applicant_name}'s application ({application_number}) has been REJECTED by superadmin. The application remains in Pending status.\nReason: {reason}"
        
        cursor.execute("""
            INSERT INTO admin_notifications 
            (id, title, message, type, relatedId, request_id, timestamp, read_status,
             admin_id, admin_area, admin_city, requested_by, requested_status, application_city)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            admin_notification_id,
            f"Request {requested_status if requested_status != 'Pending' else 'Restore'} Rejected",
            admin_notif_message,
            "request_rejected",
            app_id, req_id, ph_now_iso(), 0,
            admin_id, admin_area, admin_city, requested_by, requested_status,
            application_city
        ))
        
        general_notification_id = int(datetime.now().timestamp() * 1000) + 1
        try:
            if requested_status == "Pending":
                action_label = "restore"
            elif requested_status == "Reapply":
                action_label = "send reapply invitation"
            else:
                action_label = requested_status.lower()
            
            cursor.execute("""
                INSERT INTO notifications 
                (id, title, message, type, relatedId, timestamp, read_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                general_notification_id,
                f"Admin Request Rejected",
                f"Superadmin REJECTED {requested_by}'s request to {action_label} {applicant_name}'s application ({application_number})",
                "superadmin_action", app_id, ph_now_iso(), 0
            ))
        except Exception as gen_err:
            print(f" General notification note: {gen_err}")
        
        conn.commit()
        
        # NO EMAIL SENT TO CUSTOMER WHEN SUPERADMIN REJECTS ADMIN REQUEST
        # (in-app notification lang ang gagawin, walang email)
        
        return jsonify({
            "message": f"Request rejected, application reverted to {revert_status}"
        })

    except Exception as e:
        print(f" Error in reject_request: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    
@app.route("/api/admin/check-pending-request/<app_id>", methods=["GET"])
def check_pending_request(app_id):
    try:
        # KUHAIN ANG LAHAT NG PENDING REQUESTS (HINDI LANG REAPPLY)
        query = """
            SELECT request_id, requested_status FROM approval_requests 
            WHERE app_id = %s AND status = 'Pending'
            LIMIT 1
        """
        result = execute_query(query, (app_id,), fetch_one=True)
        
        if result:
            return jsonify({
                "hasPending": True,
                "requested_status": result.get('requested_status')
            })
        return jsonify({"hasPending": False})
    except Exception as e:
        print(f"Error checking pending request: {e}")
        return jsonify({"hasPending": False})



@app.route("/api/test-insert-notification", methods=["GET"])
def test_insert_notification():
    try:
        import time
        notification_id = int(time.time() * 1000)
        
        query = """
            INSERT INTO admin_notifications 
            (id, title, message, type, read_status, created_at)
            VALUES (%s, %s, %s, %s, %s, NOW())
        """
        result = execute_query(query, (notification_id, "Test", "Test message", "test", 0))
        
        return jsonify({
            "success": True,
            "notification_id": notification_id,
            "result": result
        })
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500    


# ===============================
# Archived Applications Page
# ===============================
@app.route("/superadmin/archived-applications")
def superadmin_archived_applications_page():
    return render_template("superadmin-archived-applications.html")


# ===============================
# GET ARCHIVED APPLICATIONS (is_archived = 1, any status)
# ===============================
@app.route("/api/superadmin/archived-applications", methods=["GET"])
def superadmin_get_archived_applications():
    try:
        limit = int(request.args.get("limit", 100))

        query = """
            SELECT application_number, first_name, last_name, email, 
                   date_submitted, time_submitted, barangay, city, birthdate, 
                   status, rejection_reason, is_archived
            FROM applications 
            WHERE is_archived = 1
            ORDER BY timestamp DESC 
            LIMIT %s
        """
        applications = execute_query(query, (limit,), fetch=True) or []

        apps = []
        for app in applications:
            # Combine date and time submitted for display
            datetime_submitted = None
            if app.get("date_submitted") and app.get("time_submitted"):
                datetime_submitted = f"{app.get('date_submitted')} {app.get('time_submitted')}"
            elif app.get("date_submitted"):
                datetime_submitted = app.get("date_submitted")

            apps.append({
                "id": app.get("application_number", ""),
                "application_number": app.get("application_number", ""),
                "first_name": app.get("first_name", ""),
                "last_name": app.get("last_name", ""),
                "email": app.get("email", ""),
                "date_submitted": datetime_submitted,
                "barangay": app.get("barangay", ""),
                "city": app.get("city", ""),
                "birthdate": app.get("birthdate", ""),
                "status": app.get("status", "Rejected"),
                "rejection_reason": app.get("rejection_reason", ""),
                "is_archived": True
            })

        return jsonify(apps)

    except Exception as e:
        print("Superadmin get archived applications error:", e)
        return jsonify({"error": str(e)}), 500


# ===============================
# UNARCHIVE APPLICATION (RESTORE TO PENDING)
# ===============================
@app.route("/api/superadmin/application/<string:app_id>/unarchive", methods=["PUT"])
def unarchive_application(app_id):
    try:
        # Check if application exists and is archived
        query = "SELECT status, is_archived FROM applications WHERE application_number = %s"
        app = execute_query(query, (app_id,), fetch_one=True)
        
        if not app:
            return jsonify({"error": "Application not found"}), 404
            
        if app.get("is_archived") != 1:
            return jsonify({"error": "Application is not archived"}), 400
        
        # Update: Unarchive and set status to Pending
        update_query = """
            UPDATE applications 
            SET is_archived = 0, 
                status = 'Pending',
                rejection_reason = NULL,
                updated_at = NOW()
            WHERE application_number = %s
        """
        execute_query(update_query, (app_id,))
        
        # Get updated application data for email
        app_data = execute_query("SELECT * FROM applications WHERE application_number = %s", (app_id,), fetch_one=True)
        
        # Send email notification to customer
        try:
            if app_data and app_data.get('email'):
                send_restore_email(
                    to_email=app_data.get('email'),
                    first_name=app_data.get('first_name'),
                    app_id=app_id
                )
                print(f" Restore email sent to {app_data.get('email')}")
        except Exception as email_err:
            print(f" Email error: {email_err}")
        
        return jsonify({
            "success": True,
            "message": "Application restored to Pending status successfully"
        })
        
    except Exception as e:
        print("Unarchive error:", e)
        return jsonify({"error": str(e)}), 500


    
# =============================== 
# View Customers (Approved applications) 
# ===============================
@app.route("/superadmin/view-customers")
def superadmin_view_customers_page():
    return render_template("superadmin-view-customers.html")


@app.route("/api/superadmin/approved-applications", methods=["GET"])
def get_approved_customers():
    try:
        limit = int(request.args.get("limit", 50))
        
        # Optional filters
        city_filter = request.args.get("city", "")
        search_term = request.args.get("search", "")
        
        # Build query for customers table - IDINAGDAG ANG plan_speed at plan_price
        query = """
            SELECT 
                c.application_number, c.contract_number, c.first_name, c.last_name, 
                c.middle_name, c.suffix, c.email, c.mobile, c.address, c.barangay, c.city, 
                c.province, c.zip, c.plan, c.plan_speed, c.plan_price, c.status, c.installation_status, 
                c.approval_date, c.billing_date, c.created_at,
                CASE WHEN u.user_id IS NOT NULL THEN 1 ELSE 0 END as user_created
            FROM customers c
            LEFT JOIN users u ON c.application_number = u.application_number
            WHERE c.status = 'Approved'
        """
        params = []
        
        # Add city filter
        if city_filter and city_filter != "all":
            query += " AND c.city = %s"
            params.append(city_filter)
        
        # Add search filter
        if search_term:
            query += """ AND (c.first_name LIKE %s OR c.last_name LIKE %s 
                       OR c.email LIKE %s OR c.application_number LIKE %s 
                       OR c.contract_number LIKE %s)"""
            search_pattern = f"%{search_term}%"
            params.extend([search_pattern, search_pattern, search_pattern, search_pattern, search_pattern])
        
        # Add order by and limit
        query += " ORDER BY c.approval_date DESC LIMIT %s"
        params.append(limit)
        
        # Execute query
        customers = execute_query(query, params, fetch=True) or []
        
        customers_list = []
        for cust in customers:
            # Build full name
            full_name_parts = []
            if cust.get('first_name'):
                full_name_parts.append(cust['first_name'])
            if cust.get('middle_name'):
                full_name_parts.append(cust['middle_name'])
            if cust.get('last_name'):
                full_name_parts.append(cust['last_name'])
            if cust.get('suffix'):
                full_name_parts.append(cust['suffix'])
            full_name = ' '.join(full_name_parts).strip() or 'N/A'
            
            customers_list.append({
                "id": cust.get('application_number'),
                "application_number": cust.get('application_number', ''),
                "contract_number": cust.get('contract_number', 'N/A'),
                "first_name": cust.get('first_name', ''),
                "last_name": cust.get('last_name', ''),
                "full_name": full_name,
                "email": cust.get('email', ''),
                "mobile": cust.get('mobile', ''),
                "plan": cust.get('plan', ''),
                "plan_speed": cust.get('plan_speed', 'N/A'),  # DIREKTA MULA SA CUSTOMERS TABLE
                "plan_price": cust.get('plan_price', 'N/A'),  # DIREKTA MULA SA CUSTOMERS TABLE
                "status": cust.get('status', 'Approved'),
                "installation_status": cust.get('installation_status', 'Pending'),
                "approval_date": cust.get('approval_date', ''),
                "billing_date": cust.get('billing_date', ''),
                "city": cust.get('city', ''),
                "barangay": cust.get('barangay', ''),
                "address": cust.get('address', ''),
                "user_created": cust.get('user_created', 0)
            })
        
        # Get total count for pagination
        count_query = "SELECT COUNT(*) as total FROM customers c WHERE c.status = 'Approved'"
        count_params = []
        
        if city_filter and city_filter != "all":
            count_query += " AND c.city = %s"
            count_params.append(city_filter)
        
        if search_term:
            count_query += """ AND (c.first_name LIKE %s OR c.last_name LIKE %s 
                       OR c.email LIKE %s OR c.application_number LIKE %s 
                       OR c.contract_number LIKE %s)"""
            count_params.extend([search_pattern, search_pattern, search_pattern, search_pattern, search_pattern])
        
        total_result = execute_query(count_query, count_params, fetch_one=True)
        total_count = total_result['total'] if total_result else 0
        
        return jsonify({
            "customers": customers_list,
            "total": total_count,
            "limit": limit
        })
        
    except Exception as e:
        print("Error fetching customers:", e)
        return jsonify({"error": str(e)}), 500
    


@app.route("/api/superadmin/installation-summary", methods=["GET"])
def get_superadmin_installation_summary():
    try:
        from datetime import datetime, timedelta

        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        area = request.args.get("area")

        print("=" * 60)
        print(" INSTALLATION SUMMARY REQUEST")
        print(f"   start_date: {start_date}")
        print(f"   end_date: {end_date}")
        print(f"   area: {area}")
        print("=" * 60)

        # Convert input to datetime
        if start_date:
            start_date = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            end_date = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1, seconds=-1)

        # ISAMA ANG APPROVED AT CANCELLED NA STATUS
        query = """
            SELECT application_number, installation_status, city, status,
                   date_pending, date_ongoing, date_installed, 
                   date_cancelled, date_terminated
            FROM customers 
            WHERE status IN ('Approved', 'Cancelled')
        """
        params = []
        
        # Add area filter if specified
        if area and area != "" and area != "all":
            query += " AND city = %s"
            params.append(area.upper())
        
        # Execute query
        customers = execute_query(query, params, fetch=True) or []

        # PARA SA PIE GRAPH - GAMITIN ANG installation_status
        installation_summary = {
            "Pending": 0,
            "Ongoing": 0,
            "Installed": 0,
            "Cancelled": 0,
            "Terminated": 0
        }

        def parse_date(d):
            if not d:
                return None
            if isinstance(d, datetime):
                return d
            if isinstance(d, str):
                try:
                    return datetime.strptime(d, "%Y-%m-%d %H:%M:%S")
                except:
                    try:
                        return datetime.strptime(d, "%Y-%m-%d")
                    except:
                        return None
            return None

        matched_customers = 0

        print("\n PROCESSING CUSTOMERS:")
        for cust in customers:
            app_num = cust.get('application_number', 'unknown')
            installation_status = cust.get('installation_status', '').strip()
            customer_status = cust.get('status', '').strip()
            
            # KUNIN ANG MGA DATES PARA SA FILTER
            dates = {
                "Pending": parse_date(cust.get("date_pending")),
                "Ongoing": parse_date(cust.get("date_ongoing")),
                "Installed": parse_date(cust.get("date_installed")),
                "Cancelled": parse_date(cust.get("date_cancelled")),
                "Terminated": parse_date(cust.get("date_terminated"))
            }
            
            # I-FILTER KUNG NASA DATE RANGE
            is_in_date_range = False
            for status, dt in dates.items():
                if dt:
                    date_in_range = True
                    if start_date and dt < start_date:
                        date_in_range = False
                    if end_date and dt > end_date:
                        date_in_range = False
                    
                    if date_in_range:
                        is_in_date_range = True
                        break
            
            # KUNG WALANG DATES, ISAMA PA RIN (para hindi mawala ang customer)
            has_any_date = any(dt is not None for dt in dates.values())
            if not has_any_date:
                is_in_date_range = True  # Isama ang customer kahit walang dates
            
            if not is_in_date_range:
                print(f"\n  Customer: {app_num} - SKIPPED (not in date range)")
                continue
            
            print(f"\n  Customer: {app_num}")
            print(f"    - customer_status: {customer_status}")
            print(f"    - installation_status: {installation_status}")
            print(f"    - date_cancelled: {cust.get('date_cancelled')}")
            
            # GAMITIN ANG installation_status PARA SA PIE GRAPH
            if installation_status in installation_summary:
                installation_summary[installation_status] += 1
                matched_customers += 1
                print(f"    Status: {installation_status}")
            else:
                print(f"    Unknown status: {installation_status}")

        print("\n" + "=" * 60)
        print(f" FINAL SUMMARY (based on installation_status): {installation_summary}")
        print(f"   Total matched: {matched_customers}")
        print("=" * 60)

        response = {
            "installation_summary": installation_summary,
            "area": area if area else "All Areas",
            "total_matched": matched_customers,
            "date_range": {
                "start": start_date.strftime("%Y-%m-%d") if start_date else None,
                "end": (end_date + timedelta(seconds=1)).strftime("%Y-%m-%d") if end_date else None
            }
        }

        return jsonify(response), 200

    except Exception as e:
        print("Error in installation summary:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Server error", "details": str(e)}), 500
    


# =============================== 
# View Single Customer Application (Superadmin) 
# ===============================
@app.route("/superadmin/view-customer-application/<customer_id>")
def superadmin_view_customer_application(customer_id):
    return render_template("superadmin-view-customer-application.html", customer_id=customer_id)


@app.route("/api/superadmin/customer/<customer_id>")
def get_single_customer(customer_id):
    try:
        # Get customer data from customers table (approved applications only)
        customer_query = """
            SELECT * FROM customers 
            WHERE application_number = %s AND status = 'Approved'
        """
        customer_data = execute_query(customer_query, (customer_id,), fetch_one=True)
        
        if not customer_data:
            return jsonify({"error": "Customer not found or not approved"}), 404
        
        # Get additional data from applications table if needed
        app_query = """
            SELECT * FROM applications 
            WHERE application_number = %s
        """
        app_data = execute_query(app_query, (customer_id,), fetch_one=True)
        
        # Merge data (customer data has priority for customer-specific fields)
        if app_data:
            # Parse JSON fields from applications
            if app_data.get('tv_qty'):
                try:
                    app_data['tv_qty'] = json.loads(app_data['tv_qty'])
                except:
                    app_data['tv_qty'] = []
            
            if app_data.get('tv_brand'):
                try:
                    app_data['tv_brand'] = json.loads(app_data['tv_brand'])
                except:
                    app_data['tv_brand'] = []
            
            if app_data.get('tv_type'):
                try:
                    app_data['tv_type'] = json.loads(app_data['tv_type'])
                except:
                    app_data['tv_type'] = []
            
            # Merge data (customer data overrides)
            result = {**app_data, **customer_data}
        else:
            result = customer_data
        
        # ✅ GUMAWA NG RANDOM PASSWORD (GALING SA BACKEND)
        random_password = generate_secure_password(8)
        
        # Convert image paths to Cloudinary URLs
        image_fields = ['profile_photo', 'id_front', 'id_back', 'proof_billing', 'signature']
        for field in image_fields:
            if result.get(field):
                result[field] = get_cloudinary_url(result[field])
                print(f" Converted {field}: {result[field][:50]}...")
        
        # Parse JSON fields in customer data if any
        if result.get('tv_qty') and isinstance(result.get('tv_qty'), str):
            try:
                result['tv_qty'] = json.loads(result['tv_qty'])
            except:
                result['tv_qty'] = []
        
        if result.get('tv_brand') and isinstance(result.get('tv_brand'), str):
            try:
                result['tv_brand'] = json.loads(result['tv_brand'])
            except:
                result['tv_brand'] = []
        
        if result.get('tv_type') and isinstance(result.get('tv_type'), str):
            try:
                result['tv_type'] = json.loads(result['tv_type'])
            except:
                result['tv_type'] = []
        
        # ✅ IDAGDAG ANG PASSWORD SA RESPONSE
        result['password'] = random_password
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error getting single customer: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500



# =============================== 
# Generate User Password
# =============================== 
def generate_secure_password(length=8):
    """Generate a secure random password with letters and numbers only"""
    # Characters: uppercase, lowercase, and digits (no special characters para madaling i-type)
    characters = string.ascii_letters + string.digits
    # Ensure at least one letter and one number
    while True:
        password = ''.join(random.choice(characters) for _ in range(length))
        # Make sure it has at least one letter and one number
        if any(c.isalpha() for c in password) and any(c.isdigit() for c in password):
            return password


# =============================== 
# Create User Account
# ===============================    
@app.route("/api/superadmin/create-user-account", methods=["POST"])
def superadmin_create_user_account():
    """Create a user account for an installed customer and send email"""
    try:
        import random
        from datetime import datetime
        
        data = request.get_json()
        application_number = data.get("application_number")
        provided_user_id = data.get("user_id")  # Get the user_id from frontend
        provided_password = data.get("password")  # ✅ KUNIN ANG PASSWORD MULA SA FRONTEND
        
        if not application_number:
            return jsonify({"error": "Application number required"}), 400
        
        # Get customer data
        customer_query = "SELECT * FROM customers WHERE application_number = %s"
        customer = execute_query(customer_query, (application_number,), fetch_one=True)
        
        if not customer:
            return jsonify({"error": "Customer not found"}), 404
        
        # Check if installation status is 'Installed'
        if customer.get('installation_status') != 'Installed':
            return jsonify({"error": "Customer installation is not yet completed"}), 400
        
        # Check if user already exists
        check_user_sql = "SELECT user_id FROM users WHERE application_number = %s OR customer_id = %s"
        existing_user = execute_query(check_user_sql, (application_number, application_number), fetch_one=True)
        
        if existing_user:
            return jsonify({"error": "User account already exists for this customer", "user_id": existing_user['user_id']}), 400
        
        # Use the provided user_id or generate a new one
        if provided_user_id:
            # Check if the provided user_id already exists
            check_provided_query = "SELECT user_id FROM users WHERE user_id = %s"
            existing_provided = execute_query(check_provided_query, (provided_user_id,), fetch_one=True)
            
            if existing_provided:
                # If exists, generate a new one
                while True:
                    new_user_id = f"CV-{random.randint(1000, 9999)}"
                    check_user_query = "SELECT user_id FROM users WHERE user_id = %s"
                    existing = execute_query(check_user_query, (new_user_id,), fetch_one=True)
                    if not existing:
                        break
                print(f" Provided user_id {provided_user_id} already exists, using {new_user_id} instead")
            else:
                new_user_id = provided_user_id
        else:
            # Generate unique user ID (CV-XXXX format)
            while True:
                new_user_id = f"CV-{random.randint(1000, 9999)}"
                check_user_query = "SELECT user_id FROM users WHERE user_id = %s"
                existing = execute_query(check_user_query, (new_user_id,), fetch_one=True)
                if not existing:
                    break
        
        # ✅ GAMITIN ANG PASSWORD MULA SA FRONTEND KUNG MERON
        if provided_password and len(provided_password) == 8:
            default_password = provided_password
            print(f" Using password from frontend: {default_password}")
        else:
            default_password = generate_secure_password(8)
            print(f" Generated fallback password: {default_password}")
        
        current_time = ph_now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Get data from customer record
        first_name = customer.get('first_name', '')
        middle_name = customer.get('middle_name', '')
        last_name = customer.get('last_name', '')
        suffix = customer.get('suffix', '')
        email = customer.get('email', '')
        contract_number = customer.get('contract_number', '')
        mobile = customer.get('mobile', '')
        address = customer.get('address', '')
        
        # Insert user into users table
        insert_user_query = """
            INSERT INTO users 
            (user_id, customer_id, application_number, email, username, password, 
             created_at, role, connection_status, contract_number, status,
             first_name, last_name, middle_name, suffix, contact_number, address)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        execute_query(insert_user_query, (
            new_user_id, application_number, application_number, email, email, default_password,
            current_time, "customer", "Connected", contract_number, "Active",
            first_name, last_name, middle_name, suffix, mobile, address
        ))
        
        print(f" User {new_user_id} created for customer {application_number}")
        
        # Send email notification
        if email:
            try:
                send_account_creation_email(
                    email, 
                    new_user_id,
                    default_password, 
                    first_name, 
                    contract_number,
                    customer
                )
                print(f" Account creation email sent to {email}")
            except Exception as email_error:
                print(f"Email error: {email_error}")
        
        return jsonify({
            "success": True,
            "message": f"User account {new_user_id} created successfully",
            "user_id": new_user_id,
            "password": default_password,
            "customer": {
                "full_name": f"{first_name} {last_name}".strip(),
                "application_number": application_number,
                "contract_number": contract_number,
                "email": email,
                "mobile": mobile,
                "address": address,
                "plan": customer.get('plan'),
                "billing_date": customer.get('billing_date')
            }
        })
        
    except Exception as e:
        print(f" Error creating user account: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500




def send_account_creation_email(to_email, user_id, password, first_name, contract_number, customer_data):
    """Send account creation email to customer using Brevo API"""

    # ==========================================
    # BREVO API CONFIGURATION
    # ==========================================

    api_key = os.getenv("BREVO_API_KEY", "")

    if not api_key:
        print(" Brevo API key not configured!")
        return False

    print(f" Sending account creation email via Brevo API to {to_email}...")

    subject = "Cablevision - Your Account Has Been Created"

    # ==========================================
    # HTML EMAIL
    # ==========================================

    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Cablevision Account Created</title>
    </head>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', 'Inter', -apple-system, BlinkMacSystemFont, Arial, sans-serif; background-color: #eef2ff;">
        <div style="max-width: 580px; margin: 0 auto; padding: 30px 20px;">
            <div style="background: #ffffff; border-radius: 32px; overflow: hidden; box-shadow: 0 20px 35px -12px rgba(0, 0, 0, 0.15);">
                <!-- HEADER -->
                <div style="background: linear-gradient(135deg, #001f3f 0%, #002b5c 100%); padding: 32px 28px; text-align: center;">
                    <div style="font-size: 44px; margin-bottom: 8px;"></div>
                    <h1 style="margin: 0; font-size: 26px; font-weight: 700; color: #ffffff;">
                        Cablevision
                    </h1>
                    <p style="margin: 6px 0 0 0; color: #93c5fd; font-size: 13px;">
                        Account Created Successfully
                    </p>
                </div>

                <!-- CONTENT -->
                <div style="padding: 28px;">
                    <h2 style="margin: 0 0 8px 0; font-size: 22px; font-weight: 700; color: #0f172a;">
                        Welcome, {first_name}!
                    </h2>
                    <p style="margin: 0 0 20px 0; font-size: 15px; color: #475569;">
                        Your Cablevision account has been successfully created.
                        You can now log in to access your account details.
                    </p>

                    <!-- LOGIN CREDENTIALS -->
                    <div style="background: #f8fafc; border-radius: 20px; padding: 18px; margin-bottom: 20px;">
                        <div style="font-size: 12px; font-weight: 600; color: #64748b; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 12px; text-align: center;">
                            LOGIN CREDENTIALS
                        </div>
                        <div style="margin-bottom: 16px; padding-bottom: 12px; border-bottom: 1px solid #e2e8f0;">
                            <div style="font-size: 11px; font-weight: 600; color: #64748b; margin-bottom: 4px;">
                                User ID
                            </div>
                            <div style="font-size: 18px; font-weight: 700; color: #0f172a; font-family: monospace;">
                                {user_id}
                            </div>
                        </div>
                        <div>
                            <div style="font-size: 11px; font-weight: 600; color: #64748b; margin-bottom: 4px;">
                                Temporary Password
                            </div>
                            <div style="font-size: 18px; font-weight: 700; color: #f59e0b; font-family: monospace;">
                                {password}
                            </div>
                        </div>
                    </div>

                    <!-- CUSTOMER DETAILS -->
                    <div style="background: #eff6ff; border-radius: 20px; padding: 18px; margin-bottom: 16px;">
                        <div style="font-size: 12px; font-weight: 600; color: #1e40af; margin-bottom: 12px;">
                            ACCOUNT INFORMATION
                        </div>
                        <div style="margin-bottom: 10px;">
                            <span style="font-size: 12px; color: #64748b;">
                                Application #:
                            </span>
                            <span style="font-size: 13px; font-weight: 500; color: #1e293b; margin-left: 8px;">
                                {customer_data.get('application_number', 'N/A')}
                            </span>
                        </div>
                        <div style="margin-bottom: 10px;">
                            <span style="font-size: 12px; color: #64748b;">
                                Contract #:
                            </span>
                            <span style="font-size: 13px; font-weight: 500; color: #1e293b; margin-left: 8px;">
                                {contract_number or 'N/A'}
                            </span>
                        </div>
                        <div style="margin-bottom: 10px;">
                            <span style="font-size: 12px; color: #64748b;">
                                Plan:
                            </span>
                            <span style="font-size: 13px; font-weight: 500; color: #1e293b; margin-left: 8px;">
                                {customer_data.get('plan', 'N/A')}
                            </span>
                        </div>
                        <div style="margin-bottom: 10px;">
                            <span style="font-size: 12px; color: #64748b;">
                                Billing Date:
                            </span>
                            <span style="font-size: 13px; font-weight: 500; color: #1e293b; margin-left: 8px;">
                                {customer_data.get('billing_date', 'N/A')}
                            </span>
                        </div>
                    </div>

                    <!-- PASSWORD WARNING -->
                    <div style="margin: 20px 0; padding: 16px; background: #fef3c7; border-radius: 12px; border-left: 4px solid #f59e0b;">
                        <div style="display: flex; align-items: center; gap: 10px;">
                            <span style="font-size: 20px;"></span>
                            <div>
                                <div style="font-size: 13px; font-weight: 700; color: #92400e; margin-bottom: 4px;">
                                    Password Change Required
                                </div>
                                <div style="font-size: 12px; color: #92400e;">
                                    For security, please change your password after your first login.
                                </div>
                            </div>
                        </div>
                    </div>

                    <!-- THANK YOU -->
                    <div style="text-align: center; padding-top: 20px; border-top: 1px solid #e2e8f0;">
                        <p style="margin: 0; font-size: 12px; color: #94a3b8;">
                            Thank you for choosing Cablevision!
                        </p>
                    </div>
                </div>

                <!-- FOOTER -->
                <div style="background: #f1f5f9; padding: 16px 28px; text-align: center;">
                    <div style="font-size: 11px; color: #64748b;">
                        2026 Cablevision Internet Service Provider. All rights reserved.
                    </div>
                </div>
            </div>
        </div>
    </body>
    </html>
    """

    # ==========================================
    # SEND VIA BREVO API
    # ==========================================

    try:
        url = "https://api.brevo.com/v3/smtp/email"

        headers = {
            "accept": "application/json",
            "api-key": api_key,
            "content-type": "application/json"
        }

        data = {
            "sender": {
                "name": "Cablevision Systems Corp.",
                "email": "cablevision.cableinternet@gmail.com"
            },
            "to": [
                {
                    "email": to_email,
                    "name": first_name
                }
            ],
            "subject": subject,
            "htmlContent": html_body
        }

        response = requests.post(url, json=data, headers=headers, timeout=30)

        # ==============================================
        # CHECK BREVO RESPONSE
        # ==============================================

        if response.status_code in [200, 201, 202]:
            print(f" Account creation email sent successfully to {to_email}")

            try:
                print(f" Brevo response: {response.json()}")
            except Exception:
                pass

            return True
        else:
            print(f" Brevo API error: {response.status_code} - {response.text}")
            return False

    except requests.exceptions.Timeout:
        print(" Brevo API request timed out.")
        return False

    except requests.exceptions.RequestException as e:
        print(f" Brevo API request error: {e}")
        return False

    except Exception as e:
        print(f" Account creation email error: {e}")
        import traceback
        traceback.print_exc()
        return False



    

import os
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta

# ==================== SHARED UPLOADS CONFIGURATION ====================
# Shared folder sa labas ng project (sa XAMPP htdocs)
SHARED_UPLOADS_BASE = r"C:\xampp\htdocs\cablevision_uploads"

# Announcements subfolder
UPLOAD_FOLDER_ANNOUNCEMENTS = os.path.join(SHARED_UPLOADS_BASE, 'announcements')
ALLOWED_EXTENSIONS_ANNOUNCEMENTS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# Create upload folder if not exists
os.makedirs(UPLOAD_FOLDER_ANNOUNCEMENTS, exist_ok=True)

# NOTE: Ang /shared-uploads route ay nasa channel logos na
# HUWAG nang i-duplicate dito!

# ============================================================
# CLOUDINARY HELPER FUNCTIONS FOR ANNOUNCEMENTS
# ============================================================

def upload_to_cloudinary_announcement(file):
    """Upload announcement image to Cloudinary and return URL"""
    try:
        print(f" Uploading announcement to Cloudinary: {file.filename}")
        
        # I-reset ang file pointer
        file.stream.seek(0)
        
        filename_without_ext = file.filename.rsplit('.', 1)[0] if hasattr(file, 'filename') else None
        
        result = cloudinary.uploader.upload(
            file,
            folder="cablevision/announcements",
            resource_type="image",
            public_id=filename_without_ext,
            overwrite=True
        )
        
        print(f" Announcement uploaded: {result['secure_url']}")
        return result['secure_url']
        
    except Exception as e:
        print(f" Cloudinary upload error: {e}")
        import traceback
        traceback.print_exc()
        return None

def delete_from_cloudinary_announcement(image_url):
    """Delete announcement image from Cloudinary"""
    if not image_url:
        return
    
    try:
        if 'cloudinary.com' in image_url:
            parts = image_url.split('/upload/')
            if len(parts) > 1:
                public_id_with_ext = parts[1]
                print(f" Public ID with extension: {public_id_with_ext}")
                
                # Remove version number if present
                if '/' in public_id_with_ext and public_id_with_ext.split('/')[0].startswith('v'):
                    public_id_with_ext = '/'.join(public_id_with_ext.split('/')[1:])
                    print(f" After removing version: {public_id_with_ext}")
                
                # Remove file extension
                public_id = public_id_with_ext.rsplit('.', 1)[0]
                print(f" Final public_id: {public_id}")
                
                result = cloudinary.uploader.destroy(public_id, resource_type="image")
                print(f" Delete result: {result}")
                
                if result.get('result') == 'ok':
                    print(f" Successfully deleted announcement: {public_id}")
                    return True
                else:
                    print(f" Delete result: {result}")
                    return False
                    
    except Exception as e:
        print(f" Cloudinary delete error: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return False

# ===============================
# ALLOWED FILE CHECK
# ===============================
def allowed_announcement_file(filename):
    """Check if file extension is allowed for announcements"""
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ===============================
# SUPERADMIN ANNOUNCEMENTS PAGE
# ===============================
@app.route("/superadmin/announcements")
def superadmin_announcements():
    return render_template("superadmin-announcements.html")

# ===============================
# CREATE ANNOUNCEMENT - CLOUDINARY
# ===============================
@app.route("/api/superadmin/announcements", methods=["POST"])
def create_announcement():
    try:
        title = request.form.get("title", "")
        message = request.form.get("message", "")
        expiration_date = request.form.get("expirationDate")
        
        # Handle image upload
        image_file = request.files.get("image")
        image_url = None
        
        if image_file and allowed_announcement_file(image_file.filename):
            # Upload to Cloudinary
            image_url = upload_to_cloudinary_announcement(image_file)
            print(f" Image URL: {image_url}")
        
        if not title and not message and not image_url:
            return jsonify({"error": "Title, message, or image required"}), 400
        
        now = ph_now()
        
        # Insert into MySQL with Cloudinary URL
        insert_query = """
            INSERT INTO announcements (title, message, image_path, date, timestamp, expirationDate, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, NOW())
        """
        announcement_id = execute_query(insert_query, (
            title,
            message,
            image_url,  # Cloudinary URL
            now.strftime("%B %d, %Y"),
            now.timestamp(),
            expiration_date
        ))
        
        return jsonify({"message": "Announcement posted", "id": announcement_id})
        
    except Exception as e:
        print(f"Error creating announcement: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# GET ANNOUNCEMENTS (with Cloudinary URLs)
# ===============================
@app.route("/api/superadmin/announcements", methods=["GET"])
def get_announcements():
    try:
        query = """
            SELECT id, title, message, image_path, date, timestamp, expirationDate, created_at
            FROM announcements 
            ORDER BY timestamp DESC
        """
        announcements = execute_query(query, fetch=True) or []
        
        print(f"DEBUG: Found {len(announcements)} announcements")
        
        result = []
        for ann in announcements:
            image_path = ann.get('image_path', '')
            
            # If image_path exists, convert to Cloudinary URL if needed
            if image_path:
                # If it's a local path (starts with /shared-uploads/)
                if image_path.startswith('/shared-uploads/') or image_path.startswith('shared-uploads/'):
                    image_path = get_cloudinary_url(image_path)
                # If it's a cablevision/ path
                elif image_path.startswith('cablevision/'):
                    image_path = get_cloudinary_url(image_path)
                # If it's already a Cloudinary URL, keep it
                # If it's empty or None, keep as is
            
            result.append({
                "id": ann['id'],
                "title": ann.get('title', ''),
                "message": ann.get('message', ''),
                "imagePath": image_path,
                "date": ann.get('date', ''),
                "timestamp": ann.get('timestamp', 0),
                "expirationDate": ann.get('expirationDate', '')
            })
        
        print(f"DEBUG: Returning {len(result)} announcements")
        return jsonify(result)
        
    except Exception as e:
        print(f"Error getting announcements: {e}")
        return jsonify([]), 500

# ===============================
# UPDATE ANNOUNCEMENT - CLOUDINARY
# ===============================
@app.route("/api/superadmin/announcements/<int:announcement_id>", methods=["PUT"])
def update_announcement(announcement_id):
    try:
        title = request.form.get("title", "")
        message = request.form.get("message", "")
        expiration_date = request.form.get("expirationDate")
        
        # Check if announcement exists
        check_query = "SELECT id, image_path FROM announcements WHERE id = %s"
        existing = execute_query(check_query, (announcement_id,), fetch_one=True)
        
        if not existing:
            return jsonify({"error": "Announcement not found"}), 404
        
        # Handle image upload (optional)
        image_file = request.files.get("image")
        image_url = existing.get('image_path')
        
        if image_file and allowed_announcement_file(image_file.filename):
            # Delete old image from Cloudinary
            if image_url and 'cloudinary.com' in image_url:
                delete_from_cloudinary_announcement(image_url)
            # Upload new image to Cloudinary
            image_url = upload_to_cloudinary_announcement(image_file)
        
        # Build update query dynamically
        update_fields = []
        params = []
        
        if title:
            update_fields.append("title = %s")
            params.append(title)
        if message:
            update_fields.append("message = %s")
            params.append(message)
        if image_url:
            update_fields.append("image_path = %s")
            params.append(image_url)
        if expiration_date:
            update_fields.append("expirationDate = %s")
            params.append(expiration_date)
        
        if not update_fields:
            return jsonify({"message": "No updates provided"}), 200
        
        params.append(announcement_id)
        update_query = f"UPDATE announcements SET {', '.join(update_fields)} WHERE id = %s"
        execute_query(update_query, params)
        
        return jsonify({"message": "Announcement updated"})
        
    except Exception as e:
        print(f"Error updating announcement: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# DELETE ANNOUNCEMENT - CLOUDINARY
# ===============================
@app.route("/api/superadmin/announcements/<int:announcement_id>", methods=["DELETE"])
def delete_announcement(announcement_id):
    try:
        # Get image path first
        check_query = "SELECT id, image_path FROM announcements WHERE id = %s"
        announcement = execute_query(check_query, (announcement_id,), fetch_one=True)
        
        if not announcement:
            return jsonify({"error": "Announcement not found"}), 404
        
        # Delete image from Cloudinary
        image_url = announcement.get('image_path')
        if image_url and 'cloudinary.com' in image_url:
            delete_from_cloudinary_announcement(image_url)
        
        # Delete from MySQL
        delete_query = "DELETE FROM announcements WHERE id = %s"
        execute_query(delete_query, (announcement_id,))
        
        return jsonify({"message": "Announcement deleted"})
        
    except Exception as e:
        print(f"Error deleting announcement: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# DELETE EXPIRED ANNOUNCEMENTS - CLOUDINARY
# ===============================
@app.route("/api/superadmin/announcements/expired", methods=["DELETE"])
def delete_expired_announcements():
    try:
        from datetime import datetime as dt
        
        current_time_utc = dt.utcnow().isoformat() + "Z"
        print(f"[EXPIRY CHECK] Current UTC: {current_time_utc}")
        
        # Get expired announcements
        query = """
            SELECT id, image_path FROM announcements 
            WHERE expirationDate IS NOT NULL AND expirationDate < %s
        """
        expired_announcements = execute_query(query, (current_time_utc,), fetch=True) or []
        
        deleted_count = 0
        for ann in expired_announcements:
            # Delete image from Cloudinary
            image_url = ann.get('image_path')
            if image_url and 'cloudinary.com' in image_url:
                delete_from_cloudinary_announcement(image_url)
            
            # Delete from database
            delete_query = "DELETE FROM announcements WHERE id = %s"
            execute_query(delete_query, (ann['id'],))
            deleted_count += 1
            print(f"[DELETED] Announcement {ann['id']} (expired)")
        
        # Update announcements without expiration date
        update_query = """
            UPDATE announcements 
            SET expirationDate = %s 
            WHERE expirationDate IS NULL
        """
        new_exp_date = (dt.utcnow() + timedelta(days=7)).isoformat() + "Z"
        execute_query(update_query, (new_exp_date,))
        print(f"[UPDATED] Added expiration date to announcements without one")
        
        return jsonify({
            "deletedCount": deleted_count,
            "message": f"Deleted {deleted_count} expired announcements",
            "currentTime": current_time_utc
        })
        
    except Exception as e:
        print(f"Error deleting expired announcements: {e}")
        return jsonify({"error": str(e)}), 500



# ===============================
# SUPERADMIN - plan requests page
# ===============================
@app.route("/superadmin/plan-requests")
def superadmin_plan_requests():
    return render_template("superadmin-plan-requests.html")



@app.route("/api/superadmin/plan-requests", methods=["GET"])
def get_plan_change_requests():
    """Get all PENDING plan change requests for superadmin"""
    
    try:
        # PENDING LANG ANG KUNIN (hindi na kasama ang Approved at Rejected)
        query = """
            SELECT 
                pcr.id,
                pcr.request_id,
                pcr.application_number,
                pcr.current_plan,
                pcr.current_speed,
                pcr.current_price,
                pcr.requested_plan,
                pcr.requested_speed,
                pcr.requested_price,
                pcr.status,
                pcr.admin_notes,
                pcr.requested_at,
                pcr.reviewed_at,
                pcr.reviewed_by,
                c.first_name, 
                c.last_name, 
                c.email, 
                c.mobile, 
                c.city, 
                c.barangay,
                c.contract_number, 
                c.billing_date
            FROM plan_change_requests pcr
            LEFT JOIN customers c ON pcr.application_number = c.application_number
            WHERE pcr.status = 'Pending'  -- PINAKA IMPORTANTE: PENDING LANG
            ORDER BY pcr.requested_at DESC
        """
        requests = execute_query(query, fetch=True) or []
        
        # Format the data
        result = []
        for req in requests:
            result.append({
                "id": req.get("id"),
                "request_id": req.get("request_id"),
                "application_number": req.get("application_number"),
                "customer_name": f"{req.get('first_name', '')} {req.get('last_name', '')}".strip() or "N/A",
                "email": req.get("email") or "N/A",
                "mobile": req.get("mobile") or "N/A",
                "city": req.get("city") or "N/A",
                "barangay": req.get("barangay") or "N/A",
                "contract_number": req.get("contract_number") or "N/A",
                "billing_date": req.get("billing_date") or "N/A",
                "current_plan": req.get("current_plan") or "N/A",
                "current_speed": req.get("current_speed") or "N/A",
                "current_price": req.get("current_price") or "0",
                "requested_plan": req.get("requested_plan") or "N/A",
                "requested_speed": req.get("requested_speed") or "N/A",
                "requested_price": req.get("requested_price") or "0",
                "status": req.get("status"),
                "admin_notes": req.get("admin_notes"),
                "requested_at": str(req.get("requested_at")) if req.get("requested_at") else None,
                "reviewed_at": str(req.get("reviewed_at")) if req.get("reviewed_at") else None,
                "reviewed_by": req.get("reviewed_by")
            })
        
        print(f" Found {len(result)} PENDING plan change requests")
        return jsonify(result)
        
    except Exception as e:
        print(f"Error in get_plan_change_requests: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# SUPERADMIN - APPROVE PLAN CHANGE REQUEST
# ===============================
@app.route("/api/superadmin/approve-plan-request", methods=["POST"])
def approve_plan_request():
    """Superadmin approves a plan change request"""
    
    data = request.get_json()
    request_id = data.get("request_id")
    
    if not request_id:
        return jsonify({"error": "Request ID required"}), 400
    
    try:
        # Get the request details
        req_query = """
            SELECT pcr.*, c.email, c.first_name, c.last_name, c.contract_number, 
                c.plan_speed as current_speed, c.plan_price as current_price
            FROM plan_change_requests pcr
            JOIN customers c ON pcr.application_number = c.application_number
            WHERE pcr.id = %s AND pcr.status = 'Pending'
        """
        plan_request = execute_query(req_query, (request_id,), fetch_one=True)
        
        if not plan_request:
            return jsonify({"error": "Request not found or already processed"}), 404
        
        application_number = plan_request.get("application_number")
        requested_plan = plan_request.get("requested_plan")
        requested_speed = plan_request.get("requested_speed")
        requested_price = plan_request.get("requested_price")
        current_plan = plan_request.get("current_plan")
        customer_email = plan_request.get("email")
        customer_name = f"{plan_request.get('first_name', '')} {plan_request.get('last_name', '')}".strip()
        request_number = plan_request.get("request_id", f"REQ-{request_id}")
        
        # UPDATE customers table with new plan
        update_customer = """
            UPDATE customers 
            SET plan = %s, plan_speed = %s, plan_price = %s
            WHERE application_number = %s
        """
        execute_query(update_customer, (requested_plan, requested_speed, requested_price, application_number))
        
        # UPDATE plan_change_requests status
        update_request = """
            UPDATE plan_change_requests 
            SET status = 'Approved', reviewed_at = NOW(), reviewed_by = 'superadmin'
            WHERE id = %s
        """
        execute_query(update_request, (request_id,))
        
        # Kunin muna ang original admin notification para malaman ang city
        original_notif_query = """
            SELECT admin_city, application_city 
            FROM admin_notifications 
            WHERE relatedId = %s AND type = 'plan_change_request'
            ORDER BY id DESC LIMIT 1
        """
        original_notif = execute_query(original_notif_query, (application_number,), fetch_one=True)

        # Update admin notification to mark as processed
        update_admin_notif = """
            UPDATE admin_notifications 
            SET action_taken_by = 'superadmin', action_status = 'Approved', read_status = 1
            WHERE relatedId = %s AND type = 'plan_change_request'
        """
        execute_query(update_admin_notif, (application_number,))

        # BAGONG NOTIFICATION PARA MALAMAN NG ADMIN NA NA-PROCESS NA
        admin_notif_id = int(datetime.now().timestamp() * 1000)
        admin_notif_query = """
            INSERT INTO admin_notifications (
                id, title, message, type, relatedId, request_id, timestamp,
                read_status, admin_city, application_city, application_id,
                action_taken_by, action_status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        execute_query(admin_notif_query, (
            admin_notif_id,
            "Plan Change Request Approved",
            f"[{request_number}] The plan change request for {customer_name} (from {current_plan or 'N/A'} to {requested_plan}) has been APPROVED by the superadmin.",
            "plan_change_processed",
            application_number,
            request_number,
            ph_now_iso(),
            0,
            original_notif.get("admin_city") if original_notif else None,
            original_notif.get("application_city") if original_notif else None,
            application_number,
            "superadmin",
            "Approved"
        ))
        print(f" Admin notification created: Plan change request {application_number} approved")
        
        # ========== SEND EMAIL NOTIFICATION FOR APPROVAL ==========
        try:
            if customer_email:
                print(f" Attempting to send approval email to {customer_email}...")
                send_plan_change_email(
                    to_email=customer_email,
                    first_name=customer_name.split()[0] if customer_name else "Customer",
                    status="Approved",
                    request_id=request_number,
                    current_plan=current_plan,
                    requested_plan=requested_plan,
                    requested_speed=requested_speed,
                    requested_price=requested_price,
                    reason=None,
                    current_speed=plan_request.get("current_speed"),      # IDAGDAG ITO
                    current_price=plan_request.get("current_price")       # IDAGDAG ITO
                )
            else:
                print(f" No email address for customer {application_number}")
        except Exception as email_err:
            print(f" Email error: {email_err}")
            import traceback
            traceback.print_exc()
        
        # ========== GET THE CORRECT USER_ID FROM USERS TABLE ==========
        user_query = """
            SELECT user_id, email, first_name, last_name
            FROM users 
            WHERE application_number = %s
            LIMIT 1
        """
        user_info = execute_query(user_query, (application_number,), fetch_one=True)
        
        if user_info:
            actual_user_id = user_info.get("user_id")
            user_email = user_info.get("email")
            user_name = f"{user_info.get('first_name', '')} {user_info.get('last_name', '')}".strip()
        else:
            actual_user_id = application_number
            user_email = plan_request.get("email")
            user_name = customer_name
        
        print(f" Creating notification for user_id: {actual_user_id}")
        
        # ========== CREATE NOTIFICATION FOR USER (APPROVED) ==========
        notification_id = int(datetime.now().timestamp() * 1000)
        notif_query = """
            INSERT INTO user_notifications (
                id, title, message, type, relatedId, user_id, user_email, user_name,
                connection_status, timestamp, read_status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        execute_query(notif_query, (
            notification_id,
            "Plan Change Approved",
            f"Your request to change your plan from {current_plan or 'N/A'} to {requested_plan} ({requested_speed}) has been APPROVED. Your new plan is now active.",
            "plan_change_approved",
            application_number,
            actual_user_id,
            user_email,
            user_name,
            "Active",
            ph_now_iso(),
            0
        ))
        print(f" User notification created for {user_name} - Plan Change Approved")
        
        return jsonify({
            "success": True,
            "message": f"Plan change request has been approved. Customer's plan updated to {requested_plan}. User notified via email and in-app notification."
        })
        
    except Exception as e:
        print(f"Error in approve_plan_request: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# SEND PLAN CHANGE STATUS EMAIL
# APPROVED / REJECTED
# BREVO API
# ===============================

def send_plan_change_email(
    to_email,
    first_name,
    status,
    request_id,
    current_plan,
    requested_plan,
    requested_speed,
    requested_price,
    reason=None,
    current_speed=None,
    current_price=None
):
    """
    Sends plan change status email using Brevo HTTP API.

    Status:
        - Approved
        - Rejected

    Returns:
        True  = email sent successfully
        False = email sending failed
    """

    import requests
    import html

    # ===============================
    # BREVO CONFIGURATION
    # ===============================

    brevo_api_key = os.getenv("BREVO_API_KEY")

    sender_email = os.getenv(
        "SMTP_FROM",
        "noreply@cablevisioncableinternet.com"
    )

    sender_name = "Cablevision Systems Corporation"

    # ===============================
    # CHECK CONFIGURATION
    # ===============================

    if not brevo_api_key:
        print(" BREVO_API_KEY is not configured!")
        return False

    if not to_email:
        print(" Customer email is empty!")
        return False

    # ===============================
    # NORMALIZE STATUS
    # ===============================

    status = str(status or "").strip().capitalize()

    if status not in ("Approved", "Rejected"):
        print(f" Invalid plan change status: {status}")
        return False

    # ===============================
    # HTML ESCAPE
    # ===============================

    def escape_html(value):
        if value is None:
            return ""
        return html.escape(str(value))

    # ===============================
    # FORMAT PRICE
    # ===============================

    def format_price(value):
        try:
            return f"₱{float(value or 0):,.2f}"
        except (ValueError, TypeError):
            return f"₱{escape_html(value or 0)}"

    # ===============================
    # CUSTOMER / PLAN VALUES
    # ===============================

    first_name = str(first_name or "Customer").strip()

    safe_request_id = escape_html(request_id or "N/A")
    safe_current_plan = escape_html(current_plan or "N/A")
    safe_current_speed = escape_html(current_speed or "N/A")
    safe_requested_plan = escape_html(requested_plan or "N/A")
    safe_requested_speed = escape_html(requested_speed or "N/A")

    formatted_current_price = format_price(current_price)
    formatted_requested_price = format_price(requested_price)

    # ===============================
    # STATUS SETTINGS
    # ===============================

    if status == "Approved":
        status_color = "#16a34a"
        status_bg = "#dcfce7"
        status_icon = ""
        action_text = "APPROVED"

        message = f"Congratulations, {first_name}!"
        message_sub = "Your plan change request has been approved by Cablevision."

    else:
        status_color = "#dc2626"
        status_bg = "#fee2e2"
        status_icon = ""
        action_text = "REJECTED"

        message = f"Plan Change Update, {first_name}"
        message_sub = "Your plan change request has been reviewed by Cablevision."

    # ===============================
    # REJECTION REASON
    # ===============================

    rejection_reason = escape_html(
        reason.strip() if isinstance(reason, str) else reason
    )

    if not rejection_reason:
        rejection_reason = "No specific reason was provided."

    # ===============================
    # STATUS MESSAGE
    # ===============================

    if status == "Approved":
        extra_message = f"""
        <div style="
            margin-top:20px;
            padding:18px;
            background:#f0fdf4;
            border:1px solid #bbf7d0;
            border-radius:14px;
        ">
            <div style="
                font-size:15px;
                font-weight:700;
                color:#166534;
                margin-bottom:8px;
            ">
                Your new plan is now active
            </div>

            <div style="
                font-size:14px;
                line-height:1.6;
                color:#14532d;
            ">
                Your request to change from
                <strong>{safe_current_plan}</strong>
                to
                <strong>{safe_requested_plan}</strong>
                has been approved.
                Your new plan will be reflected according to your
                Cablevision billing and service schedule.
            </div>
        </div>
        """

    else:
        extra_message = f"""
        <div style="
            margin-top:20px;
            padding:18px;
            background:#fef2f2;
            border:1px solid #fecaca;
            border-radius:14px;
        ">
            <div style="
                font-size:15px;
                font-weight:700;
                color:#991b1b;
                margin-bottom:8px;
            ">
                Reason for Rejection
            </div>

            <div style="
                font-size:14px;
                line-height:1.6;
                color:#7f1d1d;
            ">
                {rejection_reason}
            </div>

            <div style="
                margin-top:12px;
                font-size:13px;
                line-height:1.5;
                color:#7f1d1d;
            ">
                You may submit another plan change request if applicable.
                Please contact Cablevision support if you need assistance.
            </div>
        </div>
        """

    # ===============================
    # PLAN COMPARISON
    # ===============================

    plan_comparison_html = f"""
    <div style="
        background:#f8fafc;
        border-radius:18px;
        padding:18px;
        margin-bottom:20px;
    ">
        <div style="
            text-align:center;
            margin-bottom:16px;
        ">
            <span style="
                display:inline-block;
                background:#e2e8f0;
                color:#475569;
                padding:5px 14px;
                border-radius:20px;
                font-size:11px;
                font-weight:700;
            ">
                PLAN COMPARISON
            </span>
        </div>

        <table width="100%" cellpadding="0" cellspacing="0" border="0">
            <tr>
                <td width="45%" valign="top" style="
                    background:#ffffff;
                    border:1px solid #e2e8f0;
                    border-radius:14px;
                    padding:15px;
                    text-align:center;
                ">
                    <div style="
                        font-size:11px;
                        font-weight:700;
                        color:#dc2626;
                        margin-bottom:8px;
                    ">
                        CURRENT PLAN
                    </div>

                    <div style="
                        font-size:17px;
                        font-weight:700;
                        color:#0f172a;
                        margin-bottom:5px;
                    ">
                        {safe_current_plan}
                    </div>

                    <div style="
                        font-size:20px;
                        font-weight:800;
                        color:#dc2626;
                    ">
                        {safe_current_speed}
                    </div>

                    <div style="
                        font-size:14px;
                        color:#dc2626;
                        font-weight:600;
                    ">
                        {formatted_current_price}
                    </div>
                </td>

                <td width="10%" valign="middle" style="
                    text-align:center;
                    font-size:24px;
                    color:#94a3b8;
                ">
                    →
                </td>

                <td width="45%" valign="top" style="
                    background:#ffffff;
                    border:1px solid #e2e8f0;
                    border-radius:14px;
                    padding:15px;
                    text-align:center;
                ">
                    <div style="
                        font-size:11px;
                        font-weight:700;
                        color:#16a34a;
                        margin-bottom:8px;
                    ">
                        REQUESTED PLAN
                    </div>

                    <div style="
                        font-size:17px;
                        font-weight:700;
                        color:#0f172a;
                        margin-bottom:5px;
                    ">
                        {safe_requested_plan}
                    </div>

                    <div style="
                        font-size:20px;
                        font-weight:800;
                        color:#16a34a;
                    ">
                        {safe_requested_speed}
                    </div>

                    <div style="
                        font-size:14px;
                        color:#16a34a;
                        font-weight:600;
                    ">
                        {formatted_requested_price}
                    </div>
                </td>
            </tr>
        </table>
    </div>
    """

    # ===============================
    # HTML EMAIL
    # ===============================

    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>CableVision Plan Change</title>
    </head>

    <body style="
        margin:0;
        padding:0;
        background:#eef2ff;
        font-family:Arial, Helvetica, sans-serif;
    ">

        <div style="
            width:100%;
            padding:30px 0;
        ">
            <div style="
                max-width:580px;
                margin:0 auto;
                background:#ffffff;
                border-radius:24px;
                overflow:hidden;
                box-shadow:0 10px 30px rgba(0,0,0,0.10);
            ">

                <!-- HEADER -->
                <div style="
                    background:#001f3f;
                    padding:30px 25px;
                    text-align:center;
                ">
                    <div style="
                        font-size:28px;
                        font-weight:700;
                        color:#ffffff;
                    ">
                        Cablevision
                    </div>

                    <div style="
                        margin-top:6px;
                        font-size:13px;
                        color:#93c5fd;
                    ">
                        Internet Service Provider
                    </div>

                    <div style="
                        margin-top:16px;
                        display:inline-block;
                        background:rgba(255,255,255,0.12);
                        color:#dbeafe;
                        padding:6px 14px;
                        border-radius:20px;
                        font-size:11px;
                        font-weight:700;
                    ">
                        PLAN CHANGE
                    </div>
                </div>

                <!-- STATUS -->
                <div style="
                    text-align:center;
                    padding:22px 20px 5px;
                ">
                    <span style="
                        display:inline-block;
                        background:{status_bg};
                        color:{status_color};
                        padding:9px 22px;
                        border-radius:30px;
                        font-size:13px;
                        font-weight:700;
                    ">
                        {status_icon} REQUEST {action_text}
                    </span>
                </div>

                <!-- CONTENT -->
                <div style="
                    padding:20px 28px 30px;
                ">
                    <h2 style="
                        margin:0 0 8px;
                        font-size:22px;
                        color:#0f172a;
                    ">
                        {escape_html(message)}
                    </h2>

                    <p style="
                        margin:0 0 20px;
                        font-size:15px;
                        line-height:1.5;
                        color:#475569;
                    ">
                        {escape_html(message_sub)}
                    </p>

                    <!-- REQUEST DETAILS -->
                    <div style="
                        background:#f8fafc;
                        border-radius:16px;
                        padding:18px;
                        margin-bottom:18px;
                    ">
                        <div style="
                            font-size:11px;
                            font-weight:700;
                            color:#64748b;
                            margin-bottom:5px;
                        ">
                            REQUEST ID
                        </div>

                        <div style="
                            font-size:18px;
                            font-weight:700;
                            color:#0f172a;
                            font-family:monospace;
                            margin-bottom:16px;
                        ">
                            {safe_request_id}
                        </div>

                        <div style="
                            border-top:1px solid #e2e8f0;
                            padding-top:12px;
                        ">
                            <div style="
                                font-size:11px;
                                font-weight:700;
                                color:#64748b;
                            ">
                                STATUS
                            </div>

                            <div style="
                                margin-top:4px;
                                font-size:16px;
                                font-weight:700;
                                color:{status_color};
                            ">
                                {action_text}
                            </div>
                        </div>
                    </div>

                    <!-- PLAN COMPARISON -->
                    {plan_comparison_html}

                    <!-- STATUS MESSAGE -->
                    {extra_message}

                    <div style="
                        margin-top:28px;
                        padding-top:20px;
                        border-top:1px solid #e2e8f0;
                        text-align:center;
                    ">
                        <div style="
                            font-size:13px;
                            color:#64748b;
                        ">
                            Thank you for choosing Cablevision!
                        </div>
                    </div>
                </div>

                <!-- FOOTER -->
                <div style="
                    background:#f1f5f9;
                    padding:16px 20px;
                    text-align:center;
                ">
                    <div style="
                        font-size:11px;
                        color:#64748b;
                    ">
                        2026 Cablevision Internet Service Provider.
                        All rights reserved.
                    </div>
                </div>

            </div>
        </div>

    </body>
    </html>
    """

    # ===============================
    # PLAIN TEXT EMAIL
    # ===============================

    plain_body = (
        f"CableVision Plan Change Request {action_text}\n\n"
        f"Hello {first_name},\n\n"
        f"Your plan change request has been {status.lower()}.\n\n"
        f"Request ID: {request_id}\n\n"
        f"Current Plan:\n"
        f"{current_plan or 'N/A'}\n"
        f"Speed: {current_speed or 'N/A'}\n"
        f"Price: {formatted_current_price}\n\n"
        f"Requested Plan:\n"
        f"{requested_plan or 'N/A'}\n"
        f"Speed: {requested_speed or 'N/A'}\n"
        f"Price: {formatted_requested_price}\n"
    )

    if status == "Rejected":
        plain_body += (
            f"\nReason for Rejection:\n"
            f"{reason or 'No specific reason was provided.'}\n"
        )

    plain_body += (
        "\nThank you for choosing Cablevision!\n\n"
        "Cablevision Systems Corporation"
    )

    # ===============================
    # BREVO API PAYLOAD
    # ===============================

    payload = {
        "sender": {
            "name": sender_name,
            "email": sender_email
        },
        "to": [
            {
                "email": to_email,
                "name": first_name
            }
        ],
        "subject": (
            f"CableVision - Plan Change Request {action_text}"
        ),
        "htmlContent": html_body,
        "textContent": plain_body
    }

    # ===============================
    # BREVO API HEADERS
    # ===============================

    headers = {
        "accept": "application/json",
        "api-key": brevo_api_key,
        "content-type": "application/json"
    }

    # ===============================
    # SEND EMAIL
    # ===============================

    try:
        print(
            f" Sending plan change {status.lower()} "
            f"email via Brevo to {to_email}..."
        )

        response = requests.post(
            "https://api.brevo.com/v3/smtp/email",
            headers=headers,
            json=payload,
            timeout=30
        )

        if response.status_code not in (200, 201):
            print(
                f" Brevo API error "
                f"({response.status_code}): "
                f"{response.text}"
            )
            return False

        try:
            brevo_response = response.json()
            message_id = brevo_response.get("messageId")

            if message_id:
                print(f" Brevo Message ID: {message_id}")

        except Exception:
            pass

        print(
            f" Plan change {status.lower()} email "
            f"sent successfully to {to_email}"
        )

        return True

    except requests.exceptions.Timeout:
        print(" Brevo API request timed out")
        return False

    except requests.exceptions.RequestException as e:
        print(f" Brevo API request error: {e}")
        return False

    except Exception as e:
        print(f" Error sending plan change email: {e}")

        import traceback
        traceback.print_exc()

        return False


# ===============================
# SUPERADMIN - REJECT PLAN CHANGE REQUEST (WITH EMAIL)
# ===============================
@app.route("/api/superadmin/reject-plan-request", methods=["POST"])
def reject_plan_request():
    """Superadmin rejects a plan change request"""
    
    data = request.get_json()
    request_id = data.get("request_id")
    reason = data.get("reason", "")
    
    if not request_id:
        return jsonify({"error": "Request ID required"}), 400
    
    try:
        # Get the request details
        check_query = """
            SELECT pcr.*, c.email, c.first_name, c.last_name, c.contract_number,
                c.plan_speed as current_speed, c.plan_price as current_price
            FROM plan_change_requests pcr
            JOIN customers c ON pcr.application_number = c.application_number
            WHERE pcr.id = %s AND pcr.status = 'Pending'
        """
        plan_request = execute_query(check_query, (request_id,), fetch_one=True)
        
        if not plan_request:
            return jsonify({"error": "Request not found or already processed"}), 404
        
        application_number = plan_request.get("application_number")
        requested_plan = plan_request.get("requested_plan")
        requested_speed = plan_request.get("requested_speed")
        requested_price = plan_request.get("requested_price")
        current_plan = plan_request.get("current_plan")
        customer_email = plan_request.get("email")
        customer_name = f"{plan_request.get('first_name', '')} {plan_request.get('last_name', '')}".strip()
        request_number = plan_request.get("request_id", f"REQ-{request_id}")
        
        # UPDATE plan_change_requests status
        update_request = """
            UPDATE plan_change_requests 
            SET status = 'Rejected', reviewed_at = NOW(), reviewed_by = 'superadmin', admin_notes = %s
            WHERE id = %s
        """
        execute_query(update_request, (reason, request_id))
        
        # Kunin muna ang original admin notification para malaman ang city
        original_notif_query = """
            SELECT admin_city, application_city 
            FROM admin_notifications 
            WHERE relatedId = %s AND type = 'plan_change_request'
            ORDER BY id DESC LIMIT 1
        """
        original_notif = execute_query(original_notif_query, (application_number,), fetch_one=True)

        # Update admin notification to mark as processed
        update_admin_notif = """
            UPDATE admin_notifications 
            SET action_taken_by = 'superadmin', action_status = 'Rejected', read_status = 1
            WHERE relatedId = %s AND type = 'plan_change_request'
        """
        execute_query(update_admin_notif, (application_number,))

        # BAGONG NOTIFICATION PARA MALAMAN NG ADMIN NA NA-PROCESS NA
        admin_notif_id = int(datetime.now().timestamp() * 1000)
        admin_notif_query = """
            INSERT INTO admin_notifications (
                id, title, message, type, relatedId, request_id, timestamp,
                read_status, admin_city, application_city, application_id,
                action_taken_by, action_status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        execute_query(admin_notif_query, (
            admin_notif_id,
            "Plan Change Request Rejected",
            f"[{request_number}] The plan change request for {customer_name} (from {current_plan or 'N/A'} to {requested_plan}) has been REJECTED by the superadmin.",
            "plan_change_processed",
            application_number,
            request_number,
            ph_now_iso(),
            0,
            original_notif.get("admin_city") if original_notif else None,
            original_notif.get("application_city") if original_notif else None,
            application_number,
            "superadmin",
            "Rejected"
        ))
        print(f" Admin notification created: Plan change request {application_number} rejected")
        
        # ========== SEND EMAIL NOTIFICATION FOR REJECTION ==========
        try:
            if customer_email:
                print(f" Attempting to send rejection email to {customer_email}...")
                send_plan_change_email(
                    to_email=customer_email,
                    first_name=customer_name.split()[0] if customer_name else "Customer",
                    status="Rejected",
                    request_id=request_number,
                    current_plan=current_plan,
                    requested_plan=requested_plan,
                    requested_speed=requested_speed,
                    requested_price=requested_price,
                    reason=reason,
                    current_speed=plan_request.get("current_speed"),      # IDAGDAG ITO
                    current_price=plan_request.get("current_price")       # IDAGDAG ITO
                )
            else:
                print(f" No email address for customer {application_number}")
        except Exception as email_err:
            print(f" Email error: {email_err}")
            import traceback
            traceback.print_exc()
        
        # ========== GET THE CORRECT USER_ID FROM USERS TABLE ==========
        user_query = """
            SELECT user_id, email, first_name, last_name
            FROM users 
            WHERE application_number = %s
            LIMIT 1
        """
        user_info = execute_query(user_query, (application_number,), fetch_one=True)
        
        if user_info:
            actual_user_id = user_info.get("user_id")
            user_email = user_info.get("email")
            user_name = f"{user_info.get('first_name', '')} {user_info.get('last_name', '')}".strip()
        else:
            actual_user_id = application_number
            user_email = plan_request.get("email")
            user_name = customer_name
        
        print(f" Creating rejection notification for user_id: {actual_user_id}")
        
        # ========== CREATE NOTIFICATION FOR USER (REJECTED) ==========
        notification_id = int(datetime.now().timestamp() * 1000)
        notif_query = """
            INSERT INTO user_notifications (
                id, title, message, type, relatedId, user_id, user_email, user_name,
                connection_status, timestamp, read_status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        rejection_message = f"Your request to change your plan from {current_plan or 'N/A'} to {requested_plan} has been REJECTED."
        if reason and reason != "No reason provided" and reason != "":
            rejection_message += f" Reason: {reason}"
        
        execute_query(notif_query, (
            notification_id,
            "Plan Change Rejected",
            rejection_message,
            "plan_change_rejected",
            application_number,
            actual_user_id,
            user_email,
            user_name,
            "Active",
            ph_now_iso(),
            0
        ))
        print(f" User notification created for {user_name} (user_id: {actual_user_id}) - Plan Change Rejected")
        
        return jsonify({
            "success": True,
            "message": f"Plan change request has been rejected. Customer has been notified via email."
        })
        
    except Exception as e:
        print(f"Error in reject_plan_request: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# SUPERADMIN - TERMINATION REQUESTS PAGE
# ===============================
@app.route("/superadmin/termination-requests")
def superadmin_termination_requests():
    return render_template("superadmin-termination-requests.html")


# ===============================
# SUPERADMIN - GET TERMINATION REQUESTS
# ===============================
@app.route("/api/superadmin/termination-requests", methods=["GET"])
def get_termination_requests():
    """Get all PENDING termination requests for superadmin"""
    
    try:
        query = """
            SELECT 
                tr.id,
                tr.request_id,
                tr.application_number,
                tr.user_id,
                tr.first_name,
                tr.last_name,
                tr.email,
                tr.contact_number,
                tr.city,
                tr.contract_number,
                tr.current_plan,
                tr.current_speed,
                tr.current_price,
                tr.termination_reason,
                tr.termination_date,
                tr.status,
                tr.admin_notes,
                tr.created_at,
                tr.updated_at,
                c.plan as customer_current_plan,
                c.plan_speed as customer_current_speed,
                c.plan_price as customer_current_price
            FROM termination_requests tr
            LEFT JOIN customers c ON tr.application_number = c.application_number
            WHERE tr.status = 'Pending'
            ORDER BY tr.created_at DESC
        """
        requests = execute_query(query, fetch=True) or []
        
        result = []
        for req in requests:
            result.append({
                "id": req.get("id"),
                "request_id": req.get("request_id"),
                "application_number": req.get("application_number"),
                "customer_name": f"{req.get('first_name', '')} {req.get('last_name', '')}".strip() or "N/A",
                "email": req.get("email") or "N/A",
                "contact_number": req.get("contact_number") or "N/A",
                "city": req.get("city") or "N/A",
                "contract_number": req.get("contract_number") or "N/A",
                "current_plan": req.get("current_plan") or req.get("customer_current_plan") or "N/A",
                "current_speed": req.get("current_speed") or req.get("customer_current_speed") or "N/A",
                "current_price": req.get("current_price") or req.get("customer_current_price") or "0",
                "termination_reason": req.get("termination_reason") or "No reason provided",
                "termination_date": str(req.get("termination_date")) if req.get("termination_date") else None,
                "status": req.get("status"),
                "admin_notes": req.get("admin_notes"),
                "created_at": str(req.get("created_at")) if req.get("created_at") else None,
                "updated_at": str(req.get("updated_at")) if req.get("updated_at") else None
            })
        
        print(f" Found {len(result)} PENDING termination requests")
        return jsonify(result)
        
    except Exception as e:
        print(f"Error in get_termination_requests: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# SUPERADMIN - APPROVE TERMINATION REQUEST
# ===============================
@app.route("/api/superadmin/approve-termination", methods=["POST"])
def approve_termination_request():
    """Superadmin approves a termination request"""
    
    data = request.get_json()
    request_id = data.get("request_id")
    balance = data.get("balance", 0)
    
    if not request_id:
        return jsonify({"error": "Request ID required"}), 400
    
    try:
        # Get the request details
        req_query = """
            SELECT tr.*, c.email as customer_email, c.first_name, c.last_name, 
                   c.contract_number, c.plan, c.plan_speed, c.plan_price
            FROM termination_requests tr
            JOIN customers c ON tr.application_number = c.application_number
            WHERE tr.id = %s AND tr.status = 'Pending'
        """
        termination_request = execute_query(req_query, (request_id,), fetch_one=True)
        
        if not termination_request:
            return jsonify({"error": "Request not found or already processed"}), 404
        
        application_number = termination_request.get("application_number")
        customer_email = termination_request.get("customer_email")
        customer_name = f"{termination_request.get('first_name', '')} {termination_request.get('last_name', '')}".strip()
        request_number = termination_request.get("request_id", f"REQ-{request_id}")
        
        # KUNIN ANG CURRENT DATETIME PARA SA TERMINATION DATE
        from datetime import datetime
        termination_date = ph_now().strftime("%Y-%m-%d %H:%M:%S")
        termination_date_display = ph_now().strftime("%B %d, %Y at %I:%M %p")
        
        print(f" Termination date: {termination_date}")
        print(f" Balance to save: {balance}")
        
        # HUWAG GALAWIN ANG customers TABLE - application status lang yun
        # ANG STATUS SA users TABLE LANG ANG MAGIGING "Terminated"
        
        # UPDATE users table - set status to 'Terminated'
        update_user = """
            UPDATE users 
            SET status = 'Terminated', 
                connection_status = 'Disconnected',
                balance = %s
            WHERE application_number = %s
        """
        execute_query(update_user, (balance, application_number))
        print(f" User updated: status=Terminated, connection=Disconnected, balance={balance}")
        
        # UPDATE customers table - set installation_status to 'Terminated' AND date_terminated
        update_customer = """
            UPDATE customers 
            SET installation_status = 'Terminated',
                date_terminated = %s
            WHERE application_number = %s
        """
        execute_query(update_customer, (termination_date, application_number))
        print(f" Customer updated: installation_status=Terminated, date_terminated={termination_date}")
        
        # UPDATE applications table - set installation_status to 'Terminated'
        execute_query("UPDATE applications SET installation_status = 'Terminated' WHERE application_number = %s", (application_number,))

        # UPDATE napbox_slots - gawing 'available' ang slot ng user
        contract_number = termination_request.get("contract_number")
        update_slot_query = """
            UPDATE napbox_slots 
            SET status = 'available', 
                updated_at = NOW()
            WHERE (application_number IS NOT NULL AND application_number != '' AND application_number = %s)
               OR (contract_number IS NOT NULL AND contract_number != '' AND contract_number = %s)
        """
        execute_query(update_slot_query, (application_number, contract_number))
        print(f" NAP Box slot set to 'available' for application: {application_number}, contract: {contract_number}")

        # UPDATE termination_requests status
        update_request = """
            UPDATE termination_requests 
            SET status = 'Approved', updated_at = NOW()
            WHERE id = %s
        """
        execute_query(update_request, (request_id,))
        
        # Kunin muna ang original admin notification para malaman ang city
        original_notif_query = """
            SELECT admin_city, application_city 
            FROM admin_notifications 
            WHERE relatedId = %s AND type = 'termination_request'
            ORDER BY id DESC LIMIT 1
        """
        original_notif = execute_query(original_notif_query, (application_number,), fetch_one=True)

        # Update admin notification to mark as processed
        update_admin_notif = """
            UPDATE admin_notifications 
            SET action_taken_by = 'superadmin', action_status = 'Approved', read_status = 1
            WHERE relatedId = %s AND type = 'termination_request'
        """
        execute_query(update_admin_notif, (application_number,))

        # BAGONG NOTIFICATION PARA MALAMAN NG ADMIN NA NA-PROCESS NA
        admin_notif_id = int(datetime.now().timestamp() * 1000)
        admin_notif_query = """
            INSERT INTO admin_notifications (
                id, title, message, type, relatedId, request_id, timestamp,
                read_status, admin_city, application_city, application_id,
                action_taken_by, action_status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        execute_query(admin_notif_query, (
            admin_notif_id,
            "Termination Request Approved",
            f"[{request_number}] The termination request for {customer_name} has been APPROVED by the superadmin on {termination_date_display}.",
            "termination_processed",
            application_number,
            request_number,
            ph_now_iso(),
            0,
            original_notif.get("admin_city") if original_notif else termination_request.get("city"),
            original_notif.get("application_city") if original_notif else termination_request.get("city"),
            application_number,
            "superadmin",
            "Approved"
        ))
        print(f" Admin notification created: Termination request {application_number} approved")
        
        # ========== SEND EMAIL NOTIFICATION FOR APPROVAL WITH BALANCE ==========
        try:
            if customer_email:
                print(f" Attempting to send termination approval email to {customer_email}...")
                send_termination_email(
                    to_email=customer_email,
                    first_name=customer_name.split()[0] if customer_name else "Customer",
                    status="Approved",
                    request_id=request_number,
                    plan=termination_request.get("plan"),
                    termination_reason=termination_request.get("termination_reason"),
                    balance=balance,
                    termination_date=termination_date_display  # ← IPASA ANG DATE
                )
            else:
                print(f" No email address for customer {application_number}")
        except Exception as email_err:
            print(f" Email error: {email_err}")
            import traceback
            traceback.print_exc()
        
        # ========== GET THE CORRECT USER_ID FROM USERS TABLE ==========
        user_query = """
            SELECT user_id, email, first_name, last_name
            FROM users 
            WHERE application_number = %s
            LIMIT 1
        """
        user_info = execute_query(user_query, (application_number,), fetch_one=True)
        
        if user_info:
            actual_user_id = user_info.get("user_id")
            user_email = user_info.get("email")
            user_name = f"{user_info.get('first_name', '')} {user_info.get('last_name', '')}".strip()
        else:
            actual_user_id = application_number
            user_email = termination_request.get("email")
            user_name = customer_name
        
        print(f" Creating notification for user_id: {actual_user_id}")
        
        # ========== CREATE NOTIFICATION FOR USER (APPROVED) WITH BALANCE ==========
        notification_id = int(datetime.now().timestamp() * 1000)
        notif_query = """
            INSERT INTO user_notifications (
                id, title, message, type, relatedId, user_id, user_email, user_name,
                connection_status, timestamp, read_status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        notification_message = f"Your request to terminate your Cablevision subscription has been APPROVED on {termination_date_display}. Your account has been deactivated."
        if balance > 0:
            notification_message += f" Your outstanding balance is ₱{float(balance):,.2f}. Please settle this with our office."
        
        execute_query(notif_query, (
            notification_id,
            "Termination Request Approved",
            notification_message,
            "termination_approved",
            application_number,
            actual_user_id,
            user_email,
            user_name,
            "Disconnected",
            ph_now_iso(),
            0
        ))
        print(f" User notification created for {user_name} - Termination Approved with balance ₱{balance}")
        
        return jsonify({
            "success": True,
            "message": f"Termination request has been approved. Balance: ₱{float(balance):,.2f} has been recorded. Customer status has been set to Terminated on {termination_date_display}."
        })
        
    except Exception as e:
        print(f"Error in approve_termination_request: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# SUPERADMIN - REJECT TERMINATION REQUEST
# ===============================
@app.route("/api/superadmin/reject-termination", methods=["POST"])
def reject_termination_request():
    """Superadmin rejects a termination request"""
    
    data = request.get_json()
    request_id = data.get("request_id")
    reason = data.get("reason", "")
    balance = data.get("balance", 0)
    
    if not request_id:
        return jsonify({"error": "Request ID required"}), 400
    
    try:
        # Get the request details
        req_query = """
            SELECT tr.*, c.email as customer_email, c.first_name, c.last_name,
                   c.contract_number, c.plan
            FROM termination_requests tr
            JOIN customers c ON tr.application_number = c.application_number
            WHERE tr.id = %s AND tr.status = 'Pending'
        """
        termination_request = execute_query(req_query, (request_id,), fetch_one=True)
        
        if not termination_request:
            return jsonify({"error": "Request not found or already processed"}), 404
        
        application_number = termination_request.get("application_number")
        customer_email = termination_request.get("customer_email")
        customer_name = f"{termination_request.get('first_name', '')} {termination_request.get('last_name', '')}".strip()
        request_number = termination_request.get("request_id", f"REQ-{request_id}")
        
        print(f" Balance to save (reject): {balance}")
        
        # UPDATE users table - SAVE BALANCE ONLY (huwag baguhin ang status)
        update_user = """
            UPDATE users 
            SET balance = %s
            WHERE application_number = %s
        """
        execute_query(update_user, (balance, application_number))
        print(f" User balance updated: {balance}")
        
        # UPDATE termination_requests status
        update_request = """
            UPDATE termination_requests 
            SET status = 'Rejected', admin_notes = %s, updated_at = NOW()
            WHERE id = %s
        """
        execute_query(update_request, (reason, request_id))
        
        # Kunin muna ang original admin notification para malaman ang city
        original_notif_query = """
            SELECT admin_city, application_city 
            FROM admin_notifications 
            WHERE relatedId = %s AND type = 'termination_request'
            ORDER BY id DESC LIMIT 1
        """
        original_notif = execute_query(original_notif_query, (application_number,), fetch_one=True)

        # Update admin notification to mark as processed
        update_admin_notif = """
            UPDATE admin_notifications 
            SET action_taken_by = 'superadmin', action_status = 'Rejected', read_status = 1
            WHERE relatedId = %s AND type = 'termination_request'
        """
        execute_query(update_admin_notif, (application_number,))

        # BAGONG NOTIFICATION PARA MALAMAN NG ADMIN NA NA-PROCESS NA
        admin_notif_id = int(datetime.now().timestamp() * 1000)
        admin_notif_query = """
            INSERT INTO admin_notifications (
                id, title, message, type, relatedId, request_id, timestamp,
                read_status, admin_city, application_city, application_id,
                action_taken_by, action_status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        execute_query(admin_notif_query, (
            admin_notif_id,
            "Termination Request Rejected",
            f"[{request_number}] The termination request for {customer_name} has been REJECTED by the superadmin.",
            "termination_processed",
            application_number,
            request_number,
            ph_now_iso(),
            0,
            original_notif.get("admin_city") if original_notif else termination_request.get("city"),
            original_notif.get("application_city") if original_notif else termination_request.get("city"),
            application_number,
            "superadmin",
            "Rejected"
        ))
        print(f" Admin notification created: Termination request {application_number} rejected")
        
        # ========== SEND EMAIL NOTIFICATION FOR REJECTION WITH BALANCE ==========
        try:
            if customer_email:
                print(f" Attempting to send termination rejection email to {customer_email}...")
                send_termination_email(
                    to_email=customer_email,
                    first_name=customer_name.split()[0] if customer_name else "Customer",
                    status="Rejected",
                    request_id=request_number,
                    plan=termination_request.get("plan"),
                    termination_reason=termination_request.get("termination_reason"),
                    rejection_reason=reason,
                    balance=balance
                )
            else:
                print(f" No email address for customer {application_number}")
        except Exception as email_err:
            print(f" Email error: {email_err}")
            import traceback
            traceback.print_exc()
        
        # ========== GET THE CORRECT USER_ID FROM USERS TABLE ==========
        user_query = """
            SELECT user_id, email, first_name, last_name
            FROM users 
            WHERE application_number = %s
            LIMIT 1
        """
        user_info = execute_query(user_query, (application_number,), fetch_one=True)
        
        if user_info:
            actual_user_id = user_info.get("user_id")
            user_email = user_info.get("email")
            user_name = f"{user_info.get('first_name', '')} {user_info.get('last_name', '')}".strip()
        else:
            actual_user_id = application_number
            user_email = termination_request.get("email")
            user_name = customer_name
        
        print(f" Creating notification for user_id: {actual_user_id}")
        
        # ========== CREATE NOTIFICATION FOR USER (REJECTED) WITH BALANCE ==========
        notification_id = int(datetime.now().timestamp() * 1000)
        notif_query = """
            INSERT INTO user_notifications (
                id, title, message, type, relatedId, user_id, user_email, user_name,
                connection_status, timestamp, read_status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        rejection_message = f"Your request to terminate your Cablevision subscription has been REJECTED."
        if reason and reason != "No reason provided" and reason != "":
            rejection_message += f" Reason: {reason}"
        else:
            rejection_message += " Please contact our support team for assistance."
        if balance > 0:
            rejection_message += f" Your outstanding balance is ₱{float(balance):,.2f}."
        
        execute_query(notif_query, (
            notification_id,
            "Termination Request Rejected",
            rejection_message,
            "termination_rejected",
            application_number,
            actual_user_id,
            user_email,
            user_name,
            "Active",
            ph_now_iso(),
            0
        ))
        print(f" User notification created for {user_name} - Termination Rejected with balance ₱{balance}")
        
        return jsonify({
            "success": True,
            "message": f"Termination request has been rejected. Balance: ₱{float(balance):,.2f} has been recorded."
        })
        
    except Exception as e:
        print(f"Error in reject_termination_request: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500



# ===============================
# SEND TERMINATION STATUS EMAIL
# ===============================
def send_termination_email(to_email, first_name, status, request_id, plan, termination_reason=None, rejection_reason=None, balance=0, termination_date=None):
    """Send termination approval/rejection email using Brevo API."""

    import os, html, requests

    api_key = os.getenv("BREVO_API_KEY")
    sender_email = os.getenv("BREVO_SENDER_EMAIL", "cablevision.cableinternet@gmail.com")
    sender_name = os.getenv("BREVO_SENDER_NAME", "Cablevision")
    
    if not api_key:
        print(" BREVO_API_KEY is not configured.")
        return False

    def escape_html(text):
        return html.escape(str(text)) if text else ""

    try:
        balance = float(balance or 0)
    except (TypeError, ValueError):
        balance = 0

    approved = status == "Approved"
    status_color = "#10b981" if approved else "#ef4444"
    status_bg = "#ecfdf5" if approved else "#fef2f2"
    status_icon = "" if approved else ""
    action_text = "APPROVED" if approved else "REJECTED"

    balance_html = ""
    if balance > 0:
        balance_html = f"""
        <div style="margin-top:16px;padding:16px 20px;background:#fef3c7;border-radius:12px;border-left:4px solid #f59e0b;">
            <div style="font-size:13px;font-weight:600;color:#92400e;margin-bottom:4px;"> Outstanding Balance</div>
            <div style="font-size:24px;font-weight:700;color:#dc2626;">₱{balance:,.2f}</div>
            <div style="font-size:12px;color:#78350f;margin-top:4px;">Please settle this balance with our office.</div>
        </div>
        """

    termination_date_html = ""
    if approved and termination_date:
        termination_date_html = f"""
        <div style="margin-top:12px;padding:12px 16px;background:#e0f2fe;border-radius:10px;border-left:4px solid #0284c7;">
            <div style="font-size:13px;font-weight:600;color:#0369a1;"> Date Terminated</div>
            <div style="font-size:16px;font-weight:600;color:#0c4a6e;margin-top:2px;">{escape_html(termination_date)}</div>
        </div>
        """

    if approved:
        message_sub = "Your termination request has been approved."
        extra_message = f"""
        <div style="margin-top:20px;padding:16px;background:#f0fdf4;border-radius:12px;">
            <p style="margin:0 0 8px;color:#166534;"><strong>What's Next?</strong></p>
            <p style="margin:0;color:#14532d;font-size:14px;">
                Your Cablevision subscription has been terminated on
                <strong>{escape_html(termination_date or 'today')}</strong>.
                Your account has been deactivated.
                We appreciate your trust in Cablevision and hope to serve you again in the future.
            </p>
        </div>
        """
    else:
        message_sub = "We have an update regarding your termination request."
        extra_message = f"""
        <div style="margin-top:20px;padding:16px;background:#fef2f2;border-radius:12px;">
            <p style="margin:0 0 8px;color:#991b1b;"><strong>Reason for Rejection</strong></p>
            <p style="margin:0;color:#7f1d1d;font-size:14px;">
                {escape_html(rejection_reason) if rejection_reason else 'No specific reason provided.'}
            </p>
            <p style="margin-top:12px;color:#7f1d1d;font-size:13px;">
                Your subscription remains active. If you still wish to terminate,
                please contact our support team for assistance.
            </p>
        </div>
        """

    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Cablevision - Termination Request Update</title>
    </head>
    <body style="margin:0;padding:0;font-family:'Segoe UI','Inter',Arial,sans-serif;background:#eef2ff;">
        <div style="max-width:580px;margin:0 auto;padding:30px 20px;">
            <div style="background:#fff;border-radius:32px;overflow:hidden;box-shadow:0 20px 35px -12px rgba(0,0,0,.15);">

                <div style="background:linear-gradient(135deg,#001f3f 0%,#002b5c 100%);padding:32px 28px;text-align:center;">
                    <h1 style="margin:0;font-size:26px;font-weight:700;color:#fff;">Cablevision</h1>
                    <p style="margin:6px 0 0;color:#93c5fd;font-size:13px;">Internet Service Provider</p>
                </div>

                <div style="padding:20px 28px 0;text-align:center;">
                    <div style="display:inline-block;background:{status_bg};padding:8px 24px;border-radius:60px;">
                        <span style="font-size:14px;font-weight:600;color:{status_color};">{status_icon} REQUEST {action_text}</span>
                    </div>
                </div>

                <div style="padding:20px 28px 32px;">
                    <h2 style="margin:0 0 8px;font-size:22px;font-weight:700;color:#0f172a;">
                        Hello, {escape_html(first_name)}!
                    </h2>
                    <p style="margin:0 0 20px;font-size:15px;color:#475569;">{message_sub}</p>

                    <div style="background:#f8fafc;border-radius:20px;padding:18px;margin-bottom:16px;">
                        <div style="margin-bottom:16px;padding-bottom:12px;border-bottom:1px solid #e2e8f0;">
                            <div style="font-size:11px;font-weight:600;color:#64748b;">Request ID</div>
                            <div style="font-size:18px;font-weight:700;color:#0f172a;font-family:monospace;">{escape_html(request_id)}</div>
                        </div>
                        <div>
                            <div style="font-size:11px;font-weight:600;color:#64748b;">Current Plan</div>
                            <div style="font-size:16px;font-weight:700;color:#0f172a;">{escape_html(plan or 'N/A')}</div>
                        </div>
                        {f'''
                        <div style="margin-top:12px;padding-top:12px;border-top:1px solid #e2e8f0;">
                            <div style="font-size:11px;font-weight:600;color:#64748b;">Reason for Termination</div>
                            <div style="font-size:14px;color:#0f172a;">{escape_html(termination_reason)}</div>
                        </div>
                        ''' if termination_reason else ''}
                    </div>

                    {termination_date_html}
                    {balance_html}
                    {extra_message}

                    <div style="margin-top:28px;padding-top:20px;text-align:center;border-top:1px solid #e2e8f0;">
                        <p style="margin:0;font-size:12px;color:#94a3b8;">Thank you for being with Cablevision!</p>
                    </div>
                </div>

                <div style="background:#f1f5f9;padding:16px 28px;text-align:center;">
                    <div style="font-size:11px;color:#64748b;"> 2026 Cablevision Internet Service Provider. All rights reserved.</div>
                </div>

            </div>
        </div>
    </body>
    </html>
    """

    payload = {
        "sender": {"name": sender_name, "email": sender_email},
        "to": [{"email": to_email, "name": first_name or "Customer"}],
        "subject": "Cablevision - Termination Request Update",
        "htmlContent": html_body
    }

    try:
        print(f" Sending termination email to {to_email} via Brevo API...")

        response = requests.post(
            "https://api.brevo.com/v3/smtp/email",
            headers={
                "accept": "application/json",
                "api-key": api_key,
                "content-type": "application/json"
            },
            json=payload,
            timeout=30
        )

        if response.ok:
            result = response.json()
            print(f" Termination email sent to {to_email}")
            print(f" Brevo Message ID: {result.get('messageId')}")
            return True

        print(f" Brevo email failed: {response.status_code} - {response.text}")
        return False

    except requests.RequestException as e:
        print(f" Brevo connection error: {e}")
        return False
    except Exception as e:
        print(f" Termination email failed: {e}")
        import traceback
        traceback.print_exc()
        return False



@app.route("/api/admin/session-user", methods=["GET"])
def get_session_admin_user():
    """Get the current admin username from session using tab_id"""
    tab_id = request.args.get("tab_id")
    
    if not tab_id:
        return jsonify({"error": "No tab_id provided"}), 401
    
    # Check if session exists with this tab_id
    user_session = session.get(f"admin_{tab_id}")
    
    if user_session:
        return jsonify({
            "username": user_session.get("admin_username") or user_session.get("user_name"),
            "area": user_session.get("user_area", ""),
            "user_type": user_session.get("user_type", "admin")
        })
    else:
        return jsonify({"error": "Invalid session"}), 401
    
# =============================== ADMIN DASHBOARD FUNCTIONALITY ===============================
@app.route("/admin")
def admin_dashboard():
    """Render the admin dashboard with dynamic welcome name"""
    
    # Kunin ang username/admin_id mula sa session o query param
    username = request.args.get("username") or session.get("admin_username") or session.get("adminUsername") or "admin"
    
    # Kunin ang admin data mula sa MySQL
    query = "SELECT * FROM admins WHERE username = %s OR email = %s OR admin_id = %s"
    admin_data = execute_query(query, (username, username, username), fetch_one=True)
    
    # Gamitin ang username kung meron, otherwise default
    name = admin_data.get('username') if admin_data else username or "Admin"
    
    return render_template("admin-dashboard.html", name=name)

# ===============================
# SERVE ADMIN PROFILE PAGE
# ===============================
@app.route("/admin/profile")
def admin_profile_page():
    tab_id = request.args.get('tab_id') or session.get('active_tab')
    session_data = session.get(f"admin_{tab_id}") if tab_id else None

    if not session_data or session_data.get('user_type') != 'admin':
        return render_template("admin-profile.html", ga_enabled=False, ga_secret=None, ga_setup_uri=None)

    username = session_data.get('admin_username') or session_data.get('username')
    admin_row = execute_query(
        "SELECT ga_enabled, ga_secret FROM admins WHERE username = %s OR admin_id = %s LIMIT 1",
        (username, username),
        fetch_one=True,
    )

    ga_enabled = bool(admin_row.get('ga_enabled')) if admin_row else False
    ga_secret = admin_row.get('ga_secret') if admin_row else None
    ga_setup_uri = None

    if not ga_enabled:
        if not ga_secret:
            ga_secret = generate_ga_secret()
            execute_query("UPDATE admins SET ga_secret = %s WHERE username = %s OR admin_id = %s", (ga_secret, username, username))
        ga_setup_uri = generate_ga_provisioning_uri(username, ga_secret)

    return render_template("admin-profile.html", ga_enabled=ga_enabled, ga_secret=ga_secret, ga_setup_uri=ga_setup_uri)


# =========================================================
# LOGIN HISTORY PAGE & API ROUTES (ADMIN, SUPERADMIN, TECHNICIAN)
# =========================================================
@app.route("/admin/login-history")
def admin_login_history_page():
    return render_template("admin-login-history.html")


@app.route("/superadmin/login-history")
def superadmin_login_history_page():
    return render_template("superadmin-login-history.html")


@app.route("/technician/login-history")
def technician_login_history_page():
    return render_template("technician-login-history.html")


@app.route("/api/admin/login-history", methods=["GET"])
def get_admin_login_history_api():
    try:
        user_id, user_type, user_name, current_token = get_admin_session_user(request)

        if not user_id:
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        ensure_login_history_table()

        if current_token:
            lat = request.args.get("lat")
            lng = request.args.get("lng")
            record_login_history(user_id, user_type, current_token, lat=lat, lng=lng)

        rows = execute_query(
            "SELECT id, user_id, user_type, session_token, device_info, device_brand, browser, os, ip_address, location, "
            "DATE_FORMAT(login_time, '%b %d, %Y %h:%i %p') as formatted_login_time, "
            "DATE_FORMAT(last_active, '%b %d, %Y %h:%i %p') as formatted_last_active, "
            "status FROM login_history WHERE user_id = %s AND user_type = %s ORDER BY id DESC",
            (user_id, user_type),
            fetch=True
        ) or []

        current_device = None
        other_devices = []

        for row in rows:
            is_curr = (row.get("session_token") == current_token) or (current_token is None and current_device is None)
            row["is_current"] = is_curr
            if is_curr and not current_device:
                current_device = row
            else:
                other_devices.append(row)

        if not current_device and rows:
            current_device = rows[0]
            current_device["is_current"] = True
            other_devices = [r for r in rows if r["id"] != current_device["id"]]

        return jsonify({
            "success": True,
            "current_device": current_device,
            "other_devices": other_devices,
            "all_history": rows
        })
    except Exception as e:
        print(f"[API LOGIN HISTORY ERROR] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/admin/login-history/logout", methods=["POST"])
def logout_admin_device():
    try:
        data = request.get_json() or {}
        device_ids = data.get("device_ids", [])
        user_id, user_type, user_name, current_token = get_admin_session_user(request)

        if isinstance(device_ids, (int, str)):
            device_ids = [device_ids]

        if not user_id:
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        if not device_ids:
            return jsonify({"success": False, "error": "No devices specified"}), 400

        current_rec = None
        if current_token:
            current_rec = execute_query(
                "SELECT id FROM login_history WHERE user_id = %s AND user_type = %s AND session_token = %s LIMIT 1",
                (user_id, user_type, current_token),
                fetch_one=True
            )
        current_id = current_rec.get("id") if current_rec else None

        logout_current = any(str(did) == str(current_id) for did in device_ids)

        format_strings = ','.join(['%s'] * len(device_ids))
        query_params = [user_id, user_type] + device_ids
        execute_query(
            f"DELETE FROM login_history WHERE user_id = %s AND user_type = %s AND id IN ({format_strings})",
            query_params
        )

        tab_id = request.args.get("tab_id") or data.get("tab_id") or current_token
        if logout_current:
            if tab_id and f"admin_{tab_id}" in session:
                session.pop(f"admin_{tab_id}", None)
            else:
                for key in list(session.keys()):
                    if isinstance(key, str) and key.startswith("admin_"):
                        adm_sess = session.get(key)
                        if isinstance(adm_sess, dict) and adm_sess.get("user_id") == user_id:
                            session.pop(key, None)

            for key in ["user_id", "admin_id", "technician_id", "username", "admin_username", "technician_name", "user_type", "user_name", "user_area", "active_tab"]:
                session.pop(key, None)

        return jsonify({
            "success": True,
            "message": "Device(s) logged out and record deleted successfully",
            "logout_current": logout_current
        })
    except Exception as e:
        print(f"[LOGOUT DEVICE ERROR] {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/admin/login-history/logout-all", methods=["POST"])
def logout_all_admin_devices():
    try:
        data = request.get_json() or {}
        include_current = data.get("include_current", False)
        user_id, user_type, user_name, current_token = get_admin_session_user(request)

        if not user_id:
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        tab_id = request.args.get("tab_id") or data.get("tab_id") or current_token

        if include_current:
            execute_query("DELETE FROM login_history WHERE user_id = %s AND user_type = %s", (user_id, user_type))
            for key in list(session.keys()):
                if isinstance(key, str) and key.startswith("admin_"):
                    adm_sess = session.get(key)
                    if isinstance(adm_sess, dict) and adm_sess.get("user_id") == user_id:
                        session.pop(key, None)
            for key in ["user_id", "admin_id", "technician_id", "username", "admin_username", "technician_name", "user_type", "user_name", "user_area", "active_tab"]:
                session.pop(key, None)
            logout_current = True
        else:
            if current_token:
                execute_query("DELETE FROM login_history WHERE user_id = %s AND user_type = %s AND session_token != %s", (user_id, user_type, current_token))
            else:
                execute_query("DELETE FROM login_history WHERE user_id = %s AND user_type = %s", (user_id, user_type))
            logout_current = False

        return jsonify({
            "success": True,
            "message": "All other devices logged out and records deleted",
            "logout_current": logout_current
        })
    except Exception as e:
        print(f"[LOGOUT ALL DEVICES ERROR] {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# ===============================
# CHECK FOR NEW DEVICE LOGINS (Real-time Alert)
# ===============================
@app.route("/api/check-new-devices", methods=["GET"])
def check_new_devices():
    try:
        user_id, user_type, user_name, current_token = get_admin_session_user(request)

        if not user_id:
            print("[CHECK NEW DEVICES] Unauthorized - no user_id")
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        ensure_login_history_table()

        print(f"[CHECK NEW DEVICES] Checking for user: {user_id} ({user_type}), current_token: {current_token}")

        # Get the last check time from the request (in seconds since epoch)
        last_check_time = request.args.get("last_check", None)

        if last_check_time:
            try:
                from datetime import datetime as dt
                last_check_dt = dt.fromtimestamp(float(last_check_time))
                print(f"[CHECK NEW DEVICES] Checking devices since: {last_check_dt}")
                query = """
                    SELECT id, device_info, device_brand, browser, os, ip_address, location,
                           DATE_FORMAT(login_time, '%b %d, %Y %h:%i %p') as formatted_login_time,
                           session_token, login_time
                    FROM login_history 
                    WHERE user_id = %s AND user_type = %s AND session_token != %s AND login_time > %s
                    ORDER BY login_time DESC
                    LIMIT 1
                """
                new_devices = execute_query(query, (user_id, user_type, current_token, last_check_dt), fetch=True) or []
                print(f"[CHECK NEW DEVICES] Found {len(new_devices)} recent devices")
            except Exception as e:
                print(f"[CHECK NEW DEVICES] Timestamp parsing failed: {e}")
                new_devices = []
        else:
            # No last_check means the page has not loaded or the client is stale.
            # Do not show historical login activity as a fresh security alert.
            print("[CHECK NEW DEVICES] No last_check provided; ignoring historical logins")
            new_devices = []

        return jsonify({
            "success": True,
            "new_devices": new_devices,
            "current_timestamp": datetime.now().timestamp()
        })
    except Exception as e:
        print(f"[CHECK NEW DEVICES ERROR] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# ===============================
# DEBUG: Get Login History for Current User
# ===============================
@app.route("/api/debug/login-history", methods=["GET"])
def debug_login_history():
    """Debug endpoint to see all login history for current user"""
    try:
        user_id, user_type, user_name, current_token = get_admin_session_user(request)

        if not user_id:
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        ensure_login_history_table()
        
        query = """
            SELECT id, device_info, device_brand, browser, os, ip_address, location,
                   session_token, login_time, last_active, status
            FROM login_history 
            WHERE user_id = %s AND user_type = %s
            ORDER BY login_time DESC
            LIMIT 20
        """
        all_devices = execute_query(query, (user_id, user_type), fetch=True) or []
        
        print(f"\n[DEBUG] Login history for {user_id} ({user_type}):")
        print(f"  Current token: {current_token}")
        print(f"  Total devices: {len(all_devices)}")
        for device in all_devices:
            is_current = " CURRENT" if device.get('session_token') == current_token else ""
            print(f"    - {device.get('browser')} | {device.get('device_brand')} | Token: {device.get('session_token')[:20]}... {is_current}")
        
        return jsonify({
            "success": True,
            "current_token": current_token,
            "user_id": user_id,
            "user_type": user_type,
            "devices": all_devices
        })
    except Exception as e:
        print(f"[DEBUG ERROR] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# ===============================
# GET ADMIN PROFILE (XAMPP/MYSQL) - FIXED
# ===============================
@app.route("/api/admin/profile", methods=["GET"])
def get_admin_profile():
    try:
        identifier = request.args.get("username")
        tab_id = request.args.get("tab_id", "")

        print(f" GET ADMIN PROFILE - identifier: {identifier}, tab_id: {tab_id}")

        if not identifier:
            return jsonify({"error": "Username is required"}), 400

        if tab_id:
            user_session = session.get(f"admin_{tab_id}")
            if user_session:
                session_admin = user_session.get("admin_username") or user_session.get("user_name")
                if session_admin and session_admin != identifier:
                    print(f" Session mismatch: {session_admin} vs {identifier}")
                    identifier = session_admin

        query = """
            SELECT username, admin_id, email, mobile, area, status, profile_photo, ga_enabled, ga_secret
            FROM admins
            WHERE username = %s OR admin_id = %s OR email = %s
        """
        admin_data = execute_query(query, (identifier, identifier, identifier), fetch_one=True)

        if not admin_data:
            print(f" Admin not found: {identifier}")
            return jsonify({"error": f"Admin '{identifier}' not found"}), 404

        admin_id = admin_data.get('admin_id') or admin_data.get('username')
        area = admin_data.get('area') or "Not assigned"
        display_name = admin_data.get('username') or admin_id
        ga_enabled = bool(admin_data.get('ga_enabled', 0))
        ga_secret = admin_data.get('ga_secret')

        if not ga_enabled and not ga_secret:
            ga_secret = generate_ga_secret()
            execute_query("UPDATE admins SET ga_secret = %s WHERE username = %s OR admin_id = %s", (ga_secret, display_name, admin_id))

        ga_setup_uri = None if ga_enabled else generate_ga_provisioning_uri(display_name, ga_secret)

        profile = {
            "username": admin_data.get('username'),
            "id": admin_id,
            "name": display_name,
            "email": admin_data.get('email', ''),
            "contact": admin_data.get('mobile', ''),
            "area": area,
            "city": area,
            "status": admin_data.get('status', 'Active'),
            "ga_enabled": ga_enabled,
            "ga_secret": ga_secret,
            "ga_setup_uri": ga_setup_uri
        }

        print(f" Profile found: {profile}")

        return jsonify(profile), 200

    except Exception as e:
        print(f"Get admin profile error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Failed to fetch admin profile", "details": str(e)}), 500


@app.route("/admin/ga/enable", methods=["POST"])
def admin_enable_google_auth():
    tab_id = request.form.get('tab_id') or session.get('active_tab')
    session_data = session.get(f"admin_{tab_id}") if tab_id else None

    if not session_data or session_data.get('user_type') != 'admin':
        flash("Unauthorized access.", "danger")
        return redirect(url_for("admin_profile_page"))

    username = session_data.get('admin_username') or session_data.get('username')
    code = request.form.get("ga_code", "").strip()

    admin_row = execute_query(
        "SELECT ga_secret, ga_enabled FROM admins WHERE username = %s OR admin_id = %s LIMIT 1",
        (username, username),
        fetch_one=True,
    )

    if not admin_row:
        flash("Admin account not found.", "danger")
        return redirect(url_for("admin_profile_page"))

    secret = admin_row.get("ga_secret")
    if not secret:
        secret = generate_ga_secret()
        execute_query("UPDATE admins SET ga_secret = %s WHERE username = %s OR admin_id = %s", (secret, username, username))

    if not code:
        flash("Please enter the 6-digit code from Google Authenticator.", "danger")
        return redirect(url_for("admin_profile_page", toast="ga-missing"))

    if verify_ga_code(secret, code):
        execute_query("UPDATE admins SET ga_secret = %s, ga_enabled = 1 WHERE username = %s OR admin_id = %s", (secret, username, username))
        session_data['ga_enabled'] = True
        session[f"admin_{tab_id}"] = session_data
        flash(" Google Authenticator is now enabled!", "success")
        return redirect(url_for("admin_profile_page", toast="ga-enabled"))
    else:
        flash(" Invalid code. Please try again.", "danger")
        return redirect(url_for("admin_profile_page", toast="ga-invalid"))


@app.route("/admin/ga/disable", methods=["POST"])
def admin_disable_google_auth():
    tab_id = request.form.get('tab_id') or session.get('active_tab')
    session_data = session.get(f"admin_{tab_id}") if tab_id else None

    if not session_data or session_data.get('user_type') != 'admin':
        flash("Unauthorized access.", "danger")
        return redirect(url_for("admin_profile_page"))

    username = session_data.get('admin_username') or session_data.get('username')
    execute_query("UPDATE admins SET ga_secret = NULL, ga_enabled = 0 WHERE username = %s OR admin_id = %s", (username, username))
    session_data['ga_enabled'] = False
    session[f"admin_{tab_id}"] = session_data
    flash("Google Authenticator has been disabled.", "info")
    return redirect(url_for("admin_profile_page", toast="ga-disabled"))
    
    
# ===============================
# DELETE ADMIN PROFILE PHOTO (XAMPP/MYSQL)
# ===============================
@app.route("/api/admin/delete-profile-photo", methods=["POST"])
def delete_profile_photo():
    try:
        data = request.json
        username = data.get("username")
        
        if not username:
            return jsonify({"error": "Username required"}), 400

        # Check if admin exists
        check_query = "SELECT username FROM admins WHERE username = %s OR admin_id = %s"
        admin_exists = execute_query(check_query, (username, username), fetch_one=True)
        
        if not admin_exists:
            return jsonify({"error": "Admin not found"}), 404

        # Update profile_photo to default
        update_query = "UPDATE admins SET profile_photo = %s WHERE username = %s OR admin_id = %s"
        execute_query(update_query, ("/static/profile.jpg", username, username))
        
        return jsonify({"message": "Profile photo removed successfully"}), 200
        
    except Exception as e:
        print(f"Delete profile photo error: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# UPDATE ADMIN PROFILE (XAMPP/MYSQL)
# ===============================
@app.route("/api/update-admin-profile", methods=["POST"])
def update_admin_profile():
    try:
        data = request.get_json()
        username = data.get("username")
        email = data.get("email")
        contact = data.get("contact")
        password = data.get("password")
        current_password = data.get("current_password")
        tab_id = data.get("tab_id", "")

        if not username:
            return jsonify({"error": "Username required"}), 400
        
        # I-VERIFY MUNA ANG SESSION
        session_data = session.get(f"admin_{tab_id}")
        if not session_data or session_data.get('user_type') != 'admin':
            return jsonify({"error": "Unauthorized"}), 403

        # KUNG MAY BINAGO SA EMAIL, I-CHECK SA LAHAT NG TABLES
        if email:
            current_query = "SELECT email FROM admins WHERE username = %s OR admin_id = %s"
            current_admin = execute_query(current_query, (username, username), fetch_one=True)
            
            if current_admin and current_admin.get('email') != email:
                check_query = """
                    SELECT 
                        (SELECT COUNT(*) FROM technicians WHERE email = %s) as tech_count,
                        (SELECT COUNT(*) FROM admins WHERE email = %s AND username != %s AND admin_id != %s) as admin_count,
                        (SELECT COUNT(*) FROM superadmins WHERE email = %s) as superadmin_count
                """
                result = execute_query(check_query, (email, email, username, username, email), fetch_one=True)
                
                tech_exists = result.get('tech_count', 0) > 0 if result else False
                admin_exists = result.get('admin_count', 0) > 0 if result else False
                superadmin_exists = result.get('superadmin_count', 0) > 0 if result else False

                if tech_exists or admin_exists or superadmin_exists:
                    return jsonify({"error": f"Email '{email}' already exists"}), 400

        if password and len(password) >= 8:
            current_user = execute_query("SELECT password FROM admins WHERE username = %s OR admin_id = %s LIMIT 1", (username, username), fetch_one=True)
            if not current_user:
                return jsonify({"error": "Account not found"}), 404
            if not current_password or not verify_password(current_user.get('password'), current_password):
                return jsonify({"error": "Current password is incorrect"}), 400

        # Build update query
        update_fields = []
        params = []
        
        if contact is not None:
            clean_contact = contact.replace(" ", "")
            update_fields.append("mobile = %s")  # TAMA
            params.append(clean_contact)
        if email:
            update_fields.append("email = %s")
            params.append(email)
        if password and len(password) >= 8:
            update_fields.append("password = %s")
            params.append(hash_password(password))
        
        if not update_fields:
            return jsonify({"error": "No fields to update"}), 400
        
        params.append(username)
        params.append(username)
        
        update_query = f"UPDATE admins SET {', '.join(update_fields)} WHERE username = %s OR admin_id = %s"
        execute_query(update_query, params)
        
        print(f" Admin {username} profile updated")
        
        return jsonify({
            "success": True,
            "message": "Profile updated successfully"
        }), 200

    except Exception as e:
        print(f"Update admin profile error: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# CHECK EMAIL AVAILABILITY FOR ADMIN
# ===============================
@app.route("/api/admin/check-email", methods=["GET"])
def check_admin_email():
    """Check if email exists in technicians, admins, or superadmins tables"""
    email = request.args.get("email")
    tab_id = request.args.get("tab_id", "")
    
    if not email:
        return jsonify({"exists": False, "error": "Email required"}), 400
    
    # Get current admin username from session
    session_data = session.get(f"admin_{tab_id}")
    username = session_data.get('admin_username') if session_data else None
    
    # Check all tables
    check_query = """
        SELECT 
            (SELECT COUNT(*) FROM technicians WHERE email = %s) as tech_count,
            (SELECT COUNT(*) FROM admins WHERE email = %s) as admin_count,
            (SELECT COUNT(*) FROM superadmins WHERE email = %s) as superadmin_count
    """
    result = execute_query(check_query, (email, email, email), fetch_one=True)
    
    tech_exists = result.get('tech_count', 0) > 0 if result else False
    admin_exists = result.get('admin_count', 0) > 0 if result else False
    superadmin_exists = result.get('superadmin_count', 0) > 0 if result else False
    
    # EXCLUDE THE CURRENT ADMIN (para hindi mag-error sa sarili niyang email)
    if admin_exists and username:
        check_self_query = "SELECT email FROM admins WHERE username = %s AND email = %s"
        self_email = execute_query(check_self_query, (username, email), fetch_one=True)
        if self_email:
            admin_exists = False
    
    exists = tech_exists or admin_exists or superadmin_exists
    
    return jsonify({
        "exists": exists
    })
  


# ===============================
# ADMIN STATISTICS
# ===============================
@app.route("/api/admin/statistics", methods=["GET"])
def admin_statistics():
    try:
        username = request.args.get("username") or session.get("admin_username") or session.get("adminUsername")
        if not username:
            return jsonify({"error": "Username required"}), 400

        # ========== GET ADMIN INFO FROM MYSQL ==========
        admin_query = "SELECT area FROM admins WHERE username = %s OR admin_id = %s"
        admin_data = execute_query(admin_query, (username, username), fetch_one=True)

        if not admin_data:
            return jsonify({"error": "Admin not found"}), 404

        admin_area = str(admin_data.get("area", "")).strip().lower()

        # ========== GET APPLICATIONS FROM MYSQL ==========
        apps_query = """
            SELECT city, plan FROM applications 
            WHERE status != 'Rejected'
        """
        all_apps = execute_query(apps_query, fetch=True) or []

        total_applicants = 0
        popular_plans = {}

        for app in all_apps:
            app_city = str(app.get("city", "")).strip().lower()

            # Check if application belongs to admin's area
            if admin_area in app_city or app_city in admin_area:
                total_applicants += 1

                plan = app.get("plan", "Unknown")
                popular_plans[plan] = popular_plans.get(plan, 0) + 1

        return jsonify({
            "total_applicants": total_applicants,
            "popular_plans": popular_plans
        })

    except Exception as e:
        print(f"Admin statistics error: {e}")
        return jsonify({"error": str(e)}), 500

# ===============================
# Admin Internet Applications Page
# ===============================
@app.route("/admin/internet-applications")
def admin_internet_applications():
    return render_template("admin-internet-applications.html")

# ===============================
# GET ADMIN INTERNET APPLICATIONS - CONVERTED TO MYSQL
# ===============================
@app.route("/api/admin/internet-applications", methods=["GET"])
def get_admin_internet_applications():
    try:
        username = request.args.get("username") or session.get("admin_username") or session.get("adminUsername")
        tab_id = request.args.get("tab_id", "")
        
        if not username:
            return jsonify({"error": "Username required"}), 400
        
        if tab_id:
            user_session = session.get(f"admin_{tab_id}")
            if user_session:
                session_username = user_session.get("admin_username") or user_session.get("user_name")
                if session_username and session_username != username:
                    print(f" Session mismatch: {session_username} vs {username}")
                    username = session_username

        admin_query = "SELECT area FROM admins WHERE username = %s OR admin_id = %s"
        admin_data = execute_query(admin_query, (username, username), fetch_one=True)

        if not admin_data:
            return jsonify({"error": "Admin not found"}), 404

        admin_area = str(admin_data.get("area", "")).strip().lower()

        # IDAGDAG ANG is_archived = 0 SA WHERE CLAUSE
        apps_query = """
            SELECT application_number, first_name, last_name, email, 
                   date_submitted, time_submitted, barangay, city, birthdate, plan, 
                   status, rejection_reason, is_archived
            FROM applications
            WHERE is_archived = 0
            ORDER BY timestamp DESC
        """
        all_apps = execute_query(apps_query, fetch=True) or []

        filtered_apps = []
        for app in all_apps:
            app_city = str(app.get("city", "")).strip().lower()
            
            datetime_submitted = None
            if app.get("date_submitted") and app.get("time_submitted"):
                datetime_submitted = f"{app.get('date_submitted')} {app.get('time_submitted')}"
            elif app.get("date_submitted"):
                datetime_submitted = app.get("date_submitted")
            
            if admin_area in app_city or app_city in admin_area:
                filtered_apps.append({
                    "id": app.get("application_number"),
                    "application_number": app.get("application_number", ""),
                    "first_name": app.get("first_name", ""),
                    "last_name": app.get("last_name", ""),
                    "email": app.get("email", ""),
                    "date_submitted": datetime_submitted,
                    "barangay": app.get("barangay", ""),
                    "city": app.get("city", ""),
                    "birthdate": app.get("birthdate", ""),
                    "plan": app.get("plan", ""),
                    "status": app.get("status", "Pending"),
                    "rejection_reason": app.get("rejection_reason", ""),
                    "is_archived": app.get("is_archived", 0)
                })

        return jsonify(filtered_apps), 200
        
    except Exception as e:
        print(f"Error in get_admin_internet_applications: {e}")
        return jsonify({"error": str(e)}), 500
    


# ===============================
# ADMIN REQUEST (NOT DIRECT APPROVE)
# MYSQL / RAILWAY READY
# ===============================
@app.route("/api/admin/application/<app_id>/request", methods=["POST"])
def admin_request_application(app_id):
    conn = None
    cursor = None

    try:
        data = request.get_json() or {}
        status = data.get("status")
        reason = data.get("reason")

        # ===============================
        # GET ADMIN USERNAME
        # ===============================
        username = (
            request.args.get("username")
            or session.get("admin_username")
            or session.get("adminUsername")
        )

        if not username:
            username = request.headers.get("X-Admin-Username")

        if not username:
            username = data.get("username")

        # ===============================
        # VALIDATE STATUS
        # ===============================
        if status not in ["Approved", "Rejected", "Pending", "Reapply"]:
            return jsonify({
                "error": "Invalid status"
            }), 400

        if not username:
            return jsonify({
                "error": "Admin username required"
            }), 400

        print("=" * 60)
        print(" ADMIN REQUEST STARTED")
        print(f" Application ID: {app_id}")
        print(f" Requested Status: {status}")
        print(f" Username: {username}")

        # ===============================
        # GET APPLICATION
        # ===============================
        app_query = """
            SELECT *
            FROM applications
            WHERE application_number = %s
        """

        app_data = execute_query(
            app_query,
            (app_id,),
            fetch_one=True
        )

        if not app_data:
            return jsonify({
                "error": "Application not found"
            }), 404

        # ===============================
        # VALIDATIONS
        # ===============================

        # Restore request
        if status == "Pending" and app_data.get("status") != "Rejected":
            return jsonify({
                "error": "Only rejected applications can request restore"
            }), 400

        # Reapply request
        if status == "Reapply" and app_data.get("status") != "Rejected":
            return jsonify({
                "error": "Only rejected applications can request reapply"
            }), 400

        # Check direct reapply flag
        if status == "Reapply" and app_data.get("reapply_requested"):
            return jsonify({
                "error": "A reapply request has already been sent to the customer"
            }), 400

        # ===============================
        # CHECK EXISTING PENDING REQUEST
        # ===============================
        pending_query = """
            SELECT request_id
            FROM approval_requests
            WHERE app_id = %s
              AND status = 'Pending'
            LIMIT 1
        """

        existing_request = execute_query(
            pending_query,
            (app_id,),
            fetch_one=True
        )

        if existing_request:
            return jsonify({
                "error": "Request already sent"
            }), 400

        # ===============================
        # GET ADMIN INFORMATION
        # ===============================
        print(f" Searching for admin: '{username}'")

        clean_username = (
            str(username)
            .strip()
            .lower()
            .replace(" ", "")
        )

        print(f" Cleaned username: '{clean_username}'")

        # --------------------------------
        # Strategy 1
        # Exact username ignoring spaces
        # --------------------------------
        admin_query = """
            SELECT admin_id, username, area
            FROM admins
            WHERE LOWER(REPLACE(username, ' ', '')) = %s
            LIMIT 1
        """

        admin_info = execute_query(
            admin_query,
            (clean_username,),
            fetch_one=True
        )

        # --------------------------------
        # Strategy 2
        # Match admin_id
        # --------------------------------
        if not admin_info:

            admin_query = """
                SELECT admin_id, username, area
                FROM admins
                WHERE LOWER(admin_id) = %s
                   OR LOWER(username) = %s
                LIMIT 1
            """

            admin_info = execute_query(
                admin_query,
                (clean_username, clean_username),
                fetch_one=True
            )

        # --------------------------------
        # Strategy 3
        # Partial username match
        # --------------------------------
        if not admin_info:

            admin_query = """
                SELECT admin_id, username, area
                FROM admins
                WHERE LOWER(REPLACE(username, ' ', '')) LIKE %s
                LIMIT 1
            """

            admin_info = execute_query(
                admin_query,
                (f"%{clean_username}%",),
                fetch_one=True
            )

        # --------------------------------
        # Strategy 4
        # Superadmin fallback
        # --------------------------------
        if not admin_info:

            if "super" in clean_username and "admin" in clean_username:

                admin_query = """
                    SELECT admin_id, username, area
                    FROM admins
                    WHERE LOWER(username) LIKE '%super%admin%'
                    LIMIT 1
                """

                admin_info = execute_query(
                    admin_query,
                    fetch_one=True
                )

        # --------------------------------
        # Strategy 5
        # First active admin fallback
        # --------------------------------
        if not admin_info:

            admin_query = """
                SELECT admin_id, username, area
                FROM admins
                WHERE status = 'Active'
                LIMIT 1
            """

            admin_info = execute_query(
                admin_query,
                fetch_one=True
            )

            if admin_info:
                print(
                    f" Using fallback admin: "
                    f"{admin_info.get('admin_id')}"
                )

        print(f" Admin query result: {admin_info}")

        # ===============================
        # ADMIN NOT FOUND
        # ===============================
        if not admin_info:

            all_admins_query = """
                SELECT admin_id, username
                FROM admins
                ORDER BY admin_id
            """

            all_admins = execute_query(
                all_admins_query,
                fetch=True
            ) or []

            admin_list = [
                f"{a.get('username')} ({a.get('admin_id')})"
                for a in all_admins
            ]

            return jsonify({
                "error": f"Admin not found for username: {username}",
                "available_admins": admin_list,
                "hint": "Try using your admin_id (e.g., ACV-0001) or your exact username"
            }), 404

        # ===============================
        # ADMIN DATA
        # ===============================
        admin_id = admin_info.get("admin_id")
        admin_area = admin_info.get("area")
        admin_username = admin_info.get("username")

        # Since previous structure uses area as city
        admin_city = admin_area

        print(
            f" Admin found - "
            f"ID: {admin_id}, "
            f"Username: {admin_username}, "
            f"Area: {admin_area}"
        )

        # ===============================
        # GENERATE UNIQUE REQUEST ID
        # ===============================
        import random
        import string

        request_id = ''.join(
            random.choices(
                string.ascii_letters + string.digits,
                k=20
            )
        )

        # ===============================
        # INSERT APPROVAL REQUEST
        # ===============================
        insert_request_query = """
            INSERT INTO approval_requests
            (
                request_id,
                app_id,
                requested_by,
                requested_status,
                status,
                date_requested,
                admin_id,
                admin_area,
                admin_city,
                reason
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        request_result = execute_query(
            insert_request_query,
            (
                request_id,
                app_id,
                admin_username,
                status,
                "Pending",
                ph_now().strftime("%Y-%m-%d %H:%M:%S"),
                admin_id,
                admin_area,
                admin_city,
                reason
            )
        )

        if request_result is None:
            return jsonify({
                "error": "Failed to create approval request"
            }), 500

        print(f" Approval request created: {request_id}")

        # ===============================
        # CHANGE APPLICATION STATUS
        # ===============================
        update_app_query = """
            UPDATE applications
            SET status = 'Request Sent'
            WHERE application_number = %s
        """

        update_result = execute_query(
            update_app_query,
            (app_id,)
        )

        if update_result is None:
            return jsonify({
                "error": "Failed to update application status"
            }), 500

        print(
            f" Application {app_id} "
            f"status changed to Request Sent"
        )

        # ===============================
        # CREATE NOTIFICATION
        # ===============================
        notification_id = int(
            datetime.now().timestamp() * 1000
        )

        applicant_name = (
            f"{app_data.get('first_name', '')} "
            f"{app_data.get('last_name', '')}"
        ).strip()

        application_number = (
            app_data.get(
                "application_number",
                "N/A"
            )
        )

        # ===============================
        # NOTIFICATION MESSAGE
        # ===============================
        if status == "Pending":

            message = (
                f"{admin_username} ({admin_id}) "
                f"has requested to RESTORE "
                f"{applicant_name}'s application "
                f"({application_number}) "
                f"back to Pending status"
            )

            notif_title = "Admin Restore Request"

        elif status == "Reapply":

            message = (
                f"{admin_username} ({admin_id}) "
                f"has requested to RE-APPLY "
                f"{applicant_name}'s application "
                f"({application_number})"
            )

            if reason:
                message += f"\nMessage: {reason}"

            notif_title = "Admin Reapply Request"

        elif status == "Rejected" and reason:

            message = (
                f"{admin_username} ({admin_id}) "
                f"has requested to "
                f"{status.lower()} "
                f"{applicant_name}'s application "
                f"({application_number})"
                f"\nReason: {reason}"
            )

            notif_title = f"Admin {status} Request"

        else:

            message = (
                f"{admin_username} ({admin_id}) "
                f"has requested to "
                f"{status.lower()} "
                f"{applicant_name}'s application "
                f"({application_number})"
            )

            notif_title = f"Admin {status} Request"

        # ===============================
        # INSERT NOTIFICATION
        # ===============================
        insert_notif_query = """
            INSERT INTO notifications
            (
                id,
                title,
                message,
                type,
                relatedId,
                timestamp,
                read_status
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """

        notif_result = execute_query(
            insert_notif_query,
            (
                notification_id,
                notif_title,
                message,
                "admin_request",
                app_id,
                ph_now_iso(),
                0
            )
        )

        if notif_result is None:
            print(
                " Notification insertion failed"
            )
        else:
            print(
                f" Notification created: "
                f"{notification_id}"
            )

        # ===============================
        # SUCCESS
        # ===============================
        print("=" * 60)
        print(" ADMIN REQUEST COMPLETED")
        print("=" * 60)

        return jsonify({
            "message": "Request sent to superadmin.",
            "request_id": request_id,
            "admin_id": admin_id,
            "admin_username": admin_username
        }), 200

    except Exception as e:

        print("=" * 60)
        print(" ADMIN REQUEST ERROR")
        print(str(e))
        print("=" * 60)

        import traceback
        traceback.print_exc()

        return jsonify({
            "error": str(e)
        }), 500

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


            
# ===============================
# GET ADMIN NOTIFICATIONS
# MYSQL / RAILWAY READY
# ===============================
@app.route("/api/admin/notifications", methods=["GET"])
def get_admin_notifications():
    try:
        admin_id = (
            request.args.get("admin_id")
            or request.args.get("username")
        )

        tab_id = request.args.get("tab_id", "")

        print(
            f" GET ADMIN NOTIFICATIONS - "
            f"admin_id: {admin_id}, tab_id: {tab_id}"
        )

        # ===============================
        # VALIDATE ADMIN ID
        # ===============================
        if not admin_id:
            return jsonify({
                "error": "Admin ID required"
            }), 400

        # ===============================
        # VERIFY SESSION IF TAB ID EXISTS
        # ===============================
        if tab_id:
            user_session = session.get(
                f"admin_{tab_id}"
            )

            if user_session:
                session_admin = (
                    user_session.get("admin_username")
                    or user_session.get("user_name")
                )

                if (
                    session_admin
                    and session_admin != admin_id
                ):
                    print(
                        f" Session mismatch: "
                        f"{session_admin} vs {admin_id}"
                    )

                    admin_id = session_admin

        # ===============================
        # GET CITY / AREA FROM REQUEST
        # ===============================
        admin_city = request.args.get(
            "city",
            ""
        )

        admin_area = request.args.get(
            "area",
            ""
        )

        # ===============================
        # GET ADMIN DETAILS FROM DATABASE
        # IF CITY AND AREA WERE NOT PROVIDED
        # ===============================
        if not admin_city and not admin_area:

            admin_query = """
                SELECT area, city
                FROM admins
                WHERE admin_id = %s
                   OR id = %s
                LIMIT 1
            """

            admin_data = execute_query(
                admin_query,
                (admin_id, admin_id),
                fetch_one=True
            )

            if admin_data:
                admin_city = (
                    admin_data.get("city")
                    or ""
                )

                admin_area = (
                    admin_data.get("area")
                    or ""
                )

                print(
                    f" Admin location - "
                    f"City: {admin_city}, "
                    f"Area: {admin_area}"
                )

        # ===============================
        # GET ADMIN NOTIFICATIONS
        # ===============================
        query = """
            SELECT
                id,
                title,
                message,
                type,
                relatedId,
                request_id,
                timestamp,
                read_status,
                admin_id,
                requested_by,
                contract_number,
                billing_date,
                application_city,
                admin_city,
                admin_area
            FROM admin_notifications
            WHERE
                (
                    admin_id = %s

                    OR requested_by = %s

                    OR (
                        application_city = %s
                        AND application_city != ''
                        AND application_city IS NOT NULL
                    )

                    OR (
                        admin_city = %s
                        AND admin_city != ''
                        AND admin_city IS NOT NULL
                    )

                    OR (
                        admin_area = %s
                        AND admin_area != ''
                        AND admin_area IS NOT NULL
                    )
                )
            ORDER BY id DESC
            LIMIT 50
        """

        notifications = execute_query(
            query,
            (
                admin_id,
                admin_id,
                admin_city,
                admin_city,
                admin_area
            ),
            fetch=True
        )

        # ===============================
        # HANDLE QUERY FAILURE
        # ===============================
        if notifications is None:
            print(
                " Failed to retrieve admin notifications"
            )

            return jsonify([]), 500

        print(
            f" Found {len(notifications)} "
            f"notifications for {admin_id}"
        )

        # ===============================
        # FORMAT RESPONSE
        # ===============================
        result = []

        for n in notifications:

            result.append({
                "id": n.get("id"),

                "title": n.get(
                    "title",
                    "Notification"
                ),

                "message": n.get(
                    "message",
                    ""
                ),

                "type": n.get(
                    "type",
                    "info"
                ),

                "relatedId": n.get(
                    "relatedId"
                ),

                "request_id": n.get(
                    "request_id"
                ),

                "timestamp": str(
                    n.get("timestamp")
                ),

                "read": (
                    n.get("read_status") == 1
                ),

                "contract_number": n.get(
                    "contract_number"
                ),

                "billing_date": n.get(
                    "billing_date"
                )
            })

        # ===============================
        # RETURN NOTIFICATIONS
        # ===============================
        return jsonify(result), 200

    except Exception as e:

        print("=" * 60)
        print(" ERROR IN GET ADMIN NOTIFICATIONS")
        print(str(e))
        print("=" * 60)

        import traceback
        traceback.print_exc()

        return jsonify([]), 500


# ===============================
# MARK ADMIN NOTIFICATION AS READ (XAMPP/MYSQL)
# ===============================
@app.route("/api/admin/notifications/<int:notification_id>/read", methods=["PATCH"])
def mark_admin_notification_read(notification_id):
    try:
        admin_id = request.args.get("admin_id") or request.args.get("username")
        
        if not admin_id:
            data = request.get_json() or {}
            admin_id = data.get("admin_id") or data.get("username")
        
        if not admin_id:
            return jsonify({"error": "Admin ID or username required"}), 400
        
        # Update read_status to 1
        query = "UPDATE admin_notifications SET read_status = 1 WHERE id = %s AND read_status = 0"
        rows_affected = execute_query(query, (notification_id,))
        
        if rows_affected > 0:
            print(f" Admin {admin_id} marked notification {notification_id} as read")
        
        return jsonify({"message": "Notification marked as read"})
    
    except Exception as e:
        print(f"Error marking notification as read: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# MARK ALL ADMIN NOTIFICATIONS AS READ (XAMPP/MYSQL)
# ===============================
@app.route("/api/admin/notifications/read-all", methods=["PUT"])
def mark_all_admin_notifications_read():
    try:
        admin_id = request.args.get("admin_id") or request.args.get("username")
        
        if not admin_id:
            data = request.get_json() or {}
            admin_id = data.get("admin_id") or data.get("username")
        
        if not admin_id:
            return jsonify({"error": "Admin ID or username required"}), 400
        
        # Get admin's city/area
        admin_city = request.args.get("city", "")
        admin_area = request.args.get("area", "")
        
        if not admin_city and not admin_area:
            admin_query = """
                SELECT area, city FROM admins 
                WHERE admin_id = %s OR id = %s
            """
            admin_data = execute_query(admin_query, (admin_id, admin_id), fetch_one=True)
            if admin_data:
                admin_city = admin_data.get('city') or ''
                admin_area = admin_data.get('area') or ''
        
        # Update all unread notifications for this admin
        query = """
            UPDATE admin_notifications 
            SET read_status = 1 
            WHERE read_status = 0 
            AND (
                admin_id = %s 
                OR requested_by = %s 
                OR (admin_city = %s AND admin_city != '')
                OR (application_city = %s AND application_city != '')
                OR (admin_area = %s AND admin_area != '')
            )
        """
        rows_affected = execute_query(query, (admin_id, admin_id, admin_city, admin_city, admin_area))
        
        print(f" Admin {admin_id} marked {rows_affected} notifications as read")
        return jsonify({"message": f"Marked {rows_affected} notifications as read", "count": rows_affected})
    
    except Exception as e:
        print(f"Error marking all notifications as read: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# GET UNREAD ADMIN NOTIFICATION COUNT (XAMPP/MYSQL)
# ===============================
@app.route("/api/admin/notifications/unread/count", methods=["GET"])
def get_unread_admin_notification_count():
    try:
        admin_id = request.args.get("admin_id") or request.args.get("username")
        
        if not admin_id:
            return jsonify({"error": "Admin ID or username required"}), 400
        
        # Get admin's city/area
        admin_city = request.args.get("city", "")
        admin_area = request.args.get("area", "")
        
        if not admin_city and not admin_area:
            admin_query = """
                SELECT area, city FROM admins 
                WHERE admin_id = %s OR id = %s
            """
            admin_data = execute_query(admin_query, (admin_id, admin_id), fetch_one=True)
            if admin_data:
                admin_city = admin_data.get('city') or ''
                admin_area = admin_data.get('area') or ''
        
        # Count unread notifications
        query = """
            SELECT COUNT(*) as unread_count
            FROM admin_notifications 
            WHERE read_status = 0 
            AND (
                admin_id = %s 
                OR requested_by = %s 
                OR (admin_city = %s AND admin_city != '')
                OR (application_city = %s AND application_city != '')
                OR (admin_area = %s AND admin_area != '')
            )
        """
        result = execute_query(query, (admin_id, admin_id, admin_city, admin_city, admin_area), fetch_one=True)
        
        unread_count = result.get("unread_count", 0) if result else 0
        
        return jsonify({"unread_count": unread_count})
    
    except Exception as e:
        print(f"Error getting unread count: {e}")
        return jsonify({"unread_count": 0}), 500


# ===============================
# ADMIN - EXPORT CUSTOMERS DATA AS EXCEL (WITH AUTO COLUMN WIDTH)
# ===============================
@app.route("/api/admin/export-customers-excel", methods=["GET"])
def admin_export_customers_excel():
    """Export customers data to Excel (area based for admin) with auto column width"""
    username = request.args.get("username")
    
    if not username:
        return jsonify({"error": "Username required"}), 400
    
    try:
        # Get admin's area
        admin_query = "SELECT area FROM admins WHERE username = %s OR admin_id = %s"
        admin_info = execute_query(admin_query, (username, username), fetch_one=True)
        
        if not admin_info:
            return jsonify({"error": "Admin not found"}), 404
        
        admin_area = admin_info.get("area", "")
        
        if not admin_area:
            return jsonify({"error": "Admin area not found"}), 404
        
        # Get date filters
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        
        # Build query for customers - USE DATE() FUNCTION (NO FORMATTING)
        query = """
            SELECT 
                application_number, contract_number, first_name, last_name,
                middle_name, suffix, email, mobile, address, barangay,
                city, province, zip, plan, plan_speed, plan_price,
                status, installation_status, 
                DATE(approval_date) as approval_date,
                DATE(date_installed) as date_installed,
                billing_date,
                DATE(created_at) as created_at
            FROM customers 
            WHERE city = %s
        """
        params = [admin_area]
        
        # Add date filters
        if start_date and end_date:
            query += " AND DATE(approval_date) >= DATE(%s) AND DATE(approval_date) <= DATE(%s)"
            params.extend([start_date, end_date])
        elif start_date:
            query += " AND DATE(approval_date) >= DATE(%s)"
            params.append(start_date)
        elif end_date:
            query += " AND DATE(approval_date) <= DATE(%s)"
            params.append(end_date)
        
        query += " ORDER BY approval_date DESC"
        
        customers = execute_query(query, params, fetch=True) or []
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Customers_{admin_area}"
        
        # Define headers
        headers = [
            "Application Number", "Contract Number", "Customer Name", "First Name",
            "Last Name", "Middle Name", "Suffix", "Email", "Mobile", "Plan",
            "Speed", "Price", "Status", "Installation Status", "Address",
            "Barangay", "City", "Province", "Zip Code", "Approval Date",
            "Date Installed", "Billing Date", "Created At"
        ]
        
        # Add headers with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0047ab", end_color="0047ab", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row_idx, customer in enumerate(customers, 2):
            full_name = f"{customer.get('first_name', '')} {customer.get('last_name', '')}".strip()
            
            # Get date values
            approval_date = customer.get('approval_date')
            date_installed = customer.get('date_installed')
            created_at = customer.get('created_at')
            
            ws.cell(row=row_idx, column=1, value=str(customer.get('application_number', '')))
            ws.cell(row=row_idx, column=2, value=customer.get('contract_number', ''))
            ws.cell(row=row_idx, column=3, value=full_name)
            ws.cell(row=row_idx, column=4, value=customer.get('first_name', ''))
            ws.cell(row=row_idx, column=5, value=customer.get('last_name', ''))
            ws.cell(row=row_idx, column=6, value=customer.get('middle_name', ''))
            ws.cell(row=row_idx, column=7, value=customer.get('suffix', ''))
            ws.cell(row=row_idx, column=8, value=customer.get('email', ''))
            ws.cell(row=row_idx, column=9, value=customer.get('mobile', ''))
            ws.cell(row=row_idx, column=10, value=customer.get('plan', ''))
            ws.cell(row=row_idx, column=11, value=customer.get('plan_speed', ''))
            ws.cell(row=row_idx, column=12, value=customer.get('plan_price', ''))
            ws.cell(row=row_idx, column=13, value=customer.get('status', 'Approved'))
            ws.cell(row=row_idx, column=14, value=customer.get('installation_status', 'Pending'))
            ws.cell(row=row_idx, column=15, value=customer.get('address', ''))
            ws.cell(row=row_idx, column=16, value=customer.get('barangay', ''))
            ws.cell(row=row_idx, column=17, value=customer.get('city', ''))
            ws.cell(row=row_idx, column=18, value=customer.get('province', ''))
            ws.cell(row=row_idx, column=19, value=customer.get('zip', ''))
            
            # Format dates as string para hindi magkaproblema
            ws.cell(row=row_idx, column=20, value=str(approval_date) if approval_date else '')
            ws.cell(row=row_idx, column=21, value=str(date_installed) if date_installed else '')
            ws.cell(row=row_idx, column=22, value=customer.get('billing_date', ''))
            ws.cell(row=row_idx, column=23, value=str(created_at) if created_at else '')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        from datetime import datetime
        date_now = ph_now().strftime("%Y%m%d_%H%M%S")
        filename = f"customers_{admin_area}_{date_now}.xlsx"
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error in admin_export_customers_excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500



# ===============================
# Admin View Single Application Page
# ===============================
@app.route("/admin/view-application/<app_id>")
def admin_view_application(app_id):
    return render_template("admin-view-application.html", app_id=app_id)

# ===============================
# API Get Single Application - CONVERTED TO MYSQL
# ===============================
@app.route("/api/admin/application/<app_id>", methods=["GET"])
def admin_get_single_application(app_id):
    try:
        query = """
            SELECT application_number, first_name, last_name, middle_name, suffix,
                   email, mobile, secondary_mobile, phone, birthdate, place_of_birth,
                   sex, civil_status, citizenship, occupation, home_ownership,
                   address, billing_address, house_number, landmark, mother_maiden_name,
                   barangay, city, province, zip, employer, business_address, father_name,
                   business_phone, spouse_name, spouse_occupation, spouse_employer,
                   spouse_phone, plan, plan_price, plan_speed, service_type,
                   installation_address, installation_phone, installation_fee,
                   date_submitted, time_submitted, timestamp, status, rejection_reason,
                   signature, id_front, id_back, proof_billing, profile_photo,
                   tv_qty, tv_brand, tv_type, contract_number, billing_date,
                   approval_date, latitude, longitude,
                   reapply_requested, reapply_requested_at
            FROM applications 
            WHERE application_number = %s
        """
        data = execute_query(query, (app_id,), fetch_one=True)
        
        if not data:
            return jsonify({"error": "Application not found"}), 404
        
        # ========== CONVERT IMAGE FIELDS TO CLOUDINARY ==========
        image_fields = ['signature', 'id_front', 'id_back', 'proof_billing', 'profile_photo']
        for field in image_fields:
            if data.get(field):
                value = data.get(field)
                if value and value != 'none' and value != '':
                    if not value.startswith('http'):
                        data[field] = get_cloudinary_url(value)
                        print(f" Converted {field} to Cloudinary: {data[field][:80]}...")
        
        # Parse JSON fields (tv_qty, tv_brand, tv_type are stored as JSON strings)
        if data.get('tv_qty'):
            try:
                data['tv_qty'] = json.loads(data['tv_qty'])
            except:
                data['tv_qty'] = []
        
        if data.get('tv_brand'):
            try:
                data['tv_brand'] = json.loads(data['tv_brand'])
            except:
                data['tv_brand'] = []
        
        if data.get('tv_type'):
            try:
                data['tv_type'] = json.loads(data['tv_type'])
            except:
                data['tv_type'] = []
        
        # Add id field for frontend compatibility
        data['id'] = data.get('application_number')
        
        return jsonify(data)
        
    except Exception as e:
        print(f"Error getting single application: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# ADMIN DOWNLOAD CONTRACT PDF - CONVERTED TO MYSQL
# ===============================
@app.route("/admin/download/contract/<app_id>/<contract_number>")
def admin_download_contract_pdf(app_id, contract_number):
    """Generate and download contract PDF for admin view"""
    try:
        # Reuse the superadmin download function (already converted to MySQL)
        return download_contract_pdf(app_id, contract_number)
    except Exception as e:
        print(f"Error generating contract PDF: {e}")
        import traceback
        traceback.print_exc()
        return f"Error generating PDF: {str(e)}", 500


# ===============================
# ADMIN DOWNLOAD APPLICATION PDF - CONVERTED TO MYSQL
# ===============================
@app.route('/admin/download/pdf/<app_id>')
def admin_download_pdf(app_id):
    try:
        # Get application data from MySQL
        query = "SELECT application_number FROM applications WHERE application_number = %s"
        data = execute_query(query, (app_id,), fetch_one=True)

        if not data:
            return "Application not found", 404

        # Get application_number
        application_number = data.get("application_number")

        if not application_number:
            return "Application number missing", 400

        # Use existing PDF generator (already converted to MySQL)
        return download_pdf(application_number)

    except Exception as e:
        print(f"Admin PDF download error: {e}")
        return str(e), 500
    

# ===============================
# Admin View Customers Page
# ===============================
@app.route("/admin/view-customers")
def admin_view_customers_page():
    highlight_id = request.args.get("highlight", "")
    return render_template("admin-view-customers.html", highlight_id=highlight_id)
    

# ===============================
# GET ADMIN APPROVED APPLICATIONS - CONVERTED TO MYSQL
# ===============================
@app.route("/api/admin/approved-applications", methods=["GET"])
def get_admin_approved_applications():
    try:
        username = request.args.get("username") or session.get("admin_username") or session.get("adminUsername")
        tab_id = request.args.get("tab_id", "")  # KUNIN ANG TAB ID
        
        if not username:
            return jsonify({"error": "Username required"}), 400
        
        # KUNG MAY TAB ID, I-VERIFY NA TAMA ANG SESSION
        if tab_id:
            user_session = session.get(f"admin_{tab_id}")
            if user_session:
                session_username = user_session.get("admin_username") or user_session.get("user_name")
                if session_username and session_username != username:
                    print(f" Session mismatch: {session_username} vs {username}")
                    username = session_username

        # ========== GET ADMIN INFO FROM MYSQL ==========
        admin_query = "SELECT area FROM admins WHERE username = %s OR admin_id = %s"
        admin_data = execute_query(admin_query, (username, username), fetch_one=True)

        if not admin_data:
            return jsonify({"error": "Admin not found"}), 404

        admin_area = str(admin_data.get("area", "")).strip().lower()

        # ========== GET CUSTOMERS FROM MYSQL ==========
        customers_query = """
            SELECT application_number, contract_number, first_name, last_name, email,
                   plan, plan_speed, plan_price, status, installation_status, city,
                   approval_date, created_at
            FROM customers 
            WHERE status = 'Approved'
        """
        all_customers = execute_query(customers_query, fetch=True) or []

        approved_apps = []

        for cust in all_customers:
            cust_city = str(cust.get("city", "")).strip().lower()

            if admin_area in cust_city or cust_city in admin_area:
                approved_apps.append({
                    "id": cust.get("application_number"),
                    "application_number": cust.get("application_number", ""),
                    "contract_number": cust.get("contract_number", "N/A"),
                    "first_name": cust.get("first_name", ""),
                    "last_name": cust.get("last_name", ""),
                    "full_name": f"{cust.get('first_name', '')} {cust.get('last_name', '')}".strip(),
                    "email": cust.get("email", ""),
                    "plan": cust.get("plan", "N/A"),
                    "plan_speed": cust.get("plan_speed", "N/A"),
                    "plan_price": cust.get("plan_price", "N/A"),
                    "status": cust.get("status", "Approved"),
                    "installation_status": cust.get("installation_status", "Pending"),
                    "approval_date": cust.get("approval_date"),
                    "created_at": cust.get("created_at")
                })

        return jsonify(approved_apps), 200
        
    except Exception as e:
        print(f"Error in get_admin_approved_applications: {e}")
        return jsonify({"error": str(e)}), 500
    

    
# ===============================
# Admin View Single Customer Application Page
# ===============================
@app.route("/admin/view-customer-application/<app_id>")
def admin_view_customer_application(app_id):
    return render_template("admin-view-customer-application.html", app_id=app_id)


# ===============================
# API Get Single Customer Application (only approved applications)
# ===============================
@app.route("/api/admin/customer-application/<app_id>")
def admin_get_single_customer_application(app_id):
    try:
        # Query from applications table in MySQL
        query = """
            SELECT id, application_number, first_name, middle_name, last_name, suffix,
                   email, mobile, secondary_mobile, phone, birthdate, place_of_birth,
                   mother_maiden_name, sex, civil_status, citizenship, occupation,
                   home_ownership, address, barangay, city, province, zip,
                   employer, business_address, business_phone, spouse_name,
                   spouse_occupation, spouse_employer, spouse_phone, parents_name,
                   others, plan, plan_speed, plan_price, service_type,
                   installation_address, installation_phone, installation_fee,
                   tv_qty, tv_brand, tv_type, signature, id_front, id_back,
                   proof_billing, profile_photo, latitude, longitude,
                   status, contract_number, billing_date, approval_date,
                   installation_status, rejection_reason, date_submitted,
                   time_submitted, created_at
            FROM applications 
            WHERE application_number = %s OR id = %s
        """
        data = execute_query(query, (app_id, app_id), fetch_one=True)
        
        if not data:
            return jsonify({"error": "Customer application not found"}), 404

        # Only allow approved applications
        if data.get("status") != "Approved":
            return jsonify({"error": "Customer application is not approved"}), 403

        # ========== CONVERT IMAGE FIELDS TO CLOUDINARY ==========
        image_fields = ['signature', 'id_front', 'id_back', 'proof_billing', 'profile_photo']
        for field in image_fields:
            if data.get(field):
                value = data.get(field)
                if value and value != 'none' and value != '':
                    if not value.startswith('http'):
                        data[field] = get_cloudinary_url(value)
                        print(f" Converted {field} to Cloudinary: {data[field][:80]}...")

        # Convert datetime objects to string for JSON serialization
        if data.get('birthdate'):
            data['birthdate'] = str(data['birthdate'])
        if data.get('approval_date'):
            data['approval_date'] = str(data['approval_date'])
        if data.get('date_submitted'):
            data['date_submitted'] = str(data['date_submitted'])
        if data.get('created_at'):
            data['created_at'] = str(data['created_at'])

        # Parse JSON fields (tv_qty, tv_brand, tv_type are stored as JSON strings)
        import json
        if data.get('tv_qty'):
            try:
                data['tv_qty'] = json.loads(data['tv_qty']) if isinstance(data['tv_qty'], str) else data['tv_qty']
            except:
                data['tv_qty'] = []
        else:
            data['tv_qty'] = []
            
        if data.get('tv_brand'):
            try:
                data['tv_brand'] = json.loads(data['tv_brand']) if isinstance(data['tv_brand'], str) else data['tv_brand']
            except:
                data['tv_brand'] = []
        else:
            data['tv_brand'] = []
            
        if data.get('tv_type'):
            try:
                data['tv_type'] = json.loads(data['tv_type']) if isinstance(data['tv_type'], str) else data['tv_type']
            except:
                data['tv_type'] = []
        else:
            data['tv_type'] = []

        return jsonify(data)
        
    except Exception as e:
        print(f"Error getting customer application: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ===============================
# Admin Installation Summary - WITH CANCELLED & TERMINATED
# ===============================
@app.route("/api/admin/installation-summary", methods=["GET"])
def get_admin_installation_summary():
    try:
        from datetime import datetime, timedelta

        username = request.args.get("username") or session.get("admin_username") or session.get("adminUsername")
        if not username:
            return jsonify({"error": "Username required"}), 400

        # DATE FILTERS
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        requested_area = request.args.get("area", "").strip()

        if start_date:
            start_date = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            end_date = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1, seconds=-1)

        # ========== GET ADMIN INFO FROM MYSQL ==========
        admin_query = "SELECT area FROM admins WHERE username = %s OR admin_id = %s"
        admin_data = execute_query(admin_query, (username, username), fetch_one=True)

        if not admin_data:
            return jsonify({"error": "Admin not found"}), 404

        admin_area = str(admin_data.get("area", "") or "").strip().lower()
        area_filter = str(requested_area or admin_area).strip().lower()

        # ========== GET CUSTOMERS FROM MYSQL ==========
        # ISAMA ANG APPROVED AT CANCELLED NA STATUS
        customers_query = """
            SELECT application_number, city, status, installation_status,
                   date_pending, date_ongoing, date_installed,
                   date_cancelled, date_terminated
            FROM customers 
            WHERE status IN ('Approved', 'Cancelled')
        """
        all_customers = execute_query(customers_query, fetch=True) or []

        # IDAGDAG ANG CANCELLED AT TERMINATED
        installation_summary = {
            "Pending": 0,
            "Ongoing": 0,
            "Installed": 0,
            "Cancelled": 0,
            "Terminated": 0
        }

        def parse_date(d):
            if not d:
                return None
            if isinstance(d, datetime):
                return d
            if isinstance(d, str):
                try:
                    return datetime.strptime(d, "%Y-%m-%d %H:%M:%S")
                except:
                    try:
                        return datetime.strptime(d, "%Y-%m-%d")
                    except:
                        return None
            return None

        matched_customers = 0

        for cust in all_customers:
            cust_city = str(cust.get("city", "") or "").strip().lower()

            # AREA FILTER: allow admin area or area query parameter.
            if area_filter:
                if not (area_filter in cust_city or cust_city in area_filter):
                    continue

            # GAMITIN ANG installation_status PARA SA PIE GRAPH
            installation_status = cust.get('installation_status', '').strip()

            # DATE FIELDS (para sa date filter)
            dates = {
                "Pending": parse_date(cust.get("date_pending")),
                "Ongoing": parse_date(cust.get("date_ongoing")),
                "Installed": parse_date(cust.get("date_installed")),
                "Cancelled": parse_date(cust.get("date_cancelled")),
                "Terminated": parse_date(cust.get("date_terminated"))
            }

            # FILTER BY DATE RANGE
            is_in_date_range = False
            for status, dt in dates.items():
                if dt:
                    date_in_range = True
                    if start_date and dt < start_date:
                        date_in_range = False
                    if end_date and dt > end_date:
                        date_in_range = False
                    
                    if date_in_range:
                        is_in_date_range = True
                        break
            
            # KUNG WALANG DATES, ISAMA PA RIN
            has_any_date = any(dt is not None for dt in dates.values())
            if not has_any_date:
                is_in_date_range = True

            if not is_in_date_range:
                continue

            # GAMITIN ANG installation_status PARA SA PIE GRAPH
            if installation_status in installation_summary:
                installation_summary[installation_status] += 1
                matched_customers += 1

        # I-CONTROL ANG ORDER NG STATUSES
        ordered_statuses = ["Pending", "Ongoing", "Installed", "Cancelled", "Terminated"]
        labels = ordered_statuses
        values = [installation_summary.get(status, 0) for status in ordered_statuses]

        return jsonify({
            "installation_summary": installation_summary,
            "labels": labels,        # IDAGDAG
            "values": values,        # IDAGDAG
            "area": area_filter.title() if area_filter else "All Areas",
            "total_matched": matched_customers,
            "date_range": {
                "start": start_date.strftime("%Y-%m-%d") if start_date else None,
                "end": (end_date + timedelta(seconds=1)).strftime("%Y-%m-%d") if end_date else None
            }
        }), 200

    except Exception as e:
        print("Error in installation summary:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Server error"}), 500
    

@app.route("/api/admin/info", methods=["GET"])
def get_admin_info():
    try:
        username = request.args.get("username") or session.get("adminUsername")
        
        if not username:
            return jsonify({"error": "Username required"}), 400

        # Query from admins table in MySQL
        query = """
            SELECT admin_id, full_name, email, mobile, area, city, 
                   profile_photo, status, created_at 
            FROM admins 
            WHERE admin_id = %s OR username = %s OR id = %s
        """
        admin_data = execute_query(query, (username, username, username), fetch=True)

        if not admin_data:
            return jsonify({"error": "Admin not found"}), 404

        # Convert datetime to string if needed
        if admin_data.get('created_at'):
            admin_data['created_at'] = str(admin_data['created_at'])

        return jsonify({
            "username": admin_data.get("admin_id") or admin_data.get("username", username),
            "full_name": admin_data.get("full_name", ""),
            "email": admin_data.get("email", ""),
            "mobile": admin_data.get("mobile", ""),
            "area": admin_data.get("area", "Unknown"),
            "city": admin_data.get("city", ""),
            "profile_photo": admin_data.get("profile_photo", ""),
            "status": admin_data.get("status", "active")
        })

    except Exception as e:
        print("Error getting admin info:", e)
        return jsonify({"error": "Server error"}), 500  



@app.route("/api/admin/installation-status/<app_id>", methods=["PUT"])
def admin_update_installation_status(app_id):
    try:
        data = request.json
        new_status = data.get("installation_status")

        if not new_status:
            return jsonify({"error": "Status required"}), 400

        ref = db.reference(f"customers/{app_id}")
        customer_data = ref.get()

        if not customer_data:
            return jsonify({"error": "Customer not found"}), 404

        update_data = {
            "installation_status": new_status
        }

        # ADD DATE TRACKING
        current_time = ph_now().strftime("%Y-%m-%d %H:%M:%S")

        if new_status == "Ongoing":
            update_data["date_ongoing"] = current_time

        elif new_status == "Installed":
            update_data["date_installed"] = current_time

        ref.update(update_data)

        return jsonify({"success": True}), 200

    except Exception as e:
        print("Installation update error:", e)
        return jsonify({"error": str(e)}), 500


# ===============================
# ADMIN
# ===============================
@app.route("/admin/napbox")
def admin_napbox():
    return render_template("admin-napbox.html")



# ==================== ADVERTISEMENT MANAGEMENT ====================
import os
from werkzeug.utils import secure_filename
from datetime import datetime

# ==================== SHARED UPLOADS CONFIGURATION ====================
SHARED_UPLOADS_BASE = r"C:\xampp\htdocs\cablevision_uploads"

# Unified advertisements folder (for both PNG and MP4)
UPLOAD_FOLDER_ADS = os.path.join(SHARED_UPLOADS_BASE, 'advertisements')

# Allowed extensions - PNG and MP4 only
ALLOWED_EXTENSIONS = {'png', 'mp4'}

# Max file sizes
MAX_IMAGE_SIZE = 2 * 1024 * 1024  # 2MB for PNG
MAX_VIDEO_SIZE = 20 * 1024 * 1024  # 20MB for MP4

# Create upload folder if not exists
os.makedirs(UPLOAD_FOLDER_ADS, exist_ok=True)

# ==================== SERVE SHARED UPLOADS ====================
@app.route('/shared-uploads/<path:filename>')
def serve_shared_uploads(filename):
    """Serve files from shared uploads folder (accessible by all projects)"""
    from flask import send_from_directory
    return send_from_directory(SHARED_UPLOADS_BASE, filename)


def allowed_file(filename):
    """Check if file extension is allowed (PNG or MP4 only)"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_file_type(filename):
    """Determine if file is image or video"""
    ext = filename.rsplit('.', 1)[1].lower()
    return 'image' if ext == 'png' else 'video'


# ============================================================
# CLOUDINARY HELPER FUNCTIONS FOR ADVERTISEMENTS
# ============================================================

def upload_to_cloudinary_ad(file, file_type="image"):
    """Upload advertisement to Cloudinary and return URL"""
    try:
        print(f" Uploading advertisement to Cloudinary: {file.filename}")
        
        # I-reset ang file pointer
        file.stream.seek(0)
        
        filename_without_ext = file.filename.rsplit('.', 1)[0] if hasattr(file, 'filename') else None
        
        # Determine resource type
        resource_type = "video" if file_type == "video" else "image"
        folder = "cablevision/advertisements"
        
        result = cloudinary.uploader.upload(
            file,
            folder=folder,
            resource_type=resource_type,
            public_id=filename_without_ext,
            overwrite=True
        )
        
        print(f" Advertisement uploaded: {result['secure_url']}")
        return result['secure_url']
        
    except Exception as e:
        print(f" Cloudinary upload error: {e}")
        import traceback
        traceback.print_exc()
        return None

def delete_from_cloudinary_ad(image_url):
    """Delete advertisement from Cloudinary"""
    if not image_url:
        return
    
    try:
        if 'cloudinary.com' in image_url:
            parts = image_url.split('/upload/')
            if len(parts) > 1:
                public_id_with_ext = parts[1]
                print(f" Public ID with extension: {public_id_with_ext}")
                
                # Remove version number if present
                if '/' in public_id_with_ext and public_id_with_ext.split('/')[0].startswith('v'):
                    public_id_with_ext = '/'.join(public_id_with_ext.split('/')[1:])
                    print(f" After removing version: {public_id_with_ext}")
                
                # Remove file extension
                public_id = public_id_with_ext.rsplit('.', 1)[0]
                print(f" Final public_id: {public_id}")
                
                # Determine resource type based on URL
                resource_type = "video" if "/video/upload/" in image_url else "image"
                
                result = cloudinary.uploader.destroy(public_id, resource_type=resource_type)
                print(f" Delete result: {result}")
                
                if result.get('result') == 'ok':
                    print(f" Successfully deleted advertisement: {public_id}")
                    return True
                else:
                    print(f" Delete result: {result}")
                    return False
                    
    except Exception as e:
        print(f" Cloudinary delete error: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return False

# ============================================================
# END OF CLOUDINARY HELPER FUNCTIONS
# ============================================================

def save_ad_file(file):
    """Save advertisement file to Cloudinary and return info"""
    if not file or not allowed_file(file.filename):
        return None, None, None
    
    file_type = get_file_type(file.filename)
    
    # Check file size
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    
    if file_type == 'image' and file_size > MAX_IMAGE_SIZE:
        print(f"PNG too large: {file_size} bytes (max {MAX_IMAGE_SIZE})")
        return None, None, None
    if file_type == 'video' and file_size > MAX_VIDEO_SIZE:
        print(f"MP4 too large: {file_size} bytes (max {MAX_VIDEO_SIZE})")
        return None, None, None
    
    # For videos, check portrait orientation (using local file)
    if file_type == 'video':
        # Save temporarily to check orientation
        temp_filename = secure_filename(f"temp_{int(datetime.now().timestamp())}_{file.filename}")
        temp_path = os.path.join(UPLOAD_FOLDER_ADS, temp_filename)
        file.save(temp_path)
        
        is_portrait = is_portrait_video(temp_path)
        
        if not is_portrait:
            os.remove(temp_path)
            print("MP4 video is not portrait orientation (height must be greater than width)")
            return None, None, None
        
        # If portrait, upload to Cloudinary
        # Reset file pointer
        file.seek(0)
        file_url = upload_to_cloudinary_ad(file, file_type="video")
        
        # Delete temp file
        try:
            os.remove(temp_path)
        except:
            pass
        
        if not file_url:
            return None, None, None
            
        return file_url, file_type, file_size
    else:
        # For images, upload directly to Cloudinary
        # Reset file pointer
        file.seek(0)
        file_url = upload_to_cloudinary_ad(file, file_type="image")
        
        if not file_url:
            return None, None, None
            
        return file_url, file_type, file_size


def is_portrait_video(video_path):
    """Check if video is portrait orientation (height > width) using ffprobe"""
    try:
        import subprocess
        import json
        
        cmd = [
            'ffprobe', '-v', 'error', '-select_streams', 'v:0',
            '-show_entries', 'stream=width,height', '-of', 'json', video_path
        ]
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode == 0:
            data = json.loads(result.stdout)
            streams = data.get('streams', [])
            if streams:
                width = streams[0].get('width', 0)
                height = streams[0].get('height', 0)
                return height > width
    except Exception as e:
        print(f"Error getting video dimensions: {e}")
    return True


def delete_ad_file(file_url):
    """Delete advertisement file from Cloudinary"""
    if file_url and 'cloudinary.com' in file_url:
        return delete_from_cloudinary_ad(file_url)
    return False


# ==================== ADVERTISEMENT MANAGEMENT PAGE ====================
@app.route("/superadmin/advertisement")
def superadmin_advertisement():
    """Render the advertisement management page"""
    return render_template("superadmin-advertisement.html")

# ==================== GET ALL ADVERTISEMENTS ====================
@app.route("/api/superadmin/advertisements", methods=["GET"])
def get_advertisements():
    """Get all advertisements (both PNG and MP4) with Cloudinary URLs"""
    try:
        query = """
            SELECT id, file_path, file_type, file_size, date, timestamp, created_at
            FROM advertisements 
            ORDER BY timestamp DESC
        """
        ads = execute_query(query, fetch=True) or []
        
        ad_list = []
        for ad in ads:
            file_path = ad.get('file_path', '')
            
            # If it's a local path, convert to Cloudinary URL
            if file_path and file_path.startswith('/shared-uploads/'):
                file_type = ad.get('file_type', 'image')
                file_path = get_cloudinary_url(file_path, resource_type=file_type)
            
            ad_list.append({
                "id": ad['id'],
                "filePath": file_path,
                "file_type": ad.get('file_type', 'image'),
                "fileSize": ad.get('file_size', 0),
                "date": ad.get('date', ''),
                "timestamp": ad.get('timestamp', 0)
            })
        
        return jsonify(ad_list)
        
    except Exception as e:
        print(f"Error getting advertisements: {e}")
        return jsonify([])

# ==================== CREATE ADVERTISEMENT ====================
@app.route("/api/superadmin/advertisements", methods=["POST"])
def create_advertisement():
    """Create a new advertisement (PNG image or MP4 video) - Cloudinary"""
    try:
        ad_file = request.files.get("file")
        
        if not ad_file or not allowed_file(ad_file.filename):
            return jsonify({"error": "PNG image or MP4 video file is required"}), 400
        
        file_type = get_file_type(ad_file.filename)
        
        # Save file to Cloudinary (includes orientation check for videos)
        file_url, saved_type, file_size = save_ad_file(ad_file)
        
        if not file_url:
            if file_type == 'video':
                return jsonify({"error": "Failed to save video. Make sure it's MP4 format, portrait orientation (height > width), and under 20MB"}), 400
            else:
                return jsonify({"error": "Failed to save image. Make sure it's PNG format and under 2MB"}), 400
        
        now = ph_now()
        
        # Insert into MySQL with Cloudinary URL
        insert_query = """
            INSERT INTO advertisements (file_path, file_type, file_size, date, timestamp, created_at)
            VALUES (%s, %s, %s, %s, %s, NOW())
        """
        ad_id = execute_query(insert_query, (
            file_url,  # Cloudinary URL
            saved_type,
            file_size,
            now.strftime("%B %d, %Y"),
            now.timestamp()
        ))
        
        return jsonify({
            "message": "Advertisement uploaded successfully",
            "id": ad_id,
            "filePath": file_url,
            "fileType": saved_type,
            "fileSize": file_size
        })
        
    except Exception as e:
        print(f"Error creating advertisement: {e}")
        return jsonify({"error": str(e)}), 500


# ==================== DELETE ADVERTISEMENT ====================
@app.route("/api/superadmin/advertisements/<int:ad_id>", methods=["DELETE"])
def delete_advertisement(ad_id):
    """Delete an advertisement from Cloudinary"""
    try:
        check_query = "SELECT id, file_path FROM advertisements WHERE id = %s"
        ad = execute_query(check_query, (ad_id,), fetch_one=True)
        
        if not ad:
            return jsonify({"error": "Advertisement not found"}), 404
        
        file_path = ad.get('file_path')
        
        # Delete from Cloudinary if exists
        if file_path and 'cloudinary.com' in file_path:
            delete_from_cloudinary_ad(file_path)
        
        # Delete from MySQL
        delete_query = "DELETE FROM advertisements WHERE id = %s"
        execute_query(delete_query, (ad_id,))
        
        return jsonify({"message": "Advertisement deleted successfully"})
        
    except Exception as e:
        print(f"Error deleting advertisement: {e}")
        return jsonify({"error": str(e)}), 500


# ==================== INITIALIZE ADVERTISEMENTS TABLE ====================
def init_advertisements_table():
    """Initialize the unified advertisements table in MySQL if it doesn't exist"""
    try:
        create_table_query = """
            CREATE TABLE IF NOT EXISTS advertisements (
                id INT AUTO_INCREMENT PRIMARY KEY,
                file_path VARCHAR(500) NOT NULL,
                file_type ENUM('image', 'video') DEFAULT 'image',
                file_size BIGINT DEFAULT 0,
                date VARCHAR(50),
                timestamp BIGINT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_timestamp (timestamp),
                INDEX idx_type (file_type)
            )
        """
        execute_query(create_table_query)
        print(" advertisements table ready")
        
    except Exception as e:
        print(f"Error creating advertisements table: {e}")


# Call the initialization function when the app starts
init_advertisements_table()


# ==================== OPTIONAL: MIGRATE OLD DATA ====================
def migrate_old_logos_to_advertisements():
    """Migrate existing channel logos to new advertisements table"""
    try:
        # Check if old table exists and has data
        old_count = execute_query("SELECT COUNT(*) as count FROM channel_logos", fetch=True)
        if old_count and old_count[0].get('count', 0) > 0:
            # Get all old logos
            old_logos = execute_query("SELECT id, image_path, date, timestamp FROM channel_logos", fetch=True)
            
            for logo in old_logos:
                # Insert into new table
                insert_query = """
                    INSERT INTO advertisements (file_path, file_type, file_size, date, timestamp)
                    VALUES (%s, 'image', %s, %s, %s)
                """
                # Get file size if file exists
                file_size = 0
                if logo.get('image_path'):
                    file_path = logo.get('image_path')
                    if file_path.startswith('/shared-uploads/'):
                        relative_path = file_path.replace('/shared-uploads/', '')
                        full_path = os.path.join(SHARED_UPLOADS_BASE, relative_path.replace('/', os.sep))
                        if os.path.exists(full_path):
                            file_size = os.path.getsize(full_path)
                
                execute_query(insert_query, (
                    logo.get('image_path'),
                    file_size,
                    logo.get('date'),
                    logo.get('timestamp')
                ))
            
            print(f" Migrated {old_count[0].get('count', 0)} old logos to advertisements table")
    except Exception as e:
        print(f"Migration skipped (no old data): {e}")

# Uncomment to migrate old data
# migrate_old_logos_to_advertisements()


# ===============================
# TECHNICIAN DASHBOARD PAGE
# ===============================
@app.route("/technician/dashboard")
def technician_dashboard():
    """Render technician dashboard"""
    # Get tab_id from request
    tab_id = request.args.get('tab_id', '')
    
    # Check if technician is logged in using tab-based session
    session_data = session.get(f"admin_{tab_id}") if tab_id else None
    
    # Also check regular session as fallback
    if not session_data or session_data.get('user_type') != 'technician':
        # Check if technician_id is in session directly (legacy)
        if 'technician_id' not in session:
            return redirect('/')  # Redirect to login page
        # If found, use it
        session_data = {'user_type': 'technician', 'technician_id': session.get('technician_id')}
    
    return render_template("technician-dashboard.html")



# ===============================
# TECHNICIAN profile PAGE
# ===============================
@app.route("/technician/profile")
def technician_profile():
    tab_id = request.args.get('tab_id') or session.get('active_tab')
    session_data = session.get(f"admin_{tab_id}") if tab_id else None

    if not session_data or session_data.get('user_type') != 'technician':
        return render_template("technician-profile.html", ga_enabled=False, ga_secret=None, ga_setup_uri=None)

    technician_id = session_data.get('technician_id')
    tech_row = execute_query(
        "SELECT ga_enabled, ga_secret FROM technicians WHERE technician_id = %s LIMIT 1",
        (technician_id,),
        fetch=True,
    )

    ga_enabled = bool(tech_row[0].get('ga_enabled')) if tech_row else False
    ga_secret = tech_row[0].get('ga_secret') if tech_row else None
    ga_setup_uri = None

    if not ga_enabled:
        if not ga_secret:
            ga_secret = generate_ga_secret()
            execute_query("UPDATE technicians SET ga_secret = %s WHERE technician_id = %s", (ga_secret, technician_id))
        ga_setup_uri = generate_ga_provisioning_uri(technician_id, ga_secret)

    return render_template("technician-profile.html", ga_enabled=ga_enabled, ga_secret=ga_secret, ga_setup_uri=ga_setup_uri)


# ===============================
# TECHNICIAN get profile
# ===============================
@app.route("/api/technician/profile", methods=["GET"])
def get_technician_profile():
    try:
        technician_id = request.args.get("technician_id")

        if not technician_id:
            return jsonify({"error": "Technician ID required"}), 400

        query = """
            SELECT technician_id, name, email, contact_number, area, team_id, status, profile_photo, ga_enabled, ga_secret
            FROM technicians
            WHERE technician_id = %s
        """
        tech = execute_query(query, (technician_id,), fetch=True)

        if not tech:
            return jsonify({"error": "Technician not found"}), 404

        ga_enabled = bool(tech[0].get('ga_enabled', 0))
        ga_secret = tech[0].get('ga_secret')
        if not ga_enabled and not ga_secret:
            ga_secret = generate_ga_secret()
            execute_query("UPDATE technicians SET ga_secret = %s WHERE technician_id = %s", (ga_secret, technician_id))

        ga_setup_uri = None if ga_enabled else generate_ga_provisioning_uri(technician_id, ga_secret)

        return jsonify({
            "technician_id": tech[0].get('technician_id'),
            "name": tech[0].get('name'),
            "email": tech[0].get('email'),
            "contact_number": tech[0].get('contact_number'),
            "area": tech[0].get('area'),
            "team_id": tech[0].get('team_id'),
            "status": tech[0].get('status'),
            "profile_photo": tech[0].get('profile_photo'),
            "ga_enabled": ga_enabled,
            "ga_secret": ga_secret,
            "ga_setup_uri": ga_setup_uri
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/technician/ga/enable", methods=["POST"])
def technician_enable_google_auth():
    tab_id = request.form.get('tab_id') or session.get('active_tab')
    session_data = session.get(f"admin_{tab_id}") if tab_id else None

    if not session_data or session_data.get('user_type') != 'technician':
        flash("Unauthorized access.", "danger")
        return redirect(url_for("technician_profile"))

    technician_id = session_data.get('technician_id')
    code = request.form.get("ga_code", "").strip()

    tech_row = execute_query(
        "SELECT ga_secret, ga_enabled FROM technicians WHERE technician_id = %s LIMIT 1",
        (technician_id,),
        fetch=True,
    )

    if not tech_row:
        flash("Technician account not found.", "danger")
        return redirect(url_for("technician_profile"))

    secret = tech_row[0].get("ga_secret")
    if not secret:
        secret = generate_ga_secret()
        execute_query("UPDATE technicians SET ga_secret = %s WHERE technician_id = %s", (secret, technician_id))

    if not code:
        flash("Please enter the 6-digit code from Google Authenticator.", "danger")
        return redirect(url_for("technician_profile", toast="ga-missing"))

    if verify_ga_code(secret, code):
        execute_query("UPDATE technicians SET ga_secret = %s, ga_enabled = 1 WHERE technician_id = %s", (secret, technician_id))
        session_data['ga_enabled'] = True
        session[f"admin_{tab_id}"] = session_data
        flash(" Google Authenticator is now enabled!", "success")
        return redirect(url_for("technician_profile", toast="ga-enabled"))
    else:
        flash(" Invalid code. Please try again.", "danger")
        return redirect(url_for("technician_profile", toast="ga-invalid"))


@app.route("/technician/ga/disable", methods=["POST"])
def technician_disable_google_auth():
    tab_id = request.form.get('tab_id') or session.get('active_tab')
    session_data = session.get(f"admin_{tab_id}") if tab_id else None

    if not session_data or session_data.get('user_type') != 'technician':
        flash("Unauthorized access.", "danger")
        return redirect(url_for("technician_profile"))

    technician_id = session_data.get('technician_id')
    execute_query("UPDATE technicians SET ga_secret = NULL, ga_enabled = 0 WHERE technician_id = %s", (technician_id,))
    session_data['ga_enabled'] = False
    session[f"admin_{tab_id}"] = session_data
    flash("Google Authenticator has been disabled.", "info")
    return redirect(url_for("technician_profile", toast="ga-disabled"))


# ===============================
# UPDATE TECHNICIAN PROFILE (XAMPP/MYSQL)
# ===============================
@app.route("/api/technician/update-profile", methods=["POST"])
def update_technician_profile():
    try:
        data = request.get_json()
        technician_id = data.get("technician_id")
        email = data.get("email")
        contact = data.get("contact")
        name = data.get("name")
        password = data.get("password")
        current_password = data.get("current_password")
        tab_id = data.get("tab_id", "")

        if not technician_id:
            return jsonify({"error": "Technician ID required"}), 400

        # I-VERIFY MUNA ANG SESSION
        session_data = session.get(f"admin_{tab_id}")
        if not session_data or session_data.get('user_type') != 'technician':
            return jsonify({"error": "Unauthorized"}), 403

        # KUNG MAY BINAGO SA EMAIL, I-CHECK SA LAHAT NG TABLES
        if email:
            current_query = "SELECT email FROM technicians WHERE technician_id = %s OR email = %s"
            current_tech = execute_query(current_query, (technician_id, email), fetch=True)
            
            if current_tech and current_tech[0].get('email') != email:
                check_query = """
                    SELECT 
                        (SELECT COUNT(*) FROM technicians WHERE email = %s AND technician_id != %s) as tech_count,
                        (SELECT COUNT(*) FROM admins WHERE email = %s) as admin_count,
                        (SELECT COUNT(*) FROM superadmins WHERE email = %s) as superadmin_count
                """
                result = execute_query(check_query, (email, technician_id, email, email), fetch=True)
                
                tech_exists = result[0].get('tech_count', 0) > 0 if result else False
                admin_exists = result[0].get('admin_count', 0) > 0 if result else False
                superadmin_exists = result[0].get('superadmin_count', 0) > 0 if result else False

                if tech_exists or admin_exists or superadmin_exists:
                    return jsonify({"error": f"Email '{email}' already exists"}), 400

        if password and len(password) >= 8:
            current_user = execute_query("SELECT password FROM technicians WHERE technician_id = %s LIMIT 1", (technician_id,), fetch_one=True)
            if not current_user:
                return jsonify({"error": "Account not found"}), 404
            if not current_password or not verify_password(current_user.get('password'), current_password):
                return jsonify({"error": "Current password is incorrect"}), 400

        # Build update query
        update_fields = []
        params = []
        
        if name:
            update_fields.append("name = %s")
            params.append(name)
        if contact is not None:
            clean_contact = contact.replace(" ", "")
            update_fields.append("contact_number = %s")
            params.append(clean_contact)
        if email:
            update_fields.append("email = %s")
            params.append(email)
        if password and len(password) >= 8:
            update_fields.append("password = %s")
            params.append(hash_password(password))
        
        if not update_fields:
            return jsonify({"error": "No fields to update"}), 400
        
        params.append(technician_id)
        
        update_query = f"UPDATE technicians SET {', '.join(update_fields)} WHERE technician_id = %s"
        execute_query(update_query, params)
        
        print(f" Technician {technician_id} profile updated")
        
        return jsonify({
            "success": True,
            "message": "Profile updated successfully"
        }), 200

    except Exception as e:
        print(f"Update technician profile error: {e}")
        return jsonify({"error": str(e)}), 500


# ===============================
# CHECK EMAIL AVAILABILITY FOR TECHNICIAN
# ===============================
@app.route("/api/technician/check-email", methods=["GET"])
def check_technician_email():
    """Check if email exists in technicians, admins, or superadmins tables"""
    email = request.args.get("email")
    tab_id = request.args.get("tab_id", "")
    
    if not email:
        return jsonify({"exists": False, "error": "Email required"}), 400
    
    # Get current technician ID from session
    session_data = session.get(f"admin_{tab_id}")
    technician_id = session_data.get('technician_id') if session_data else None
    
    # Check all tables
    check_query = """
        SELECT 
            (SELECT COUNT(*) FROM technicians WHERE email = %s) as tech_count,
            (SELECT COUNT(*) FROM admins WHERE email = %s) as admin_count,
            (SELECT COUNT(*) FROM superadmins WHERE email = %s) as superadmin_count
    """
    result = execute_query(check_query, (email, email, email), fetch=True)
    
    tech_exists = result[0].get('tech_count', 0) > 0 if result else False
    admin_exists = result[0].get('admin_count', 0) > 0 if result else False
    superadmin_exists = result[0].get('superadmin_count', 0) > 0 if result else False
    
    # EXCLUDE THE CURRENT TECHNICIAN (para hindi mag-error sa sarili niyang email)
    if tech_exists and technician_id:
        check_self_query = "SELECT email FROM technicians WHERE technician_id = %s AND email = %s"
        self_email = execute_query(check_self_query, (technician_id, email), fetch=True)
        if self_email:
            tech_exists = False
    
    exists = tech_exists or admin_exists or superadmin_exists
    
    return jsonify({
        "exists": exists
    })


# ===============================
# DELETE TECHNICIAN PROFILE PHOTO
# ===============================
@app.route("/api/technician/delete-profile-photo", methods=["POST"])
def delete_technician_profile_photo():
    try:
        data = request.json
        technician_id = data.get("technician_id")
        
        if not technician_id:
            return jsonify({"error": "Technician ID required"}), 400

        # Check if technician exists
        check_query = "SELECT technician_id FROM technicians WHERE technician_id = %s"
        tech_exists = execute_query(check_query, (technician_id,), fetch=True)
        
        if not tech_exists:
            return jsonify({"error": "Technician not found"}), 404

        # Update profile_photo to default
        update_query = "UPDATE technicians SET profile_photo = %s WHERE technician_id = %s"
        execute_query(update_query, ("/static/profile.jpg", technician_id))
        
        return jsonify({"message": "Profile photo removed successfully"}), 200
        
    except Exception as e:
        print(f"Delete profile photo error: {e}")
        return jsonify({"error": str(e)}), 500  

# ===============================
# TECHNICIAN NAP BOX SLOTS PAGE
# ===============================
@app.route("/technician/technician-napbox")
def technician_napbox():
    """Render NAP Box Slots page for technician"""
    return render_template("technician-napbox.html")

# ===============================
# HELPER: GET BOUNDARY FROM GEORISK (MUNICIPAL LEVEL)
# ===============================
def get_municipal_boundary_from_georisk(city_name):
    """Get municipal boundary GeoJSON from GeoRisk API"""
    import requests
    
    try:
        georisk_url = "https://portal.georisk.gov.ph/arcgis/rest/services/PSA/Barangay/MapServer/4/query"
        
        query_params = {
            "where": f"city_name = '{city_name.upper()}'",
            "outFields": "city_name,brgy_name,prov_name",
            "returnGeometry": "true",
            "f": "geojson",
            "outSR": "4326"
        }
        
        response = requests.get(georisk_url, params=query_params, timeout=15)
        
        if response.status_code == 200:
            data = response.json()
            
            if data.get("features") and len(data["features"]) > 0:
                combined = {
                    "type": "FeatureCollection",
                    "features": data["features"]
                }
                print(f" Got municipal boundary for {city_name} with {len(data['features'])} barangays")
                return combined
            else:
                print(f" No boundary data found for {city_name}")
                return None
        else:
            print(f" GeoRisk API error: {response.status_code}")
            return None
            
    except Exception as e:
        print(f"Error getting boundary from GeoRisk: {e}")
        return None

# ===============================
# API: GET TECHNICIAN AREA INFORMATION
# ===============================
@app.route("/api/technician/area", methods=["GET"])
def get_technician_area():
    """Get technician's assigned area information with coordinates"""
    technician_id = request.args.get("technician_id")
    
    if not technician_id:
        return jsonify({"error": "Technician ID required"}), 400
    
    # Get technician info
    tech_query = """
        SELECT technician_id, name, area
        FROM technicians 
        WHERE technician_id = %s
        LIMIT 1
    """
    technician = execute_query(tech_query, (technician_id,), fetch=True)
    
    if not technician:
        return jsonify({"error": "Technician not found"}), 404
    
    technician_area_raw = technician[0].get('area', '')
    
    print(f" Technician area: '{technician_area_raw}'")
    
    # Get actual count from napboxes table (case-insensitive)
    count_query = """
        SELECT COUNT(*) as count FROM napboxes 
        WHERE UPPER(area) = UPPER(%s)
    """
    count_result = execute_query(count_query, (technician_area_raw,), fetch=True)
    actual_napbox_count = count_result[0].get('count', 0) if count_result else 0
    
    # Get total slots count
    slots_count_query = """
        SELECT COUNT(*) as total_slots 
        FROM napbox_slots s
        JOIN napboxes n ON s.napbox_id = n.id
        WHERE UPPER(n.area) = UPPER(%s)
    """
    slots_result = execute_query(slots_count_query, (technician_area_raw,), fetch=True)
    actual_slots_count = slots_result[0].get('total_slots', 0) if slots_result else 0
    
    print(f" Actual NAP boxes: {actual_napbox_count}, Actual slots: {actual_slots_count}")
    
    # Get or create area_mapping
    check_area_query = "SELECT id FROM area_mapping WHERE UPPER(area_name) = UPPER(%s)"
    existing_area = execute_query(check_area_query, (technician_area_raw,), fetch=True)
    
    if existing_area:
        # Update existing area_mapping with actual counts
        update_query = """
            UPDATE area_mapping 
            SET napbox_count = %s, total_slots = %s, updated_at = NOW()
            WHERE UPPER(area_name) = UPPER(%s)
        """
        execute_query(update_query, (actual_napbox_count, actual_slots_count, technician_area_raw))
        print(f" Updated area_mapping for '{technician_area_raw}'")
    else:
        # Create new area_mapping
        insert_query = """
            INSERT INTO area_mapping (area_name, district, latitude, longitude, napbox_count, total_slots, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
        """
        execute_query(insert_query, (technician_area_raw, technician_area_raw, 14.5995, 120.9842, actual_napbox_count, actual_slots_count))
        print(f" Created area_mapping for '{technician_area_raw}'")
    
    return jsonify({
        "technician_id": technician[0].get('technician_id'),
        "name": technician[0].get('name'),
        "area": technician_area_raw,
        "district": technician_area_raw,
        "latitude": 14.5995,
        "longitude": 120.9842,
        "napbox_count": actual_napbox_count,
        "total_slots": actual_slots_count
    })

# ===============================
# API: GET MUNICIPAL BOUNDARY
# ===============================
@app.route("/api/technician/boundary/<city_name>", methods=["GET"])
def get_municipal_boundary(city_name):
    """Get municipal boundary GeoJSON"""
    boundary_data = get_municipal_boundary_from_georisk(city_name)
    
    if boundary_data:
        return jsonify(boundary_data)
    else:
        return jsonify({"error": "Boundary not found"}), 404

# ===============================
# API: GET BARANGAYS FOR TECHNICIAN'S AREA
# ===============================
@app.route("/api/technician/barangays", methods=["GET"])
def get_technician_barangays():
    """Get list of barangays in technician's area"""
    technician_id = request.args.get("technician_id")
    
    if not technician_id:
        return jsonify({"error": "Technician ID required"}), 400
    
    area_query = "SELECT area FROM technicians WHERE technician_id = %s LIMIT 1"
    technician = execute_query(area_query, (technician_id,), fetch_one=True)
    
    if not technician:
        return jsonify({"error": "Technician not found"}), 404
    
    technician_area = technician.get('area', '').upper().strip()
    
    if not technician_area:
        return jsonify({"error": "Technician has no assigned area"}), 400
    
    query = "SELECT DISTINCT barangay FROM areas WHERE city = %s ORDER BY barangay"
    barangays = execute_query(query, (technician_area,), fetch_all=True) or []
    
    barangay_list = [b.get('barangay') for b in barangays if b.get('barangay')]
    
    return jsonify(barangay_list)

# ===============================
# API: GET NAP BOX SLOTS
# ===============================
@app.route("/api/technician/technician-napbox", methods=["GET"])
def get_technician_napbox():
    """Get NAP box slots for technician's area"""
    technician_id = request.args.get("technician_id")
    
    if not technician_id:
        return jsonify({"error": "Technician ID required"}), 400
    
    # Get technician area
    area_query = "SELECT area FROM technicians WHERE technician_id = %s LIMIT 1"
    technician = execute_query(area_query, (technician_id,), fetch_one=True)
    
    if not technician:
        return jsonify({"error": "Technician not found"}), 404
    
    technician_area = technician.get('area', '')
    
    print(f" Technician area: '{technician_area}'")
    
    # Get napboxes - case-insensitive comparison with created_at
    napbox_query = """
        SELECT n.id, n.napbox_name, n.location, n.latitude, n.longitude, n.area,
               n.coverage_radius, n.barangay, n.created_at, n.updated_at
        FROM napboxes n
        WHERE UPPER(n.area) = UPPER(%s)
        ORDER BY n.napbox_name
    """
    napboxes = execute_query(napbox_query, (technician_area,), fetch=True) or []
    
    print(f" Found {len(napboxes)} NAP boxes for area '{technician_area}'")
    
    napbox_ids = [nb.get('id') for nb in napboxes if nb.get('id')]
    
    slots_data = []
    status_counts = {'total': 0, 'available': 0, 'occupied': 0}
    
    if napbox_ids:
        placeholders = ','.join(['%s'] * len(napbox_ids))
        slots_query = f"""
            SELECT s.id, s.slot_number, s.status, s.napbox_id, 
                   s.customer_name, s.customer_phone, s.installation_date, 
                   s.updated_at, s.created_at,
                   s.contract_number, s.application_number,
                   n.napbox_name as napbox_name, n.barangay, n.area
            FROM napbox_slots s
            LEFT JOIN napboxes n ON s.napbox_id = n.id
            WHERE s.napbox_id IN ({placeholders})
            ORDER BY n.napbox_name, CAST(s.slot_number AS UNSIGNED)
        """
        slots = execute_query(slots_query, napbox_ids, fetch=True) or []
        
        for slot in slots:
            slot_data = {
                'id': slot.get('id'),
                'slot_number': slot.get('slot_number'),
                'status': slot.get('status', 'available'),
                'napbox_id': slot.get('napbox_id'),
                'napbox_name': slot.get('napbox_name'),
                'barangay': slot.get('barangay'),
                'customer_name': slot.get('customer_name'),
                'customer_phone': slot.get('customer_phone'),
                'contract_number': slot.get('contract_number'),
                'application_number': slot.get('application_number'),
                'installation_date': str(slot.get('installation_date')) if slot.get('installation_date') else None,
                'created_at': str(slot.get('created_at')) if slot.get('created_at') else None,
                'updated_at': str(slot.get('updated_at')) if slot.get('updated_at') else None
            }
            slots_data.append(slot_data)
            
            status = slot.get('status', 'available')
            status_counts['total'] += 1
            if status in status_counts:
                status_counts[status] += 1
    
    # I-RETURN ANG MGA NAPBOXES NA MAY CREATED_AT
    napboxes_for_map = []
    for nb in napboxes:
        lat = nb.get('latitude')
        lng = nb.get('longitude')
        
        napbox_data = {
            'id': nb.get('id'),
            'name': nb.get('napbox_name'),
            'location': nb.get('location') or nb.get('napbox_name'),
            'area': nb.get('area'),
            'barangay': nb.get('barangay'),
            'coverage_radius': nb.get('coverage_radius') or 500,
            'created_at': str(nb.get('created_at')) if nb.get('created_at') else None,
            'updated_at': str(nb.get('updated_at')) if nb.get('updated_at') else None
        }
        
        if lat and lng:
            try:
                napbox_data['latitude'] = float(lat)
                napbox_data['longitude'] = float(lng)
            except (ValueError, TypeError):
                print(f" Invalid coordinates for NAP box {nb.get('id')}")
                napbox_data['latitude'] = None
                napbox_data['longitude'] = None
        else:
            napbox_data['latitude'] = None
            napbox_data['longitude'] = None
            
        napboxes_for_map.append(napbox_data)
    
    return jsonify({
        'slots': slots_data,
        'stats': status_counts,
        'napboxes': napboxes_for_map,
        'area': technician_area
    })

# ===============================
# GET BARANGAYS BY CITY FROM AREAS TABLE
# ===============================
@app.route("/api/areas/by-city/<city>", methods=["GET"])
def get_areas_by_city(city):
    """Get barangays by city name from areas table"""
    try:
        # GAMITIN ANG get_db_connection() (hindi localhost)
        conn = get_db_connection()
        if not conn:
            print(" Database connection failed")
            return jsonify([]), 500
            
        cursor = conn.cursor(dictionary=True)
        
        query = """
            SELECT DISTINCT barangay, zip 
            FROM areas 
            WHERE LOWER(city) = LOWER(%s)
            ORDER BY barangay
        """
        cursor.execute(query, (city,))
        areas_data = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        result = []
        for area in areas_data:
            result.append({
                "barangay": area.get("barangay", ""),
                "zip": area.get("zip", "")
            })
        
        print(f" Found {len(result)} barangays for city: {city}")
        return jsonify(result)
        
    except Exception as e:
        print(f"Error getting areas by city: {e}")
        import traceback
        traceback.print_exc()
        return jsonify([]), 500    

@app.route("/api/technician/napbox", methods=["POST"])
def add_or_update_napbox():
    try:
        data = request.get_json()
        napbox_id = data.get("id")
        napbox_name = data.get("napbox_name")
        latitude = data.get("latitude")
        longitude = data.get("longitude")
        area = data.get("area")
        coverage_radius = data.get("coverage_radius", 500)
        num_slots = data.get("num_slots", 8)
        barangay = data.get("barangay")
        
        print(f" Adding NAP box - Name: {napbox_name}, Area: '{area}', Coords: {latitude}, {longitude}, Slots: {num_slots}")
        
        if not napbox_name or not latitude or not longitude or not area:
            return jsonify({"error": "Missing required fields"}), 400
        
        # Get barangay from coordinates if not provided
        if not barangay:
            try:
                import requests
                geo_response = requests.get(
                    f"https://nominatim.openstreetmap.org/reverse?format=json&lat={latitude}&lon={longitude}&zoom=18&addressdetails=1",
                    headers={"User-Agent": "CableVision-App"},
                    timeout=5
                )
                geo_data = geo_response.json()
                barangay = geo_data.get('address', {}).get('village') or \
                          geo_data.get('address', {}).get('suburb') or \
                          geo_data.get('address', {}).get('neighbourhood') or \
                          'Unknown'
            except:
                barangay = 'Unknown'
        
        if napbox_id:
            # Update existing
            update_query = """
                UPDATE napboxes 
                SET latitude = %s, longitude = %s, location = %s, 
                    coverage_radius = %s, barangay = %s, updated_at = NOW()
                WHERE id = %s
            """
            execute_query(update_query, (latitude, longitude, napbox_name, coverage_radius, barangay, napbox_id))
            message = "NAP box location updated"
            print(f" Updated NAP box ID: {napbox_id}")
        else:
            # GAMITIN ANG get_db_connection() (hindi localhost)
            conn = get_db_connection()
            if not conn:
                return jsonify({"error": "Database connection failed"}), 500
                
            cursor = conn.cursor(dictionary=True)
            
            try:
                # 1. INSERT NAP BOX
                insert_query = """
                    INSERT INTO napboxes (napbox_name, location, latitude, longitude, area, coverage_radius, barangay, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """
                cursor.execute(insert_query, (napbox_name, napbox_name, latitude, longitude, area, coverage_radius, barangay))
                napbox_id = cursor.lastrowid
                print(f" Inserted NAP box ID: {napbox_id}, Area: '{area}'")
                
                # 2. CREATE SLOTS
                for slot_num in range(1, num_slots + 1):
                    slot_query = """
                        INSERT INTO napbox_slots (slot_number, status, napbox_id, barangay)
                        VALUES (%s, 'available', %s, %s)
                    """
                    cursor.execute(slot_query, (str(slot_num), napbox_id, barangay))
                    print(f"   Created slot {slot_num} for NAP box {napbox_id}")
                
                # 3. COMMIT ALL CHANGES
                conn.commit()
                print(f" COMMITTED {num_slots} slots to napbox_slots table")
                
            except Exception as db_error:
                print(f" Database error: {db_error}")
                conn.rollback()
                raise db_error
            finally:
                cursor.close()
                conn.close()
            
            # Update area_mapping
            count_query = "SELECT COUNT(*) as count FROM napboxes WHERE LOWER(area) = LOWER(%s)"
            count_result = execute_query(count_query, (area,), fetch_one=True)
            new_count = count_result.get('count', 0) if count_result else 0
            
            slots_count_query = """
                SELECT COUNT(*) as total_slots 
                FROM napbox_slots s
                JOIN napboxes n ON s.napbox_id = n.id
                WHERE LOWER(n.area) = LOWER(%s)
            """
            slots_result = execute_query(slots_count_query, (area,), fetch_one=True)
            total_slots = slots_result.get('total_slots', 0) if slots_result else 0
            
            update_area_query = """
                UPDATE area_mapping 
                SET napbox_count = %s, total_slots = %s, updated_at = NOW()
                WHERE LOWER(area_name) = LOWER(%s)
            """
            execute_query(update_area_query, (new_count, total_slots, area))
            print(f" Updated area_mapping for '{area}': napbox_count={new_count}, total_slots={total_slots}")
            
            message = "NAP box added successfully"
        
        return jsonify({
            "message": message, 
            "id": napbox_id,
            "napbox_name": napbox_name,
            "area": area
        })
        
    except Exception as e:
        print(f" Error adding/updating NAP box: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

        
    
@app.route('/api/technician/update-slot-status', methods=['POST'])
def update_slot_status():
    try:
        data = request.json
        slot_id = data.get('slot_id')
        new_status = data.get('status')
        technician_id = data.get('technician_id')
        
        print(f" Updating slot {slot_id} to status: {new_status}")
        
        if not slot_id or not new_status:
            return jsonify({'success': False, 'error': 'Missing slot_id or status'}), 400
        
        # GET CURRENT SLOT DATA BEFORE UPDATE
        current_slot_query = """
            SELECT id, status, customer_name, customer_phone, application_number, 
                   installation_date, slot_number, napbox_id, barangay
            FROM napbox_slots 
            WHERE id = %s
        """
        current_slot = execute_query(current_slot_query, (slot_id,), fetch_one=True)
        
        if not current_slot:
            return jsonify({'success': False, 'error': 'Slot not found'}), 404
        
        current_status = current_slot.get('status')
        
        print(f" Current status: {current_status}, New status: {new_status}")
        
        # UPDATE STATUS LANG - HUWAG BURAHIN ANG CUSTOMER DATA
        # Keep all customer details, only change the status
        query = """
            UPDATE napbox_slots 
            SET status = %s, updated_at = NOW()
            WHERE id = %s
        """
        execute_query(query, (new_status, slot_id))
        
        print(f" Slot {slot_id} status updated from {current_status} to {new_status}")
        print(f"   Customer data preserved: {current_slot.get('customer_name')}, {current_slot.get('customer_phone')}, {current_slot.get('application_number')}")
        
        # Optional: If slot became available, you might want to update the customer's installation status
        # This is optional - you can remove or keep as needed
        if current_status == 'occupied' and new_status == 'available':
            app_number = current_slot.get('application_number')
            if app_number:
                print(f" Slot {slot_id} is now available but still has customer data from application {app_number}")
        
        return jsonify({
            'success': True, 
            'message': f'Slot status updated to {new_status}',
            'customer_data': {
                'customer_name': current_slot.get('customer_name'),
                'customer_phone': current_slot.get('customer_phone'),
                'application_number': current_slot.get('application_number'),
                'installation_date': str(current_slot.get('installation_date')) if current_slot.get('installation_date') else None
            }
        })
        
    except Exception as e:
        print(f" Error updating slot status: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
    
# ===============================
# CREATE TABLES (Run once)
# ===============================
def create_napbox_tables():
    """Create NAP boxes and slots tables if not exists"""
    try:
        # CREATE TABLE napboxes - WALANG DATA INSERT
        napbox_query = """
            CREATE TABLE IF NOT EXISTS napboxes (
                id INT AUTO_INCREMENT PRIMARY KEY,
                napbox_name VARCHAR(100) NOT NULL,
                location VARCHAR(255),
                latitude DECIMAL(10, 8),
                longitude DECIMAL(11, 8),
                area VARCHAR(100),
                barangay VARCHAR(100),
                coverage_radius INT DEFAULT 500,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
        """
        execute_query(napbox_query)
        
        # CREATE TABLE napbox_slots - WALANG DATA INSERT
        slots_query = """
            CREATE TABLE IF NOT EXISTS napbox_slots (
                id INT AUTO_INCREMENT PRIMARY KEY,
                slot_number VARCHAR(10) NOT NULL,
                status ENUM('available', 'occupied') DEFAULT 'available',
                napbox_id INT NOT NULL,
                barangay VARCHAR(100),
                customer_name VARCHAR(200),
                customer_phone VARCHAR(50),
                installation_date DATE,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (napbox_id) REFERENCES napboxes(id) ON DELETE CASCADE,
                UNIQUE KEY unique_slot_per_napbox (napbox_id, slot_number)
            )
        """
        execute_query(slots_query)
        
        print(" NAP box tables ready - NO DEFAULT DATA INSERTED")
        
    except Exception as e:
        print(f"Error creating NAP box tables: {e}")

# Call initialization
create_napbox_tables()

# ===============================
# API: DELETE NAP BOX
# ===============================
@app.route("/api/technician/napbox/<int:napbox_id>", methods=["DELETE"])
def delete_napbox(napbox_id):
    """Delete a NAP box and all its slots"""
    try:
        print(f" Deleting NAP box ID: {napbox_id}")
        
        # Get napbox info first (including area)
        check_query = "SELECT id, napbox_name, area FROM napboxes WHERE id = %s"
        napbox = execute_query(check_query, (napbox_id,), fetch_one=True)
        
        if not napbox:
            return jsonify({"error": "NAP Box not found"}), 404
        
        napbox_name = napbox.get('napbox_name')
        area = napbox.get('area')
        
        # Delete slots first
        delete_slots_query = "DELETE FROM napbox_slots WHERE napbox_id = %s"
        execute_query(delete_slots_query, (napbox_id,))
        print(f"   - Deleted slots for NAP box {napbox_id}")
        
        # Delete the napbox
        delete_query = "DELETE FROM napboxes WHERE id = %s"
        execute_query(delete_query, (napbox_id,))
        print(f"   - Deleted NAP box {napbox_id}")
        
        # ========== UPDATE area_mapping napbox_count (decrement) ==========
        if area:
            # Get updated count of napboxes for this area
            count_query = "SELECT COUNT(*) as count FROM napboxes WHERE LOWER(area) = LOWER(%s)"
            count_result = execute_query(count_query, (area,), fetch_one=True)
            new_count = count_result.get('count', 0) if count_result else 0
            
            # Get total slots for this area
            slots_count_query = """
                SELECT COUNT(*) as total_slots 
                FROM napbox_slots s
                JOIN napboxes n ON s.napbox_id = n.id
                WHERE LOWER(n.area) = LOWER(%s)
            """
            slots_result = execute_query(slots_count_query, (area,), fetch_one=True)
            total_slots = slots_result.get('total_slots', 0) if slots_result else 0
            
            # Update area_mapping
            update_area_query = """
                UPDATE area_mapping 
                SET napbox_count = %s, total_slots = %s, updated_at = NOW()
                WHERE LOWER(area_name) = LOWER(%s)
            """
            execute_query(update_area_query, (new_count, total_slots, area))
            print(f" Updated area_mapping for '{area}': napbox_count={new_count}, total_slots={total_slots}")
        
        return jsonify({
            "success": True,
            "message": f"NAP Box '{napbox_name}' deleted successfully"
        })
        
    except Exception as e:
        print(f" Error deleting NAP box: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/technician/napbox/delete", methods=["POST"])
def delete_napbox_post():
    """Delete NAP box via POST (fallback)"""
    try:
        data = request.get_json()
        napbox_id = data.get("napbox_id")
        
        if not napbox_id:
            return jsonify({"error": "NAP Box ID is required"}), 400
        
        # Get napbox info first (including area)
        check_query = "SELECT id, napbox_name, area FROM napboxes WHERE id = %s"
        napbox = execute_query(check_query, (napbox_id,), fetch_one=True)
        
        if not napbox:
            return jsonify({"error": "NAP Box not found"}), 404
        
        napbox_name = napbox.get('napbox_name')
        area = napbox.get('area')
        
        execute_query("DELETE FROM napbox_slots WHERE napbox_id = %s", (napbox_id,))
        execute_query("DELETE FROM napboxes WHERE id = %s", (napbox_id,))
        
        # ========== UPDATE area_mapping napbox_count (decrement) ==========
        if area:
            count_query = "SELECT COUNT(*) as count FROM napboxes WHERE LOWER(area) = LOWER(%s)"
            count_result = execute_query(count_query, (area,), fetch_one=True)
            new_count = count_result.get('count', 0) if count_result else 0
            
            slots_count_query = """
                SELECT COUNT(*) as total_slots 
                FROM napbox_slots s
                JOIN napboxes n ON s.napbox_id = n.id
                WHERE LOWER(n.area) = LOWER(%s)
            """
            slots_result = execute_query(slots_count_query, (area,), fetch_one=True)
            total_slots = slots_result.get('total_slots', 0) if slots_result else 0
            
            update_area_query = """
                UPDATE area_mapping 
                SET napbox_count = %s, total_slots = %s, updated_at = NOW()
                WHERE LOWER(area_name) = LOWER(%s)
            """
            execute_query(update_area_query, (new_count, total_slots, area))
            print(f" Updated area_mapping for '{area}': napbox_count={new_count}, total_slots={total_slots}")
        
        print(f" NAP Box '{napbox_name}' deleted via POST fallback")
        
        return jsonify({
            "success": True,
            "message": f"NAP Box '{napbox_name}' deleted successfully"
        })
        
    except Exception as e:
        print(f" Error deleting NAP box (POST): {e}")
        return jsonify({"error": str(e)}), 500
        


# ===============================
# TECHNICIAN SLOT ASSIGNMENT PAGE
# ===============================
@app.route("/technician/slot-assignments")
def technician_slot_assignments():
    """Technician page for assigning slots to approved customers"""
    return render_template("technician-slot-assignments.html")


# ===============================
# TECHNICIAN GET PENDING CUSTOMER
# ===============================
@app.route("/api/technician/pending-customers", methods=["GET"])
def get_pending_customers_for_technician():
    """Get ALL approved customers assigned to technician's team"""
    try:
        technician_id = request.args.get("technician_id")
        
        if not technician_id:
            return jsonify({"error": "Technician ID required"}), 400
        
        # Get technician's team_id
        tech_query = "SELECT team_id, area FROM technicians WHERE technician_id = %s LIMIT 1"
        tech_result = execute_query(tech_query, (technician_id,), fetch_one=True)
        
        if not tech_result:
            return jsonify({"error": "Technician not found"}), 404
        
        team_id = tech_result.get('team_id')
        technician_area = tech_result.get('area', '')
        
        print(f" Technician {technician_id} belongs to team: {team_id}")
        
        # If technician has no team, fallback to area-based
        if not team_id:
            print(f" Technician has no team, falling back to area-based: {technician_area}")
            query = """
                SELECT 
                    c.application_number, c.first_name, c.last_name, c.middle_name, 
                    c.suffix, c.email, c.mobile, c.address, c.barangay, c.city, 
                    c.plan, c.approval_date, c.contract_number, c.installation_status,
                    c.assigned_team_id,
                    c.installation_date,
                    c.latitude, c.longitude,
                    a.plan_speed, a.preferred_napbox_id, a.preferred_napbox_name,
                    a.installation_address, a.landmark,
                    ns.id as slot_id, ns.slot_number, ns.status as slot_status, nb.napbox_name as assigned_napbox
                FROM customers c
                JOIN applications a ON c.application_number = a.application_number
                LEFT JOIN users u ON u.application_number = c.application_number
                LEFT JOIN napbox_slots ns ON c.application_number = ns.application_number AND ns.status = 'occupied'
                LEFT JOIN napboxes nb ON ns.napbox_id = nb.id
                WHERE c.status = 'Approved'
                AND LOWER(c.city) = LOWER(%s)
                AND (a.is_archived = 0 OR a.is_archived IS NULL)
                AND (c.installation_status IS NULL OR c.installation_status NOT IN ('Cancelled', 'Terminated'))
                AND (u.status IS NULL OR u.status != 'Terminated' OR (u.status = 'Terminated' AND c.installation_status IN ('Pending', 'Slot Assigned', 'Ongoing')))
                ORDER BY 
                    CASE WHEN ns.id IS NULL THEN 0 ELSE 1 END,
                    c.approval_date ASC
            """
            params = [technician_area]
        else:
            # TEAM-BASED: Get customers assigned to this team
            query = """
                SELECT 
                    c.application_number, c.first_name, c.last_name, c.middle_name, 
                    c.suffix, c.email, c.mobile, c.address, c.barangay, c.city, 
                    c.plan, c.approval_date, c.contract_number, c.installation_status,
                    c.assigned_team_id,
                    c.installation_date,
                    c.latitude, c.longitude,
                    a.plan_speed, a.preferred_napbox_id, a.preferred_napbox_name,
                    a.installation_address, a.landmark,
                    ns.id as slot_id, ns.slot_number, ns.status as slot_status, nb.napbox_name as assigned_napbox
                FROM customers c
                JOIN applications a ON c.application_number = a.application_number
                LEFT JOIN users u ON u.application_number = c.application_number
                LEFT JOIN napbox_slots ns ON c.application_number = ns.application_number AND ns.status = 'occupied'
                LEFT JOIN napboxes nb ON ns.napbox_id = nb.id
                WHERE c.status = 'Approved'
                AND c.assigned_team_id = %s
                AND (a.is_archived = 0 OR a.is_archived IS NULL)
                AND (c.installation_status IS NULL OR c.installation_status NOT IN ('Cancelled', 'Terminated'))
                AND (u.status IS NULL OR u.status != 'Terminated' OR (u.status = 'Terminated' AND c.installation_status IN ('Pending', 'Slot Assigned', 'Ongoing')))
                ORDER BY 
                    CASE WHEN ns.id IS NULL THEN 0 ELSE 1 END,
                    c.approval_date ASC
            """
            params = [team_id]
        
        customers = execute_query(query, params, fetch=True) or []
        
        # I-format ang response
        customers_list = []
        for cust in customers:
            customers_list.append({
                "application_number": cust.get('application_number'),
                "first_name": cust.get('first_name'),
                "last_name": cust.get('last_name'),
                "middle_name": cust.get('middle_name'),
                "suffix": cust.get('suffix'),
                "email": cust.get('email'),
                "mobile": cust.get('mobile'),
                "address": cust.get('address'),
                "barangay": cust.get('barangay'),
                "city": cust.get('city'),
                "plan": cust.get('plan'),
                "approval_date": cust.get('approval_date'),
                "contract_number": cust.get('contract_number'),
                "installation_status": cust.get('installation_status'),
                "assigned_team_id": cust.get('assigned_team_id'),
                "installation_date": cust.get('installation_date'),
                "latitude": cust.get('latitude'),      # IDINAGDAG
                "longitude": cust.get('longitude'),    # IDINAGDAG
                "plan_speed": cust.get('plan_speed'),
                "preferred_napbox_id": cust.get('preferred_napbox_id'),
                "preferred_napbox_name": cust.get('preferred_napbox_name'),
                "installation_address": cust.get('installation_address'),
                "landmark": cust.get('landmark'),
                "assigned_slot": {
                    "id": cust.get('slot_id'),
                    "slot_number": cust.get('slot_number'),
                    "status": cust.get('slot_status'),
                    "napbox_name": cust.get('assigned_napbox')
                } if cust.get('slot_id') else None
            })
        
        print(f" Found {len(customers_list)} customers for team {team_id or technician_area}")
        
        return jsonify({
            "customers": customers_list,
            "technician_area": technician_area,
            "team_id": team_id,
            "total": len(customers_list)
        })
        
    except Exception as e:
        print("Error fetching customers:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# TECHNICIAN UGET AVAILABLE NAPBOX
# ===============================
@app.route("/api/technician/available-napboxes", methods=["GET"])
def get_available_napboxes_for_technician():
    """Get NAP boxes in technician's area with available slots"""
    try:
        technician_id = request.args.get("technician_id")
        
        if not technician_id:
            return jsonify({"error": "Technician ID required"}), 400
        
        # Get technician's assigned area
        tech_query = "SELECT area FROM technicians WHERE technician_id = %s LIMIT 1"
        tech_result = execute_query(tech_query, (technician_id,), fetch_one=True)
        
        if not tech_result:
            return jsonify({"error": "Technician not found"}), 404
        
        technician_area = tech_result.get('area', '')
        
        # Get NAP boxes in the same area (case-insensitive) with latitude, longitude, and barangay
        query = """
            SELECT 
                nb.id, nb.napbox_name, nb.barangay, nb.area, nb.coverage_radius,
                nb.latitude, nb.longitude,
                (SELECT COUNT(*) FROM napbox_slots ns 
                 WHERE ns.napbox_id = nb.id AND ns.status = 'available') as available_slots
            FROM napboxes nb
            WHERE LOWER(nb.area) = LOWER(%s)
            AND EXISTS (
                SELECT 1 FROM napbox_slots ns 
                WHERE ns.napbox_id = nb.id AND ns.status = 'available'
            )
            ORDER BY nb.barangay, nb.napbox_name
        """
        napboxes = execute_query(query, (technician_area,), fetch=True) or []
        
        print(f" Found {len(napboxes)} NAP boxes in area: {technician_area}")
        for nb in napboxes:
            print(f"   - {nb.get('napbox_name')} | Barangay: {nb.get('barangay')} | Slots: {nb.get('available_slots')}")
        
        return jsonify({"napboxes": napboxes})
        
    except Exception as e:
        print("Error fetching available napboxes:", e)
        return jsonify({"error": str(e)}), 500


# ===============================
# TECHNICIAN ASSIGN SLOT
# ===============================
@app.route("/api/technician/assign-slot-to-customer", methods=["POST"])
def assign_slot_to_customer():
    """Assign a specific slot to a customer (technician action)"""
    try:
        data = request.get_json()
        
        application_number = data.get("application_number")
        slot_id = data.get("slot_id")
        napbox_id = data.get("napbox_id")
        installation_date = data.get("installation_date")
        technician_id = data.get("technician_id")
        
        print("=" * 60)
        print(" ASSIGN SLOT - START")
        print(f"   Application: {application_number}")
        print(f"   Slot ID: {slot_id}")
        print(f"   Technician ID: {technician_id}")
        print("=" * 60)
        
        if not application_number or not slot_id:
            return jsonify({"error": "Application number and slot ID are required"}), 400
        
        # Get technician's team
        tech_query = "SELECT team_id FROM technicians WHERE technician_id = %s LIMIT 1"
        tech_result = execute_query(tech_query, (technician_id,), fetch_one=True)
        
        if not tech_result:
            return jsonify({"error": "Technician not found"}), 404
        
        technician_team_id = tech_result.get('team_id')
        
        # Get customer details (including assigned_team_id)
        customer_query = """
            SELECT first_name, last_name, mobile, email, barangay, city, contract_number, assigned_team_id
            FROM customers 
            WHERE application_number = %s
        """
        customer = execute_query(customer_query, (application_number,), fetch_one=True)
        
        if not customer:
            return jsonify({"error": "Customer not found"}), 404
        
        customer_team_id = customer.get('assigned_team_id')
        
        # Verify that technician's team matches customer's assigned team
        if technician_team_id and customer_team_id and technician_team_id != customer_team_id:
            print(f" Team mismatch: Technician team={technician_team_id}, Customer team={customer_team_id}")
            return jsonify({"error": "You are not authorized to assign slots for this customer. This customer belongs to a different team."}), 403
        
        customer_name = f"{customer.get('first_name', '')} {customer.get('last_name', '')}".strip()
        customer_phone = customer.get('mobile', '')
        customer_city = customer.get('city', '')
        contract_number = customer.get('contract_number', '')
        
        print(f" Customer Name: {customer_name}")
        print(f" Customer Team: {customer_team_id}")
        print(f" Technician Team: {technician_team_id}")
        print(f" Contract Number: '{contract_number}'")
        
        # Check if slot is still available
        check_slot_query = """
            SELECT id, status, slot_number, napbox_id
            FROM napbox_slots 
            WHERE id = %s AND status = 'available'
        """
        slot = execute_query(check_slot_query, (slot_id,), fetch_one=True)
        
        if not slot:
            return jsonify({"error": "Slot is no longer available"}), 400
        
        # ========== UPDATE NAPBOX SLOTS TABLE ==========
        update_slot_query = """
            UPDATE napbox_slots 
            SET status = 'occupied',
                customer_name = %s,
                customer_phone = %s,
                application_number = %s,
                contract_number = %s,
                installation_date = %s,
                updated_at = NOW()
            WHERE id = %s
        """
        execute_query(update_slot_query, (
            customer_name, 
            customer_phone, 
            application_number, 
            contract_number,
            installation_date or ph_now().strftime("%Y-%m-%d"), 
            slot_id
        ))
        print(f" Slot {slot['slot_number']} assigned to {customer_name} (Contract: {contract_number})")
        
        # ========== UPDATE CUSTOMERS TABLE - Installation Status to "Slot Assigned" ==========
        try:
            update_customer_query = """
                UPDATE customers 
                SET installation_status = 'Slot Assigned'
                WHERE application_number = %s
            """
            execute_query(update_customer_query, (application_number,))
            print(f" Customer {application_number} installation_status updated to 'Slot Assigned'")
        except Exception as cust_err:
            print(f" Error updating customer installation_status: {cust_err}")
            # Continue even if this fails - the slot assignment is the main action
        
        # ========== UPDATE APPLICATIONS TABLE ==========
        try:
            update_app_query = """
                UPDATE applications 
                SET installation_status = 'Slot Assigned'
                WHERE application_number = %s
            """
            execute_query(update_app_query, (application_number,))
            print(f" Application {application_number} installation_status updated to 'Slot Assigned'")
        except Exception as app_err:
            print(f" Error updating application installation_status: {app_err}")
        
        # ========== CREATE NOTIFICATION FOR ADMIN (by city/area) ==========
        try:
            print(f" Looking for admin with area: '{customer_city}'")
            
            admin_query = """
                SELECT admin_id, username, area 
                FROM admins 
                WHERE UPPER(area) = UPPER(%s)
                LIMIT 1
            """
            city_admin = execute_query(admin_query, (customer_city,), fetch_one=True)
            print(f" Admin found: {city_admin}")
            
            if city_admin:
                admin_id = city_admin.get('admin_id')
                admin_username = city_admin.get('username')
                admin_notification_id = int(datetime.now().timestamp() * 1000) + 1
                
                print(f" Creating admin notification for admin_id: {admin_id} ({admin_username})")
                print(f" Admin area: {customer_city}")
                
                admin_notif_query = """
                    INSERT INTO admin_notifications 
                    (id, admin_id, admin_city, title, message, type, relatedId, 
                     timestamp, read_status, application_city, application_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                execute_query(admin_notif_query, (
                    admin_notification_id,
                    admin_id,
                    customer_city,
                    "Slot Assigned",
                    f"Slot #{slot['slot_number']} assigned to {customer_name} (Application: {application_number}, Contract: {contract_number})",
                    "slot_assigned",
                    application_number,
                    ph_now_iso(),
                    0,
                    customer_city,
                    application_number
                ))
                print(f" Admin notification created for {admin_id} in {customer_city}")
            else:
                print(f" No admin found for area: {customer_city}")
                print(f" Tip: Insert an admin with area = '{customer_city}' into admins table")
                
        except Exception as admin_err:
            print(f" Admin notification error: {admin_err}")
            import traceback
            traceback.print_exc()
        
        # ========== CREATE NOTIFICATION FOR SUPERADMIN ==========
        try:
            notification_id = int(datetime.now().timestamp() * 1000)
            notification_query = """
                INSERT INTO notifications (id, title, message, type, relatedId, timestamp, read_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            execute_query(notification_query, (
                notification_id,
                "Slot Assigned",
                f"Slot #{slot['slot_number']} has been assigned to {customer_name} (Application: {application_number}, Contract: {contract_number}). Installation scheduled.",
                "slot_assigned",
                application_number,
                ph_now_iso(),
                0
            ))
            print(f" Superadmin notification created")
        except Exception as notif_err:
            print(f"Notification error: {notif_err}")
        
        return jsonify({
            "success": True,
            "message": f"Slot #{slot['slot_number']} assigned to {customer_name}",
            "slot_number": slot['slot_number']
        })
        
    except Exception as e:
        print("Error assigning slot:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# TECHNICIAN CANCEL INSTALLATION
# ===============================
@app.route("/api/technician/cancel-installation", methods=["POST"])
def cancel_installation():
    """Cancel an installation - archives the application and marks it cancelled"""
    try:
        data = request.get_json()

        application_number = data.get("application_number")
        technician_id = data.get("technician_id")
        reason = (data.get("reason") or "").strip()

        print("=" * 60)
        print(" CANCEL INSTALLATION - START")
        print(f"   Application: {application_number}")
        print(f"   Technician ID: {technician_id}")
        print("=" * 60)

        if not application_number or not technician_id:
            return jsonify({"error": "Application number and technician ID are required"}), 400

        # Get technician's team
        tech_query = "SELECT team_id FROM technicians WHERE technician_id = %s LIMIT 1"
        tech_result = execute_query(tech_query, (technician_id,), fetch_one=True)

        if not tech_result:
            return jsonify({"error": "Technician not found"}), 404

        technician_team_id = tech_result.get('team_id')

        # Get customer details
        customer_query = """
            SELECT first_name, last_name, contract_number, assigned_team_id, city, installation_status
            FROM customers
            WHERE application_number = %s
        """
        customer = execute_query(customer_query, (application_number,), fetch_one=True)

        if not customer:
            return jsonify({"error": "Customer not found"}), 404

        customer_team_id = customer.get('assigned_team_id')

        # Same team check gaya ng assign-slot
        if technician_team_id and customer_team_id and technician_team_id != customer_team_id:
            return jsonify({"error": "You are not authorized to cancel this customer's installation. This customer belongs to a different team."}), 403

        customer_name = f"{customer.get('first_name', '')} {customer.get('last_name', '')}".strip()
        contract_number = customer.get('contract_number', '')
        customer_city = customer.get('city', '')
        current_status = customer.get('installation_status', 'Pending')

        # Optional guard: huwag na ulit i-cancel kung Installed/completed na
        if current_status == 'Installed':
            return jsonify({"error": "This installation is already completed and cannot be cancelled."}), 400

        # KUNIN ANG CURRENT DATETIME PARA SA CANCELLATION DATE
        from datetime import datetime
        cancellation_date = ph_now().strftime("%Y-%m-%d %H:%M:%S")
        cancellation_date_display = ph_now().strftime("%B %d, %Y at %I:%M %p")

        print(f" Cancellation date: {cancellation_date}")

        # ========== RELEASE THE SLOT (kung may naka-assign na) ==========
        try:
            occupied_slot_query = """
                SELECT id, slot_number
                FROM napbox_slots
                WHERE application_number = %s AND status = 'occupied'
                LIMIT 1
            """
            occupied_slot = execute_query(occupied_slot_query, (application_number,), fetch_one=True)

            if occupied_slot:
                release_slot_query = """
                    UPDATE napbox_slots
                    SET status = 'available',
                        customer_name = NULL,
                        customer_phone = NULL,
                        application_number = NULL,
                        contract_number = NULL,
                        installation_date = NULL,
                        updated_at = NOW()
                    WHERE id = %s
                """
                execute_query(release_slot_query, (occupied_slot['id'],))
                print(f" Slot {occupied_slot['slot_number']} released back to available")
        except Exception as slot_err:
            print(f" Error releasing slot: {slot_err}")
            # Continue pa rin - hindi dapat mag-block sa cancellation

        # ========== UPDATE APPLICATIONS TABLE ==========
        update_app_query = """
            UPDATE applications
            SET status = 'Cancelled',
                installation_status = 'Cancelled',
                is_archived = 1,
                rejection_reason = %s
            WHERE application_number = %s
        """
        execute_query(update_app_query, (reason if reason else 'Cancelled by technician', application_number))
        print(f" Application {application_number} status = Cancelled, is_archived = 1, reason = {reason}")

        # ========== UPDATE CUSTOMERS TABLE WITH date_cancelled ==========
        update_customer_query = """
            UPDATE customers
            SET status = 'Cancelled',
                installation_status = 'Cancelled',
                date_cancelled = %s
            WHERE application_number = %s
        """
        execute_query(update_customer_query, (cancellation_date, application_number))
        print(f" Customer {application_number} installation_status = Cancelled, date_cancelled = {cancellation_date}")

        # ========== NOTIFY ADMIN (by city/area) ==========
        try:
            admin_query = """
                SELECT admin_id, username FROM admins
                WHERE UPPER(area) = UPPER(%s) LIMIT 1
            """
            city_admin = execute_query(admin_query, (customer_city,), fetch_one=True)

            if city_admin:
                admin_notif_id = int(datetime.now().timestamp() * 1000) + 3
                msg = f"Installation for {customer_name} (Application: {application_number}, Contract: {contract_number}) has been cancelled by technician on {cancellation_date_display}."
                if reason:
                    msg += f" Reason: {reason}"

                admin_notif_query = """
                    INSERT INTO admin_notifications
                    (id, admin_id, admin_city, title, message, type, relatedId,
                     timestamp, read_status, application_city, application_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                execute_query(admin_notif_query, (
                    admin_notif_id,
                    city_admin.get('admin_id'),
                    customer_city,
                    "Installation Cancelled",
                    msg,
                    "installation_cancelled",
                    application_number,
                    ph_now_iso(),
                    0,
                    customer_city,
                    application_number
                ))
        except Exception as admin_err:
            print(f" Admin notification error: {admin_err}")

        # ========== NOTIFY SUPERADMIN ==========
        try:
            notification_id = int(datetime.now().timestamp() * 1000) + 4
            msg = f"Installation for {customer_name} (Application: {application_number}) was cancelled on {cancellation_date_display}."
            if reason:
                msg += f" Reason: {reason}"
            notification_query = """
                INSERT INTO notifications (id, title, message, type, relatedId, timestamp, read_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            execute_query(notification_query, (
                notification_id,
                "Installation Cancelled",
                msg,
                "installation_cancelled",
                application_number,
                ph_now_iso(),
                0
            ))
        except Exception as notif_err:
            print(f"Notification error: {notif_err}")

        return jsonify({
            "success": True,
            "message": f"Installation for {customer_name} has been cancelled on {cancellation_date_display}."
        })

    except Exception as e:
        print("Error cancelling installation:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ===============================
# TECHNICIAN GET AVAILABLE SLOT FOR TECH
# ===============================
@app.route("/api/technician/available-slots-for-napbox/<int:napbox_id>", methods=["GET"])
def get_available_slots_for_technician(napbox_id):
    """Get available slots for a specific NAP box"""
    try:
        query = """
            SELECT id, slot_number, status, napbox_id
            FROM napbox_slots
            WHERE napbox_id = %s AND status = 'available'
            ORDER BY CAST(slot_number AS UNSIGNED)
        """
        slots = execute_query(query, (napbox_id,), fetch=True) or []
        
        return jsonify({"slots": slots})
        
    except Exception as e:
        print("Error fetching available slots:", e)
        return jsonify({"error": str(e)}), 500


@app.route("/api/technician/debug-customer/<app_number>", methods=["GET"])
def debug_customer(app_number):
    """Debug endpoint to check customer and slot status"""
    try:
        # Check customer
        customer_query = "SELECT * FROM customers WHERE application_number = %s"
        customer = execute_query(customer_query, (app_number,), fetch_one=True)
        
        # Check if slot exists
        slot_query = "SELECT * FROM napbox_slots WHERE application_number = %s"
        slot = execute_query(slot_query, (app_number,), fetch_one=True)
        
        # Check technician
        tech_query = "SELECT * FROM technicians"
        techs = execute_query(tech_query, fetch=True)
        
        return jsonify({
            "customer": customer,
            "slot_assigned": slot,
            "technicians": techs
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ===============================
# TECHNICIAN UPDATE INSTALLATION STATUS
# ===============================
@app.route("/api/technician/update-installation-status", methods=["POST"])
def technician_update_installation_status():
    """Update installation status for a customer (technician action)"""
    try:
        data = request.get_json()
        
        application_number = data.get("application_number")
        new_status = data.get("installation_status")
        technician_id = data.get("technician_id")
        
        if not application_number or not new_status:
            return jsonify({"error": "Application number and status are required"}), 400
        
        if new_status not in ["Ongoing", "Installed"]:
            return jsonify({"error": "Invalid status. Allowed: Ongoing, Installed"}), 400
        
        # Get technician's team
        tech_query = "SELECT team_id FROM technicians WHERE technician_id = %s LIMIT 1"
        tech_result = execute_query(tech_query, (technician_id,), fetch_one=True)
        
        if not tech_result:
            return jsonify({"error": "Technician not found"}), 404
        
        technician_team_id = tech_result.get('team_id')
        
        current_time = ph_now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Check if customer exists and has a slot assigned
        customer_query = """
            SELECT c.*, ns.id as slot_id, ns.slot_number
            FROM customers c
            LEFT JOIN napbox_slots ns ON c.application_number = ns.application_number AND ns.status = 'occupied'
            WHERE c.application_number = %s
        """
        customer_data = execute_query(customer_query, (application_number,), fetch_one=True)
        
        if not customer_data:
            return jsonify({"error": "Customer not found"}), 404
        
        customer_team_id = customer_data.get('assigned_team_id')
        
        # Verify that technician's team matches customer's assigned team
        if technician_team_id and customer_team_id and technician_team_id != customer_team_id:
            print(f" Team mismatch: Technician team={technician_team_id}, Customer team={customer_team_id}")
            return jsonify({"error": "You are not authorized to update this customer. This customer belongs to a different team."}), 403
        
        # For 'Ongoing' status, check if slot is assigned
        if new_status == "Ongoing" and not customer_data.get('slot_id'):
            return jsonify({"error": "Cannot set to Ongoing. Please assign a slot first."}), 400
        
        # Update customers table
        update_fields = ["installation_status = %s"]
        params = [new_status]
        
        if new_status == "Ongoing":
            update_fields.append("date_ongoing = %s")
            params.append(current_time)
        elif new_status == "Installed":
            update_fields.append("date_installed = %s")
            params.append(current_time)
            # Awtomatikong i-update ang users table kapag Installed na (status Active / Connected)
            # Kasama na rin ang pag-clear ng has_pending_reconnect at reconnect_requested_at
            # dahil dito lang talaga natatapos ang buong reconnection process (may bagong
            # na-assign nang slot ang technician). Bago nito, may nakabinbing reconnect pa rin
            # ang user kahit na-approve na ng superadmin, kaya dapat naka-lock pa ang
            # "Request Reconnect" button sa user dashboard hanggang dito.
            try:
                # Kunin muna ang user info AT ang dating status BAGO i-update,
                # para malaman natin kung ito ba ay "reconnect after slot
                # reassignment" (dating Terminated) o normal na first-time
                # install, at para may makuhang user_id/email/name para sa notif.
                user_before = execute_query(
                    """SELECT user_id, email, first_name, last_name, status
                       FROM users WHERE application_number = %s LIMIT 1""",
                    (application_number,), fetch_one=True
                )

                execute_query(
                    """UPDATE users 
                       SET status = 'Active', 
                           connection_status = 'Connected',
                           has_pending_reconnect = 0,
                           reconnect_requested_at = NULL
                       WHERE application_number = %s""",
                    (application_number,)
                )
                print(f" User {application_number} status updated to Active / Connected, has_pending_reconnect cleared.")

                # ========== NOTIFY THE USER (reconnect after slot reassignment) ==========
                if user_before:
                    try:
                        was_terminated = user_before.get('status') == 'Terminated'
                        target_user_id = user_before.get('user_id')
                        target_user_email = user_before.get('email') or ''
                        target_user_name = f"{user_before.get('first_name', '')} {user_before.get('last_name', '')}".strip() or 'User'

                        if was_terminated:
                            notif_title = "Reconnection Complete"
                            notif_message = "Great news! Your NAP Box slot has been reassigned and your reconnection is now complete. Your account is active and your internet connection has been restored."
                        else:
                            notif_title = "Installation Complete"
                            notif_message = "Your installation is now complete. Your account is active and your internet connection is ready to use."

                        user_notif_id = int(datetime.now().timestamp() * 1000) + 2
                        user_notif_query = """
                            INSERT INTO user_notifications 
                            (id, title, message, type, relatedId, user_id, user_email, user_name, connection_status, timestamp, read_status)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """
                        execute_query(user_notif_query, (
                            user_notif_id,
                            notif_title,
                            notif_message,
                            "reconnection_complete" if was_terminated else "installation_complete",
                            application_number,
                            target_user_id,
                            target_user_email,
                            target_user_name,
                            "Connected",
                            ph_now_iso(),
                            0
                        ))
                        print(f" User notification sent to {target_user_id} ({'reconnection' if was_terminated else 'installation'} complete)")
                    except Exception as user_notif_err:
                        print(f" Error creating user notification: {user_notif_err}")
                        import traceback
                        traceback.print_exc()

            except Exception as user_act_err:
                print(f" Error activating user status on installation: {user_act_err}")

        
        params.append(application_number)
        update_query = f"UPDATE customers SET {', '.join(update_fields)} WHERE application_number = %s"
        execute_query(update_query, params)
        print(f" Technician updated customer {application_number} to {new_status}")
        
        # Update applications table
        app_update_query = "UPDATE applications SET installation_status = %s WHERE application_number = %s"
        execute_query(app_update_query, (new_status, application_number))
        
        customer_name = f"{customer_data.get('first_name', '')} {customer_data.get('last_name', '')}".strip()
        customer_city = customer_data.get('city', '')
        slot_number = customer_data.get('slot_number', '')
        
        status_text = "started" if new_status == "Ongoing" else "completed"
        
        # ========== CREATE NOTIFICATION FOR ADMIN (by city/area) - FIXED ==========
        try:
            print(f" Looking for admin with area: '{customer_city}'")
            
            admin_query = """
                SELECT admin_id, username, area 
                FROM admins 
                WHERE UPPER(area) = UPPER(%s)
                LIMIT 1
            """
            city_admin = execute_query(admin_query, (customer_city,), fetch_one=True)
            print(f" Admin found: {city_admin}")
            
            if city_admin:
                admin_id = city_admin.get('admin_id')
                admin_notification_id = int(datetime.now().timestamp() * 1000) + 1
                
                admin_notif_query = """
                    INSERT INTO admin_notifications 
                    (id, admin_id, admin_city, title, message, type, relatedId, 
                     timestamp, read_status, application_city, application_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                execute_query(admin_notif_query, (
                    admin_notification_id,
                    admin_id,
                    customer_city,
                    f"Installation {new_status}",
                    f"Installation {status_text} for {customer_name} (Application: {application_number}) - Slot #{slot_number}",
                    "installation_update",
                    application_number,
                    ph_now_iso(),
                    0,
                    customer_city,
                    application_number
                ))
                print(f" Admin notification created for {admin_id} in {customer_city}")
            else:
                print(f" No admin found for area: {customer_city}")
                
        except Exception as admin_err:
            print(f" Admin notification error: {admin_err}")
            import traceback
            traceback.print_exc()
        
        # ========== CREATE NOTIFICATION FOR SUPERADMIN ==========
        try:
            notification_id = int(datetime.now().timestamp() * 1000)
            notification_query = """
                INSERT INTO notifications (id, title, message, type, relatedId, timestamp, read_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            execute_query(notification_query, (
                notification_id,
                f"Installation {new_status}",
                f"Installation status for {customer_name} (Application: {application_number}) has been updated to {new_status}.",
                "installation_update",
                application_number,
                ph_now_iso(),
                0
            ))
            print(f" Superadmin notification created")
        except Exception as notif_err:
            print(f"Notification error: {notif_err}")
        
        return jsonify({
            "success": True,
            "message": f"Installation status updated to {new_status}",
            "status": new_status
        })
        
    except Exception as e:
        print("Error updating installation status:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def create_technician_notification(technician_id, title, message, notif_type, related_id=None, application_number=None, customer_name=None):
    """Create a notification for a specific technician"""
    try:
        import time
        notification_id = int(time.time() * 1000)
        
        # Get technician's area
        tech_query = "SELECT area FROM technicians WHERE technician_id = %s"
        tech = execute_query(tech_query, (technician_id,), fetch_one=True)
        technician_area = tech.get('area') if tech else None    
        
        query = """
            INSERT INTO technician_notifications 
            (id, technician_id, technician_area, title, message, type, relatedId, 
             application_number, customer_name, timestamp, read_status, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, NOW())
        """
        execute_query(query, (
            notification_id, technician_id, technician_area, title, message, 
            notif_type, related_id, application_number, customer_name,
            ph_now_iso()
        ))
        
        print(f" Technician notification created for {technician_id}: {title}")
        return True
        
    except Exception as e:
        print(f" Error creating technician notification: {e}")
        return False


def create_technician_notifications_by_area(area, title, message, notif_type, related_id=None, application_number=None, customer_name=None):
    """Create notifications for ALL technicians in a specific area"""
    try:
        import time
        
        # Get all technicians in this area
        tech_query = "SELECT technician_id FROM technicians WHERE UPPER(area) = UPPER(%s) AND status = 'Active'"
        technicians = execute_query(tech_query, (area,), fetch=True) or []
        
        print(f" Found {len(technicians)} technicians in area: {area}")
        
        if not technicians:
            print(f" No active technicians found in area: {area}")
            return False
        
        success_count = 0
        for tech in technicians:
            technician_id = tech.get('technician_id')
            if technician_id:
                notification_id = int(time.time() * 1000) + success_count
                
                query = """
                    INSERT INTO technician_notifications 
                    (id, technician_id, technician_area, title, message, type, relatedId, 
                     application_number, customer_name, timestamp, read_status, created_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, NOW())
                """
                execute_query(query, (
                    notification_id, technician_id, area, title, message, 
                    notif_type, related_id, application_number, customer_name,
                    ph_now_iso()
                ))
                success_count += 1
        
        print(f" Created {success_count} technician notifications in area: {area}")
        return True
        
    except Exception as e:
        print(f" Error creating technician notifications by area: {e}")
        return False



# ==================== TECHNICIAN NOTIFICATION API ROUTES ====================

@app.route("/api/technician/notifications", methods=["GET"])
def get_technician_notifications():
    """Get notifications for a specific technician"""
    try:
        technician_id = request.args.get("technician_id")
        
        if not technician_id:
            return jsonify({"error": "Technician ID required"}), 400
        
        print(f" Fetching notifications for technician: {technician_id}")
        
        query = """
            SELECT id, title, message, type, relatedId, application_number, 
                   customer_name, timestamp, read_status, created_at
            FROM technician_notifications
            WHERE technician_id = %s
            ORDER BY id DESC
            LIMIT 50
        """
        notifications = execute_query(query, (technician_id,), fetch=True) or []
        
        result = []
        for n in notifications:
            result.append({
                "id": n.get("id"),
                "title": n.get("title", "Notification"),
                "message": n.get("message", ""),
                "type": n.get("type", "info"),
                "relatedId": n.get("relatedId"),
                "application_number": n.get("application_number"),
                "customer_name": n.get("customer_name"),
                "timestamp": str(n.get("timestamp")),
                "read": n.get("read_status") == 1,
                "created_at": str(n.get("created_at"))
            })
        
        print(f" Found {len(result)} notifications for technician {technician_id}")
        return jsonify(result)
        
    except Exception as e:
        print(f" Error fetching technician notifications: {e}")
        return jsonify([]), 500


@app.route("/api/technician/notifications/<int:notification_id>/read", methods=["PATCH"])
def mark_technician_notification_read(notification_id):
    """Mark a single notification as read"""
    try:
        technician_id = request.args.get("technician_id")
        
        if not technician_id:
            data = request.get_json() or {}
            technician_id = data.get("technician_id")
        
        if not technician_id:
            return jsonify({"error": "Technician ID required"}), 400
        
        query = "UPDATE technician_notifications SET read_status = 1 WHERE id = %s AND read_status = 0"
        execute_query(query, (notification_id,))
        
        print(f" Technician {technician_id} marked notification {notification_id} as read")
        return jsonify({"message": "Notification marked as read"})
    
    except Exception as e:
        print(f"Error marking notification as read: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/technician/notifications/read-all", methods=["PUT"])
def mark_all_technician_notifications_read():
    """Mark all notifications for a technician as read"""
    try:
        technician_id = request.args.get("technician_id")
        
        if not technician_id:
            data = request.get_json() or {}
            technician_id = data.get("technician_id")
        
        if not technician_id:
            return jsonify({"error": "Technician ID required"}), 400
        
        query = "UPDATE technician_notifications SET read_status = 1 WHERE technician_id = %s AND read_status = 0"
        rows_affected = execute_query(query, (technician_id,))
        
        print(f" Technician {technician_id} marked {rows_affected} notifications as read")
        return jsonify({"message": f"Marked {rows_affected} notifications as read", "count": rows_affected})
    
    except Exception as e:
        print(f"Error marking all notifications as read: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/technician/notifications/unread/count", methods=["GET"])
def get_unread_technician_notification_count():
    """Get unread notification count for a technician"""
    try:
        technician_id = request.args.get("technician_id")
        
        if not technician_id:
            return jsonify({"error": "Technician ID required"}), 400
        
        query = "SELECT COUNT(*) as unread_count FROM technician_notifications WHERE technician_id = %s AND read_status = 0"
        result = execute_query(query, (technician_id,), fetch_one=True)
        
        unread_count = result.get("unread_count", 0) if result else 0
        
        return jsonify({"unread_count": unread_count})
    
    except Exception as e:
        print(f"Error getting unread count: {e}")
        return jsonify({"unread_count": 0}), 500


@app.route("/api/technician/update-slot", methods=["POST"])
def technician_update_slot():
    """Technician updates slot customer information and status"""
    try:
        data = request.get_json()

        slot_id = data.get("slot_id")
        customer_name = data.get("customer_name", "")
        contract_number = data.get("contract_number", "")
        customer_phone = data.get("customer_phone", "")
        status = data.get("status")

        if not slot_id:
            return jsonify({"error": "Slot ID required"}), 400

        current_query = """
            SELECT customer_name, contract_number, customer_phone, status
            FROM napbox_slots 
            WHERE id = %s
        """
        current_slot = execute_query(current_query, (slot_id,), fetch_one=True)

        if not current_slot:
            return jsonify({"error": "Slot not found"}), 404

        # CHECK: KUNG MAY CONTRACT NUMBER AT HINDI EMPTY
        if contract_number and contract_number.strip():
            # CHECK KUNG MAY IBANG SLOT NA GAMIT ANG CONTRACT NUMBER NA ITO (EXCLUDING CURRENT SLOT)
            check_query = """
                SELECT id, slot_number, customer_name 
                FROM napbox_slots 
                WHERE contract_number = %s AND id != %s
            """
            existing_slot = execute_query(check_query, (contract_number, slot_id), fetch_one=True)
            
            if existing_slot:
                return jsonify({
                    "error": f"Contract number '{contract_number}' is already used in Slot #{existing_slot['slot_number']} (Customer: {existing_slot.get('customer_name', 'Unknown')})"
                }), 400

        final_status = status if status is not None else ('occupied' if customer_name else 'available')

        # Look up new owner's application_number from customers if contract_number is passed
        app_num = data.get("application_number")
        if not app_num and contract_number and contract_number.strip():
            c_data = execute_query("SELECT application_number FROM customers WHERE contract_number = %s LIMIT 1", (contract_number.strip(),), fetch_one=True)
            if c_data:
                app_num = c_data.get('application_number')

        update_query = """
            UPDATE napbox_slots 
            SET customer_name = %s,
                contract_number = %s,
                customer_phone = %s,
                application_number = %s,
                status = %s,
                updated_at = NOW()
            WHERE id = %s
        """
        execute_query(update_query, (customer_name, contract_number, customer_phone, app_num, final_status, slot_id))


        print(f" [Technician] Slot {slot_id} updated: Status={final_status.upper()}")

        return jsonify({
            "success": True,
            "message": f"Slot updated successfully. Status: {final_status.upper()}",
            "status": final_status,
            "preserved_data": {
                "previous_customer": current_slot.get('customer_name'),
                "previous_contract": current_slot.get('contract_number')
            }
        })

    except Exception as e:
        print(f"Error updating slot (technician): {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/technician/check-contract-number', methods=['POST'])
def technician_check_contract_number():
    """Check if contract number already exists in napbox_slots table"""
    try:
        data = request.json
        contract_number = data.get('contract_number')
        exclude_slot_id = data.get('exclude_slot_id')  # Para ma-exclude ang kasalukuyang slot
        technician_id = data.get('technician_id')
        tab_id = data.get('tab_id')
        
        if not contract_number:
            return jsonify({'exists': False, 'error': 'Contract number is required'}), 400
        
        # I-CONFIGURE ANG DATABASE CONNECTION
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        # CHECK IF CONTRACT NUMBER EXISTS (EXCLUDING CURRENT SLOT)
        if exclude_slot_id:
            query = """
                SELECT id, slot_number, napbox_id, customer_name 
                FROM napbox_slots 
                WHERE contract_number = %s AND id != %s
            """
            cursor.execute(query, (contract_number, exclude_slot_id))
        else:
            query = """
                SELECT id, slot_number, napbox_id, customer_name 
                FROM napbox_slots 
                WHERE contract_number = %s
            """
            cursor.execute(query, (contract_number,))
        
        result = cursor.fetchone()
        cursor.close()
        db.close()
        
        if result:
            return jsonify({
                'exists': True,
                'slot_id': result['id'],
                'slot_number': result['slot_number'],
                'napbox_id': result['napbox_id'],
                'customer_name': result.get('customer_name', 'Unknown')
            })
        else:
            return jsonify({'exists': False})
            
    except Exception as e:
        print(f" Error checking contract number: {e}")
        return jsonify({'exists': False, 'error': str(e)}), 500


# ===============================
# TECHNICIAN DOWNLOAD APPLICATION PDF
# ===============================
@app.route('/technician/download/pdf/<app_id>')
def technician_download_pdf(app_id):
    """Technician download application PDF - checks if technician has access to this application"""
    try:
        # Get technician_id from session or request
        technician_id = request.args.get('technician_id') or session.get('technician_id')
        
        if not technician_id:
            return "Technician ID required", 400
        
        # Get application data
        query = "SELECT application_number, technician_id, team_id FROM applications WHERE application_number = %s"
        data = execute_query(query, (app_id,), fetch_one=True)
        
        if not data:
            return "Application not found", 404
        
        application_number = data.get("application_number")
        
        # VERIFY: Technician has access to this application
        # Check if technician is assigned to this application or belongs to the same team
        has_access = False
        
        # Check 1: Technician is directly assigned to this application
        if data.get('technician_id') == technician_id:
            has_access = True
        
        # Check 2: Technician belongs to the same team as the application
        if not has_access:
            # Get technician's team
            tech_query = "SELECT team_id FROM technicians WHERE technician_id = %s"
            tech_data = execute_query(tech_query, (technician_id,), fetch_one=True)
            
            if tech_data and tech_data.get('team_id'):
                tech_team_id = tech_data.get('team_id')
                app_team_id = data.get('team_id')
                
                if tech_team_id == app_team_id:
                    has_access = True
        
        # Check 3: Technician's area matches application's city
        if not has_access:
            tech_query = "SELECT area FROM technicians WHERE technician_id = %s"
            tech_data = execute_query(tech_query, (technician_id,), fetch_one=True)
            
            if tech_data and tech_data.get('area'):
                tech_area = tech_data.get('area').lower().strip()
                
                # Get application's city
                city_query = "SELECT city FROM applications WHERE application_number = %s"
                city_data = execute_query(city_query, (app_id,), fetch_one=True)
                
                if city_data and city_data.get('city'):
                    app_city = city_data.get('city').lower().strip()
                    
                    # Check if tech area matches app city
                    if tech_area == app_city:
                        has_access = True
        
        if not has_access:
            return "Access denied: You are not authorized to view this application", 403
        
        # Use existing PDF generator
        return download_pdf(application_number)
        
    except Exception as e:
        print(f"Technician PDF download error: {e}")
        import traceback
        traceback.print_exc()
        return str(e), 500





@app.route("/api/admin/update-slot", methods=["POST"])
def update_slot():
    """Admin updates slot customer information and status"""
    try:
        data = request.get_json()
        
        slot_id = data.get("slot_id")
        customer_name = data.get("customer_name", "")
        contract_number = data.get("contract_number", "")
        customer_phone = data.get("customer_phone", "")
        status = data.get("status")
        
        if not slot_id:
            return jsonify({"error": "Slot ID required"}), 400
        
        # GET CURRENT SLOT DATA
        current_query = """
            SELECT customer_name, contract_number, customer_phone, status
            FROM napbox_slots 
            WHERE id = %s
        """
        current_slot = execute_query(current_query, (slot_id,), fetch_one=True)
        
        if not current_slot:
            return jsonify({"error": "Slot not found"}), 404
        
        # CHECK: KUNG MAY CONTRACT NUMBER AT HINDI EMPTY
        if contract_number and contract_number.strip():
            # CHECK KUNG MAY IBANG SLOT NA GAMIT ANG CONTRACT NUMBER NA ITO (EXCLUDING CURRENT SLOT)
            check_query = """
                SELECT id, slot_number, customer_name 
                FROM napbox_slots 
                WHERE contract_number = %s AND id != %s
            """
            existing_slot = execute_query(check_query, (contract_number, slot_id), fetch_one=True)
            
            if existing_slot:
                return jsonify({
                    "error": f"Contract number '{contract_number}' is already used in Slot #{existing_slot['slot_number']} (Customer: {existing_slot.get('customer_name', 'Unknown')})"
                }), 400
        
        # DETERMINE FINAL STATUS
        if status is not None:
            final_status = status
        else:
            final_status = 'occupied' if customer_name else 'available'
        
        update_query = """
            UPDATE napbox_slots 
            SET customer_name = %s,
                contract_number = %s,
                customer_phone = %s,
                status = %s,
                updated_at = NOW()
            WHERE id = %s
        """
        execute_query(update_query, (customer_name, contract_number, customer_phone, final_status, slot_id))
        
        print(f" Slot {slot_id} updated:")
        print(f"   Customer: {customer_name or '(empty)'}")
        print(f"   Contract: {contract_number or '(empty)'}")
        print(f"   Status: {final_status.upper()}")
        
        return jsonify({
            "success": True,
            "message": f"Slot updated successfully. Status: {final_status.upper()}",
            "status": final_status,
            "preserved_data": {
                "previous_customer": current_slot.get('customer_name'),
                "previous_contract": current_slot.get('contract_number')
            }
        })
        
    except Exception as e:
        print(f"Error updating slot: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500



@app.route('/api/check-contract-number-exists', methods=['POST'])
def check_contract_number_exists():
    """Check if contract number already exists in napbox_slots table"""
    try:
        data = request.json
        contract_number = data.get('contract_number')
        exclude_slot_id = data.get('exclude_slot_id')
        
        if not contract_number:
            return jsonify({'exists': False, 'error': 'Contract number is required'}), 400
        
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        if exclude_slot_id:
            query = """
                SELECT id, slot_number, napbox_id, customer_name 
                FROM napbox_slots 
                WHERE contract_number = %s AND id != %s
            """
            cursor.execute(query, (contract_number, exclude_slot_id))
        else:
            query = """
                SELECT id, slot_number, napbox_id, customer_name 
                FROM napbox_slots 
                WHERE contract_number = %s
            """
            cursor.execute(query, (contract_number,))
        
        result = cursor.fetchone()
        cursor.close()
        db.close()
        
        if result:
            return jsonify({
                'exists': True,
                'slot_id': result['id'],
                'slot_number': result['slot_number'],
                'napbox_id': result['napbox_id'],
                'customer_name': result.get('customer_name', 'Unknown')
            })
        else:
            return jsonify({'exists': False})
            
    except Exception as e:
        print(f" Error checking contract number: {e}")
        return jsonify({'exists': False, 'error': str(e)}), 500


# ===============================
# CLEAR SLOT DATA (TECHNICIAN)
# ===============================
@app.route("/api/technician/clear-slot", methods=["POST"])
def clear_slot_data():
    """Clear all customer data from a slot (make it fully available)"""
    try:
        # KUNIN ANG DATA MULA SA REQUEST
        data = request.get_json()
        if not data:
            return jsonify({"error": "Invalid request data", "success": False}), 400
        
        # KUNIN ANG TAB ID MULA SA REQUEST
        tab_id = data.get("tab_id")
        
        # KUNIN ANG TECHNICIAN ID
        technician_id = None
        
        # Paraan 1: Kung may tab_id, gamitin ito para makuha ang session
        if tab_id:
            tech_session = session.get(f"technician_{tab_id}")
            if tech_session:
                technician_id = tech_session.get("technician_id")
                print(f" Found technician via tab_id: {technician_id}")
            else:
                # Subukan kunin mula sa session storage
                tech_session = session.get("technician_session")
                if tech_session:
                    technician_id = tech_session.get("technician_id")
                    print(f" Found technician via technician_session: {technician_id}")
        
        # Paraan 2: Direktang kunin mula sa session
        if not technician_id:
            technician_id = session.get("technician_id")
            if technician_id:
                print(f" Found technician via session['technician_id']: {technician_id}")
        
        # Paraan 3: Kunin mula sa request body (fallback)
        if not technician_id:
            technician_id = data.get("technician_id")
            if technician_id:
                print(f" Found technician via request body: {technician_id}")
        
        # Kung wala pa rin, mag-error
        if not technician_id:
            print(" No technician_id found in session or request")
            return jsonify({
                "error": "Invalid session. Please login again.",
                "success": False
            }), 401
        
        # KUNIN ANG SLOT ID
        slot_id = data.get("slot_id")
        
        if not slot_id:
            return jsonify({
                "error": "Slot ID is required",
                "success": False
            }), 400
        
        # I-VERIFY NA EXIST ANG SLOT
        verify_query = """
            SELECT ns.id, ns.napbox_id, ns.status, ns.customer_name, 
                   nb.area, nb.barangay
            FROM napbox_slots ns
            JOIN napboxes nb ON ns.napbox_id = nb.id
            WHERE ns.id = %s
        """
        slot_data = execute_query(verify_query, (slot_id,), fetch_one=True)
        
        if not slot_data:
            return jsonify({
                "error": "Slot not found",
                "success": False
            }), 404
        
        # I-CHECK KUNG AVAILABLE ANG SLOT
        if slot_data.get('status') != 'available':
            return jsonify({
                "error": "Slot is not available. Only available slots can be cleared.",
                "success": False
            }), 400
        
        # I-CLEAR ANG CUSTOMER DATA
        clear_query = """
            UPDATE napbox_slots 
            SET 
                customer_name = NULL,
                customer_phone = NULL,
                application_number = NULL,
                contract_number = NULL,
                installation_date = NULL,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """
        execute_query(clear_query, (slot_id,))
        
        print(f" Slot {slot_id} cleared by technician {technician_id}")
        
        return jsonify({
            "success": True,
            "message": f"Slot #{slot_id} cleared successfully"
        })
        
    except Exception as e:
        print(f" Error clearing slot: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "error": str(e),
            "success": False
        }), 500



import base64
import os
import time
import hmac
import hashlib
import struct

def generate_ga_secret(length=20):
    """Generate a Google Authenticator secret key"""
    return base64.b32encode(os.urandom(length)).decode("utf-8").rstrip("=")

def generate_ga_provisioning_uri(username, secret):
    """Generate provisioning URI for QR code"""
    from urllib.parse import quote
    safe_username = quote(str(username), safe="")
    return f"otpauth://totp/Cablevision:{safe_username}?secret={secret}&issuer=Cablevision"

def verify_ga_code(secret, code):
    """
    Verify Google Authenticator code with proper time sync handling
    """
    if not secret or not code:
        print(f" GA verify: Missing secret or code - secret: {bool(secret)}, code: {bool(code)}")
        return False

    try:
        code = ''.join(filter(str.isdigit, str(code)))
        if len(code) != 6:
            print(f" GA verify: Invalid code length - {len(code)}")
            return False

        secret = str(secret).strip().replace(" ", "").upper()
        if not secret:
            return False

        missing_padding = (-len(secret)) % 8
        if missing_padding:
            secret += "=" * missing_padding

        decoded_key = base64.b32decode(secret.encode("utf-8"), casefold=True)
        current_timestamp = int(time.time()) // 30

        for offset in range(-3, 4):
            timestamp = current_timestamp + offset
            msg = struct.pack(">Q", timestamp)
            hmac_obj = hmac.new(decoded_key, msg, hashlib.sha1)
            digest = hmac_obj.digest()
            digest_offset = digest[-1] & 0x0F
            binary_code = struct.unpack(">I", digest[digest_offset:digest_offset + 4])[0] & 0x7FFFFFFF
            expected = str(binary_code % 1000000).zfill(6)

            if expected == code:
                return True

        return False

    except Exception as e:
        print(f" GA verification error: {e}")
        import traceback
        traceback.print_exc()
        return False


def ensure_ga_columns(table_name):
    try:
        columns = execute_query(f"SHOW COLUMNS FROM {table_name}", fetch=True) or []
        existing_fields = {col.get("Field") for col in columns if col.get("Field")}

        if "ga_secret" not in existing_fields:
            execute_query(f"ALTER TABLE {table_name} ADD COLUMN ga_secret VARCHAR(64) NULL")
            print(f" Added ga_secret column to {table_name}")

        if "ga_enabled" not in existing_fields:
            execute_query(f"ALTER TABLE {table_name} ADD COLUMN ga_enabled TINYINT(1) NOT NULL DEFAULT 0")
            print(f" Added ga_enabled column to {table_name}")

    except Exception as e:
        print(f"Could not ensure Google Auth columns for {table_name}: {e}")


# Call this during app startup
for table_name in ["superadmins", "admins", "technicians"]:
    ensure_ga_columns(table_name)    

@app.route("/superadmin/ga/enable", methods=["POST"])
def superadmin_enable_google_auth():
    """Enable Google Authenticator for superadmin"""
    print("=" * 60)
    print(" GA ENABLE - START")
    print("=" * 60)
    
    tab_id = request.args.get('tab_id') or request.form.get('tab_id')
    if not tab_id:
        tab_id = session.get('active_tab')
    
    session_data = session.get(f"admin_{tab_id}")
    
    if not session_data or session_data.get('user_type') != 'superadmin':
        flash("Unauthorized access.", "danger")
        return redirect(url_for("superadmin_profile"))
    
    username = session_data.get('username')
    
    if not username:
        flash("Not logged in.", "danger")
        return redirect(url_for("login"))

    code = request.form.get("ga_code", "").strip()
    
    # Get secret from database
    user_row = execute_query(
        "SELECT ga_secret, ga_enabled FROM superadmins WHERE username = %s LIMIT 1",
        (username,),
        fetch_one=True,
    )
    
    if not user_row:
        flash(f"User '{username}' not found.", "danger")
        return redirect(url_for("superadmin_profile"))
    
    secret = user_row.get("ga_secret")
    ga_already_enabled = user_row.get("ga_enabled") == 1

    if ga_already_enabled:
        flash("Google Authenticator is already enabled.", "info")
        return redirect(url_for("superadmin_profile"))

    if not secret:
        secret = generate_ga_secret()
        rows = execute_query(
            "UPDATE superadmins SET ga_secret = %s WHERE username = %s", 
            (secret, username)
        )
        print(f" Generated new secret, rows affected: {rows}")

    if not code:
        flash("Please enter the 6-digit code from Google Authenticator.", "danger")
        return redirect(url_for("superadmin_profile", toast="ga-missing"))

    is_valid = verify_ga_code(secret, code)
    print(f" Verification result: {is_valid}")
    
    if is_valid:
        # I-STORE ANG RESULT NG UPDATE
        rows_affected = execute_query(
            "UPDATE superadmins SET ga_secret = %s, ga_enabled = 1 WHERE username = %s", 
            (secret, username)
        )
        print(f" Rows affected: {rows_affected}")
        
        # I-VERIFY KUNG NA-UPDATE
        verify = execute_query(
            "SELECT ga_enabled FROM superadmins WHERE username = %s",
            (username,),
            fetch_one=True
        )
        print(f" Verified ga_enabled: {verify.get('ga_enabled') if verify else 'N/A'}")
        
        if verify and verify.get('ga_enabled') == 1:
            session_data['ga_enabled'] = True
            session[f"admin_{tab_id}"] = session_data
            flash(" Google Authenticator is now enabled!", "success")
            print(f" GA enabled for {username}")
            return redirect(url_for("superadmin_profile", toast="ga-enabled"))
        else:
            flash(" Failed to save GA settings. Please try again.", "danger")
            print(f" GA NOT saved to database!")
            return redirect(url_for("superadmin_profile", toast="ga-invalid"))
        
    else:
        flash(" Invalid code. Please try again.", "danger")
        return redirect(url_for("superadmin_profile", toast="ga-invalid"))

@app.route("/superadmin/ga/disable", methods=["POST"])
def superadmin_disable_google_auth():
    """Disable Google Authenticator for superadmin"""
    tab_id = request.args.get('tab_id') or request.form.get('tab_id') or session.get('active_tab')
    
    session_data = session.get(f"admin_{tab_id}")
    
    if not session_data or session_data.get('user_type') != 'superadmin':
        flash("Unauthorized access.", "danger")
        return redirect(url_for("superadmin_profile"))
    
    username = session_data.get('username')
    
    if not username:
        flash("Not logged in.", "danger")
        return redirect(url_for("login"))

    execute_query(
        "UPDATE superadmins SET ga_secret = NULL, ga_enabled = 0 WHERE username = %s", 
        (username,)
    )
    
    # I-UPDATE ANG SESSION
    session_data['ga_enabled'] = False
    session[f"admin_{tab_id}"] = session_data
    
    flash("Google Authenticator has been disabled.", "info")
    print(f" GA disabled for superadmin {username}")
    
    return redirect(url_for("superadmin_profile", toast="ga-disabled"))

# ===============================
# Run Flask App
# ===============================
if __name__ == "__main__":
    import os

    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)